import logging
import re

import pandas as pd

from logic.gefen_processor import normalize_amount

logger = logging.getLogger(__name__)

# PaySchool keeps a superseded "חשבונית עסקה" number in parentheses next to the
# new tax-invoice number, e.g. "27203 (3646)". The Gefen "דיווח ביצוע" report still
# carries the original (parenthesised) number, so reconciliation must key on that.
# Strict: something before, then "(<digits>)" at the very end of the string.
_PAYSCOOL_PAREN_INVOICE_RE = re.compile(r"^(.+?)\s*\((\d+)\)\s*$")


def canonical_payscool_invoice(value) -> str:
    """Return the parenthesised original invoice number when the PaySchool
    "<new> (<original>)" pattern is present; otherwise the value unchanged."""
    if value is None:
        return ""
    s = str(value).strip()
    m = _PAYSCOOL_PAREN_INVOICE_RE.match(s)
    return m.group(2) if m else s


def load_payscool(filepath: str) -> tuple[pd.DataFrame, int]:
    import openpyxl
    wb = openpyxl.load_workbook(filepath, read_only=True)
    sheet_name = None
    for s in wb.sheetnames:
        ws = wb[s]
        rows = list(ws.iter_rows(min_row=4, max_row=4, values_only=True))
        if rows and "סעיף" in {str(v).strip() for v in rows[0] if v is not None}:
            sheet_name = s
            break
    wb.close()
    df = pd.read_excel(filepath, sheet_name=sheet_name or 0, header=None)
    df.columns = df.iloc[3]
    df = df.iloc[4:].reset_index(drop=True)

    df["report_code"] = df["סעיף"].apply(_extract_report_code)
    df = df[df["report_code"].notna()].copy()

    cancelled_count = int((df["סטטוס חשבונית"] == "מבוטלת").sum())
    df = df[df["סטטוס חשבונית"] != "מבוטלת"].copy()

    df["amount"] = df['סה"כ לסעיף'].apply(normalize_amount)
    canon_invoice = df["מספר חשבונית"].apply(canonical_payscool_invoice)
    paren_count = int((canon_invoice != df["מספר חשבונית"].apply(
        lambda v: "" if v is None else str(v).strip()
    )).sum())
    if paren_count:
        logger.info(
            "payscool: normalized %d invoice(s) of the 'NEW (ORIG)' form for reconciliation",
            paren_count,
        )
    df["ichud"] = (
        df["ח.פ"].apply(normalize_amount)
        + "-"
        + canon_invoice.apply(normalize_amount)
        + "-"
        + df["report_code"].astype(str)
        + "-"
        + df["amount"]
    )
    return df, cancelled_count


def _extract_report_code(value) -> str | None:
    match = re.search(r"\((\d+)\)", str(value))
    return match.group(1) if match else None
