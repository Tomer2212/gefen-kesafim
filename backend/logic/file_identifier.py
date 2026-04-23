from pathlib import Path

import openpyxl


def identify_file(filename: str) -> str:
    """Return 'gefen', 'kesafim2000', 'payscool', or 'unknown'."""
    ext = Path(filename).suffix.lower()

    if ext == ".xls":
        return _identify_xls(filename)

    if ext == ".xlsx":
        return _identify_xlsx(filename)

    return "unknown"


def _identify_xls(filename: str) -> str:
    """XLS files are either Kesafim2000 (TSV disguised as XLS) or unknown."""
    try:
        with open(filename, "r", encoding="iso-8859-8") as f:
            first_line = f.readline().rstrip("\r\n")
        parts = first_line.split("\t")
        a1 = parts[0].strip() if len(parts) > 0 else ""
        d1 = parts[3].strip() if len(parts) > 3 else ""
        if a1 == "קוד גפן" and d1 == "סוג תקציב":
            return "kesafim2000"
    except Exception:
        pass
    return "unknown"


def _identify_xlsx(filename: str) -> str:
    """Identify XLSX files by sheet name + cell content."""
    try:
        wb = openpyxl.load_workbook(filename, read_only=True)
    except Exception:
        return "unknown"

    sheets = wb.sheetnames

    # --- Gefen ---
    if "דיווח ביצוע" in sheets:
        ws = wb["דיווח ביצוע"]
        row1 = _get_row(ws, 1)
        if (
            _cell(row1, 0) == "מסלול רכישה"
            and _cell(row1, 2) == "סוג מענה"
            and _cell(row1, 3) == "שם מענה"
            and _cell(row1, 13) == "האם קיים קובץ"
        ):
            wb.close()
            return "gefen"

    # --- PaySchool ---
    if "Data" in sheets:
        ws = wb["Data"]
        row4 = _get_row(ws, 4)
        if (
            _cell(row4, 0) == "סעיף"
            and _cell(row4, 4) == "תאריך חשבונית"
            and _cell(row4, 9) == "סוג חשבונית"
        ):
            wb.close()
            return "payscool"

    # --- Kesafim2000 saved as XLSX ---
    for sheet_name in sheets:
        ws = wb[sheet_name]
        row1 = _get_row(ws, 1)
        if _cell(row1, 0) == "קוד גפן" and _cell(row1, 3) == "סוג תקציב":
            wb.close()
            return "kesafim2000"

    wb.close()
    return "unknown"


def _get_row(ws, row_num: int) -> list:
    rows = list(ws.iter_rows(min_row=row_num, max_row=row_num, values_only=True))
    return list(rows[0]) if rows else []


def _cell(row: list, index: int) -> str:
    if index < len(row) and row[index] is not None:
        return str(row[index]).strip()
    return ""
