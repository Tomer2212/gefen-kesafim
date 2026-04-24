from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

orange_fill = PatternFill("solid", fgColor="FFA500")
blue_fill = PatternFill("solid", fgColor="4472C4")
red_fill = PatternFill("solid", fgColor="FF0000")
green_fill = PatternFill("solid", fgColor="70AD47")
header_font = Font(bold=True, color="FFFFFF", name="Arial")


def export(
    df_gefen: pd.DataFrame,
    df_finance: pd.DataFrame,
    in_finance_not_gefen: pd.DataFrame,
    in_gefen_not_finance: pd.DataFrame,
    output_path: str,
    finance_label: str = "כספים",
) -> str:
    wb = Workbook()
    wb.remove(wb.active)

    _add_result_sheet(wb, f"קיים ב{finance_label} אך לא בגפן", in_finance_not_gefen, red_fill)
    _add_result_sheet(wb, f"משויך בגפן אך לא ב{finance_label}", in_gefen_not_finance, red_fill)

    wb.save(output_path)
    return output_path


def _add_sheet(wb: Workbook, title: str, df: pd.DataFrame, fill: PatternFill) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True
    _write_df(ws, df, fill)


def _add_result_sheet(wb: Workbook, title: str, df: pd.DataFrame, fill: PatternFill) -> None:
    ws = wb.create_sheet(title=title)
    ws.sheet_view.rightToLeft = True

    if df.empty:
        ws.cell(row=1, column=1, value="✓ אין פערים").fill = green_fill
        ws.cell(row=1, column=1).font = header_font
    else:
        _write_df(ws, df, fill)


def _write_df(ws, df: pd.DataFrame, fill: PatternFill) -> None:
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = fill
        cell.font = header_font

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
