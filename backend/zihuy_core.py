import re
from openpyxl import load_workbook
import openpyxl

CODES_JSON = {
    'יסודי': [43,44,45,46,47,51,52,53,56,57,60,63,64,65,67,68,69,70,71,72,73,74,75,77,78,80,81,83,84,85,88,89,90,126,128,129,130,131,132,133,134,147,151,153,155,157,159,161,163,166,168],
    'חטיבת ביניים': [43,44,45,46,47,49,50,51,52,53,56,57,60,63,64,65,67,68,69,70,71,72,73,74,75,77,78,80,81,83,84,85,88,89,90,126,128,129,130,131,132,133,134,135,147,151,153,155,157,159,161,163,166,168],
    'תיכון': [48,54,55,58,59,61,62,66,76,87,91,92,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,127,136,137,138,139,140,141,142,148,150,152,154,156,157,158,159,160,162,163,164,165,167,169]
}

STAGE_FROM_PLAN = {
    'עליונה בלבד': 'תיכון',
    'חט"ב בלבד': 'חטיבת ביניים',
    'יסודי בלבד': 'יסודי',
}

UNIQUE_TYPE = {
    'מייקרוסופט': 'תקציב גפ"ן',
    'תוכנה לניהול פדגוגי': 'תקציב גפ"ן',
    'תוכנה לניהול מערכת שעות': 'תקציב גפ"ן',
    "לימודי עזרה ראשונה לכיתות י'": 'תקציב גפ"ן',
    'שעות בודדות תקציב סל דוקאטי': 'תקציב דוקאטי',
}


BUDGET_NAME_MAP = [
    (['חירום מחוזי', 'גפן חירום', 'חירום'], 'גפן חירום'),
    (['גפ\"ן', 'גפן'], 'גפן'),
    (['תנופה לצפון', 'תנופה'], 'תנופה'),
    (['תקומה'], 'תקומה'),
    (['דוקאטי', 'סל דוקאטי'], 'דוקאטי'),
    (['חינוך לסובלנות'], 'חינוך לסובלנות'),
    (['קולות קוראים', 'קול קורא'], 'קולות קוראים'),
    (['פל\"ג', 'פלג'], 'פל"ג'),
]

def _norm_key_amount(val):
    if not val:
        return ''
    s = str(val).replace(',', '').strip()
    try:
        f = float(s)
        return str(int(f)) if f == int(f) else f'{f:.2f}'.rstrip('0').rstrip('.')
    except Exception:
        return s

def normalize_budget_name(raw_name):
    if not raw_name: return raw_name
    name = str(raw_name).strip()
    for keys, normalized in BUDGET_NAME_MAP:
        for key in keys:
            if key in name:
                return normalized
    return name

def get_stage_from_json(code):
    try: c = int(code)
    except: return None
    in_junior = c in CODES_JSON['חטיבת ביניים'] or c in CODES_JSON['יסודי']
    in_senior = c in CODES_JSON['תיכון']
    if in_senior and not in_junior: return 'תיכון'
    if in_junior and not in_senior: return 'חטיבת ביניים'
    return None

def get_stage_junior_ambiguous(code):
    try: c = int(code)
    except: return False
    return c in CODES_JSON['יסודי'] and c in CODES_JSON['חטיבת ביניים']

def resolve_stage(code, plan_stage_val):
    json_stage = get_stage_from_json(code)
    if json_stage == 'תיכון': return 'תיכון'
    if json_stage is None: return None
    if get_stage_junior_ambiguous(code) and plan_stage_val:
        plan_resolved = STAGE_FROM_PLAN.get(str(plan_stage_val).strip())
        if plan_resolved: return plan_resolved
    return json_stage

def extract_plan_num(text):
    if not text: return None
    m = re.match(r'^\s*(\d+)\s*-', str(text))
    return m.group(1) if m else None

def extract_supplier_num(text):
    if not text: return None
    m = re.match(r'^\s*(\d+)\s*-', str(text))
    return m.group(1) if m else None

def load_plan_file(fpath):
    wb = load_workbook(fpath, read_only=True)
    ws_hkl = wb['הכל']
    hkl_rows = list(ws_hkl.iter_rows(values_only=True))
    budgets = []
    for row in hkl_rows[1:]:
        if row[0] and row[0] != "סה''כ תקציבים במערכת גפן":
            if row[0] not in budgets: budgets.append(row[0])
            else: break

    ws_plan = wb['פירוט המענים']
    plan_rows = list(ws_plan.iter_rows(values_only=True))

    stage_values = set(str(row[4]).strip() for row in plan_rows[1:] if row[4])
    detected_level = None
    if 'עליונה בלבד' in stage_values and 'חט"ב בלבד' not in stage_values and 'יסודי בלבד' not in stage_values:
        detected_level = 'עליונה'
    elif 'חט"ב בלבד' in stage_values and 'עליונה בלבד' not in stage_values:
        detected_level = 'ביניים'
    elif 'יסודי בלבד' in stage_values and 'עליונה בלבד' not in stage_values:
        detected_level = 'יסודי'

    plan_num_to_budgets = {}
    code_to_budgets = {}
    plan_num_to_plan_stage = {}
    code_to_plan_stage = {}
    type_name_to_budget = {}

    for row in plan_rows[1:]:
        budget = row[0]; plan_num = row[9]
        code = str(row[17]) if row[17] is not None else None
        plan_stage_val = row[4]
        type_name = row[7]

        if plan_num is not None:
            key = str(int(plan_num)) if isinstance(plan_num, float) else str(plan_num)
            if key not in plan_num_to_budgets: plan_num_to_budgets[key] = set()
            plan_num_to_budgets[key].add(budget)
            if plan_stage_val and key not in plan_num_to_plan_stage:
                plan_num_to_plan_stage[key] = plan_stage_val

        if code:
            if code not in code_to_budgets: code_to_budgets[code] = set()
            code_to_budgets[code].add(budget)
            if plan_stage_val and code not in code_to_plan_stage:
                code_to_plan_stage[code] = plan_stage_val

        if type_name and budget:
            type_name_to_budget[str(type_name).strip()] = budget

    return budgets, plan_num_to_budgets, code_to_budgets, plan_num_to_plan_stage, code_to_plan_stage, detected_level, type_name_to_budget

def try_identify_row(r, results, plan_num_to_budgets, code_to_budgets,
                     plan_num_to_plan_stage, code_to_plan_stage, type_name_to_budget):
    col_code = r['orig'][1]
    col_type = r['orig'][2]
    plan_num = r['plan_num']
    cur_idx = r['_idx']

    budget_result = None
    stage_result = None

    # שלב 1 - מספר מענה
    if plan_num:
        bf = plan_num_to_budgets.get(plan_num, set())
        psv = plan_num_to_plan_stage.get(plan_num)
        if len(bf) == 1:
            budget_result = list(bf)[0]
            stage_result = resolve_stage(col_code, psv)
        elif len(bf) > 1:
            last_budget_for_pn = None; last_idx_for_pn = None
            for idx, prev_r in enumerate(results[:cur_idx]):
                if prev_r['plan_num'] == plan_num and prev_r['budget']:
                    last_budget_for_pn = prev_r['budget']; last_idx_for_pn = idx
            if last_budget_for_pn is not None:
                diff = any(results[idx]['budget'] and results[idx]['budget'] != last_budget_for_pn
                           for idx in range(last_idx_for_pn + 1, cur_idx))
                if diff:
                    remaining = bf - {last_budget_for_pn}
                    if len(remaining) == 1:
                        budget_result = list(remaining)[0]
                        stage_result = resolve_stage(col_code, psv)
                else:
                    has_certain_after = any(
                        results[k]['budget'] and results[k]['stage']
                        for k in range(r['_idx'] + 1, len(results))
                    )
                    if has_certain_after:
                        budget_result = last_budget_for_pn
                        stage_result = resolve_stage(col_code, psv)

    # שלב 2 - סוג מענה
    if budget_result is None and col_type in UNIQUE_TYPE:
        budget_result = UNIQUE_TYPE[col_type]
        psv = code_to_plan_stage.get(col_code)
        stage_result = resolve_stage(col_code, psv)

    # שלב 3 - קוד דיווח
    if budget_result is None and col_code:
        bf = code_to_budgets.get(col_code, set())
        psv = code_to_plan_stage.get(col_code)
        if len(bf) == 1:
            budget_result = list(bf)[0]
            stage_result = resolve_stage(col_code, psv)
        elif len(bf) > 1:
            last_budget_for_code = None; last_idx_for_code = None
            for idx, prev_r in enumerate(results[:cur_idx]):
                if prev_r['orig'][1] == col_code and prev_r['budget']:
                    last_budget_for_code = prev_r['budget']; last_idx_for_code = idx
            if last_budget_for_code is not None:
                diff = any(results[idx]['budget'] and results[idx]['budget'] != last_budget_for_code
                           for idx in range(last_idx_for_code + 1, cur_idx))
                if diff:
                    remaining = bf - {last_budget_for_code}
                    if len(remaining) == 1:
                        budget_result = list(remaining)[0]
                        stage_result = resolve_stage(col_code, psv)
                else:
                    has_certain_after = any(
                        results[k]['budget'] and results[k]['stage']
                        for k in range(r['_idx'] + 1, len(results))
                    )
                    if has_certain_after:
                        budget_result = last_budget_for_code
                        stage_result = resolve_stage(col_code, psv)

    if stage_result is None and col_code:
        psv = code_to_plan_stage.get(col_code)
        stage_result = resolve_stage(col_code, psv)

    return budget_result, stage_result

def identify(doch_paths, plan_fpaths):
    all_budgets = []
    plan_num_to_budgets = {}; code_to_budgets = {}
    plan_num_to_plan_stage = {}; code_to_plan_stage = {}
    type_name_to_budget = {}

    for fpath in plan_fpaths:
        budgets, pn_map, c_map, pn_stage, c_stage, level, tn_map = load_plan_file(fpath)
        for b in budgets:
            if b not in all_budgets: all_budgets.append(b)
        for k, v in pn_map.items():
            if k not in plan_num_to_budgets: plan_num_to_budgets[k] = set()
            plan_num_to_budgets[k] |= v
        for k, v in c_map.items():
            if k not in code_to_budgets: code_to_budgets[k] = set()
            code_to_budgets[k] |= v
        for k, v in pn_stage.items():
            if k not in plan_num_to_plan_stage: plan_num_to_plan_stage[k] = v
        for k, v in c_stage.items():
            if k not in code_to_plan_stage: code_to_plan_stage[k] = v
        for k, v in tn_map.items():
            if k not in type_name_to_budget: type_name_to_budget[k] = v

    possible_combos = set()
    for code, bfs in code_to_budgets.items():
        psv = code_to_plan_stage.get(code)
        stage = resolve_stage(code, psv)
        if stage:
            for b in bfs: possible_combos.add((b, stage))
    dominant_stage = None
    for code, psv in code_to_plan_stage.items():
        s = resolve_stage(code, psv)
        if s == 'תיכון': dominant_stage = 'תיכון'; break
        elif s == 'חטיבת ביניים': dominant_stage = 'חטיבת ביניים'
        elif s == 'יסודי' and dominant_stage is None: dominant_stage = 'יסודי'
    if dominant_stage:
        for b in all_budgets: possible_combos.add((b, dominant_stage))

    if isinstance(doch_paths, str):
        doch_paths = [doch_paths]

    seen_rows = set()
    header_row = None
    unique_rows = []
    for doch_path in doch_paths:
        wb_doch = load_workbook(doch_path, read_only=True)
        ws_doch = wb_doch['דיווח ביצוע']
        rows = list(ws_doch.iter_rows(values_only=True))
        if rows and header_row is None:
            header_row = rows[0]
        for row in rows[1:]:
            key = tuple(row)
            if key not in seen_rows:
                seen_rows.add(key)
                unique_rows.append(row)

    doch_rows_all = ([header_row] if header_row else []) + unique_rows

    plan_stages = set()
    for code in code_to_budgets:
        psv = code_to_plan_stage.get(code)
        s = resolve_stage(code, psv)
        if s: plan_stages.add(s)

    doch_stages = set()
    for row in doch_rows_all[1:]:
        code = str(row[1]) if row[1] is not None else ''
        s = get_stage_from_json(code)
        if s: doch_stages.add(s)

    warnings = []
    extra_stages = doch_stages - plan_stages
    if extra_stages:
        warnings.append(f'קובץ הדיווח מכיל קודים של {extra_stages} שאינם בקובץ התכנון שהועלה.')

    results = []
    for i, row in enumerate(doch_rows_all[1:], start=2):
        name = str(row[3]) if row[3] is not None else ''
        if name.startswith('קול קורא חינוך לסובלנות פתוח עד'):
            continue
        col_a = str(row[0]) if row[0] is not None else ''
        col_code = str(row[1]) if row[1] is not None else ''
        col_type = str(row[2]) if row[2] is not None else ''
        col_name = row[3]; col_inv = row[4]; col_date = row[5]; col_sup = row[6]
        col_mah = row[7]; col_item = row[8]; col_qty = row[9]
        col_desc = row[10]; col_amt = row[11]; col_stat = row[12]; col_file = row[13]
        plan_num = extract_plan_num(str(col_name)) if col_name else None
        sup_num = extract_supplier_num(str(col_sup)) if col_sup else None
        union_key = f'{sup_num}-{col_inv}-{col_code}-{_norm_key_amount(col_amt)}' if sup_num else f'_unresolved_{len(results)}'
        results.append({
            'row': i, 'budget': None, 'stage': None,
            'col_a': col_a, 'col_type': col_type,
            'orig': (col_a, col_code, col_type, col_name, col_inv, col_date, col_sup,
                     col_mah, col_item, col_qty, col_desc, col_amt, col_stat, col_file),
            'plan_num': plan_num, 'sup_num': sup_num, 'union_key': union_key,
            '_idx': len(results)
        })

    for i, r in enumerate(results):
        r['_idx'] = i

    def apply_microsoft_rule(results):
        n = len(results); i = 0
        while i < n:
            if results[i]['col_a'].startswith('מענה משרדי'):
                start = i; end = i
                while end + 1 < n and results[end + 1]['col_a'].startswith('מענה משרדי'): end += 1
                if any(results[j]['col_type'] == 'מייקרוסופט' for j in range(start, end + 1)):
                    last_certain = None
                    for j in range(start - 1, -1, -1):
                        if results[j]['budget'] and results[j]['stage']:
                            last_certain = (results[j]['budget'], results[j]['stage']); break
                    if last_certain:
                        for j in range(start, end + 1):
                            if not results[j]['budget']:
                                results[j]['budget'] = last_certain[0]
                                results[j]['stage'] = last_certain[1]
                i = end + 1
            else: i += 1

    max_iterations = 20
    for iteration in range(max_iterations):
        changed = False

        for r in results:
            if not r['budget'] or not r['stage']:
                b, s = try_identify_row(r, results, plan_num_to_budgets, code_to_budgets,
                                        plan_num_to_plan_stage, code_to_plan_stage, type_name_to_budget)
                if b:
                    r['budget'] = b; r['stage'] = s; changed = True

        prev_known = sum(1 for r in results if r['budget'])
        apply_microsoft_rule(results)
        if sum(1 for r in results if r['budget']) > prev_known: changed = True

        n = len(results)
        for i in range(n):
            if not results[i]['budget'] or not results[i]['stage']:
                prev = nxt = None
                for j in range(i - 1, -1, -1):
                    if results[j]['budget'] and results[j]['stage']:
                        prev = (results[j]['budget'], results[j]['stage']); break
                for j in range(i + 1, n):
                    if results[j]['budget'] and results[j]['stage']:
                        nxt = (results[j]['budget'], results[j]['stage']); break
                if prev is not None and nxt is not None and prev == nxt:
                    results[i]['budget'] = prev[0]; results[i]['stage'] = prev[1]; changed = True

        if not changed: break

    n = len(results)

    def get_closed_combos_for_row(row_idx):
        last_seen = {}
        for i in range(row_idx):
            if results[i]['budget'] and results[i]['stage']:
                combo = (results[i]['budget'], results[i]['stage'])
                last_seen[combo] = i

        closed = set()
        for combo, last_idx in last_seen.items():
            for k in range(last_idx + 1, row_idx):
                if results[k]['budget'] and results[k]['stage']:
                    other_combo = (results[k]['budget'], results[k]['stage'])
                    if other_combo != combo:
                        closed.add(combo)
                        break
        return closed

    for edge_start, edge_end, direction in [(n - 1, -1, -1), (0, n, 1)]:
        uncertain_edge = []
        for i in range(edge_start, edge_end, direction):
            if not results[i]['budget'] or not results[i]['stage']: uncertain_edge.append(i)
            else: break
        if not uncertain_edge: continue
        for i in uncertain_edge:
            closed_for_row = get_closed_combos_for_row(i)
            open_combos = possible_combos - closed_for_row
            if len(open_combos) == 1:
                combo = list(open_combos)[0]
                psv = code_to_plan_stage.get(results[i]['orig'][1])
                row_stage = resolve_stage(results[i]['orig'][1], psv)
                if row_stage and row_stage != combo[1]: continue
                results[i]['budget'] = combo[0]; results[i]['stage'] = combo[1]

    groups = []
    cur_b = results[0]['budget']; cur_s = results[0]['stage']; cur_rows = [results[0]]
    for r in results[1:]:
        if r['budget'] == cur_b and r['stage'] == cur_s: cur_rows.append(r)
        else:
            groups.append({'budget': cur_b, 'stage': cur_s, 'rows': cur_rows})
            cur_b = r['budget']; cur_s = r['stage']; cur_rows = [r]
    groups.append({'budget': cur_b, 'stage': cur_s, 'rows': cur_rows})

    rows_to_remove = set()
    for i, g1 in enumerate(groups):
        for j, g2 in enumerate(groups):
            if i >= j: continue
            if g1['budget'] == g2['budget'] and g1['stage'] == g2['stage'] and len(g1['rows']) == len(g2['rows']):
                if [tuple(r['orig']) for r in g1['rows']] == [tuple(r['orig']) for r in g2['rows']]:
                    for r in g2['rows']: rows_to_remove.add(r['row'])

    results_clean = [r for r in results if r['row'] not in rows_to_remove]

    combo_sequence = []
    for r in results_clean:
        if r['budget'] and r['stage']:
            combo = (r['budget'], r['stage'])
            if not combo_sequence or combo_sequence[-1] != combo:
                combo_sequence.append(combo)
    seen_combos = set()
    for combo in combo_sequence:
        if combo in seen_combos:
            warnings.append(f'שגיאת תקינות: שילוב {combo[0]}+{combo[1]} מופיע שוב לאחר שנסגר.')
        seen_combos.add(combo)

    found_budgets = set(r['budget'] for r in results_clean if r['budget'])
    missing_budgets = [b for b in all_budgets if b not in found_budgets]

    return results_clean, warnings, missing_budgets, all_budgets

def save_output(results_clean, output_path, all_identified=False):
    wb_out = openpyxl.Workbook(); ws_out = wb_out.active
    ws_out.title = 'zihuy'; ws_out.sheet_view.rightToLeft = True
    headers = ['תקציב','שלב','איחוד','מסלול רכישה','קוד דווח','סוג מענה','מספר מענה',
               'שם מענה','מספר חשבונית','תאריך חשבונית','מספר ספק','קוד ושם ספק',
               'מהות ההוצאה','מספר פריט','כמות','תאור פריט','סכום פריט','סטטוס חשבונית','האם קיים קובץ']
    ws_out.append(headers)
    for r in results_clean:
        o = r['orig']
        normalized_budget = normalize_budget_name(r['budget']) if r['budget'] else r['budget']
        union_key = r['union_key']
        if all_identified and normalized_budget and union_key:
            union_key = f"{union_key}-{normalized_budget}"
        ws_out.append([normalized_budget, r['stage'], union_key,
                       o[0], o[1], o[2], r['plan_num'], o[3], o[4], o[5],
                       r['sup_num'], o[6], o[7], o[8], o[9], o[10], o[11], o[12], o[13]])
    wb_out.save(output_path)
