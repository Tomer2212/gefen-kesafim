"""
Standalone plan-roster extraction for the "Compare Checks" feature (מענים section).

Deliberately independent from backend/logic/tikhnun_processor.py and from
analyze_router.py::_finalize_tikhnun_metrics / _compute_multi_budget_tikhnun.
Those power the existing "דיווח חסר" flow and must not be touched. This module
duplicates only the small pieces of parsing logic needed (sheet identification,
budget list from "הכל", plan grouping from "פירוט המענים") but WITHOUT the
doch/report cross-reference and WITHOUT the hefresh-based filtering — it
returns every plan found in the planning file, regardless of report status.
"""

from openpyxl import load_workbook

from zihuy_core import normalize_budget_name


def _to_f(v):
    if v is None:
        return 0.0
    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return 0.0


def _find_tikhnun_sheets(wb):
    """Locate the "הכל" and "פירוט המענים" sheets by name, falling back to
    the same header-cell signature used by logic/file_identifier.py.
    Returns (hakol_rows, perut_rows) or None if not a recognizable planning file.
    """
    hakol_ws_name = perut_ws_name = None
    for sh in wb.sheetnames:
        sh_s = sh.strip()
        if sh_s == "הכל":
            hakol_ws_name = sh
        elif sh_s == "פירוט המענים":
            perut_ws_name = sh

    if not hakol_ws_name or not perut_ws_name:
        for sh in wb.sheetnames:
            if sh == hakol_ws_name or sh == perut_ws_name:
                continue
            peek = list(wb[sh].iter_rows(min_row=1, max_row=1, values_only=True))
            if not peek:
                continue
            hdr = peek[0]
            if not hakol_ws_name and len(hdr) > 14 and hdr[10] == "השתתפות רשות/ בעלות":
                hakol_ws_name = sh
            elif not perut_ws_name and len(hdr) > 15 and hdr[14] == "עלות מענה כוללת":
                perut_ws_name = sh

    if not hakol_ws_name or not perut_ws_name:
        return None

    hakol_rows = [list(r) for r in wb[hakol_ws_name].iter_rows(values_only=True)]
    perut_rows = [list(r) for r in wb[perut_ws_name].iter_rows(values_only=True)]
    return hakol_rows, perut_rows


def extract_budgets(hakol_rows: list) -> dict:
    """Distinct budgets from the "הכל" sheet. {norm_name: {"raw_name", "H"}}"""
    budgets: dict = {}
    seen_names: set = set()
    for row in hakol_rows[1:]:
        name = row[0] if row else None
        if not name:
            continue
        name_s = str(name).strip()
        if not name_s or "סה''כ" in name_s:
            continue
        if name_s in seen_names:
            continue
        seen_names.add(name_s)
        h_num = _to_f(row[7]) if len(row) > 7 else 0.0
        if h_num <= 0:
            continue
        norm_name = normalize_budget_name(name_s)
        budgets[norm_name] = {"raw_name": name_s, "H": h_num}
    return budgets


def _plan_key(rcode: str, mispnum: str, name: str) -> str:
    if mispnum and mispnum != "אין":
        return mispnum
    return f"{rcode}-{name}"


def extract_plan_roster(plan_fpath: str) -> dict:
    """
    Parse a planning file and return every plan ("מענה") it contains, grouped by
    normalized budget name — regardless of report/hefresh status.

    Returns: { budget_norm: [ {"key", "mispnum", "rcode", "name", "tikhnun"}, ... ] }
    """
    wb = load_workbook(plan_fpath, read_only=True)
    try:
        found = _find_tikhnun_sheets(wb)
        if not found:
            return {}
        hakol_rows, perut_rows = found
    finally:
        wb.close()

    budgets = extract_budgets(hakol_rows)
    if not budgets:
        return {}

    plan_groups: dict = {}
    for r in perut_rows[1:]:
        raw_bname = str(r[0]).strip() if r and r[0] else ""
        budget_norm = normalize_budget_name(raw_bname)
        if budget_norm not in budgets:
            continue

        gk = tuple(
            str(r[i]).strip() if (len(r) > i and r[i] is not None) else ""
            for i in [0, 1, 2, 3, 4, 6, 7, 8, 9, 14, 17]
        )
        if gk not in plan_groups:
            rcode = str(r[17]).strip() if len(r) > 17 and r[17] else ""
            name = str(r[8]).strip() if len(r) > 8 and r[8] else ""
            j_val = str(r[9]).strip() if len(r) > 9 and r[9] else ""
            mispnum = j_val if (j_val and j_val != " ") else "אין"
            plan_groups[gk] = {
                "tikhnun": 0.0,
                "rcode": rcode,
                "name": name,
                "mispnum": mispnum,
                "budget_norm": budget_norm,
                "key": _plan_key(rcode, mispnum, name),
            }
        col13 = r[13] if len(r) > 13 else None
        amount = _to_f(col13) if col13 is not None else _to_f(r[15] if len(r) > 15 else None)
        plan_groups[gk]["tikhnun"] += amount

    roster: dict = {}
    for g in plan_groups.values():
        roster.setdefault(g["budget_norm"], []).append({
            "key": g["key"],
            "mispnum": g["mispnum"],
            "rcode": g["rcode"],
            "name": g["name"],
            "tikhnun": g["tikhnun"],
        })
    return roster
