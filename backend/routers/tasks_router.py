import base64
import hmac
import html
import io
import logging
import os
import re
import secrets
import shutil
import tempfile
import time
import urllib.parse
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Annotated

from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File
from fastapi.responses import Response
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel

import booking_token_logic
import graph_client
import task_logic
import whatsapp_twilio
from routers import schools_router as _schools_router
from academic_years import DEFAULT_ACADEMIC_YEAR
from auth import get_current_user
from booking_logic import get_org_mailbox_capability, format_ranges_html, format_ranges_text
from email_resend import send_resend_email
from supabase_client import get_admin_client, reset_admin_client

logger = logging.getLogger(__name__)
router = APIRouter()

CRON_SECRET = os.getenv("CRON_SECRET", "")
APP_URL = os.getenv("APP_URL", "http://localhost:5173")


_ROLE_CONTACT_FIELDS = {
    "principal": ("principal_name", "principal_email", "principal_phone"),
    "principal_chativa": ("principal_chativa_name", "principal_chativa_email", "principal_chativa_phone"),
    "secretary": ("secretary_name", "secretary_email", "secretary_phone"),
    "finance_contact": ("finance_contact_name", "finance_contact_email", "finance_contact_phone"),
}
_PARTICIPANT_ROLE_LABELS = {
    "principal": "מנהל/ת", "principal_chativa": 'מנהל/ת חט"ב',
    "secretary": "מנהלנ/ית", "finance_contact": "אחראי/ת כספים",
}
_MEETING_TYPE_LABELS = {"gefen": "גפן", "current": "שוטף", "district": "מחוז"}
_TYPED_ADVISOR_TABLES = {"gefen": "school_advisors_gefen", "current": "school_advisors_current", "district": "school_advisors_district"}


def _principal_slots_for_school(school: dict, stage_scope: str | None) -> list[str]:
    """Round 8 — which of ["principal", "principal_chativa"] a "מנהל/ת" participant actually
    resolves to for this school, given the meeting requirement's chosen stage_scope
    ('tichon'/'chativa'/'both', same convention as the existing meetings.stage_scope feature —
    see StageScopeModal.jsx). Six-year schools (stage == "sheshshnati") with
    principal_same_person=True collapse to a single slot regardless of stage_scope — there's
    only one real person to invite either way. Non-six-year schools always just ["principal"]."""
    if school.get("stage") != "sheshshnati" or school.get("principal_same_person"):
        return ["principal"]
    scope = stage_scope or "both"
    slots = []
    # "separate" (round 16 — two independent meetings, one per principal) still needs both
    # principals' contact details resolved/checked, exactly like "both" (one merged meeting) —
    # the two scopes only differ in how _build_meeting_booking_link groups the resolved
    # participants into one range vs two, not in who needs a valid contact.
    if scope in ("tichon", "both", "separate"):
        slots.append("principal")
    if scope in ("chativa", "both", "separate"):
        slots.append("principal_chativa")
    return slots


def _principal_role_label(school: dict, slot: str) -> str:
    """The display label for a principal slot in THIS school's context — six-year schools (with
    a real second principal) get the חט"ע/חט"ב-qualified labels; everyone else just "מנהל/ת"."""
    if slot == "principal_chativa":
        return 'מנהל/ת חט"ב'
    if school.get("stage") == "sheshshnati" and not school.get("principal_same_person"):
        return 'מנהל/ת חט"ע'
    return "מנהל/ת"


def _build_school_contacts(school: dict, roles: list[str], stage_scope: str | None = None) -> list[dict]:
    """Round-5 meeting-requirement participants — role-level selection resolved to that
    school's actual named contact, reusing _ROLE_CONTACT_FIELDS (same 3 roles
    RECIPIENT_ROLE_OPTIONS/taskShared.js already use for message recipients). Mirrors
    frontend/src/components/meetings/schoolContacts.js's buildSchoolContacts, but keyed by the
    tasks-domain role names (principal/secretary/finance_contact) instead of the
    meetings-domain ones (principal/secretary/finance) to stay consistent with the rest of
    this file — deliberately does NOT include extra_contacts (the participant-role picker only
    offers the 3 fixed roles, unlike the single-school "תיאום ישיר" modal). Round 8: "principal"
    expands to 1-2 concrete slots via _principal_slots_for_school before resolution, so a
    six-year school's "מנהל/ת" participant correctly becomes the tichon and/or chativa
    principal depending on stage_scope — every other role is unaffected."""
    contacts = []
    for role in roles:
        slots = _principal_slots_for_school(school, stage_scope) if role == "principal" else [role]
        for slot in slots:
            fields = _ROLE_CONTACT_FIELDS.get(slot)
            if not fields:
                continue
            name_field, email_field, _phone_field = fields
            name = school.get(name_field)
            if name:
                contacts.append({"key": slot, "name": name, "email": school.get(email_field) or None})
    return contacts


def _require_manager(user: dict):
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")


def _resolve_recipient(school: dict, recipient_role: str) -> dict:
    """Resolves a recipient_role ('meeting_coordinator'|'principal'|'secretary'|
    'finance_contact'|'extra:<index>') against a school row into {name, email, phone} —
    mirrors schools_router._resolve_meeting_coordinator's reference-resolution convention.
    'meeting_coordinator' is a level of indirection: it follows school['meeting_coordinator']
    (itself one of the other role refs, as configured in "פרטי בית הספר") rather than naming
    a fixed field — same value schools_router surfaces as meeting_coordinator_contact."""
    if recipient_role == "meeting_coordinator":
        ref = school.get("meeting_coordinator")
        return _resolve_recipient(school, ref) if ref else {"name": None, "email": None, "phone": None}
    if recipient_role in _ROLE_CONTACT_FIELDS:
        name_f, email_f, phone_f = _ROLE_CONTACT_FIELDS[recipient_role]
        return {"name": school.get(name_f), "email": school.get(email_f), "phone": school.get(phone_f)}
    if recipient_role and recipient_role.startswith("extra:"):
        try:
            idx = int(recipient_role.split(":", 1)[1])
        except ValueError:
            return {"name": None, "email": None, "phone": None}
        extra = (school.get("extra_contacts") or [])
        if 0 <= idx < len(extra):
            c = extra[idx]
            return {"name": c.get("name"), "email": c.get("email"), "phone": c.get("phone")}
    return {"name": None, "email": None, "phone": None}


def _channel_missing_contact(channel: str, recipient: dict) -> bool:
    if channel == "whatsapp_twilio":
        return not recipient.get("phone")
    return not recipient.get("email")


# Round-2 redesign: cascading fallback order used by the pre-creation contact-resolution flow
# (TaskContactResolutionModal) when the primary/chosen recipient_role has no contact info for
# a given school — matches the priority the product owner specified.
MEETING_TASK_ROLE_PRIORITY = ["meeting_coordinator", "secretary", "principal", "finance_contact"]
GENERAL_TASK_ROLE_PRIORITY = ["principal", "secretary", "finance_contact"]


def _resolve_recipient_with_cascade(
    school: dict, primary_role: str, is_meeting_task: bool, channel: str, alternate_role: str | None = None,
) -> dict:
    """Tries alternate_role (if given) or primary_role first, then cascades through the fixed
    priority order above until a contact with the field the channel needs (email/phone) is
    found. Returns {"recipient": {name,email,phone}, "resolved_via": role|None} — resolved_via
    is None only if no contact was found anywhere in the cascade (the "zero contacts" case
    that routes to the Excel export/import flow)."""
    priority = MEETING_TASK_ROLE_PRIORITY if is_meeting_task else GENERAL_TASK_ROLE_PRIORITY
    candidates = [alternate_role or primary_role]
    for role in priority:
        if role not in candidates:
            candidates.append(role)
    tried = set()
    for role in candidates:
        if not role or role in tried:
            continue
        tried.add(role)
        recipient = _resolve_recipient(school, role)
        if not _channel_missing_contact(channel, recipient):
            return {"recipient": recipient, "resolved_via": role}
    return {"recipient": {"name": None, "email": None, "phone": None}, "resolved_via": None}


def _parse_dt(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except ValueError:
        raise HTTPException(status_code=422, detail="פורמט תאריך/שעה לא תקין")


class ConditionsIn(BaseModel):
    groups: list[dict] = []


class AttachmentKeyIn(BaseModel):
    key: str
    filename: str | None = None


class MessageConfigIn(BaseModel):
    recipient_role: str
    channel: str  # 'email_resend' | 'email_outlook' | 'whatsapp_twilio'
    subject: str | None = None
    body_template: str = ""
    attachment_keys: list[AttachmentKeyIn] = []


class TaskCreateIn(BaseModel):
    name: str
    criteria: ConditionsIn
    message_config: MessageConfigIn
    academic_year: str = DEFAULT_ACADEMIC_YEAR
    scheduled_for: str | None = None  # ISO datetime; if set (and in the future), the school
    # list is left empty and evaluated later by process-scheduled-tasks instead of now — lets
    # a manager set up a task ("קביעת פגישות רבעון 3") ahead of time, matched against whichever
    # schools meet the criteria on that future date rather than today's.
    manual_school_ids: list[str] | None = None  # bypasses criteria matching — the meeting-task
    # wizard fast-path's "pick schools manually" mode (round-2 redesign).
    success_criteria: ConditionsIn | None = None  # None = auto-derive from `criteria` when
    # track_success is on (task_logic.invert_criteria); explicit = evaluated directly instead.
    track_success: bool = True  # False = "just track sends" — no success/progress concept.
    is_meeting_task: bool = False  # cosmetic only (which wizard path created this task) — no
    # backend logic branches on it; behavior is driven purely by the fields above.
    confirm_outlook_limit: bool = False  # set after the manager confirms past the warning below
    excluded_school_ids: list[str] | None = None  # round-5 TaskMeetingResolutionModal's
    # per-row "הסרה מהמשימה" — subtracted from matched_school_ids right after matching,
    # regardless of whether the audience came from criteria or manual_school_ids.
    meeting_overrides: dict | None = None  # round-6 "קבע אחר לפגישה זו בלבד" —
    # {school_id: {meeting_service_type: {"advisor_ids":[...], "duration_minutes": int}}},
    # consulted by _build_meeting_booking_link ahead of the school's own card defaults but
    # never written back to the school itself (see _build_meeting_booking_link's docstring).


class TaskPatchIn(BaseModel):
    name: str | None = None
    status: str | None = None
    message_config: MessageConfigIn | None = None


class PreviewIn(BaseModel):
    criteria: ConditionsIn
    academic_year: str = DEFAULT_ACADEMIC_YEAR
    manual_school_ids: list[str] | None = None


class SendIn(BaseModel):
    only_school_ids: list[str] | None = None  # None = send to everyone not-yet-done
    scheduled_at: str | None = None  # ISO datetime; if set (and in the future), messages are
    # queued now but held until process-message-queue's scheduled_at filter releases them.
    confirm_outlook_limit: bool = False  # set after the manager confirms past the warning below


# Real documented Microsoft Graph/Exchange Online limits: 30 messages/minute (fixed, not
# adjustable) and 10,000 recipients/day (cumulative). This constant is a per-task *warning*
# threshold, not the hard Graph limit itself — deliberately set well below 10,000 so a manager
# is warned before a single task's bulk-send would meaningfully eat into the daily cap shared
# by every other Outlook send in the org that day. Configurable since org size varies.
OUTLOOK_SEND_WARN_THRESHOLD = int(os.getenv("OUTLOOK_SEND_WARN_THRESHOLD", "300"))

# Round 12 — real enforcement of Microsoft's actual cumulative daily cap (10,000 recipients/day,
# shared by every Outlook send in the org that day), not just the per-task warning above.
# Deliberately set below the real 10,000 so this stops sends before Microsoft itself would.
OUTLOOK_DAILY_SEND_CAP = int(os.getenv("OUTLOOK_DAILY_SEND_CAP", "9500"))

# Real enforcement of Microsoft's actual per-minute Graph sendMail cap (30/minute, fixed) —
# process_message_queue's BATCH_SIZE/cron-cadence combo only gave incidental protection (no
# guaranteed spacing between individual sends within one drained batch). 2.2s gives ~27/min
# actual throughput, a real safety margin under 30, not sitting exactly on the limit.
OUTLOOK_SEND_MIN_INTERVAL_SECONDS = float(os.getenv("OUTLOOK_SEND_MIN_INTERVAL_SECONDS", "2.2"))


def _outlook_daily_count_ok_and_increment(org_id: str) -> bool:
    """Checked/incremented against org_outlook_send_counts (org_id, send_date) right before every
    real email_outlook send — this is live DB state, not in-memory (Architecture Invariant #1),
    so it stays correct across the cron-drained queue and the immediate-send path alike."""
    today = date.today().isoformat()
    for attempt in range(2):
        try:
            db2 = get_admin_client()
            rows = (
                db2.table("org_outlook_send_counts").select("count")
                .eq("org_id", org_id).eq("send_date", today).execute().data or []
            )
            current = rows[0]["count"] if rows else 0
            if current >= OUTLOOK_DAILY_SEND_CAP:
                return False
            db2.table("org_outlook_send_counts").upsert(
                {"org_id": org_id, "send_date": today, "count": current + 1},
                on_conflict="org_id,send_date",
            ).execute()
            return True
        except Exception as exc:
            if attempt == 0:
                logger.warning("_outlook_daily_count_ok_and_increment attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("_outlook_daily_count_ok_and_increment failed after 2 attempts: %s", exc, exc_info=True)
                return True  # fail-open: never block a real send over a transient DB hiccup


@router.get("/field-options")
def get_field_options(user: Annotated[dict, Depends(get_current_user)]):
    # No manager gate — this is a static label/options catalog (field names, goal
    # planning/reporting percentages, control-letter status labels, etc.), not per-school data.
    # An advisor viewing their own field-kind person-task in אזור אישי needs this catalog too
    # (to render the "מדד הצלחה" column's title/options), so gating it manager-only left every
    # advisor's fetch silently 403ing (swallowed by the frontend's .catch), which meant they
    # never got real labels — just raw goal_key/field-key fallback text.
    return {
        **task_logic.field_options(user["org_id"]),
        "meeting_types": task_logic.MEETING_SERVICE_TYPE_OPTIONS,
    }


class TemplateIn(BaseModel):
    name: str
    subject: str | None = None
    body_template: str = ""


@router.get("/templates")
def list_templates(user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            return (
                db.table("org_task_message_templates")
                .select("*")
                .eq("org_id", user["org_id"])
                .order("created_at", desc=True)
                .execute()
                .data or []
            )
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_templates attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_templates failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/templates")
def create_template(body: TemplateIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            row = (
                db.table("org_task_message_templates")
                .insert({
                    "org_id": user["org_id"],
                    "created_by": user["id"],
                    "name": body.name,
                    "subject": body.subject,
                    "body_template": body.body_template,
                })
                .execute()
            )
            return row.data[0]
        except Exception as exc:
            if attempt == 0:
                logger.warning("create_template attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("create_template failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.delete("/templates/{template_id}")
def delete_template(template_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("org_task_message_templates").delete().eq("id", template_id).eq("org_id", user["org_id"]).execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("delete_template attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("delete_template failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


class AudienceIn(BaseModel):
    name: str
    criteria: ConditionsIn
    academic_year: str | None = None


@router.get("/audiences")
def list_audiences(user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            return (
                db.table("org_task_audiences")
                .select("*")
                .eq("org_id", user["org_id"])
                .order("created_at", desc=True)
                .execute()
                .data or []
            )
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_audiences attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_audiences failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/audiences")
def create_audience(body: AudienceIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            row = (
                db.table("org_task_audiences")
                .insert({
                    "org_id": user["org_id"],
                    "created_by": user["id"],
                    "name": body.name,
                    "criteria": body.criteria.model_dump(),
                    "academic_year": body.academic_year,
                })
                .execute()
            )
            return row.data[0]
        except Exception as exc:
            if attempt == 0:
                logger.warning("create_audience attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("create_audience failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.patch("/audiences/{audience_id}")
def update_audience(audience_id: str, body: AudienceIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            row = (
                db.table("org_task_audiences")
                .update({"name": body.name, "criteria": body.criteria.model_dump()})
                .eq("id", audience_id)
                .eq("org_id", user["org_id"])
                .execute()
            )
            if not row.data:
                raise HTTPException(status_code=404, detail="קהל לא נמצא")
            return row.data[0]
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("update_audience attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("update_audience failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.delete("/audiences/{audience_id}")
def delete_audience(audience_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("org_task_audiences").delete().eq("id", audience_id).eq("org_id", user["org_id"]).execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("delete_audience attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("delete_audience failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.get("/channel-availability")
def get_channel_availability(user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    outlook = get_org_mailbox_capability(user["org_id"])
    twilio = whatsapp_twilio.get_org_twilio_connection(user["org_id"])
    return {"email_outlook": outlook["connected"], "whatsapp_twilio": twilio["connected"]}


class TwilioSettingsIn(BaseModel):
    account_sid: str | None = None
    auth_token: str | None = None
    from_number: str | None = None


@router.get("/twilio-settings")
def get_twilio_settings(user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    rows = db.table("twilio_connections").select("*").eq("org_id", user["org_id"]).execute().data or []
    row = rows[0] if rows else {}
    return {
        "status": row.get("status", "disconnected"),
        "has_credentials": bool(row.get("account_sid") and row.get("auth_token")),
        "from_number": row.get("from_number"),
    }


@router.put("/twilio-settings")
def put_twilio_settings(body: TwilioSettingsIn, user: Annotated[dict, Depends(get_current_user)]):
    """Infra-only (see plan): stores credentials but never flips status to 'connected' —
    whatsapp_twilio.send_whatsapp_message always raises until real API wiring is built."""
    _require_manager(user)
    db = get_admin_client()
    db.table("twilio_connections").upsert({
        "org_id": user["org_id"],
        "account_sid": body.account_sid,
        "auth_token": body.auth_token,
        "from_number": body.from_number,
        "status": "disconnected",
    }, on_conflict="org_id").execute()
    return {"ok": True}


@router.post("/preview")
def preview_task(body: PreviewIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    matched = task_logic.find_matching_schools(
        user["org_id"], body.criteria.model_dump(), body.academic_year, manual_school_ids=body.manual_school_ids,
    )
    return {"count": len(matched), "schools": matched}


class DeriveSuccessIn(BaseModel):
    criteria: ConditionsIn


@router.post("/derive-success-criteria")
def derive_success_criteria(body: DeriveSuccessIn, user: Annotated[dict, Depends(get_current_user)]):
    """Powers the wizard's "מה נחשב הצלחה?" step, option (a) — auto-derive. `invertible=False`
    means the criteria isn't cleanly invertible (multi-group, or a `contains` op present) and
    the wizard must disable that option, forcing (b) custom or (c) track_success=False."""
    _require_manager(user)
    inverted = task_logic.invert_criteria(body.criteria.model_dump())
    return {"success_criteria": inverted, "invertible": inverted is not None}


class ContactsCheckIn(BaseModel):
    criteria: ConditionsIn | None = None
    manual_school_ids: list[str] | None = None
    recipient_role: str
    channel: str
    is_meeting_task: bool = False
    alternate_role: str | None = None
    academic_year: str = DEFAULT_ACADEMIC_YEAR


def _missing_contact_detail(school: dict) -> dict:
    """Shared by /contacts/check and /meetings/check's coordinator problem — surfaces a
    school's existing contact rows (even partially-filled ones, e.g. a name with no email) so
    a resolution modal can offer to complete/pick one instead of jumping straight to "add a
    brand-new contact" — most "missing" schools already have a name on file, just not the
    field the chosen channel needs."""
    return {
        "contacts": [
            {"role": role, "role_label": _PARTICIPANT_ROLE_LABELS[role],
             "name": school.get(name_f), "email": school.get(email_f), "phone": school.get(phone_f)}
            for role, (name_f, email_f, phone_f) in _ROLE_CONTACT_FIELDS.items()
        ],
        "meeting_coordinator": school.get("meeting_coordinator"),
    }


def _check_contact_problems(
    org_id: str, criteria: dict, manual_school_ids: list[str] | None, recipient_role: str,
    channel: str, is_meeting_task: bool, alternate_role: str | None, academic_year: str,
) -> dict:
    """Extracted from the /contacts/check endpoint handler so it can be reused for a scheduled
    general (non-meeting) task's re-validation right before activation (see
    _try_activate_scheduled_task) — same reasoning as _check_meeting_problems for the meeting
    track. Resolves each matched school's recipient through the role-priority cascade and
    reports who has a usable contact and who doesn't."""
    matched = task_logic.find_matching_schools(org_id, criteria, academic_year, manual_school_ids=manual_school_ids)
    school_ids = [m["school_id"] for m in matched]
    db = get_admin_client()
    schools = db.table("schools").select("*").in_("id", school_ids).execute().data or [] if school_ids else []
    schools_by_id = {s["id"]: s for s in schools}

    resolved_by_school = {}
    entries_by_school = {}
    for m in matched:
        school = schools_by_id.get(m["school_id"])
        if not school:
            continue
        res = _resolve_recipient_with_cascade(school, recipient_role, is_meeting_task, channel, alternate_role)
        entry = {
            "school_id": m["school_id"], "school_name": m["school_name"],
            "symbol": m.get("symbol"), "authority": m.get("authority"),
            "resolved_via": res["resolved_via"], "has_contact": res["resolved_via"] is not None,
            "opted_out": None,
        }
        if not entry["has_contact"]:
            entry.update(_missing_contact_detail(school))
        resolved_by_school[m["school_id"]] = res["recipient"].get("email")
        entries_by_school[m["school_id"]] = entry

    # Purely informational, never affects has_contact/missing_count — an opted-out school still
    # "has a contact", it's just currently suppressed from receiving messages (see task_logic.
    # opted_out_recipients). TaskContactResolutionModal is confirmed non-blocking regardless.
    try:
        opted_out_map = task_logic.opted_out_recipients(db, academic_year, resolved_by_school)
    except Exception as exc:
        logger.warning("_check_contact_problems: opt-out lookup failed (non-fatal): %s", exc)
        opted_out_map = {}
    for school_id, email in opted_out_map.items():
        entries_by_school[school_id]["opted_out"] = {"email": email}

    results = list(entries_by_school.values())
    return {"schools": results, "missing_count": sum(1 for r in results if not r["has_contact"])}


@router.post("/contacts/check")
def check_contacts(body: ContactsCheckIn, user: Annotated[dict, Depends(get_current_user)]):
    """Pre-creation contact-resolution check (TaskContactResolutionModal) — for the task's
    audience (by criteria or manual_school_ids), resolves each school's recipient through the
    role-priority cascade and reports who has a usable contact and who doesn't. Round 6: for
    meeting-scheduling tasks specifically, this coordinator check has been folded into the
    unified POST /tasks/meetings/check + TaskMeetingResolutionModal instead — this endpoint
    (and TaskContactResolutionModal) now only runs for general/non-meeting tasks."""
    _require_manager(user)
    criteria = body.criteria.model_dump() if body.criteria else {}
    return _check_contact_problems(
        user["org_id"], criteria, body.manual_school_ids, body.recipient_role,
        body.channel, body.is_meeting_task, body.alternate_role, body.academic_year,
    )


_CONTACT_EXPORT_COLUMNS = [
    ("שם מנהל/ת", "principal_name"), ("טלפון מנהל/ת", "principal_phone"), ("מייל מנהל/ת", "principal_email"),
    ("שם מנהלנ/ית", "secretary_name"), ("טלפון מנהלנ/ית", "secretary_phone"), ("מייל מנהלנ/ית", "secretary_email"),
    ("שם אחראי/ת כספים", "finance_contact_name"), ("טלפון אחראי/ת כספים", "finance_contact_phone"), ("מייל אחראי/ת כספים", "finance_contact_email"),
]


class ExportMissingContactsIn(BaseModel):
    school_ids: list[str]


@router.post("/contacts/export-missing")
def export_missing_contacts(body: ExportMissingContactsIn, user: Annotated[dict, Depends(get_current_user)]):
    """Generates a ready-to-fill Excel for schools with zero usable contact info at all (the
    tail end of the contact-resolution cascade) — filled columns get written back to `schools`
    via POST /tasks/contacts/import. `meeting_coordinator` is intentionally not a column here:
    it's an indirection pointing at one of the three roles below, configured separately in
    "פרטי בית הספר" — filling in any one of the three roles gives the cascade something to
    resolve to regardless of that pointer."""
    _require_manager(user)
    db = get_admin_client()
    schools = db.table("schools").select("*").in_("id", body.school_ids).eq("org_id", user["org_id"]).execute().data or []

    wb = Workbook()
    ws = wb.active
    ws.title = "בתי ספר חסרי פרטי קשר"
    ws.sheet_view.rightToLeft = True
    ws.append(["שם מוסד", "סמל מוסד", "בעלות"] + [label for label, _ in _CONTACT_EXPORT_COLUMNS])
    for s in schools:
        ws.append([s.get("name"), s.get("symbol"), s.get("authority")] + [s.get(field) for _, field in _CONTACT_EXPORT_COLUMNS])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "בתי_ספר_חסרי_פרטי_קשר.xlsx"
    encoded_name = urllib.parse.quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"missing-contacts.xlsx\"; filename*=UTF-8''{encoded_name}"},
    )


@router.post("/contacts/import")
async def import_missing_contacts(user: Annotated[dict, Depends(get_current_user)], file: UploadFile = File(...)):
    """Parses the completed export-missing file back and writes the filled contact fields into
    `schools`, matched by סמל מוסד (symbol) — mirrors the header-detection/per-row-error
    convention of schools_router.import_schools, but updates existing schools instead of
    creating new ones."""
    _require_manager(user)
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "יש להעלות קובץ Excel בלבד (.xlsx)")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"updated": 0, "errors": ["הקובץ ריק"]}

    first_cell = str(rows[0][0] or "").strip()
    start = 1 if first_cell in ("שם מוסד", "שם בית ספר", "שם", "name", "Name") else 0

    db = get_admin_client()
    updated = 0
    errors = []
    for i, row in enumerate(rows[start:], start=start + 2):
        symbol = str(row[1] or "").strip().split(".")[0] if len(row) > 1 else ""
        if not symbol:
            continue
        school_rows = db.table("schools").select("id").eq("org_id", user["org_id"]).eq("symbol", symbol).execute().data or []
        if not school_rows:
            errors.append(f"שורה {i}: לא נמצא בית ספר עם סמל מוסד '{symbol}'")
            continue
        patch = {}
        for col_idx, (_label, field) in enumerate(_CONTACT_EXPORT_COLUMNS):
            cell_idx = 3 + col_idx
            if len(row) > cell_idx and row[cell_idx]:
                patch[field] = str(row[cell_idx]).strip()
        if patch:
            db.table("schools").update(patch).eq("id", school_rows[0]["id"]).execute()
            updated += 1
    return {"updated": updated, "errors": errors}


class MeetingRequirementIn(BaseModel):
    meeting_service_type: str
    date_from: str | None = None
    date_to: str | None = None
    advisor_mode: str = "default"
    advisor_ids: list[str] = []
    duration_mode: str = "default"
    duration_minutes: int | None = None
    participant_roles: list[str] = []
    stage_scope: str | None = None  # round 8: 'tichon' | 'chativa' | 'both' — only meaningful
    # when "principal" is a participant_role and meeting_service_type != "current" (see
    # _principal_slots_for_school); ignored for non-six-year schools either way.


class MeetingsCheckIn(BaseModel):
    criteria: ConditionsIn | None = None
    manual_school_ids: list[str] | None = None
    meeting_requirements: list[MeetingRequirementIn]
    channel: str
    academic_year: str = DEFAULT_ACADEMIC_YEAR


def _coordinator_problem(school: dict, channel: str) -> dict | None:
    """The 'who receives the scheduling email itself' check for a meeting task — always
    recipient_role='meeting_coordinator'/is_meeting_task=True, unlike /contacts/check's
    general-purpose role param. Returns None when resolved."""
    res = _resolve_recipient_with_cascade(school, "meeting_coordinator", True, channel)
    if res["resolved_via"] is not None:
        return None
    return _missing_contact_detail(school)


def _check_meeting_problems(
    org_id: str, criteria: dict, manual_school_ids: list[str] | None,
    meeting_requirements: list[MeetingRequirementIn], channel: str, academic_year: str,
) -> dict:
    """Round-6 unified pre-creation check (TaskMeetingResolutionModal) for 'קביעת פגישות'
    tasks — replaces both the old split-out advisor/participant-name-only version of this
    endpoint AND a separate call to /tasks/contacts/check for the coordinator. Per matched
    school, reports up to four independent problem kinds:
    - coordinator: who actually receives the "pick a time" email — same cascade /contacts/check
      uses, fixed to recipient_role="meeting_coordinator".
    - participants: named contacts invited to the meeting itself, deduped across every meeting
      requirement's participant_roles for this school (not one row per requirement) — requires
      BOTH a name and the channel-specific field (email for email channels, phone for
      whatsapp), not just a name like the old version only checked.
    - meeting_defaults: a "default"-mode advisor or duration left unset on the school's own
      card for a division actually being requested — never checked before round 6 (silently
      fell back to 60 minutes at send time instead). One entry per meeting_service_type that
      has either gap, not two separate entries.
    - advisor_access: a manually-chosen advisor who no longer has access to the school's card.

    Extracted out of the /meetings/check endpoint handler so it can be reused both for the
    pre-creation wizard check AND for re-validating an already-created task right before a
    scheduled send (see /tasks/{task_id}/meetings/check and _try_activate_scheduled_task) —
    behavior is byte-for-byte identical to the original inline version, just parameterized."""
    matched = task_logic.find_matching_schools(
        org_id, criteria, academic_year, manual_school_ids=manual_school_ids,
    )
    school_ids = [m["school_id"] for m in matched]
    if not school_ids:
        return {"schools": [], "total_schools": 0, "ok_schools": 0}

    db = get_admin_client()
    schools = db.table("schools").select("*").in_("id", school_ids).execute().data or []
    schools_by_id = {s["id"]: s for s in schools}

    advisor_map: dict[str, dict[str, list[str]]] = {}
    for service_type, table_name in _TYPED_ADVISOR_TABLES.items():
        rows = db.table(table_name).select("school_id, advisor_id").in_("school_id", school_ids).execute().data or []
        for r in rows:
            advisor_map.setdefault(r["school_id"], {}).setdefault(service_type, []).append(r["advisor_id"])

    duration_rows = (
        db.table("school_year_admin_data")
        .select("school_id, meeting_duration_gefen, meeting_duration_current, meeting_duration_district")
        .eq("academic_year", academic_year).in_("school_id", school_ids).execute().data or []
    )
    duration_by_school = {r["school_id"]: r for r in duration_rows}
    needs_phone = channel == "whatsapp_twilio"

    entries_by_school: dict[str, dict] = {}
    coordinator_emails: dict[str, str | None] = {}
    for m in matched:
        school = schools_by_id.get(m["school_id"])
        if not school:
            continue

        coordinator = _coordinator_problem(school, channel)
        coordinator_emails[school["id"]] = _resolve_recipient_with_cascade(
            school, "meeting_coordinator", True, channel,
        )["recipient"].get("email")

        # Round 8: "principal" expands to 1-2 concrete slots (tichon/chativa) per this
        # school's stage_scope-aware resolution — every other role stays a flat 1:1 pass-through.
        participant_roles_needed: set[str] = set()
        for req in meeting_requirements:
            for role in req.participant_roles:
                if role == "principal":
                    participant_roles_needed.update(_principal_slots_for_school(school, req.stage_scope))
                else:
                    participant_roles_needed.add(role)
        bad_roles = []
        for role in participant_roles_needed:
            fields = _ROLE_CONTACT_FIELDS.get(role)
            if not fields:
                continue
            name_f, email_f, phone_f = fields
            channel_field_ok = bool(school.get(phone_f)) if needs_phone else bool(school.get(email_f))
            if not (school.get(name_f) and channel_field_ok):
                bad_roles.append(role)
        participants = None
        if bad_roles:
            participants = {
                "roles": bad_roles,
                "role_labels": {role: _principal_role_label(school, role) if role.startswith("principal") else _PARTICIPANT_ROLE_LABELS.get(role, role) for role in bad_roles},
                "contacts": {
                    role: {"name": school.get(_ROLE_CONTACT_FIELDS[role][0]),
                           "email": school.get(_ROLE_CONTACT_FIELDS[role][1]),
                           "phone": school.get(_ROLE_CONTACT_FIELDS[role][2])}
                    for role in bad_roles
                },
            }

        duration_row = duration_by_school.get(school["id"], {})
        type_needs: dict[str, dict] = {}
        for req in meeting_requirements:
            service_type = req.meeting_service_type
            if service_type not in _MEETING_TYPE_LABELS:
                continue
            entry = type_needs.setdefault(service_type, {"missing_advisor": False, "missing_duration": False})
            if req.advisor_mode != "manual" and not advisor_map.get(school["id"], {}).get(service_type):
                entry["missing_advisor"] = True
            if req.duration_mode != "manual" and not duration_row.get(f"meeting_duration_{service_type}"):
                entry["missing_duration"] = True
        meeting_defaults = [
            {"meeting_service_type": t, **flags}
            for t, flags in type_needs.items() if flags["missing_advisor"] or flags["missing_duration"]
        ]

        # Advisor-access check — the "the manually picked advisor can't actually reach this
        # school's card" case from the temp-access-grant feature. Only relevant for manual-mode
        # requirements (a "default" advisor is resolved from the school's own card, which
        # implicitly already has access via advisors_gefen/current/district assignment).
        advisor_access = []
        for req in meeting_requirements:
            if req.advisor_mode != "manual" or not req.advisor_ids:
                continue
            for advisor_id in req.advisor_ids:
                if not _schools_router._advisor_has_school_access(db, advisor_id, school):
                    advisor_access.append({
                        "meeting_service_type": req.meeting_service_type,
                        "advisor_id": advisor_id,
                        "date_from": req.date_from,
                        "date_to": req.date_to,
                    })

        if coordinator or participants or meeting_defaults or advisor_access:
            entries_by_school[school["id"]] = {
                "school_id": school["id"], "school_name": school["name"],
                "symbol": school.get("symbol"), "authority": school.get("authority"),
                "coordinator": coordinator, "participants": participants, "meeting_defaults": meeting_defaults,
                "advisor_access": advisor_access, "opted_out": None,
            }

    # Purely informational (never a blocking leaf) — an opted-out coordinator is added even to
    # an otherwise-clean school so the manager sees it, but never contributes to ok_schools math
    # below in a way that would block "צור משימה" (TaskMeetingResolutionModal keeps this field
    # out of its leaves/blocking system entirely).
    try:
        opted_out_map = task_logic.opted_out_recipients(db, academic_year, coordinator_emails)
    except Exception as exc:
        logger.warning("_check_meeting_problems: opt-out lookup failed (non-fatal): %s", exc)
        opted_out_map = {}
    for school_id, email in opted_out_map.items():
        school = schools_by_id.get(school_id)
        if not school:
            continue
        entry = entries_by_school.setdefault(school_id, {
            "school_id": school_id, "school_name": school["name"],
            "symbol": school.get("symbol"), "authority": school.get("authority"),
            "coordinator": None, "participants": None, "meeting_defaults": [],
            "advisor_access": [], "opted_out": None,
        })
        entry["opted_out"] = {"email": email}

    results = list(entries_by_school.values())
    blocking_count = sum(1 for r in results if r["coordinator"] or r["participants"] or r["meeting_defaults"] or r["advisor_access"])
    return {"schools": results, "total_schools": len(matched), "ok_schools": len(matched) - blocking_count}


@router.post("/meetings/check")
def check_meetings(body: MeetingsCheckIn, user: Annotated[dict, Depends(get_current_user)]):
    """Pre-creation check, driven by TaskMeetingResolutionModal in the task-creation wizard.
    See _check_meeting_problems for the actual problem-detection logic."""
    _require_manager(user)
    criteria = body.criteria.model_dump() if body.criteria else {}
    return _check_meeting_problems(
        user["org_id"], criteria, body.manual_school_ids, body.meeting_requirements,
        body.channel, body.academic_year,
    )


def _meeting_requirements_from_task(task: dict) -> list[MeetingRequirementIn]:
    """Reconstructs the meeting_requirements list an already-created task was built with,
    straight from its persisted success_criteria — same source _meeting_conditions_from_
    success_criteria (used by _queue_messages_for_schools) already reads from. Extra keys
    like "type" on each condition dict are silently ignored by MeetingRequirementIn."""
    return [MeetingRequirementIn(**c) for c in _meeting_conditions_from_success_criteria(task)]


@router.get("/{task_id}/meetings/check")
def check_task_meeting_problems(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Live re-check of an already-created meeting-scheduling task's current problems — used
    by the 'בעיות' screen opened from a red task-card badge or a task_send_problems
    notification (see has_meeting_send_problems / _try_activate_scheduled_task)."""
    _require_manager(user)
    db = get_admin_client()
    task_row = db.table("org_tasks").select("*").eq("id", task_id).eq("org_id", user["org_id"]).execute().data
    if not task_row:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    task = task_row[0]
    meeting_requirements = _meeting_requirements_from_task(task)
    channel = (task.get("message_config") or {}).get("channel") or "email_resend"
    return _check_meeting_problems(
        user["org_id"], task.get("criteria") or {}, task.get("manual_school_ids"),
        meeting_requirements, channel, task.get("academic_year") or DEFAULT_ACADEMIC_YEAR,
    )


class MeetingProblemRowIn(BaseModel):
    school_id: str
    school_name: str
    symbol: str | None = None
    authority: str | None = None
    kind: str  # "coordinator" | "participant" | "advisor" | "duration"
    detail: str | None = None  # role (participant) or meeting_service_type (advisor/duration)


class ExportMissingMeetingsIn(BaseModel):
    rows: list[MeetingProblemRowIn]


_MEETING_PROBLEM_KIND_LABELS = {
    "coordinator": "אחראי/ת תיאום פגישות חסר/ה", "participant": "איש קשר חסר",
    "advisor": "יועץ מבצע חסר", "duration": "משך פגישה חסר",
}


@router.post("/meetings/export-missing")
def export_missing_meetings(body: ExportMissingMeetingsIn, user: Annotated[dict, Depends(get_current_user)]):
    """Excel export of the TaskMeetingResolutionModal's unified problem list — one row per
    issue (a school can appear more than once if it has several gaps), mirrors
    export_missing_contacts' pattern."""
    _require_manager(user)
    wb = Workbook()
    ws = wb.active
    ws.title = "בעיות תיאום פגישה"
    ws.sheet_view.rightToLeft = True
    ws.append(["שם מוסד", "סמל מוסד", "בעלות", "סוג בעיה", "פרט"])
    for r in body.rows:
        kind_label = _MEETING_PROBLEM_KIND_LABELS.get(r.kind, r.kind)
        if r.kind in ("advisor", "duration"):
            detail_label = _MEETING_TYPE_LABELS.get(r.detail, r.detail)
        elif r.kind == "participant":
            detail_label = _PARTICIPANT_ROLE_LABELS.get(r.detail, r.detail)
        else:
            detail_label = ""
        ws.append([r.school_name, r.symbol, r.authority, kind_label, detail_label])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "בתי_ספר_עם_בעיות_תיאום_פגישה.xlsx"
    encoded_name = urllib.parse.quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"missing-meetings.xlsx\"; filename*=UTF-8''{encoded_name}"},
    )


@router.post("/")
def create_task(body: TaskCreateIn, user: Annotated[dict, Depends(get_current_user)]):
    """Round-2 redesign: creation now immediately queues messages to whoever matches (see
    plan "יצירת משימה = שליחה מיידית") — the old flow only computed matched_school_ids and
    left sending to a separate manual click. The interactive contact-resolution flow
    (TaskContactResolutionModal, driven by POST /tasks/contacts/check) runs client-side
    BEFORE this endpoint is called, so by the time we get here the manager has already
    resolved/skipped/completed missing contacts — this endpoint doesn't need to know about
    that flow, it just queues to whatever final audience it's given."""
    _require_manager(user)
    criteria = body.criteria.model_dump()
    manual_ids = body.manual_school_ids
    if not (criteria.get("groups") or manual_ids):
        raise HTTPException(status_code=422, detail="יש לבחור קריטריון סינון או לבחור בתי ספר ידנית")
    success_criteria = body.success_criteria.model_dump() if body.success_criteria else None

    scheduled_for_dt = _parse_dt(body.scheduled_for)
    is_future_scheduled = bool(scheduled_for_dt and scheduled_for_dt > datetime.now(timezone.utc))
    if is_future_scheduled:
        matched_school_ids = []
        status = "scheduled"
    else:
        matched = task_logic.find_matching_schools(user["org_id"], criteria, body.academic_year, manual_school_ids=manual_ids)
        excluded = set(body.excluded_school_ids or [])
        matched_school_ids = [m["school_id"] for m in matched if m["school_id"] not in excluded]
        status = "active"

    channel = body.message_config.channel
    if not is_future_scheduled and channel == "email_outlook" and len(matched_school_ids) > OUTLOOK_SEND_WARN_THRESHOLD and not body.confirm_outlook_limit:
        raise HTTPException(status_code=409, detail={
            "outlook_limit_exceeded": True,
            "message": (
                f"פעולה זו תשלח {len(matched_school_ids)} מיילים דרך Outlook הארגוני, מעל לסף האזהרה "
                f"המוגדר ({OUTLOOK_SEND_WARN_THRESHOLD}). מגבלות Microsoft בפועל: עד 30 הודעות "
                f"בדקה ועד 10,000 נמענים ביום (משותף לכל השליחות בארגון, לא רק למשימה זו)."
            ),
            "count": len(matched_school_ids),
            "threshold": OUTLOOK_SEND_WARN_THRESHOLD,
        })

    task = None
    for attempt in range(2):
        try:
            db = get_admin_client()
            row = (
                db.table("org_tasks")
                .insert({
                    "org_id": user["org_id"],
                    "created_by": user["id"],
                    "name": body.name,
                    "status": status,
                    "criteria": criteria,
                    "matched_school_ids": matched_school_ids,
                    "message_config": body.message_config.model_dump(),
                    "scheduled_for": scheduled_for_dt.isoformat() if is_future_scheduled else None,
                    "academic_year": body.academic_year,
                    "manual_school_ids": manual_ids,
                    "success_criteria": success_criteria,
                    "track_success": body.track_success,
                    "is_meeting_task": body.is_meeting_task,
                    "meeting_overrides": body.meeting_overrides,
                })
                .execute()
            )
            task = row.data[0]
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("create_task attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
            else:
                logger.error("create_task failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    missing_contact_school_ids = []
    if not is_future_scheduled and matched_school_ids:
        if body.track_success:
            # Never send a "please do the thing" reminder to a school that already satisfies
            # success on day one (e.g. a manually-picked school that happens to already have
            # the meeting) — see plan decision #2.
            progress = task_logic.compute_task_progress(user["org_id"], task)
            send_targets = [r["school_id"] for r in progress["schools"] if not r["done"]]
            _mark_already_done_schools(db, task["id"], matched_school_ids, send_targets)
        else:
            send_targets = matched_school_ids
        if send_targets:
            missing_contact_school_ids, _skipped_broadcast = _queue_messages_for_schools(db, task, user["org_id"], send_targets)

    return {**task, "missing_contact_school_ids": missing_contact_school_ids}


@router.get("/")
def list_tasks(
    user: Annotated[dict, Depends(get_current_user)],
    status: str | None = None,
    is_meeting_task: bool | None = None,
    created_by: str | None = None,
    track_success: bool | None = None,
    has_meeting_send_problems: bool | None = None,
    channel: str | None = None,
    created_from: str | None = None,
    created_to: str | None = None,
    academic_year: str | None = None,
):
    """Redesigned "משימות" table (see plan doc) — the list view is a pure read of stored
    columns (status, cached_total_schools/cached_actions_needed/cached_actions_completed/
    cached_progress_pct), never a live compute_task_progress loop over every org task
    (Architecture Invariant #7). Those cached columns are kept correct by task_logic.
    recompute_task_status_and_cache, called at the mutation touchpoints in this file, exactly
    like has_meeting_send_problems already was. Server-side filters below are all either plain
    column `.eq()`/range chains or (for `channel`, nested inside the message_config JSONB) a
    post-fetch Python filter over the single already org-scoped result set — never a per-row
    query loop."""
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            query = db.table("org_tasks").select("*").eq("org_id", user["org_id"])
            if status:
                query = query.eq("status", status)
            if is_meeting_task is not None:
                query = query.eq("is_meeting_task", is_meeting_task)
            if created_by:
                query = query.eq("created_by", created_by)
            if track_success is not None:
                query = query.eq("track_success", track_success)
            if has_meeting_send_problems is not None:
                query = query.eq("has_meeting_send_problems", has_meeting_send_problems)
            if created_from:
                query = query.gte("created_at", created_from)
            if created_to:
                query = query.lte("created_at", created_to)
            if academic_year:
                query = query.eq("academic_year", academic_year)
            rows = query.order("created_at", desc=True).execute().data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_tasks attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
            else:
                logger.error("list_tasks failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    if channel:
        rows = [r for r in rows if (r.get("message_config") or {}).get("channel") == channel]

    try:
        db = get_admin_client()
        creator_ids = list({r["created_by"] for r in rows if r.get("created_by")})
        names_map = {}
        if creator_ids:
            p_rows = db.table("profiles").select("id, full_name").in_("id", creator_ids).execute().data or []
            names_map = {p["id"]: p["full_name"] for p in p_rows}
        for r in rows:
            r["created_by_name"] = names_map.get(r.get("created_by"))
    except Exception as exc:
        logger.warning("list_tasks creator-name enrichment failed (non-fatal): %s", exc)

    try:
        db = get_admin_client()
        task_ids = [r["id"] for r in rows]
        pins_map = {}
        if task_ids:
            pin_rows = (
                db.table("task_pins").select("task_id, pinned_at")
                .eq("user_id", user["id"]).in_("task_id", task_ids).execute().data or []
            )
            pins_map = {p["task_id"]: p["pinned_at"] for p in pin_rows}
        for r in rows:
            r["pinned_at"] = pins_map.get(r["id"])
    except Exception as exc:
        logger.warning("list_tasks pin enrichment failed (non-fatal): %s", exc)

    return rows


def _get_task_or_404(db, task_id: str, org_id: str) -> dict:
    rows = db.table("org_tasks").select("*").eq("id", task_id).eq("org_id", org_id).execute().data or []
    if not rows:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    return rows[0]


class TaskOptOutIn(BaseModel):
    email: str
    token: str


# Registered before "/{task_id}" (a literal path segment would otherwise be swallowed by
# that catch-all pattern under the same HTTP method — same reason /field-options, /templates,
# /preview etc. all sit above the "/{task_id}" routes too).
@router.patch("/opt-out")
def opt_out_task_contact(body: TaskOptOutIn):
    """Public endpoint (no auth) — same HMAC-token convention as
    signup_router.unsubscribe_marketing/PATCH /signup/unsubscribe. A contact who clicks the
    opt-out link in a task message (only sent to inactive-client contacts, see
    _queue_messages_for_schools) lands here via TaskOptOutPage.jsx and is suppressed from all
    future task messages, across every task, until manually removed from the table."""
    email = body.email.strip().lower()
    expected = task_logic.make_optout_token(email)
    if not hmac.compare_digest(expected, body.token):
        raise HTTPException(status_code=400, detail="קישור לא תקין")
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("task_opted_out_contacts").upsert({"email": email}, on_conflict="email").execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("opt_out_task_contact attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("opt_out_task_contact failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.get("/{task_id}")
def get_task(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            task = _get_task_or_404(db, task_id, user["org_id"])
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_task attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("get_task failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")
    progress = task_logic.compute_task_progress(user["org_id"], task)

    # Round-2 redesign: success_criteria is now an independent tree from criteria (targeting).
    # The frontend needs to know the ACTUAL conditions condition_results reflects (whichever
    # tree compute_task_progress walked) to label the results-table columns correctly —
    # describing them from task["criteria"] would show the wrong (targeting) conditions once
    # the two trees diverge. None when track_success is off (no success tree at all).
    effective_success_criteria = None
    if task.get("track_success", True):
        effective_success_criteria = task.get("success_criteria") or task_logic.invert_criteria(task.get("criteria") or {})

    # Real send-status per school (bug #3): org_task_messages.status is already truthful
    # (process_message_queue only marks "sent" after the channel call succeeds) — the gap was
    # that nothing surfaced it to the manager. Non-fatal: a failure here must not break the
    # main progress view (Architecture Invariant #6).
    message_summary = {"queued": 0, "sent": 0, "failed": 0, "pending": 0, "skipped": 0, "outlook_pending": 0}
    try:
        msg_rows = (
            db.table("org_task_messages")
            .select("school_id, status, sent_at, error, channel, booking_token_id, fallback_used, outlook_sent_at")
            .eq("task_id", task_id)
            .order("created_at", desc=True)
            .execute()
            .data or []
        )
        latest_by_school = {}
        for m in msg_rows:
            latest_by_school.setdefault(m["school_id"], m)

        # Round 7: "link clicked" / "meeting booked" — the two reliable (non-pixel) send-status
        # signals for TaskPanel's tooltip. Batched, not N+1.
        token_ids = list({m["booking_token_id"] for m in latest_by_school.values() if m.get("booking_token_id")})
        tokens_by_id = {}
        if token_ids:
            try:
                token_rows = (
                    db.table("meeting_booking_tokens")
                    .select("id, first_viewed_at, booked_ranges, date_ranges")
                    .in_("id", token_ids).execute().data or []
                )
                tokens_by_id = {t["id"]: t for t in token_rows}
            except Exception as exc:
                logger.warning("get_task booking-token enrichment failed (non-fatal): %s", exc)

        for row in progress["schools"]:
            latest = latest_by_school.get(row["school_id"])
            if not latest:
                row["send_status"] = None
                continue
            send_status = {
                "status": latest["status"],
                "sent_at": latest.get("sent_at"),
                "error": latest.get("error"),
                "channel": latest.get("channel"),
                "link_viewed_at": None,
                "meeting_progress": None,
                # Round 15 — "outlook_pending" means Graph accepted the send but delivery isn't
                # confirmed yet (bounce-check grace window); fallback_used means a bounce was
                # detected and Resend sent it instead.
                "fallback_used": bool(latest.get("fallback_used")),
                "outlook_sent_at": latest.get("outlook_sent_at"),
            }
            token = tokens_by_id.get(latest.get("booking_token_id"))
            if token:
                send_status["link_viewed_at"] = token.get("first_viewed_at")
                total = len(token.get("date_ranges") or [])
                done = len(token.get("booked_ranges") or [])
                send_status["meeting_progress"] = {"done": done, "total": total}
            row["send_status"] = send_status
        for m in latest_by_school.values():
            message_summary["queued"] += 1
            if m["status"] in message_summary:
                message_summary[m["status"]] += 1
    except Exception as exc:
        logger.warning("get_task message-status enrichment failed (non-fatal): %s", exc)

    # Per-school notes/excluded-emails (bugs #9/#10) — non-fatal, same reasoning as above.
    try:
        school_ids = [r["school_id"] for r in progress["schools"]]
        notes_map = _fetch_school_notes_map(db, task_id, school_ids)
        _today = date.today()
        for row in progress["schools"]:
            note_row = notes_map.get(row["school_id"])
            row["note"] = (note_row or {}).get("note")
            row["excluded_emails"] = (note_row or {}).get("excluded_emails") or []
            row["skip_reason"] = (note_row or {}).get("skip_reason")
            row["broadcast_skip"] = {
                "mode": (note_row or {}).get("broadcast_skip_mode"),
                "until": (note_row or {}).get("broadcast_skip_until"),
                "active": _broadcast_skip_active(note_row, _today),
            }
    except Exception as exc:
        logger.warning("get_task notes/exclusions enrichment failed (non-fatal): %s", exc)

    # Self-heals the cached status/progress columns list_tasks reads (see task_logic.
    # recompute_task_status_and_cache) every time a manager actually opens this task — on top
    # of the write-time touchpoints elsewhere, this guarantees the Tasks table can never drift
    # from what this exact detail view is showing for more than one open of the task.
    cache_patch = task_logic.recompute_task_status_and_cache(db, user["org_id"], task)
    if cache_patch:
        task = {**task, **cache_patch}

    return {**task, "progress": progress, "message_summary": message_summary, "effective_success_criteria": effective_success_criteria}


@router.patch("/{task_id}")
def patch_task(task_id: str, body: TaskPatchIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    _get_task_or_404(db, task_id, user["org_id"])
    patch = {}
    if body.name is not None:
        patch["name"] = body.name
    if body.status is not None:
        patch["status"] = body.status
    if body.message_config is not None:
        patch["message_config"] = body.message_config.model_dump()
    if not patch:
        return _get_task_or_404(db, task_id, user["org_id"])
    row = db.table("org_tasks").update(patch).eq("id", task_id).eq("org_id", user["org_id"]).execute()
    return row.data[0]


@router.delete("/{task_id}")
def delete_task(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    _get_task_or_404(db, task_id, user["org_id"])
    db.table("org_tasks").delete().eq("id", task_id).eq("org_id", user["org_id"]).execute()
    return {"ok": True}


def _render_template(template: str, school: dict, booking_link: str | None = None, opt_out_link: str | None = None,
                      meetings_list: str | None = None, recipient_name: str | None = None,
                      advisor_names: str | None = None, keep_tokens: bool = False) -> str:
    """keep_tokens=True leaves {meetings_list}/{booking_link} untouched in the output — used when
    the caller (_wrap_email_html) still needs to substitute them itself (meetings_list as trusted
    pre-built HTML, booking_link as a styled button rather than a raw URL). {advisor_names} is
    plain text (like {recipient_name}/{school_name}) so it's always substituted directly here,
    regardless of keep_tokens — just the raw comma-joined advisor name(s) ("עם" belongs in the
    template text itself, e.g. "מבוקש לתאם עם {advisor_names} את"). Effectively always non-empty
    for meeting tasks in practice — _build_meeting_booking_link refuses to build a booking link at
    all when no advisor resolves for any range, so ranges_data never reaches this function empty."""
    text = (template or "").replace("{school_name}", school.get("name") or "")
    if "{recipient_name}" in text:
        text = text.replace("{recipient_name}", recipient_name or "")
    if "{advisor_names}" in text:
        text = text.replace("{advisor_names}", advisor_names or "")
    if not keep_tokens:
        if "{meetings_list}" in text:
            text = text.replace("{meetings_list}", meetings_list or "")
        if "{booking_link}" in text:
            text = text.replace("{booking_link}", booking_link or "")
    if opt_out_link:
        text = f"{text}\n\nלהסרה מרשימת התפוצה: {opt_out_link}"
    return text


_EMAIL_BODY_TOKEN_RE = re.compile(r"(\{meetings_list\}|\{booking_link\})")


def _wrap_email_html(rendered_text: str, booking_url: str | None, meetings_list_html: str | None, branded: bool = True) -> str:
    """Wraps a manager-authored body (rendered via _render_template(..., keep_tokens=True), so
    {meetings_list}/{booking_link} are still literal tokens here). Free-text segments are
    HTML-escaped and \\n\\n/\\n converted to real paragraph/line breaks — fixing the long-standing
    bug where plain-text task bodies were sent as literal (unescaped) HTML, so newlines never
    rendered for the recipient. {meetings_list}'s value is trusted, pre-built HTML (from
    booking_logic.format_ranges_html) and is never escaped.

    branded=True (default, used for email_resend) wraps the body in the same branded card used
    by booking_logic.build_direct_coordination_email_html — "גפן AI" header bar + footer.
    branded=False (used for email_outlook) skips that chrome entirely — an Outlook send goes out
    under the actual sender's own name/mailbox and is meant to read as a personal email from
    them, not as automated system mail; the blue "גפן AI" branding looked out of place there
    (confirmed live)."""
    parts = _EMAIL_BODY_TOKEN_RE.split(rendered_text)
    html_parts = []
    for part in parts:
        if part == "{meetings_list}":
            if meetings_list_html:
                html_parts.append(meetings_list_html)
        elif part == "{booking_link}":
            if booking_url:
                html_parts.append(
                    f'<div style="text-align: center; margin: 8px 0 16px 0;">'
                    f'<a href="{booking_url}" style="display: inline-block; background: #0070F3; color: white; '
                    f'font-size: 14px; font-weight: 700; padding: 12px 28px; border-radius: 8px; '
                    f'text-decoration: none;">קביעת מועד</a></div>'
                )
        elif part:
            for paragraph in part.split("\n\n"):
                if not paragraph.strip():
                    continue
                escaped = html.escape(paragraph).replace("\n", "<br/>")
                html_parts.append(f'<p style="margin: 0 0 16px 0; color: #334155; line-height: 1.8;">{escaped}</p>')
    body_html = "".join(html_parts)

    if not branded:
        return f"""
<html>
<body dir="rtl" style="font-family: Arial, sans-serif; font-size: 14px; color: #1e293b; margin: 0; padding: 0;">
  {body_html}
</body>
</html>"""

    return f"""
<html>
<body dir="rtl" style="font-family: Arial, sans-serif; font-size: 14px; color: #1e293b;
                       background: #f8fafc; margin: 0; padding: 24px;">
  <div style="max-width: 560px; margin: 0 auto; background: white;
              border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden;">
    <div style="background: #0070F3; padding: 20px 24px;">
      <p style="margin: 0; color: white; font-size: 14px; font-weight: 700;">גפן AI</p>
    </div>
    <div style="padding: 28px 24px;">{body_html}</div>
    <div style="background: #f1f5f9; padding: 12px 24px; text-align: center;">
      <p style="margin: 0; font-size: 11px; color: #94a3b8;">נשלח אוטומטית מגפן AI</p>
    </div>
  </div>
</body>
</html>"""


def _first_meeting_condition(criteria: dict) -> dict | None:
    for group in (criteria or {}).get("groups") or []:
        for c in group.get("conditions") or []:
            if c.get("type") == "meeting":
                return c
    return None


def _build_booking_link(db, org_id: str, task: dict, school_id: str) -> str | None:
    """Generic single-link fallback — only used when a task references {booking_link} without
    being a structured meeting-scheduling task (round-5's _build_meeting_booking_link handles
    those; see its call site in _queue_messages_for_schools). Reuses the existing
    unique-scheduling-link machinery instead of building a new one."""
    advisors = db.table("advisor_schools").select("advisor_id").eq("school_id", school_id).execute().data or []
    if not advisors:
        return None
    advisor_id = advisors[0]["advisor_id"]
    meeting_cond = _first_meeting_condition(task.get("criteria") or {}) or {}
    months = []
    if meeting_cond.get("date_from"):
        months = [meeting_cond["date_from"][:7]]
    window = {"days_of_week": [0, 1, 2, 3, 4], "start_hour": 8, "end_hour": 16, "duration_minutes": 60}
    # draft_id is nullable on meeting_booking_tokens (FK -> meeting_booking_drafts, which this
    # generalized Tasks feature intentionally does not create rows in — see plan §"extend, don't
    # fork"). Passing None reuses/creates a token keyed only on (school, advisor), which is fine:
    # the underlying link is meant to be reusable per school+advisor regardless of which task
    # asked for it.
    token_row = booking_token_logic.get_or_create_booking_token(db, org_id, school_id, advisor_id, None, months, window)
    return f"{os.getenv('APP_URL', '')}/book/{token_row['token']}"


def _meeting_conditions_from_success_criteria(task: dict) -> list[dict]:
    success_criteria = task.get("success_criteria") or {}
    groups = success_criteria.get("groups") or []
    if not groups:
        return []
    return [c for c in (groups[0].get("conditions") or []) if c.get("type") == "meeting"]


def _build_meeting_booking_link(
    db, org_id: str, school: dict, meeting_conditions: list[dict],
    advisor_map: dict[str, list[str]], duration_row: dict, school_overrides: dict | None = None,
) -> tuple[str | None, str | None, list[dict] | None]:
    """Structured, per-meeting-requirement booking link for 'קביעת פגישות' tasks (round 5) —
    mirrors schools_router.send_direct_coordination_request's ranges_data construction
    (label/key building, participants fixed up front), but resolved per-school from each
    meeting requirement's advisor_mode/duration_mode/participant_roles instead of a manager
    interactively picking advisors/participants for one specific school. advisor_map =
    {"gefen":[ids],"current":[ids],"district":[ids]} — that school's typed advisor
    assignments; duration_row = that school's school_year_admin_data row (for
    meeting_duration_gefen/current/district). school_overrides (round 6) =
    {service_type: {"advisor_ids":[...], "duration_minutes": int}} — the task's own
    per-school "קבע אחר לפגישה זו בלבד" overrides (org_tasks.meeting_overrides[school_id]),
    which never get written back to the school's own card. Resolution priority per field:
    condition-level "manual" value (shared by every school) > this school's override > the
    school's own card default. Returns (None, None) (falls back to no link in the message) if
    nothing usable could be resolved — should not normally happen since POST /tasks/meetings/
    check runs first and lets the manager fix/override/skip every gap before creation; the
    60-minute duration fallback below is a defensive last resort only, not a real UX path.
    Round 7: also returns the minted token's id (for TaskPanel's send-status tracking — see
    _queue_messages_for_schools' booking_token_id column) and attaches each range's own
    resolved advisor_ids (round-7 bug fix — the token-level advisor list used to be one flat
    union shared across every range, so a school booking one meeting type ended up inviting
    every advisor from every OTHER meeting type too; meeting_booking_router.py now prefers a
    range's own advisor_ids over the token-level union when present).
    Round 9: also returns ranges_data itself (previously built here and then discarded once the
    token was minted) so the caller can render the same rich per-meeting HTML email that
    "תיאום ישיר" already sends, via booking_logic.build_direct_coordination_email_html — instead
    of duplicating that template."""
    school_overrides = school_overrides or {}
    valid_conditions = [c for c in meeting_conditions if c.get("meeting_service_type") in _MEETING_TYPE_LABELS]
    type_counts: dict[str, int] = {}
    for c in valid_conditions:
        type_counts[c["meeting_service_type"]] = type_counts.get(c["meeting_service_type"], 0) + 1

    type_seen: dict[str, int] = {}
    ranges_data = []
    advisor_ids_union: list[str] = []
    for c in valid_conditions:
        service_type = c["meeting_service_type"]
        override = school_overrides.get(service_type) or {}
        start_date, end_date = c.get("date_from"), c.get("date_to")
        if not start_date or not end_date:
            continue
        if c.get("advisor_mode") == "manual":
            advisor_ids = c.get("advisor_ids")
        else:
            advisor_ids = override.get("advisor_ids") or advisor_map.get(service_type, [])
        if not advisor_ids:
            continue
        if c.get("duration_mode") == "manual":
            duration_minutes = c.get("duration_minutes") or 60
        else:
            duration_minutes = override.get("duration_minutes") or duration_row.get(f"meeting_duration_{service_type}") or 60

        roles = c.get("participant_roles") or []
        stage_scope = c.get("stage_scope")
        # Round 16 — "separate" builds TWO independent ranges (one per principal) instead of one
        # merged range with both principals as participants, mirroring StageScopeModal.jsx's
        # existing "שתי פגישות נפרדות" option for the general (non-task) scheduling flow.
        split_scopes = ["tichon", "chativa"] if (stage_scope == "separate" and "principal" in roles) else [stage_scope]

        for split_scope in split_scopes:
            participants = _build_school_contacts(school, roles, split_scope)
            if not participants:
                continue

            type_seen[service_type] = type_seen.get(service_type, 0) + 1
            base_label = f"פגישת {_MEETING_TYPE_LABELS[service_type]}"
            if len(split_scopes) > 1:
                principal_slot = "principal" if split_scope == "tichon" else "principal_chativa"
                label = f"{base_label} — {_principal_role_label(school, principal_slot)}"
            elif type_counts[service_type] == 1:
                label = base_label
            else:
                label = f"{base_label} ({type_seen[service_type]})"
            ranges_data.append({
                "key": f"r{len(ranges_data)}-{secrets.token_urlsafe(6)}",
                "start_date": start_date, "end_date": end_date,
                "service_type": service_type, "duration_minutes": duration_minutes,
                "label": label, "participants": participants, "advisor_ids": advisor_ids,
                # Round 17 — tags this range with which principal slot it's for (or the
                # condition's original stage_scope when not splitting), so the booked `meetings`
                # row can carry it too — without this, two ranges from the same "separate"
                # condition are indistinguishable at the DB level and booking either one
                # incorrectly satisfies (and duplicate-blocks) the other.
                "stage_scope": split_scope,
            })
            for aid in advisor_ids:
                if aid not in advisor_ids_union:
                    advisor_ids_union.append(aid)

    if not ranges_data or not advisor_ids_union:
        return None, None, None

    token_row = booking_token_logic.create_direct_booking_token(db, org_id, school["id"], advisor_ids_union, ranges_data)
    return f"{os.getenv('APP_URL', '')}/book/{token_row['token']}", token_row["id"], ranges_data


def _broadcast_skip_active(note_row: dict | None, today: date) -> bool:
    """Single source of truth for "is this school currently excluded from a broadcast reminder".
    'once'  — always active until it's consumed by the next broadcast send.
    'until' — active while broadcast_skip_until >= today (inclusive).
    'forever' — always.
    Note: this is purely about *reminder delivery* — it has ZERO effect on task
    progress/completion (compute_task_progress / recompute_task_status_and_cache never read it).
    """
    if not note_row:
        return False
    mode = note_row.get("broadcast_skip_mode")
    if mode in ("once", "forever"):
        return True
    if mode == "until":
        u = note_row.get("broadcast_skip_until")
        return bool(u) and str(u)[:10] >= today.isoformat()
    return False


def _fetch_school_notes_map(db, task_id: str, school_ids: list[str]) -> dict[str, dict]:
    """Non-fatal by design at the call sites — a lookup failure here should degrade to
    'no exclusions known' rather than block sending, per Architecture Invariant #6."""
    if not school_ids:
        return {}
    rows = (
        db.table("org_task_school_notes")
        .select("school_id, note, excluded_emails, skip_reason, broadcast_skip_mode, broadcast_skip_until")
        .eq("task_id", task_id)
        .in_("school_id", school_ids)
        .execute()
        .data or []
    )
    return {r["school_id"]: r for r in rows}


def _send_message_now(db, org_id: str, school_id: str, channel: str, subject: str, body: str,
                       recipient_email: str | None, recipient_phone: str | None,
                       attachment_keys: list[dict] | None, created_by: str | None = None) -> tuple[str, str | None]:
    """Round-7 — actually dispatches one message through its channel. Shared by
    process_message_queue's cron-drained batches AND _queue_messages_for_schools' own
    immediate-send path below (email_resend/whatsapp no longer wait for the cron at all — see
    that call site's comment for why). Returns ("sent"|"outlook_pending", None) or
    ("failed", error_str); never raises, so callers can always safely write the returned status
    straight into a row.

    Round 15: email_outlook returns "outlook_pending", not "sent" — Graph accepting the send
    only means the request was queued for delivery, not that it arrived (a real bounce/NDR can
    still come back minutes later, confirmed live). The caller (process_message_queue) is
    responsible for later confirming/falling-back via graph_client.find_bounce_for_subject; this
    function itself never blocks waiting for that.

    created_by: the task's creator (org_tasks.created_by) — for email_outlook this is WHOSE
    mailbox actually sends the message. Previously this looked up the school's assigned advisor
    instead (advisor_schools), which sent from a possibly-unrelated, possibly low-reputation/
    newly-created mailbox instead of the (usually senior, established) manager who authored the
    task — confirmed live to be the wrong mailbox. Required for email_outlook; other channels
    ignore it.

    attachment_keys: list of {"key": storage_key, "filename": original_filename} — the display
    filename travels alongside the (ASCII-only) storage key rather than being derived from it
    (round 13.1 — see upload_task_attachment's docstring for why the key can't carry it)."""
    try:
        attachments = None
        if attachment_keys:
            attachments = []
            for a in attachment_keys:
                content = db.storage.from_("check-files").download(a["key"])
                attachments.append({"filename": a.get("filename") or a["key"].rsplit("/", 1)[-1], "content_b64": base64.b64encode(content).decode("ascii")})

        if channel == "email_outlook":
            capability = get_org_mailbox_capability(org_id)
            if not capability["connected"]:
                raise RuntimeError("הארגון אינו מחובר ל-Outlook")
            if not _outlook_daily_count_ok_and_increment(org_id):
                raise RuntimeError("מגבלת השליחה היומית של Outlook הארגוני הושגה להיום — ההודעה תישלח מחר או שכדאי לעבור זמנית לערוץ מייל אחר")
            if not created_by:
                raise RuntimeError("לא ידוע יוצר המשימה לצורך שליחה מ-Outlook")
            graph_client.send_mail_as_advisor(db, org_id, created_by, subject or "", body or "", recipient_email, attachments=attachments)
            return "outlook_pending", None
        elif channel == "whatsapp_twilio":
            whatsapp_twilio.send_whatsapp_message(org_id, recipient_phone, body or "", attachment_keys)
        else:  # email_resend (default)
            send_resend_email(recipient_email, subject or "", body or "", attachments=attachments)
        return "sent", None
    except Exception as exc:
        return "failed", str(exc)


def _mark_already_done_schools(db, task_id: str, matched_school_ids: list[str], send_targets: list[str]) -> None:
    """Schools excluded from send_targets because they already satisfied success at the moment
    send_targets was computed (see track_success filtering in create_task/
    _try_activate_scheduled_task) get no org_task_messages row at all — from TaskPanel.jsx's
    point of view that's indistinguishable from "not sent yet"/"missing contact". Persists the
    same skip_reason marker _queue_messages_for_schools already uses for its own skip cases, so
    get_task can surface it and the UI can show "אין צורך" instead of a bare "—"."""
    already_done = set(matched_school_ids) - set(send_targets)
    for school_id in already_done:
        try:
            db.table("org_task_school_notes").upsert(
                {"task_id": task_id, "school_id": school_id, "skip_reason": "already_done"},
                on_conflict="task_id,school_id",
            ).execute()
        except Exception as exc:
            logger.warning("_mark_already_done_schools: failed to persist skip_reason (non-fatal): %s", exc)


def _queue_messages_for_schools(db, task: dict, org_id: str, school_ids: list[str],
                                 scheduled_at: datetime | None = None) -> tuple[list[str], list[dict]]:
    """Builds and inserts org_task_messages rows. Returns
    ``(missing_contact_ids, broadcast_skipped)`` where:
    - ``missing_contact_ids`` — school_ids missing the contact info required by the chosen
      channel/recipient_role, or whose resolved recipient is excluded/opted-out (surfaced the
      same way as a missing contact so the manager sees them in the same "couldn't send" list).
    - ``broadcast_skipped`` — ``[{"id", "name"}]`` of schools deliberately skipped because they
      have an active broadcast-reminder exclusion (see _broadcast_skip_active). NOT an error —
      reported so the caller can tell the manager "sent, but N schools were excluded". A 'once'
      exclusion is consumed here (cleared) since this send is its "next broadcast".

    Round 7: only email_outlook (real 30/min Graph rate limit) and future-scheduled messages
    (scheduled_at set) actually wait for process-message-queue's cron drain — email_resend and
    whatsapp_twilio have no such constraint and are sent synchronously, right here, exactly
    like the existing "תיאום ישיר" feature already does. Previously every channel queued
    uniformly regardless of rate-limit need, which is what caused a single-school Resend
    message to sit "pending" for however long it took the next GitHub Actions cron tick to
    fire (up to ~5 min nominally, more in practice) instead of sending immediately."""
    message_config = task.get("message_config") or {}
    recipient_role = message_config.get("recipient_role")
    channel = message_config.get("channel")
    body_template = message_config.get("body_template") or ""
    needs_booking_link = "{booking_link}" in body_template
    academic_year = task.get("academic_year") or DEFAULT_ACADEMIC_YEAR

    schools = db.table("schools").select("*").in_("id", school_ids).execute().data or []
    year_rows = (
        db.table("school_year_admin_data")
        .select("school_id, client_status, meeting_duration_gefen, meeting_duration_current, meeting_duration_district")
        .eq("academic_year", academic_year).in_("school_id", school_ids).execute().data or []
    )
    client_status_map = {r["school_id"]: r.get("client_status") for r in year_rows}
    duration_map = {r["school_id"]: r for r in year_rows}

    # Only fetched when actually needed — a structured meeting-scheduling task's booking link
    # (round 5) needs each school's typed advisor assignments, batched once here rather than
    # per-school inside the loop below (Architecture Invariant #7 — no N+1 queries).
    is_meeting_task = task.get("is_meeting_task", False)
    meeting_conditions = _meeting_conditions_from_success_criteria(task) if is_meeting_task else []
    meeting_overrides = task.get("meeting_overrides") or {}
    advisor_map: dict[str, dict[str, list[str]]] = {}
    if needs_booking_link and meeting_conditions:
        for service_type, table_name in _TYPED_ADVISOR_TABLES.items():
            rows = db.table(table_name).select("school_id, advisor_id").in_("school_id", school_ids).execute().data or []
            for r in rows:
                advisor_map.setdefault(r["school_id"], {}).setdefault(service_type, []).append(r["advisor_id"])

    try:
        notes_map = _fetch_school_notes_map(db, task["id"], school_ids)
    except Exception as exc:
        logger.warning("_queue_messages_for_schools: exclusion lookup failed (non-fatal, treated as no exclusions): %s", exc)
        notes_map = {}

    # Cascade (round-2 redesign) — must match exactly what the pre-creation contact-resolution
    # flow (POST /tasks/contacts/check) showed the manager, or a school that resolved via a
    # fallback role there would fail to find a contact again here at actual send time.
    resolved = {s["id"]: _resolve_recipient_with_cascade(s, recipient_role, is_meeting_task, channel) for s in schools}
    resolved_recipients = {sid: r["recipient"] for sid, r in resolved.items()}
    try:
        opted_out_map = task_logic.opted_out_recipients(
            db, academic_year, {sid: r.get("email") for sid, r in resolved_recipients.items()},
        )
    except Exception as exc:
        logger.warning("_queue_messages_for_schools: opt-out lookup failed (non-fatal, treated as none opted out): %s", exc)
        opted_out_map = {}

    missing = []
    skipped_broadcast = []
    queue_rows = []
    _today = date.today()
    for school in schools:
        # Broadcast-reminder exclusion — deliberate manager choice, checked FIRST so an excluded
        # school is never counted as a missing-contact problem and never flips a "בעיות" flag.
        if _broadcast_skip_active(notes_map.get(school["id"]), _today):
            skipped_broadcast.append({"id": school["id"], "name": school.get("name")})
            continue
        recipient = resolved_recipients[school["id"]]
        if _channel_missing_contact(channel, recipient):
            missing.append(school["id"])
            try:
                db.table("org_task_school_notes").upsert(
                    {"task_id": task["id"], "school_id": school["id"], "skip_reason": "missing_contact"},
                    on_conflict="task_id,school_id",
                ).execute()
            except Exception as exc:
                logger.warning("_queue_messages_for_schools: failed to persist skip_reason (non-fatal): %s", exc)
            continue
        recipient_email = (recipient.get("email") or "").strip().lower()
        excluded_emails = {e.lower() for e in (notes_map.get(school["id"], {}).get("excluded_emails") or [])}
        if recipient_email and school["id"] in opted_out_map:
            missing.append(school["id"])
            try:
                db.table("org_task_school_notes").upsert(
                    {"task_id": task["id"], "school_id": school["id"], "skip_reason": "opted_out"},
                    on_conflict="task_id,school_id",
                ).execute()
            except Exception as exc:
                logger.warning("_queue_messages_for_schools: failed to persist skip_reason (non-fatal): %s", exc)
            continue
        if recipient_email and recipient_email in excluded_emails:
            missing.append(school["id"])
            continue

        opt_out_link = None
        if client_status_map.get(school["id"]) != "active" and recipient_email:
            opt_out_link = f"{APP_URL}/tasks/opt-out?email={recipient_email}&token={task_logic.make_optout_token(recipient_email)}"

        ranges_data = None
        if not needs_booking_link:
            booking_link, booking_token_id = None, None
        elif meeting_conditions:
            booking_link, booking_token_id, ranges_data = _build_meeting_booking_link(
                db, org_id, school, meeting_conditions,
                advisor_map.get(school["id"], {}), duration_map.get(school["id"], {}),
                meeting_overrides.get(school["id"]),
            )
            if booking_link is None:
                # Defensive hardening (round 6) — POST /tasks/meetings/check + the resolution
                # modal's blocking gate should make this unreachable in practice, but a school
                # must never receive a message with a dead {booking_link} placeholder.
                missing.append(school["id"])
                try:
                    db.table("org_task_school_notes").upsert(
                        {"task_id": task["id"], "school_id": school["id"], "skip_reason": "incomplete_meeting_config"},
                        on_conflict="task_id,school_id",
                    ).execute()
                except Exception as exc:
                    logger.warning("_queue_messages_for_schools: failed to persist skip_reason (non-fatal): %s", exc)
                continue
        else:
            booking_link, booking_token_id = _build_booking_link(db, org_id, task, school["id"]), None

        # Round 12: the manager's typed body_template is now always what's actually sent — a new
        # {meetings_list} placeholder is substituted with the same per-range date/duration/
        # participants block "תיאום ישיר" already renders (booking_logic.format_ranges_html/
        # format_ranges_text), instead of silently discarding body_template in favor of a fixed
        # HTML template (round 9's behavior). Email channels (Resend/Outlook) get the free text
        # wrapped in the branded HTML card via _wrap_email_html (which also fixes a pre-existing
        # bug where plain-text bodies were sent as literal unescaped HTML, so \n never rendered).
        # WhatsApp doesn't render HTML, so it keeps the plain-text path with plain bullets.
        advisor_names_value = ""
        if ranges_data:
            advisor_ids_for_names = list({aid for r in ranges_data for aid in (r.get("advisor_ids") or [])})
            if advisor_ids_for_names:
                try:
                    prof_rows = db.table("profiles").select("id, full_name").in_("id", advisor_ids_for_names).execute().data or []
                    names = [p["full_name"] for p in prof_rows if p.get("full_name")]
                    if names:
                        advisor_names_value = ", ".join(names)
                except Exception as exc:
                    logger.warning("_queue_messages_for_schools: advisor-name lookup failed (non-fatal): %s", exc)

        if channel == "whatsapp_twilio":
            meetings_list_text = format_ranges_text(ranges_data) if ranges_data else None
            body = _render_template(body_template, school, booking_link, opt_out_link,
                                     meetings_list=meetings_list_text, recipient_name=recipient.get("name"),
                                     advisor_names=advisor_names_value)
        else:
            meetings_list_html = format_ranges_html(ranges_data) if ranges_data else None
            rendered = _render_template(body_template, school, opt_out_link=opt_out_link,
                                         recipient_name=recipient.get("name"), advisor_names=advisor_names_value,
                                         keep_tokens=True)
            body = _wrap_email_html(rendered, booking_link, meetings_list_html, branded=(channel != "email_outlook"))

        row = {
            "task_id": task["id"],
            "school_id": school["id"],
            "recipient_name": recipient.get("name"),
            "recipient_email": recipient.get("email"),
            "recipient_phone": recipient.get("phone"),
            "recipient_role": recipient_role,
            "channel": channel,
            "subject": _render_template(message_config.get("subject") or "", school),
            "body": body,
            "attachment_keys": message_config.get("attachment_keys") or [],
            "status": "pending",
            "booking_token_id": booking_token_id,
        }
        if scheduled_at:
            row["scheduled_at"] = scheduled_at.isoformat()
        elif channel != "email_outlook":
            status, error = _send_message_now(
                db, org_id, school["id"], channel, row["subject"], row["body"],
                row.get("recipient_email"), row.get("recipient_phone"), row.get("attachment_keys"),
                created_by=task.get("created_by"),
            )
            row["status"] = status
            if status == "sent":
                row["sent_at"] = datetime.now(timezone.utc).isoformat()
            else:
                row["error"] = error
        queue_rows.append(row)

    if queue_rows:
        db.table("org_task_messages").insert(queue_rows).execute()

    # Consume 'once' exclusions — this broadcast was their "next" one. Cleared for every school
    # in this batch that carried mode='once' (whether or not it ended up skipped above — a
    # done/not-targeted school simply won't be in `school_ids`). Non-fatal.
    once_ids = [s["id"] for s in schools if (notes_map.get(s["id"]) or {}).get("broadcast_skip_mode") == "once"]
    if once_ids:
        try:
            db.table("org_task_school_notes").update(
                {"broadcast_skip_mode": None, "broadcast_skip_until": None},
            ).eq("task_id", task["id"]).in_("school_id", once_ids).execute()
        except Exception as exc:
            logger.warning("_queue_messages_for_schools: failed to consume 'once' exclusions (non-fatal): %s", exc)

    return missing, skipped_broadcast


@router.post("/{task_id}/send")
def send_task_bulk(task_id: str, body: SendIn, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    task = _get_task_or_404(db, task_id, user["org_id"])
    scheduled_at = _parse_dt(body.scheduled_at)

    if body.only_school_ids:
        target_ids = body.only_school_ids
    else:
        progress = task_logic.compute_task_progress(user["org_id"], task)
        target_ids = [r["school_id"] for r in progress["schools"] if not r["done"]]

    if not target_ids:
        return {"queued": 0, "missing_contact_school_ids": []}

    channel = (task.get("message_config") or {}).get("channel")
    if channel == "email_outlook" and len(target_ids) > OUTLOOK_SEND_WARN_THRESHOLD and not body.confirm_outlook_limit:
        raise HTTPException(status_code=409, detail={
            "outlook_limit_exceeded": True,
            "message": (
                f"פעולה זו תשלח {len(target_ids)} מיילים דרך Outlook הארגוני, מעל לסף האזהרה "
                f"המוגדר ({OUTLOOK_SEND_WARN_THRESHOLD}). מגבלות Microsoft בפועל: עד 30 הודעות "
                f"בדקה ועד 10,000 נמענים ביום (משותף לכל השליחות בארגון, לא רק למשימה זו)."
            ),
            "count": len(target_ids),
            "threshold": OUTLOOK_SEND_WARN_THRESHOLD,
        })

    missing, broadcast_skipped = _queue_messages_for_schools(db, task, user["org_id"], target_ids, scheduled_at)
    queued = len(target_ids) - len(missing) - len(broadcast_skipped)
    if task.get("needs_outlook_confirmation"):
        # Manual confirmation via this endpoint (the banner's "אשר שליחה" button, see
        # process_scheduled_tasks) resolves the flag regardless of channel — a manager who
        # explicitly clicked send has made their decision either way.
        db.table("org_tasks").update({"needs_outlook_confirmation": False}).eq("id", task_id).execute()
    task_logic.recompute_task_status_and_cache(db, user["org_id"], task)
    if missing:
        raise HTTPException(status_code=409, detail={
            "message": "לחלק מבתי הספר חסרים פרטי קשר מתאימים לערוץ שנבחר",
            "missing_contact_school_ids": missing,
            "broadcast_skipped_schools": broadcast_skipped,
            "queued": queued,
        })
    return {"queued": queued, "missing_contact_school_ids": [], "broadcast_skipped_schools": broadcast_skipped}


@router.post("/{task_id}/schools/{school_id}/send")
def send_task_single(task_id: str, school_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    task = _get_task_or_404(db, task_id, user["org_id"])

    # A manual per-row send respects an active broadcast-reminder exclusion (does NOT consume a
    # 'once' — consumption is tied to an actual broadcast send). The manager must clear the
    # exclusion in the "החרגות" cell first.
    skip_note = (
        db.table("org_task_school_notes").select("broadcast_skip_mode, broadcast_skip_until")
        .eq("task_id", task_id).eq("school_id", school_id).execute().data or []
    )
    if skip_note and _broadcast_skip_active(skip_note[0], date.today()):
        raise HTTPException(status_code=409, detail={
            "message": "בית הספר מוחרג מתזכורות גורפות — יש להסיר את ההחרגה כדי לשלוח לו תזכורת",
            "broadcast_skipped": True,
            "mode": skip_note[0].get("broadcast_skip_mode"),
        })

    missing, _skipped_broadcast = _queue_messages_for_schools(db, task, user["org_id"], [school_id])
    task_logic.recompute_task_status_and_cache(db, user["org_id"], task)
    if missing:
        note_rows = (
            db.table("org_task_school_notes").select("skip_reason")
            .eq("task_id", task_id).eq("school_id", school_id).execute().data or []
        )
        if note_rows and note_rows[0].get("skip_reason") == "opted_out":
            raise HTTPException(status_code=409, detail={
                "message": "בית הספר ביקש הסרה מרשימת התפוצה — ההודעה לא נשלחה, עד ששדה 'סטטוס לקוח' שלו יהפוך ל'פעיל'",
                "opted_out": True,
            })
        raise HTTPException(status_code=409, detail={
            "message": "לבית הספר חסרים פרטי קשר מתאימים לערוץ שנבחר",
            "missing_contact_school_ids": missing,
        })
    return {"ok": True}


class ContactInfoIn(BaseModel):
    email: str | None = None
    phone: str | None = None


@router.put("/{task_id}/schools/{school_id}/contact-info")
def put_school_contact_info(task_id: str, school_id: str, body: ContactInfoIn, user: Annotated[dict, Depends(get_current_user)]):
    """Fills in the email/phone for whichever contact role the task's recipient_role actually
    resolves to on this school — needed because 'meeting_coordinator' is an indirection (it
    points at a *different* fixed role per school), so the missing-contact modal can't know
    which schools.<role>_email/<role>_phone column to write to without asking the backend."""
    _require_manager(user)
    db = get_admin_client()
    task = _get_task_or_404(db, task_id, user["org_id"])
    school_rows = db.table("schools").select("*").eq("id", school_id).eq("org_id", user["org_id"]).execute().data or []
    if not school_rows:
        raise HTTPException(status_code=404, detail="בית הספר לא נמצא")
    school = school_rows[0]

    recipient_role = (task.get("message_config") or {}).get("recipient_role")
    resolved_role = school.get("meeting_coordinator") if recipient_role == "meeting_coordinator" else recipient_role
    if resolved_role not in _ROLE_CONTACT_FIELDS:
        raise HTTPException(status_code=422, detail="לא ניתן לקבוע לאיזה איש קשר לשייך את הפרטים — יש להגדיר קודם 'אחראי/ת לתיאום פגישות' בפרטי בית הספר")

    _name_f, email_f, phone_f = _ROLE_CONTACT_FIELDS[resolved_role]
    patch = {}
    if body.email:
        patch[email_f] = body.email
    if body.phone:
        patch[phone_f] = body.phone
    if patch:
        db.table("schools").update(patch).eq("id", school_id).eq("org_id", user["org_id"]).execute()
    return {"ok": True, "resolved_role": resolved_role}


class TaskSchoolNoteIn(BaseModel):
    note: str | None = None
    excluded_emails: list[str] | None = None


@router.put("/{task_id}/schools/{school_id}/note")
def put_task_school_note(task_id: str, school_id: str, body: TaskSchoolNoteIn, user: Annotated[dict, Depends(get_current_user)]):
    """Upserts per-(task, school) free-text notes and/or excluded-email addresses. Both fields
    share one row (org_task_school_notes) since they're both small per-task-per-school
    manager-entered annotations — only the fields provided in the body are updated."""
    _require_manager(user)
    db = get_admin_client()
    _get_task_or_404(db, task_id, user["org_id"])
    patch = {"task_id": task_id, "school_id": school_id, "updated_by": user["id"], "updated_at": "now()"}
    if body.note is not None:
        patch["note"] = body.note
    if body.excluded_emails is not None:
        patch["excluded_emails"] = [e.strip().lower() for e in body.excluded_emails if e.strip()]
    db.table("org_task_school_notes").upsert(patch, on_conflict="task_id,school_id").execute()
    return {"ok": True}


class BroadcastSkipIn(BaseModel):
    mode: str | None = None   # 'once' | 'until' | 'forever' | None (= clear the exclusion)
    until: str | None = None  # 'YYYY-MM-DD' — required when mode == 'until'


@router.put("/{task_id}/schools/{school_id}/broadcast-skip")
def put_broadcast_skip(task_id: str, school_id: str, body: BroadcastSkipIn, user: Annotated[dict, Depends(get_current_user)]):
    """Sets (or clears, mode=None) a per-(task, school) exclusion from broadcast reminder sends.
    Separate from PUT .../note so a NULL can be written explicitly. Purely about reminder
    delivery — never affects task progress/completion."""
    _require_manager(user)
    if body.mode not in (None, "once", "until", "forever"):
        raise HTTPException(status_code=422, detail="ערך לא חוקי")
    until_val = None
    if body.mode == "until":
        try:
            until_val = date.fromisoformat((body.until or "")[:10])
        except ValueError:
            raise HTTPException(status_code=422, detail="יש לבחור תאריך תקין")
        if until_val < date.today():
            raise HTTPException(status_code=422, detail="התאריך שנבחר כבר עבר")
    db = get_admin_client()
    _get_task_or_404(db, task_id, user["org_id"])
    db.table("org_task_school_notes").upsert(
        {
            "task_id": task_id, "school_id": school_id,
            "broadcast_skip_mode": body.mode,
            "broadcast_skip_until": until_val.isoformat() if until_val else None,
            "updated_by": user["id"], "updated_at": "now()",
        },
        on_conflict="task_id,school_id",
    ).execute()
    return {"ok": True}


class TaskSchoolFilterIn(BaseModel):
    criteria: ConditionsIn


@router.post("/{task_id}/schools/filter")
def filter_task_schools(task_id: str, body: TaskSchoolFilterIn, user: Annotated[dict, Depends(get_current_user)]):
    """Backs the redesigned Tasks table's in-row "סנן בתי ספר" — full parity with every field
    category the audience-criteria builder supports (school/contact/financial/goals/control-
    letters/meetings), which is why this needs a real server-side evaluation rather than a
    client-side filter over already-fetched (partial) row data. Reuses find_matching_schools
    exactly as-is (same evaluate_tree pass over every org school) and intersects the result with
    this task's own matched_school_ids — no duplicated evaluation logic, no per-school-in-a-loop
    queries beyond what find_matching_schools already does for a single compute_task_progress-
    equivalent call."""
    _require_manager(user)
    db = get_admin_client()
    task = _get_task_or_404(db, task_id, user["org_id"])
    matched_ids = set(task.get("matched_school_ids") or [])
    if not matched_ids:
        return {"school_ids": []}
    academic_year = task.get("academic_year") or DEFAULT_ACADEMIC_YEAR
    all_matches = task_logic.find_matching_schools(user["org_id"], body.criteria.model_dump(), academic_year)
    return {"school_ids": [m["school_id"] for m in all_matches if m["school_id"] in matched_ids]}


@router.post("/{task_id}/pin")
def pin_task(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Personal per-manager pin (task_pins) — floats a task to the very top of the redesigned
    Tasks table regardless of its status tier, visible only to the manager who pinned it."""
    _require_manager(user)
    db = get_admin_client()
    _get_task_or_404(db, task_id, user["org_id"])
    db.table("task_pins").upsert({"task_id": task_id, "user_id": user["id"]}, on_conflict="task_id,user_id").execute()
    return {"ok": True}


@router.delete("/{task_id}/pin")
def unpin_task(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    db = get_admin_client()
    db.table("task_pins").delete().eq("task_id", task_id).eq("user_id", user["id"]).execute()
    return {"ok": True}


@router.post("/attachments/upload")
async def upload_task_attachment(
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    """Round 13 — deliberately NOT scoped to an existing task_id (unlike the old
    /{task_id}/attachments this replaces, the only caller of which was TaskCreateWizard.jsx).
    Task creation queues/sends the first wave of messages immediately (round 2), so the wizard
    must be able to upload attachments and get real storage_keys *before* POST /tasks/ — otherwise
    that first wave goes out with an empty attachment_keys list.

    Round 13.1 fix: the storage_key itself must stay ASCII-only — Supabase Storage rejects keys
    containing non-ASCII characters (confirmed directly: a Hebrew filename raised
    storage3.exceptions.StorageApiError 400 "Invalid key"), which is exactly what a filename in
    Hebrew (routine in this app) produces. The real filename is returned separately in the
    response and threaded through as message_config.attachment_keys' own {key, filename} shape
    (see _send_message_now) instead of being embedded in the storage path."""
    _require_manager(user)
    db = get_admin_client()

    run_dir = Path(tempfile.mkdtemp(prefix="task_attachment_"))
    try:
        suffix = Path(file.filename or "").suffix
        dest = run_dir / f"attachment{suffix}"
        dest.write_bytes(await file.read())
        storage_key = f"tasks/pending/{secrets.token_hex(8)}{suffix}"
        db.storage.from_("check-files").upload(storage_key, dest.read_bytes())
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

    return {"storage_key": storage_key, "filename": file.filename}


@router.post("/process-message-queue")
def process_message_queue(request: Request):
    """Cron-triggered (same GitHub Actions convention as process-booking-email-queue in
    schools_router.py:2432) — drains a small batch of pending org_task_messages rows."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=401, detail="unauthorized")

    db = get_admin_client()
    BATCH_SIZE = 20
    rows = (
        db.table("org_task_messages")
        .select("*")
        .eq("status", "pending")
        .lte("scheduled_at", datetime.now(timezone.utc).isoformat())
        .order("created_at")
        .limit(BATCH_SIZE)
        .execute()
        .data or []
    )

    sent, failed, skipped = 0, 0, 0
    touched_task_ids: set[str] = set()  # recomputed once each at the end (not per-row — see below)
    last_outlook_send_at = None  # monotonic clock — paces real email_outlook sends only (see
    # OUTLOOK_SEND_MIN_INTERVAL_SECONDS); other channels in the same batch are unaffected.
    for row in rows:
        try:
            task = db.table("org_tasks").select("org_id, created_by, academic_year").eq("id", row["task_id"]).execute().data
            org_id = task[0]["org_id"] if task else None
            created_by = task[0]["created_by"] if task else None
            task_academic_year = (task[0].get("academic_year") if task else None) or DEFAULT_ACADEMIC_YEAR
            if not org_id:
                raise ValueError(f"task {row['task_id']} not found for queued message {row['id']}")
            touched_task_ids.add(row["task_id"])

            # Belt-and-suspenders re-check (bugs #10/#13): a row can have been queued before an
            # exclusion/opt-out was recorded, or the school's client_status could have changed
            # since queue-time, so this is checked again right before the actual send, not only
            # at queue-time in _queue_messages_for_schools.

            # Broadcast-reminder exclusion — a 'forever'/'until' exclusion may have been added
            # after this row was queued, or an 'until' date may have "activated" between queue
            # time and now (never 'once' — those never enter the queue). No email needed.
            skip_note_rows = (
                db.table("org_task_school_notes").select("broadcast_skip_mode, broadcast_skip_until")
                .eq("task_id", row["task_id"]).eq("school_id", row["school_id"]).execute().data or []
            )
            if skip_note_rows and _broadcast_skip_active(skip_note_rows[0], date.today()):
                db.table("org_task_messages").update({"status": "skipped"}).eq("id", row["id"]).execute()
                skipped += 1
                continue

            recipient_email = (row.get("recipient_email") or "").strip().lower()
            if recipient_email:
                is_opted_out = bool(task_logic.opted_out_recipients(
                    db, task_academic_year, {row["school_id"]: recipient_email},
                ))
                note_rows = (
                    db.table("org_task_school_notes").select("excluded_emails")
                    .eq("task_id", row["task_id"]).eq("school_id", row["school_id"]).execute().data or []
                )
                excluded = {e.lower() for e in (note_rows[0].get("excluded_emails") or [])} if note_rows else set()
                if is_opted_out or recipient_email in excluded:
                    db.table("org_task_messages").update({"status": "skipped"}).eq("id", row["id"]).execute()
                    skipped += 1
                    continue

            if row["channel"] == "email_outlook" and last_outlook_send_at is not None:
                # Real pacing (not just the incidental cron-cadence/batch-size protection) —
                # measured from actual elapsed time so Graph's own response latency counts
                # toward the interval instead of stacking on top of a blind sleep.
                wait = OUTLOOK_SEND_MIN_INTERVAL_SECONDS - (time.monotonic() - last_outlook_send_at)
                if wait > 0:
                    time.sleep(wait)

            status, error = _send_message_now(
                db, org_id, row["school_id"], row["channel"], row["subject"], row["body"],
                row.get("recipient_email"), row.get("recipient_phone"), row.get("attachment_keys"),
                created_by=created_by,
            )
            if row["channel"] == "email_outlook":
                last_outlook_send_at = time.monotonic()
            if status == "sent":
                db.table("org_task_messages").update({"status": "sent", "sent_at": "now()"}).eq("id", row["id"]).execute()
                sent += 1
            elif status == "outlook_pending":
                # Round 15 — Graph accepted the send, but that's not confirmed delivery (a bounce
                # can still arrive minutes later). Held in this interim status until the
                # grace-window check below either confirms it or falls back to Resend.
                db.table("org_task_messages").update({"status": "outlook_pending", "outlook_sent_at": "now()"}).eq("id", row["id"]).execute()
                sent += 1
            else:
                logger.warning("process_message_queue: failed to send row %s: %s", row["id"], error)
                db.table("org_task_messages").update({"status": "failed", "error": error}).eq("id", row["id"]).execute()
                failed += 1
        except Exception as exc:
            logger.warning("process_message_queue: failed to send row %s: %s", row["id"], exc)
            try:
                db.table("org_task_messages").update({"status": "failed", "error": str(exc)}).eq("id", row["id"]).execute()
            except Exception as log_exc:
                logger.error("process_message_queue: failed to mark row %s failed: %s", row["id"], log_exc)
            failed += 1

    # Round 15 — second pass: confirm or fall back any Outlook sends that cleared their grace
    # window (find_bounce_for_subject relies on a real NDR that empirically arrives within ~1
    # minute of send, so 5 minutes leaves generous margin without materially delaying
    # confirmation for the common case).
    OUTLOOK_BOUNCE_GRACE_MINUTES = 5
    confirmed, fell_back = 0, 0
    outlook_pending_rows = (
        db.table("org_task_messages")
        .select("*")
        .eq("status", "outlook_pending")
        .lte("outlook_sent_at", (datetime.now(timezone.utc) - timedelta(minutes=OUTLOOK_BOUNCE_GRACE_MINUTES)).isoformat())
        .order("outlook_sent_at")
        .limit(BATCH_SIZE)
        .execute()
        .data or []
    )
    for row in outlook_pending_rows:
        try:
            task = db.table("org_tasks").select("org_id, created_by").eq("id", row["task_id"]).execute().data
            org_id = task[0]["org_id"] if task else None
            created_by = task[0]["created_by"] if task else None
            if not org_id:
                raise ValueError(f"task {row['task_id']} not found for outlook_pending message {row['id']}")
            touched_task_ids.add(row["task_id"])

            if not created_by:
                # Can't check for a bounce without a mailbox to inspect — best-effort finalize.
                db.table("org_task_messages").update({"status": "sent", "sent_at": "now()"}).eq("id", row["id"]).execute()
                confirmed += 1
                continue

            # The message was actually sent from the task creator's own mailbox (see
            # _send_message_now) — so that's whose Inbox the NDR would land in, not the
            # school's assigned advisor's.
            bounced = graph_client.find_bounce_for_subject(db, org_id, created_by, row.get("subject") or "", row["outlook_sent_at"])
            if bounced is None:
                continue  # couldn't check (missing Mail.Read consent / transient error) — retried next tick

            if not bounced:
                db.table("org_task_messages").update({"status": "sent", "sent_at": "now()"}).eq("id", row["id"]).execute()
                confirmed += 1
                continue

            attachments = None
            if row.get("attachment_keys"):
                attachments = []
                for a in row["attachment_keys"]:
                    content = db.storage.from_("check-files").download(a["key"])
                    attachments.append({"filename": a.get("filename") or a["key"].rsplit("/", 1)[-1], "content_b64": base64.b64encode(content).decode("ascii")})
            try:
                send_resend_email(row.get("recipient_email"), row.get("subject") or "", row.get("body") or "", attachments=attachments)
                db.table("org_task_messages").update({"status": "sent", "sent_at": "now()", "fallback_used": True}).eq("id", row["id"]).execute()
                fell_back += 1
            except Exception as exc:
                logger.warning("process_message_queue: Outlook bounced and Resend fallback also failed for row %s: %s", row["id"], exc)
                db.table("org_task_messages").update({
                    "status": "failed", "error": f"השליחה נכשלה גם דרך Outlook (bounce) וגם דרך הגיבוי (Resend): {exc}",
                }).eq("id", row["id"]).execute()
                failed += 1
        except Exception as exc:
            logger.warning("process_message_queue: outlook_pending confirmation failed for row %s (non-fatal, retried next tick): %s", row["id"], exc)

    # Recompute status/cache once PER DISTINCT TASK touched in this tick (never per message row —
    # Architecture Invariant #7) — bounded by BATCH_SIZE (<=20 rows across both passes, so at
    # most 20 distinct tasks, typically far fewer).
    if touched_task_ids:
        try:
            touched_tasks = db.table("org_tasks").select("*").in_("id", list(touched_task_ids)).execute().data or []
            for t in touched_tasks:
                task_logic.recompute_task_status_and_cache(db, t["org_id"], t)
        except Exception as exc:
            logger.warning("process_message_queue: batch status/cache recompute failed (non-fatal): %s", exc)

    return {
        "ok": True, "sent": sent, "failed": failed, "skipped": skipped, "batch_size": len(rows),
        "outlook_confirmed": confirmed, "outlook_fallback": fell_back,
    }


def _hold_back_scheduled_task(db, task: dict, meeting_wording: bool) -> dict:
    """Shared false→true/true→false transition + one-time notification logic for both problem
    kinds (meeting-requirement problems and general-task contact problems) — see
    _try_activate_scheduled_task. Reuses the same has_meeting_send_problems boolean column for
    both task types: a given org_tasks row is always exclusively meeting or general, never both,
    so there's no collision risk in sharing one column rather than adding a second."""
    if not task.get("has_meeting_send_problems"):
        db.table("org_tasks").update({"has_meeting_send_problems": True}).eq("id", task["id"]).execute()
        title = (
            "קיימות בעיות שמונעות ביצוע משימה לקביעת פגישות שתוזמנה להיום. אנא היכנס בכדי לפתור אותן."
            if meeting_wording else
            "קיימות בעיות שמונעות ביצוע משימה שתוזמנה להיום. אנא היכנס בכדי לפתור אותן."
        )
        try:
            _schools_router._create_notifications(db, [{
                "recipient_id": task["created_by"],
                "type": "task_send_problems",
                "data": {"title": title, "task_id": task["id"], "task_name": task.get("name")},
            }])
        except Exception as exc:
            logger.warning("_try_activate_scheduled_task: notification failed (non-fatal) for task %s: %s", task["id"], exc)
    return {"activated": False, "has_problems": True}


def _try_activate_scheduled_task(db, task: dict) -> dict:
    """Attempts to activate a single 'scheduled' task whose scheduled_for has arrived (also
    reused for a manager-triggered immediate retry — see /retry-now below). Before activating,
    re-validates whatever would block the send against CURRENT data (not the stale snapshot the
    wizard checked at creation time — this is the whole point: something can have broken in
    between): for meeting-scheduling tasks, coordinator/participant/meeting-default/advisor-
    access problems (_check_meeting_problems); for general tasks, recipient-contact
    resolvability (_check_contact_problems) — the exact same cascade /tasks/contacts/check
    already uses pre-creation, just re-run live. If any matched school currently has a problem,
    activation is held entirely: status stays 'scheduled', nothing is sent to ANY school (not
    just the problem ones), and — only on the false→true transition of has_meeting_send_problems,
    to avoid renotifying every 15-minute tick — a single notification is created for the task's
    creator. Once the underlying data is fixed (via the dedicated 'בעיות' screen for meeting
    tasks, a direct fix on the school card, or /retry-now), the very next call naturally proceeds
    past this check and sends — no separate 'resolved' bookkeeping needed, it self-heals by
    construction. Returns {"activated": bool, "has_problems": bool}."""
    org_id = task["org_id"]

    if task.get("is_meeting_task"):
        meeting_requirements = _meeting_requirements_from_task(task)
        channel = (task.get("message_config") or {}).get("channel") or "email_resend"
        academic_year = task.get("academic_year") or DEFAULT_ACADEMIC_YEAR
        check = _check_meeting_problems(
            org_id, task.get("criteria") or {}, task.get("manual_school_ids"),
            meeting_requirements, channel, academic_year,
        )
        # opted_out-only entries are informational (see _check_meeting_problems) and must never
        # hold back activation on their own — only real coordinator/participants/meeting_defaults/
        # advisor_access problems do.
        if any(s["coordinator"] or s["participants"] or s["meeting_defaults"] or s["advisor_access"] for s in check["schools"]):
            return _hold_back_scheduled_task(db, task, meeting_wording=True)
        elif task.get("has_meeting_send_problems"):
            db.table("org_tasks").update({"has_meeting_send_problems": False}).eq("id", task["id"]).execute()
    else:
        message_config = task.get("message_config") or {}
        recipient_role = message_config.get("recipient_role")
        channel = message_config.get("channel") or "email_resend"
        academic_year = task.get("academic_year") or DEFAULT_ACADEMIC_YEAR
        check = _check_contact_problems(
            org_id, task.get("criteria") or {}, task.get("manual_school_ids"), recipient_role,
            channel, False, None, academic_year,
        )
        if any(not s["has_contact"] for s in check["schools"]):
            return _hold_back_scheduled_task(db, task, meeting_wording=False)
        elif task.get("has_meeting_send_problems"):
            db.table("org_tasks").update({"has_meeting_send_problems": False}).eq("id", task["id"]).execute()

    matched = task_logic.find_matching_schools(
        org_id, task.get("criteria") or {}, task.get("academic_year") or DEFAULT_ACADEMIC_YEAR,
        manual_school_ids=task.get("manual_school_ids"),
    )
    matched_school_ids = [m["school_id"] for m in matched]
    db.table("org_tasks").update({
        "matched_school_ids": matched_school_ids,
        "status": "active",
    }).eq("id", task["id"]).execute()
    task["matched_school_ids"] = matched_school_ids
    task["status"] = "active"

    if matched_school_ids:
        if task.get("track_success", True):
            progress = task_logic.compute_task_progress(org_id, task)
            send_targets = [r["school_id"] for r in progress["schools"] if not r["done"]]
            _mark_already_done_schools(db, task["id"], matched_school_ids, send_targets)
        else:
            send_targets = matched_school_ids

        channel = (task.get("message_config") or {}).get("channel")
        if send_targets and channel == "email_outlook" and len(send_targets) > OUTLOOK_SEND_WARN_THRESHOLD:
            db.table("org_tasks").update({"needs_outlook_confirmation": True}).eq("id", task["id"]).execute()
        elif send_targets:
            _queue_messages_for_schools(db, task, org_id, send_targets)

    task_logic.recompute_task_status_and_cache(db, org_id, task)
    return {"activated": True, "has_problems": False}


@router.post("/process-scheduled-tasks")
def process_scheduled_tasks(request: Request):
    """Cron-triggered (same convention as process-message-queue). Evaluates tasks that were
    created with a future scheduled_for — computing the school-match snapshot now (on/after the
    scheduled date) instead of at creation time, then flipping them to 'active' so they behave
    like any other task from that point on. Round-2 redesign: also immediately queues the
    first-wave messages on activation (same "creation = send" principle as create_task), since
    there's no separate manual click for a scheduled task either.

    No user is present at cron time, so the Outlook-limit interactive 3-way choice (continue/
    switch channel/cancel) is impossible here — if activation would exceed the warning
    threshold on an Outlook-channel task, messages are deliberately left unqueued and
    `needs_outlook_confirmation` is set so TaskPanel.jsx can surface a banner requiring a
    manager to explicitly confirm before anything sends (see plan decision #1).

    Per-task activation logic lives in _try_activate_scheduled_task (also reused by the
    manager-triggered /retry-now endpoint for an immediate re-attempt after fixing a problem
    that was blocking a scheduled meeting-task's send — see that function's docstring)."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=401, detail="unauthorized")

    db = get_admin_client()

    try:
        _schools_router.process_temp_advisor_access(db)
    except Exception as exc:
        logger.warning("process_scheduled_tasks: temp advisor access processing failed (non-fatal): %s", exc)

    # Housekeeping — clear expired 'until' broadcast-reminder exclusions. Display/send already
    # treat them as inactive live (via _broadcast_skip_active), so this only keeps the data tidy
    # and covers the "check the date once a day" requirement. Non-fatal.
    try:
        db.table("org_task_school_notes").update(
            {"broadcast_skip_mode": None, "broadcast_skip_until": None},
        ).eq("broadcast_skip_mode", "until").lt("broadcast_skip_until", date.today().isoformat()).execute()
    except Exception as exc:
        logger.warning("process_scheduled_tasks: expired broadcast-skip cleanup failed (non-fatal): %s", exc)

    now_iso = datetime.now(timezone.utc).isoformat()
    rows = (
        db.table("org_tasks")
        .select("*")
        .eq("status", "scheduled")
        .lte("scheduled_for", now_iso)
        .execute()
        .data or []
    )

    evaluated = 0
    for task in rows:
        try:
            _try_activate_scheduled_task(db, task)
            evaluated += 1
        except Exception as exc:
            logger.warning("process_scheduled_tasks: failed to evaluate task %s: %s", task["id"], exc)

    return {"ok": True, "evaluated": evaluated, "batch_size": len(rows)}


@router.post("/{task_id}/retry-now")
def retry_task_now(task_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Manual immediate retry for a scheduled task currently held back by has_meeting_send_
    problems — called right after the manager fixes the underlying issue via the 'בעיות' screen,
    so they don't have to wait for the next cron tick even though the scheduled time already
    passed. No-op (returns current state) if the task isn't in 'scheduled' status."""
    _require_manager(user)
    db = get_admin_client()
    task_row = db.table("org_tasks").select("*").eq("id", task_id).eq("org_id", user["org_id"]).execute().data
    if not task_row:
        raise HTTPException(status_code=404, detail="משימה לא נמצאה")
    task = task_row[0]
    if task["status"] != "scheduled":
        task_logic.recompute_task_status_and_cache(db, user["org_id"], task)
        return {"activated": task["status"] == "active", "has_problems": False}
    return _try_activate_scheduled_task(db, task)
