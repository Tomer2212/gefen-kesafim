import io
import json
import logging
import os
import re
import secrets
import smtplib
import time
from concurrent.futures import ThreadPoolExecutor, wait as futures_wait, FIRST_COMPLETED
from datetime import datetime, timedelta, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
from typing import Annotated

import httpx
from bidi.algorithm import get_display
from fastapi import APIRouter, Depends, HTTPException, Request, UploadFile, File, Form
from fastapi.responses import Response
from openpyxl import load_workbook, Workbook
from pydantic import BaseModel
from reportlab.lib import colors as rl_colors
from reportlab.lib.pagesizes import A4, landscape as rl_landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

import graph_client
from academic_years import ACADEMIC_YEARS, DEFAULT_ACADEMIC_YEAR
from auth import get_current_user, invalidate_profile_cache
from email_resend import send_resend_email
from meeting_upload_logic import build_upload_checklist, get_or_create_upload_token
from supabase_client import get_admin_client, reset_admin_client

logger = logging.getLogger(__name__)
router = APIRouter()

# ---------------------------------------------------------------------------
# Schools list PDF export helpers
# ---------------------------------------------------------------------------

_EXPORT_FONTS_DIR = Path(__file__).parent.parent / "logic" / "fonts"
_EXPORT_FONT = "SchExportHeb"
_EXPORT_FONT_BOLD = "SchExportHebBold"
_EXPORT_FONTS_OK = False


def _ensure_export_fonts() -> tuple[str, str]:
    global _EXPORT_FONTS_OK, _EXPORT_FONT, _EXPORT_FONT_BOLD
    if _EXPORT_FONTS_OK:
        return _EXPORT_FONT, _EXPORT_FONT_BOLD
    try:
        pdfmetrics.registerFont(TTFont(_EXPORT_FONT, str(_EXPORT_FONTS_DIR / "NotoSansHebrew-Regular.ttf")))
        pdfmetrics.registerFont(TTFont(_EXPORT_FONT_BOLD, str(_EXPORT_FONTS_DIR / "NotoSansHebrew-Bold.ttf")))
    except Exception:
        _EXPORT_FONT = "Helvetica"
        _EXPORT_FONT_BOLD = "Helvetica-Bold"
    _EXPORT_FONTS_OK = True
    return _EXPORT_FONT, _EXPORT_FONT_BOLD


def _he(text: str) -> str:
    s = str(text) if text else ""
    if any("א" <= c <= "ת" for c in s):
        return get_display(s)
    return s


def _build_schools_pdf(title: str, headers: list[str], rows: list[list[str]]) -> bytes:
    font, font_bold = _ensure_export_fonts()
    buf = io.BytesIO()
    page_size = rl_landscape(A4) if len(headers) > 4 else A4
    doc = SimpleDocTemplate(
        buf, pagesize=page_size,
        rightMargin=1.5 * cm, leftMargin=1.5 * cm,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm,
    )
    page_w = page_size[0] - 3 * cm
    col_w = page_w / max(len(headers), 1)
    t_headers = [_he(h) for h in reversed(headers)]
    t_rows = [[_he(str(c)) for c in reversed(row)] for row in rows]
    tbl = Table([t_headers, *t_rows], colWidths=[col_w] * len(headers), repeatRows=1)
    tbl.setStyle(TableStyle([
        ("FONTNAME",       (0, 0), (-1, 0),  font_bold),
        ("FONTNAME",       (0, 1), (-1, -1), font),
        ("FONTSIZE",       (0, 0), (-1, -1), 9),
        ("BACKGROUND",     (0, 0), (-1, 0),  rl_colors.HexColor("#1e3a5f")),
        ("TEXTCOLOR",      (0, 0), (-1, 0),  rl_colors.white),
        ("ALIGN",          (0, 0), (-1, -1), "RIGHT"),
        ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#f8fafc")]),
        ("GRID",           (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#e2e8f0")),
        ("TOPPADDING",     (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING",  (0, 0), (-1, -1), 6),
        ("LEFTPADDING",    (0, 0), (-1, -1), 8),
        ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
    ]))
    title_style = ParagraphStyle("sch_title", fontName=font_bold, fontSize=13, alignment=2)
    doc.build([Paragraph(_he(title), title_style), Spacer(1, 0.4 * cm), tbl])
    return buf.getvalue()

# ---------------------------------------------------------------------------

DIVISION_LABELS = {
    "tikkon": "חטיבה עליונה",
    "beinayim": "חטיבת ביניים",
    "yesodi": "יסודי",
    "other": "אחר",
}

# Fixed goal definitions set by the Ministry of Education — identical for every school,
# every division ("all"), and evaluated separately per budget (תקציב) selected in the UI.
# "kind" is "planning"/"reporting" (shown in the עמידה-tracked "יעדים" table, with goal_type/goal_number
# stored per-school in school_goals) or "date" (shown read-only in "תאריכים חשובים", never saved/toggled).
# Order here is chronological (target_date ascending) — keep it that way.
GOAL_DEFINITIONS = [
    {"key": "training_days_scholarships_deadline",  "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לבקשת ימי הדרכה, בקשות לתוכניות משרדיות, פעימה ראשונה לבקשת שעות בודדות והמרת שכל\"מ. כמו כן, מועד אחרון לתכנון מלגות למסעות לפולין", "target_date": "2025-06-30"},
    {"key": "menagim_activation_start",             "division": "all", "kind": "date",      "goal_number": None, "label": "תחילת הפעלה של מענים שנבחרו (כמו תוכן דיגיטלי, תוכנות ניהול והעסקת כוח אדם). בנוסף, ייפתח חלון נוסף להגשת בקשה להמרת שכל\"מ", "target_date": "2025-09-01"},
    {"key": "planning_40_sep",                      "division": "all", "kind": "planning",  "goal_number": 40, "label": "יעד תכנון: לפחות 40% מהתקציב.",                         "target_date": "2025-09-15", "date_overrides": {"תשפ\"ז": "2026-09-01"}},
    {"key": "bank_account_forms_nov",                "division": "all", "kind": "date",      "goal_number": None, "label": "מועד הגשת טפסים תקינים לפתיחת חשבון בנק בית ספרי (למוסדות חדשים), לצורך קבלת מקדמה ראשונה בחודש נובמבר", "target_date": "2025-09-15"},
    {"key": "digital_content_cancel_deadline",       "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לביטול רכישת ספקי תוכן דיגיטלי ללא צורך בתשלום קנס או חיוב", "target_date": "2025-09-20"},
    {"key": "advance_payment_1_oct",                 "division": "all", "kind": "date",      "goal_number": None, "label": "פעימת תשלום מקדמה ראשונה (עד 40%), המותנית בהגשת 40% מהתכנון במועד שנקבע, ובכך שלבית הספר קיים חשבון בנק ורישיון בתוקף", "target_date": "2025-10-01"},
    {"key": "bank_account_forms_dec",                "division": "all", "kind": "date",      "goal_number": None, "label": "מועד הגשת טפסים תקינים לפתיחת חשבון בנק בית ספרי (למוסדות חדשים), לצורך קבלת מקדמה ראשונה בחודש דצמבר", "target_date": "2025-10-15"},
    {"key": "planning_70_oct",                      "division": "all", "kind": "planning",  "goal_number": 70, "label": "יעד תכנון: לפחות 70% מתקציב הגפ\"ן.",                    "target_date": "2025-10-31"},
    {"key": "pedagogical_committee_deadline",        "division": "all", "kind": "date",      "goal_number": None, "label": "המועד אחרון לקיום הוועדה המלווה הפדגוגית ומילוי פרטיה במערכת", "target_date": "2025-10-31"},
    {"key": "workplan_approval_supervision_deadline","division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לאישור תוכניות עבודה על ידי הפיקוח", "target_date": "2025-11-13"},
    {"key": "workplan_approval_authority_deadline",  "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לאישור תוכניות עבודה על ידי הרשות או הבעלות", "target_date": "2025-11-25"},
    {"key": "workplan_changes_biweekly_start",       "division": "all", "kind": "date",      "goal_number": None, "label": "החל מחודש דצמבר, כל עדכון או שינוי בתוכנית העבודה \"יוקפץ\" אוטומטית לסבב אישורים (פיקוח ורשות) פעמיים בחודש – ב-1 וב-15 לחודש", "target_date": "2025-12-01"},
    {"key": "reporting_valid_10_dec",                "division": "all", "kind": "reporting", "goal_number": 10, "label": "יעד דיווח: ביצוע תקין של לפחות 10% מהתקציב",             "target_date": "2025-12-31"},
    {"key": "bank_account_final_deadline",           "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון מוחלט להסדרת חשבון בנק בית ספרי ורישיון. מוסד שלא יסדיר זאת עד לתאריך זה, לא יוכל להשתתף בתוכנית הגפ\"ן השנה", "target_date": "2025-12-31"},
    {"key": "selections_lock_deadline",              "division": "all", "kind": "date",      "goal_number": None, "label": "נעילת בחירות: לא ניתן יותר לשנות בחירות שנעשו לגבי מרבית המענים המשרדיים. כמו כן, מועד אחרון למחיקת ימי הדרכה שלא אוישו", "target_date": "2025-12-31"},
    {"key": "supplier_payments_deadline",            "division": "all", "kind": "date",      "goal_number": None, "label": "תשלומים לספקים: מועד אחרון לביטול רכישת פריט תוכן דיגיטלי (אחרת תחויבו ב-100% תשלום). כמו כן, חובה להעביר תשלום ראשון (או מלא, תלוי בתקציב) לתוכנות ניהול פדגוגי", "target_date": "2025-12-31"},
    {"key": "second_payment_eligibility_check",      "division": "all", "kind": "date",      "goal_number": None, "label": "זכאות לתשלום: בדיקת זכאות לפעימת התשלום השנייה (מותנה בתכנון של 70% והגשת דוח מהשנה הקודמת)", "target_date": "2025-12-31"},
    {"key": "planning_90_jan",                      "division": "all", "kind": "planning",  "goal_number": 90, "label": "יעד תכנון: 90% מהתקציב.",                               "target_date": "2026-01-31"},
    {"key": "reporting_valid_25_feb",                "division": "all", "kind": "reporting", "goal_number": 25, "label": "יעד דיווח: ביצוע תקין של לפחות 25% מהתקציב",             "target_date": "2026-02-28"},
    {"key": "scholarships_deadline_mar",             "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון להזנת מלגות בגפ\"ן",                          "target_date": "2026-03-31"},
    {"key": "third_payment_apr",                     "division": "all", "kind": "date",      "goal_number": None, "label": "קבלת פעימת תשלום שלישית (השלמה ל-100% מהתקציב), בכפוף לעמידה ביעדי התכנון והדיווח", "target_date": "2026-04-01"},
    {"key": "reporting_valid_70_may",                "division": "all", "kind": "reporting", "goal_number": 70, "label": "יעד דיווח: ביצוע תקין של לפחות 70% מהתקציב",             "target_date": "2026-05-31"},
    {"key": "work_plan_changes_deadline_jun",        "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לשינויים בתוכנית העבודה",                     "target_date": "2026-07-10"},
    {"key": "feedback_deadline_jul",                 "division": "all", "kind": "date",      "goal_number": None, "label": "המועד האחרון להזנת משוב (פידבק) על תוכניות שצרכתם ממאגר התוכניות", "target_date": "2026-07-10"},
    {"key": "final_changes_approval_supervision",    "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לאישור סופי של השינויים שהוטמעו על ידי הפיקוח. לאחר מכן, תוכנית העבודה ננעלת סופית לשינויים", "target_date": "2026-07-13"},
    {"key": "reporting_valid_85_jul",                "division": "all", "kind": "reporting", "goal_number": 85, "label": "יעד דיווח: ביצוע תקין של לפחות 85% מהתקציב",             "target_date": "2026-07-15"},
    {"key": "final_changes_approval_authority",      "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון לאישור סופי של השינויים שהוטמעו על ידי הרשות. לאחר מכן, תוכנית העבודה ננעלת סופית לשינויים", "target_date": "2026-07-25"},
    {"key": "receipts_deadline_aug",                 "division": "all", "kind": "date",      "goal_number": None, "label": "מועד אחרון להנפקת אסמכתאות עבור גפן תשפ\"ו",             "target_date": "2026-08-31"},
    {"key": "year_close_critical_deadline",          "division": "all", "kind": "date",      "goal_number": None, "label": "מועד סגירת שנה קריטי: מועד אחרון להעלאת כל החשבוניות, כרטסות הנהלת החשבונות והגשת דו\"ח ביצוע שנתי מאושר וחתום ע\"י המנהל. אי הגשת הדוח כראוי תוביל לקיזוזי תקציב", "target_date": "2026-11-15"},
    {"key": "final_ministry_reconciliation",         "division": "all", "kind": "date",      "goal_number": None, "label": "סיכום והתחשבנות סופית מטעם המשרד", "target_date": "2027-02-01"},
]


def _shift_goal_date(goal_def: dict, academic_year: str) -> str:
    """target_date/label above are written for DEFAULT_ACADEMIC_YEAR — shift the year forward
    (day/month unchanged) by however many academic years ahead `academic_year` is, unless an
    explicit one-off override exists for that academic year in `date_overrides`."""
    override = goal_def.get("date_overrides", {}).get(academic_year)
    if override:
        return override
    iso_date = goal_def["target_date"]
    offset = ACADEMIC_YEARS.index(academic_year) - ACADEMIC_YEARS.index(DEFAULT_ACADEMIC_YEAR)
    if offset == 0:
        return iso_date
    y, m, d = iso_date.split("-")
    return f"{int(y) + offset}-{m}-{d}"


# ---------------------------------------------------------------------------
# Request models
# ---------------------------------------------------------------------------

class SchoolIn(BaseModel):
    name: str
    symbol: str | None = None
    city: str | None = None
    authority: str | None = None
    stage: str | None = None
    finance_software: str | None = None
    principal_name: str | None = None
    principal_phone: str | None = None
    secretary_name: str | None = None
    secretary_phone: str | None = None
    finance_contact_name: str | None = None
    finance_contact_phone: str | None = None
    finance_contact_email: str | None = None
    principal_email: str | None = None
    secretary_email: str | None = None
    school_phone: str | None = None
    address: str | None = None
    district: str | None = None
    notes: str | None = None
    restrict_access_to: list[str] | None = None
    extra_contacts: list[dict] | None = None
    principal_day_off: list[str] | None = None
    secretary_day_off: list[str] | None = None
    finance_contact_day_off: list[str] | None = None
    meeting_coordinator: str | None = None
    principal_chativa_name: str | None = None
    principal_chativa_phone: str | None = None
    principal_chativa_email: str | None = None
    principal_chativa_day_off: list[str] | None = None
    principal_same_person: bool | None = None


class SchoolYearAdminDataIn(BaseModel):
    service_type: str | None = None
    client_status: str | None = None
    requested_price: float | None = None
    order_method: list[str] | None = None
    order_amount_gefen: float | None = None
    hours_ordered: float | None = None
    rate: float | None = None
    payment_received: float | None = None
    payment_requests_sent: float | None = None
    contract_sent: bool | None = None
    contract_received: bool | None = None
    receipts_sent: float | None = None
    closure_parents_status: bool | None = None
    closure_parents_notes: str | None = None
    closure_authority_status: bool | None = None
    closure_authority_notes: str | None = None
    meeting_allocation_gefen: float | None = None
    meeting_allocation_current: float | None = None
    meeting_allocation_district: float | None = None
    meeting_duration_gefen: int | None = None
    meeting_duration_current: int | None = None
    meeting_duration_district: int | None = None
    invoice_transaction_status: str | None = None
    payment_method: str | None = None
    invoice_numbers: list[str] | None = None
    deposit_dates: list[str] | None = None
    amount_paid: float | None = None


class GefenAccountIn(BaseModel):
    division_type: str
    custom_label: str | None = None
    finance_software: str | None = None
    tmura_model: bool | None = None


class GoalStatusIn(BaseModel):
    division_type: str
    budget_name: str
    goal_key: str
    academic_year: str
    met: bool | None = None


class ControlLetterIn(BaseModel):
    received_date: str | None = None
    days_to_answer: int | None = None
    status: str | None = None
    notes: str | None = None


class AdvisorAssignIn(BaseModel):
    advisor_id: str


class MeetingIn(BaseModel):
    meeting_date: str | None = None
    status: str | None = "scheduled"
    start_time: str | None = None
    end_time: str | None = None
    advisor_id: str | None = None      # legacy single-advisor field (kept for compat)
    advisor_ids: list[str] | None = None  # multi-advisor array
    participants: list[dict] | None = None
    meeting_type: str | None = None
    meeting_service_type: str | None = None
    actual_duration: str | None = None
    notes: str | None = None
    reminder_enabled: bool | None = False
    academic_year: str | None = None
    primary_contact_key: str | None = None  # which participant's phone to use in the Outlook event subject
    stage_scope: str | None = None  # six-year schools only: 'tichon' | 'chativa' | 'both'


class MeetingStatusPatchIn(BaseModel):
    status: str | None = None
    notes: str | None = None
    start_time: str | None = None
    end_time: str | None = None


class DirectCoordinationParticipantIn(BaseModel):
    key: str
    name: str
    email: str | None = None


class DirectCoordinationRangeIn(BaseModel):
    start_date: str
    end_date: str
    meeting_service_type: str  # "gefen" | "current"
    duration_minutes: int      # 30..180, step 15
    participants: list[DirectCoordinationParticipantIn]


class DirectCoordinationIn(BaseModel):
    advisor_ids: list[str]
    ranges: list[DirectCoordinationRangeIn]


class UserInviteIn(BaseModel):
    email: str
    full_name: str | None = None
    role: str = "advisor"
    control_domains: list[str] = []
    work_phone: str | None = None


class UserRoleIn(BaseModel):
    role: str


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _require_manager(user: dict):
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")


def _validate_work_phone(phone: str | None) -> str | None:
    if phone is None or phone == "":
        return phone
    if not re.fullmatch(r"05\d{8}", phone):
        raise HTTPException(status_code=400, detail="טלפון עבודה חייב להיות 10 ספרות המתחילות ב-05")
    return phone


def _require_owner(user: dict):
    if user["role"] != "owner":
        raise HTTPException(status_code=403, detail="פעולה זו מיועדת לבעלים בלבד")


def _create_notifications(db, notifications: list[dict], pref_key: str | None = None):
    """Insert notification rows, optionally filtering by per-recipient preferences. Non-fatal."""
    if not notifications:
        return
    try:
        if pref_key:
            ids = [n["recipient_id"] for n in notifications]
            prefs_rows = (db.table("profiles")
                          .select("id, notification_preferences")
                          .in_("id", ids).execute().data or [])
            prefs_map = {r["id"]: (r.get("notification_preferences") or {}) for r in prefs_rows}
            notifications = [n for n in notifications
                             if prefs_map.get(n["recipient_id"], {}).get(pref_key, True)]
        if notifications:
            db.table("notifications").insert(notifications).execute()
    except Exception as exc:
        logger.warning("_create_notifications failed (non-fatal): %s", exc)


class NotificationPreferencesIn(BaseModel):
    meeting_reminder: bool | None = None
    meeting_reminder_minutes: int | None = None
    notify_update_request_submitted: bool | None = None
    notify_update_request_reviewed: bool | None = None
    notify_update_request_result: bool | None = None
    notify_school_created: bool | None = None
    notify_school_deleted: bool | None = None
    notify_advisor_assignment: bool | None = None
    notify_role_changed: bool | None = None
    notify_mention: bool | None = None
    notify_task_assigned: bool | None = None
    notify_call_contact_ambiguous: bool | None = None


# ---------------------------------------------------------------------------
# Schools
# ---------------------------------------------------------------------------

_MEETING_COORDINATOR_ROLE_LABEL = {
    "principal": "מנהל/ת",
    "principal_chativa": "מנהל/ת חט\"ב",
    "secretary": "מנהלנ/ית",
    "finance_contact": "אחראי/ת כספים",
}


def _resolve_meeting_coordinator(school: dict) -> dict | None:
    """Resolves school['meeting_coordinator'] (a reference, e.g. 'secretary' or 'extra:1')
    into the actual contact's current name/phone/email. Returns None if unset or the
    reference no longer points to an existing contact (e.g. that extra contact was removed) —
    never raises, so this stays non-fatal enrichment."""
    ref = school.get("meeting_coordinator")
    if not ref:
        return None
    if ref in _MEETING_COORDINATOR_ROLE_LABEL:
        name = school.get(f"{ref}_name")
        if not name:
            return None
        return {
            "role": ref,
            "role_label": _MEETING_COORDINATOR_ROLE_LABEL[ref],
            "name": name,
            "phone": school.get(f"{ref}_phone"),
            "email": school.get(f"{ref}_email"),
        }
    if ref.startswith("extra:"):
        try:
            idx = int(ref.split(":", 1)[1])
        except ValueError:
            return None
        extras = school.get("extra_contacts") or []
        if idx < 0 or idx >= len(extras):
            return None
        ec = extras[idx]
        if not ec.get("name"):
            return None
        return {
            "role": ref,
            "role_label": ec.get("role") or "איש קשר נוסף",
            "name": ec.get("name"),
            "phone": ec.get("phone"),
            "email": ec.get("email"),
        }
    return None


@router.get("/")
def list_schools(
    user: Annotated[dict, Depends(get_current_user)],
    include_deleted: bool = False,
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    is_advisor = user["role"] not in ("owner", "manager")
    if include_deleted and is_advisor:
        include_deleted = False
    schools = []

    for attempt in range(2):
        try:
            db = get_admin_client()

            if is_advisor:
                # Q_pre: fetch only this advisor's assigned school IDs (fast, indexed)
                assigned = db.table("advisor_schools").select("school_id").eq("advisor_id", user["id"]).execute()
                assigned_ids = [r["school_id"] for r in (assigned.data or [])]

                # Q1: open-to-all schools + directly assigned schools
                # Note: restrict_access_to.cs.[uuid] is intentionally kept OUT of or_() here.
                # PostgREST's or= parser misinterprets JSON array brackets as grouping syntax,
                # causing an APIError. Instead we use a separate direct .filter() call (Q2).
                open_filters = ["restrict_access_to.is.null"]
                if assigned_ids:
                    open_filters.append(f"id.in.({','.join(assigned_ids)})")
                q1_builder = (
                    db.table("schools")
                    .select("*, gefen_accounts(*), advisor_schools(advisor_id)")
                    .eq("org_id", user["org_id"])
                    .or_(",".join(open_filters))
                    .order("name")
                )
                if not include_deleted:
                    q1_builder = q1_builder.eq("status", "active")
                q1 = q1_builder.execute()

                # Q2: schools with an explicit allow-list that includes this advisor.
                # Uses direct .filter() (not inside or_()) so PostgREST parses the JSON value correctly.
                q2_builder = (
                    db.table("schools")
                    .select("*, gefen_accounts(*), advisor_schools(advisor_id)")
                    .eq("org_id", user["org_id"])
                    .filter("restrict_access_to", "cs", json.dumps([user["id"]]))
                )
                if not include_deleted:
                    q2_builder = q2_builder.eq("status", "active")
                q2 = q2_builder.execute()

                # Merge Q1 + Q2, deduplicate by ID, re-sort by name
                seen_ids = {s["id"] for s in (q1.data or [])}
                schools = (q1.data or []) + [s for s in (q2.data or []) if s["id"] not in seen_ids]
                schools.sort(key=lambda s: s.get("name") or "")
            else:
                # Q1 (manager/owner): fetch all schools within their org
                all_builder = (
                    db.table("schools")
                    .select("*, gefen_accounts(*), advisor_schools(advisor_id)")
                    .eq("org_id", user["org_id"])
                    .order("name")
                )
                if not include_deleted:
                    all_builder = all_builder.eq("status", "active")
                all_res = all_builder.execute()
                schools = all_res.data or []

            break  # success
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_schools attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_schools failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    # Enrich Q3 (profiles) and Q4 (meetings stats) sequentially — safer with 3s timeout
    # db is the singleton already acquired inside the retry loop above — still valid.
    all_advisor_ids = list({
        row["advisor_id"]
        for s in schools
        for row in (s.get("advisor_schools") or [])
        if row.get("advisor_id")
    })
    school_ids = [s["id"] for s in schools]

    profiles_map: dict = {}
    m_stats: dict = {}

    if all_advisor_ids:
        try:
            p_rows = db.table("profiles").select("id, full_name, email").in_("id", all_advisor_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
        except Exception as exc:
            logger.warning("profiles enrichment failed (non-fatal): %s", exc)

    if school_ids:
        try:
            stats_res = db.rpc("get_meetings_stats", {"school_ids": school_ids}).execute()
            m_stats = {
                r["school_id"]: {"completed": r["completed"], "total_minutes": r["total_minutes"]}
                for r in (stats_res.data or [])
            }
        except Exception as exc:
            logger.warning("meetings stats enrichment failed (non-fatal): %s", exc)

    for school in schools:
        for row in (school.get("advisor_schools") or []):
            row["profiles"] = profiles_map.get(row["advisor_id"])
        school["meetings_stats"] = m_stats.get(school["id"])

    # Enrich Q5 (check_metrics) and Q6 (school_goals) — used by the dashboard's advanced
    # filter (checks/goals). Non-fatal: on failure the schools list still returns, the
    # advanced filter fields simply find no matches.
    metrics_by_school: dict = {}
    goals_by_school: dict = {}
    if school_ids:
        try:
            metrics_rows = (
                db.table("check_metrics")
                .select(
                    "school_id, division_type, budget_name, pct_plan, pct_divuach, pct_tanuz, "
                    "budget_amount, planned_amount, fixed_gap_abs, flexible_remaining, sum_chayav, sum_divuach, "
                    "rejected_count, rejected_sum, no_pdf_count, no_pdf_sum, partial_count, partial_sum, "
                    "finance_not_gefen_count, finance_not_gefen_sum, gefen_not_finance_count, gefen_not_finance_sum"
                )
                .eq("academic_year", academic_year)
                .in_("school_id", school_ids)
                .execute()
            )
            for r in (metrics_rows.data or []):
                metrics_by_school.setdefault(r["school_id"], []).append(r)
        except Exception as exc:
            logger.warning("check_metrics enrichment failed (non-fatal): %s", exc)

        try:
            goal_rows = (
                db.table("school_goals")
                .select("school_id, division_type, budget_name, goal_key, met")
                .eq("academic_year", academic_year)
                .in_("school_id", school_ids)
                .execute()
            )
            for r in (goal_rows.data or []):
                goals_by_school.setdefault(r["school_id"], []).append(r)
        except Exception as exc:
            logger.warning("school_goals enrichment failed (non-fatal): %s", exc)

    # Enrich Q7 ("סגירת שנה" — closure status/notes with parents and with the authority) for
    # the dashboard and admin schools tables. Only the 4 closure fields are selected here
    # (not the whole school_year_admin_data row) so advisors never receive financial fields
    # (pricing, payment status, etc.) they aren't otherwise granted access to.
    closure_by_school: dict = {}
    if school_ids:
        try:
            closure_rows = (
                db.table("school_year_admin_data")
                .select("school_id, closure_parents_status, closure_parents_notes, "
                        "closure_authority_status, closure_authority_notes, "
                        "meeting_allocation_gefen, meeting_allocation_current, meeting_allocation_district, "
                        "meeting_duration_gefen, meeting_duration_current, meeting_duration_district")
                .eq("academic_year", academic_year)
                .in_("school_id", school_ids)
                .execute()
            )
            closure_by_school = {r["school_id"]: r for r in (closure_rows.data or [])}
        except Exception as exc:
            logger.warning("year_closure enrichment failed (non-fatal): %s", exc)

    # Enrich Q7b ("מכתב בקרה" — control letters, one fixed row per division_type) for the
    # dashboard and admin schools tables.
    control_letters_by_school: dict = {}
    if school_ids:
        try:
            cl_rows = (
                db.table("control_letters")
                .select("school_id, division_type, received_date, days_to_answer, status, notes, "
                        "original_letter_file_name, response_letter_file_name")
                .in_("school_id", school_ids)
                .execute()
            )
            for r in (cl_rows.data or []):
                control_letters_by_school.setdefault(r["school_id"], []).append(r)
        except Exception as exc:
            logger.warning("control_letters enrichment failed (non-fatal): %s", exc)

    # Enrich Q8 (per-service-type advisor lists — school_advisors_gefen/current/district) for
    # the dashboard/admin "יועץ מלווה [גפן/שוטף/מחוז]" columns. Non-fatal: on failure schools
    # still return, just without these lists populated.
    typed_advisors_by_school: dict = {"gefen": {}, "current": {}, "district": {}}
    if school_ids:
        for service_type, table_name in (
            ("gefen", "school_advisors_gefen"),
            ("current", "school_advisors_current"),
            ("district", "school_advisors_district"),
        ):
            try:
                rows = db.table(table_name).select("school_id, advisor_id").in_("school_id", school_ids).execute()
                for r in (rows.data or []):
                    typed_advisors_by_school[service_type].setdefault(r["school_id"], []).append(r["advisor_id"])
            except Exception as exc:
                logger.warning("%s enrichment failed (non-fatal): %s", table_name, exc)

        typed_advisor_ids = list({
            aid
            for by_school in typed_advisors_by_school.values()
            for ids in by_school.values()
            for aid in ids
        })
        if typed_advisor_ids:
            try:
                tp_rows = db.table("profiles").select("id, full_name, email").in_("id", typed_advisor_ids).execute()
                for p in (tp_rows.data or []):
                    if p["id"] not in profiles_map:
                        profiles_map[p["id"]] = p
            except Exception as exc:
                logger.warning("typed advisors profile enrichment failed (non-fatal): %s", exc)

    for school in schools:
        school["check_metrics"] = metrics_by_school.get(school["id"], [])
        school["goal_statuses"] = goals_by_school.get(school["id"], [])
        school["year_closure"] = closure_by_school.get(school["id"]) or {}
        school["control_letters"] = control_letters_by_school.get(school["id"], [])
        school["meeting_coordinator_contact"] = _resolve_meeting_coordinator(school)
        for service_type in ("gefen", "current", "district"):
            advisor_ids = typed_advisors_by_school[service_type].get(school["id"], [])
            school[f"advisors_{service_type}"] = [profiles_map[aid] for aid in advisor_ids if aid in profiles_map]

    return schools


@router.post("/")
def create_school(
    body: SchoolIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if user["role"] not in ("owner", "manager", "advisor"):
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")
    db = get_admin_client()
    if not _check_permission(db, user, "can_add_school"):
        raise HTTPException(status_code=403, detail="אין הרשאה להוסיף בתי ספר")
    if not body.meeting_coordinator:
        raise HTTPException(status_code=400, detail="יש להגדיר אחראי/ת לתיאום פגישות")
    payload = body.model_dump(exclude_none=True)
    payload["org_id"] = user["org_id"]
    row = db.table("schools").insert(payload).execute()
    new_school = row.data[0]
    new_school_id = new_school["id"]
    # Auto-assign advisor creators so subsequent account creation is authorized
    if user["role"] == "advisor":
        try:
            db.table("advisor_schools").upsert({"advisor_id": user["id"], "school_id": new_school_id}).execute()
        except Exception as e:
            logger.warning("Failed to auto-assign advisor to new school (non-fatal): %s", e)
    # Notify owners (and managers when advisor creates) about the new school
    try:
        if user["role"] in ("advisor", "manager"):
            school_name = new_school.get("name", "בית ספר")
            notif_title = f'{user.get("full_name", "יועץ")} הוסיף את בית הספר {school_name}'
            if user["role"] == "advisor":
                # Notify all owners + all managers
                recipients = db.table("profiles").select("id").eq("org_id", user["org_id"]).in_("role", ["owner", "manager"]).execute()
            else:
                # Manager created — notify owners only
                recipients = db.table("profiles").select("id").eq("org_id", user["org_id"]).eq("role", "owner").execute()
            notif_rows = [{
                "recipient_id": r["id"],
                "type": "school_created",
                "school_id": new_school_id,
                "data": {
                    "title": notif_title,
                    "school_name": school_name,
                    "sender_name": user.get("full_name", ""),
                    "deeplink": f"/school/{new_school_id}",
                }
            } for r in (recipients.data or []) if r["id"] != user["id"]]
            _create_notifications(db, notif_rows, pref_key="notify_school_created")
    except Exception as exc:
        logger.warning("school_created notification failed (non-fatal): %s", exc)
    return new_school


@router.put("/{school_id}")
def update_school(
    school_id: str,
    body: SchoolIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager"):
        if not _check_permission(db, user, "can_edit_school_directly"):
            raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")
        assigned = (
            db.table("advisor_schools")
            .select("school_id")
            .eq("advisor_id", user["id"])
            .eq("school_id", school_id)
            .execute()
        )
        if not assigned.data:
            raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
    # Fetch current restrict_access_to before update (for diff → notifications)
    old_school = db.table("schools").select("name, restrict_access_to").eq("id", school_id).eq("org_id", user["org_id"]).execute()
    old_restrict = set(old_school.data[0].get("restrict_access_to") or []) if old_school.data else set()
    school_name = old_school.data[0]["name"] if old_school.data else "בית ספר"

    update_data = body.model_dump(exclude_none=True)
    # Allow explicitly clearing restrict_access_to (null = כולם)
    if "restrict_access_to" in body.model_fields_set and body.restrict_access_to is None:
        update_data["restrict_access_to"] = None
    row = (
        db.table("schools")
        .update(update_data)
        .eq("id", school_id)
        .eq("org_id", user["org_id"])
        .execute()
    )
    if not row.data:
        raise HTTPException(status_code=404, detail="בית הספר לא נמצא")

    # Non-fatal: notify advisors added/removed from restrict_access_to
    if "restrict_access_to" in body.model_fields_set:
        try:
            new_restrict = set(body.restrict_access_to or [])
            added = new_restrict - old_restrict
            removed = old_restrict - new_restrict
            notif_rows = [
                *[{
                    "recipient_id": aid,
                    "type": "advisor_assigned",
                    "school_id": school_id,
                    "data": {
                        "title": f"קיבלת גישה לבית הספר {school_name}",
                        "school_name": school_name,
                        "sender_name": user.get("full_name", ""),
                        "deeplink": f"/school/{school_id}",
                    }
                } for aid in added],
                *[{
                    "recipient_id": aid,
                    "type": "advisor_removed",
                    "school_id": school_id,
                    "data": {
                        "title": f"הגישה שלך לבית הספר {school_name} הוסרה",
                        "school_name": school_name,
                        "sender_name": user.get("full_name", ""),
                    }
                } for aid in removed],
            ]
            if notif_rows:
                _create_notifications(db, notif_rows, pref_key="notify_advisor_assignment")
        except Exception as exc:
            logger.warning("restrict_access_to change notification failed (non-fatal): %s", exc)

    return row.data[0]


@router.delete("/{school_id}")
def delete_school(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    if not _check_permission(db, user, "can_delete_schools"):
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק בתי ספר")
    if user["role"] == "advisor":
        assigned = db.table("advisor_schools").select("school_id").eq("advisor_id", user["id"]).eq("school_id", school_id).execute()
        if not assigned.data:
            raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
    school_row = db.table("schools").select("name").eq("id", school_id).eq("org_id", user["org_id"]).execute()
    school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
    from datetime import datetime, timezone
    db.table("schools").update({
        "status": "pending_deletion",
        "deleted_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", school_id).eq("org_id", user["org_id"]).execute()

    # Cancel the Outlook events for this school's future meetings — a deleted school's
    # meetings shouldn't keep sitting on advisors' calendars. Restoring the school
    # (recycle-bin flow) re-creates them; the meeting DB rows themselves are untouched.
    try:
        today = datetime.now(timezone.utc).date().isoformat()
        future = db.table("meetings").select("id, calendar_sync").eq("school_id", school_id).gte("meeting_date", today).eq("status", "scheduled").execute()
        for m in (future.data or []):
            if m.get("calendar_sync"):
                with graph_client.calendar_sync_lock(db, m["id"]) as acquired:
                    if acquired:
                        graph_client.sync_meeting_cancel(db, user["org_id"], m["calendar_sync"])
                        graph_client.persist_calendar_sync(db, m["id"], {})
    except Exception as exc:
        logger.warning("calendar cancel-sync failed for deleted school %s (non-fatal): %s", school_id, exc)

    try:
        owner_ids = [r["id"] for r in (db.table("profiles").select("id").eq("role", "owner").eq("org_id", user["org_id"]).execute().data or [])]
        notif_rows = [{
            "recipient_id": oid,
            "type": "school_deleted",
            "data": {
                "title": f'{user.get("full_name", "מנהל")} העביר את בית הספר {school_name} לסל המחזור',
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
            }
        } for oid in owner_ids if oid != user["id"]]
        _create_notifications(db, notif_rows, pref_key="notify_school_deleted")
    except Exception as exc:
        logger.warning("school_deleted notification failed (non-fatal): %s", exc)
    return {"ok": True}



@router.post("/{school_id}/restore")
def restore_school(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    if user["role"] != "owner":
        raise HTTPException(status_code=403, detail="רק בעלים יכולים לשחזר בית ספר")
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("schools").update({
                "status": "active",
                "deleted_at": None,
            }).eq("id", school_id).eq("org_id", user["org_id"]).execute()
            break
        except Exception as exc:
            if attempt == 0:
                reset_admin_client()
                time.sleep(0.1)
            else:
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשחזור בית הספר — נסה שוב")

    # Non-fatal: notify assigned advisors that the school was restored
    try:
        db = get_admin_client()
        school_row = db.table("schools").select("name").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
        advisor_rows = db.table("advisor_schools").select("advisor_id").eq("school_id", school_id).execute()
        advisor_ids = [r["advisor_id"] for r in (advisor_rows.data or [])]
        notif_rows = [{
            "recipient_id": aid,
            "type": "advisor_assigned",
            "school_id": school_id,
            "data": {
                "title": f"בית הספר {school_name} שוחזר על ידי הבעלים והוחזר לטיפולך",
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "deeplink": f"/school/{school_id}",
            }
        } for aid in advisor_ids]
        _create_notifications(db, notif_rows, pref_key="notify_advisor_assignment")
    except Exception as exc:
        logger.warning("restore_school advisor notification failed (non-fatal): %s", exc)

    # Re-create Outlook events for this school's future meetings that were cancelled
    # when the school was deleted (calendar_sync was cleared to {} at that point).
    try:
        from datetime import datetime, timezone
        today = datetime.now(timezone.utc).date().isoformat()
        future = db.table("meetings").select("*").eq("school_id", school_id).gte("meeting_date", today).eq("status", "scheduled").execute()
        for m in (future.data or []):
            if m.get("calendar_sync"):
                continue  # already synced (wasn't touched by the delete flow)
            with graph_client.calendar_sync_lock(db, m["id"]) as acquired:
                if acquired:
                    subject = _build_meeting_subject(db, school_id, m.get("participants"), m.get("primary_contact_key"))
                    sync_map = graph_client.sync_meeting_create(db, user["org_id"], m, subject)
                    if sync_map:
                        graph_client.persist_calendar_sync(db, m["id"], sync_map)
    except Exception as exc:
        logger.warning("calendar re-sync failed for restored school %s (non-fatal): %s", school_id, exc)

    return {"ok": True}


# ---------------------------------------------------------------------------
# School year admin data (מסך ניהול → בתי ספר: מחירים, חוזים, תשלומים)
# ---------------------------------------------------------------------------

def _advisor_has_school_access(db, user: dict, school_id: str) -> bool:
    """Whether an advisor (non-manager) may view/edit this school's data."""
    school_row = (
        db.table("schools")
        .select("restrict_access_to")
        .eq("id", school_id)
        .eq("org_id", user["org_id"])
        .execute()
    )
    if not school_row.data:
        return False
    restrict = school_row.data[0].get("restrict_access_to")
    if restrict is None:
        return True
    if user["id"] in restrict:
        return True
    assigned = (
        db.table("advisor_schools")
        .select("school_id")
        .eq("advisor_id", user["id"])
        .eq("school_id", school_id)
        .execute()
    )
    return bool(assigned.data)


def _attach_updater_names(db, rows: list[dict]) -> None:
    """Non-fatal enrichment: resolve order_amount_gefen_updated_by → full_name."""
    updater_ids = list({r["order_amount_gefen_updated_by"] for r in rows if r.get("order_amount_gefen_updated_by")})
    if not updater_ids:
        return
    try:
        p_rows = db.table("profiles").select("id, full_name").in_("id", updater_ids).execute()
        names_map = {p["id"]: p["full_name"] for p in (p_rows.data or [])}
        for r in rows:
            uid = r.get("order_amount_gefen_updated_by")
            r["order_amount_gefen_updated_by_name"] = names_map.get(uid) if uid else None
    except Exception as exc:
        logger.warning("profile enrichment for year_admin_data failed (non-fatal): %s", exc)


@router.get("/year-admin-data")
def list_year_admin_data(
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    _require_manager(user)
    data = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("school_year_admin_data")
                .select("*")
                .eq("academic_year", academic_year)
                .execute()
            )
            data = rows.data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_year_admin_data attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("list_year_admin_data failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    _attach_updater_names(db, data)
    return {r["school_id"]: r for r in data}


@router.get("/{school_id}/year-admin-data")
def get_year_admin_data(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    row = (
        db.table("school_year_admin_data")
        .select("*")
        .eq("school_id", school_id)
        .eq("academic_year", academic_year)
        .execute()
    )
    data = row.data[0] if row.data else {}
    if data:
        _attach_updater_names(db, [data])
    return data


@router.put("/{school_id}/year-admin-data")
def upsert_year_admin_data(
    school_id: str,
    body: SchoolYearAdminDataIn,
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    from datetime import datetime, timezone

    for attempt in range(2):
        try:
            db = get_admin_client()
            is_manager = user["role"] in ("owner", "manager")
            if not is_manager:
                if not _advisor_has_school_access(db, user, school_id):
                    raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
                # Advisors (non-manager) may only update the Gefen order amount and the
                # "סגירת שנה" fields from the school card — every other admin-table field is
                # manager+ only.
                allowed_fields = {
                    "order_amount_gefen",
                    "closure_parents_status", "closure_parents_notes",
                    "closure_authority_status", "closure_authority_notes",
                }
                submitted_fields = set(body.model_fields_set)
                if submitted_fields - allowed_fields:
                    raise HTTPException(status_code=403, detail="אין הרשאה לעדכן שדות אלו")

            update_data = body.model_dump(exclude_unset=True)

            old_row = (
                db.table("school_year_admin_data")
                .select("order_amount_gefen")
                .eq("school_id", school_id)
                .eq("academic_year", academic_year)
                .execute()
            )
            old_amount = old_row.data[0]["order_amount_gefen"] if old_row.data else None

            if "order_amount_gefen" in update_data and update_data["order_amount_gefen"] != old_amount:
                update_data["order_amount_gefen_updated_by"] = user["id"]
                update_data["order_amount_gefen_updated_at"] = datetime.now(timezone.utc).isoformat()

            update_data["school_id"] = school_id
            update_data["academic_year"] = academic_year
            update_data["updated_at"] = datetime.now(timezone.utc).isoformat()

            row = (
                db.table("school_year_admin_data")
                .upsert(update_data, on_conflict="school_id,academic_year")
                .execute()
            )
            data = row.data[0] if row.data else update_data
            _attach_updater_names(db, [data])
            return data
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("upsert_year_admin_data attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("upsert_year_admin_data failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/{school_id}/year-admin-data/contract-file")
async def upload_contract_file(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    import secrets
    import shutil
    import tempfile

    _require_manager(user)
    db = get_admin_client()

    run_dir = Path(tempfile.mkdtemp(prefix=f"contract_{school_id}_"))
    try:
        suffix = Path(file.filename or "").suffix
        dest = run_dir / f"contract{suffix}"
        dest.write_bytes(await file.read())
        storage_key = f"contracts/{school_id}/{academic_year}/{secrets.token_hex(8)}{suffix}"
        db.storage.from_("check-files").upload(storage_key, dest.read_bytes())
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

    row = (
        db.table("school_year_admin_data")
        .upsert(
            {
                "school_id": school_id,
                "academic_year": academic_year,
                "contract_file_storage_key": storage_key,
                "contract_file_name": file.filename,
            },
            on_conflict="school_id,academic_year",
        )
        .execute()
    )
    return row.data[0] if row.data else {"contract_file_storage_key": storage_key, "contract_file_name": file.filename}


@router.get("/{school_id}/year-admin-data/contract-file")
def download_contract_file(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    row = (
        db.table("school_year_admin_data")
        .select("contract_file_storage_key, contract_file_name")
        .eq("school_id", school_id)
        .eq("academic_year", academic_year)
        .execute()
    )
    if not row.data or not row.data[0].get("contract_file_storage_key"):
        raise HTTPException(status_code=404, detail="לא הועלה קובץ חוזה")

    storage_key = row.data[0]["contract_file_storage_key"]
    filename = row.data[0].get("contract_file_name") or "contract"
    content = db.storage.from_("check-files").download(storage_key)
    return Response(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{school_id}/control-letters")
def list_control_letters(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    for attempt in range(2):
        try:
            db = get_admin_client()
            if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
                raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
            rows = db.table("control_letters").select("*").eq("school_id", school_id).execute()
            return rows.data or []
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_control_letters attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("list_control_letters failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.put("/{school_id}/control-letters/{division_type}")
def upsert_control_letter(
    school_id: str,
    division_type: str,
    body: ControlLetterIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    update_data = body.model_dump(exclude_unset=True)
    update_data["school_id"] = school_id
    update_data["division_type"] = division_type

    row = (
        db.table("control_letters")
        .upsert(update_data, on_conflict="school_id,division_type")
        .execute()
    )
    return row.data[0] if row.data else update_data


def _upload_control_letter_file(school_id: str, division_type: str, user: dict, file: UploadFile, kind: str):
    import shutil
    import tempfile

    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    suffix = Path(file.filename or "").suffix.lower()
    if suffix != ".pdf":
        raise HTTPException(status_code=400, detail="ניתן להעלות קובצי PDF בלבד")

    run_dir = Path(tempfile.mkdtemp(prefix=f"control_letter_{school_id}_"))
    try:
        dest = run_dir / f"{kind}.pdf"
        content = file.file.read()
        dest.write_bytes(content)
        storage_key = f"control-letters/{school_id}/{division_type}/{kind}/{secrets.token_hex(8)}.pdf"
        db.storage.from_("check-files").upload(storage_key, dest.read_bytes())
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

    update_data = {
        "school_id": school_id,
        "division_type": division_type,
        f"{kind}_letter_storage_key": storage_key,
        f"{kind}_letter_file_name": file.filename,
    }
    row = (
        db.table("control_letters")
        .upsert(update_data, on_conflict="school_id,division_type")
        .execute()
    )
    return row.data[0] if row.data else update_data


def _download_control_letter_file(school_id: str, division_type: str, user: dict, kind: str):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    row = (
        db.table("control_letters")
        .select(f"{kind}_letter_storage_key, {kind}_letter_file_name")
        .eq("school_id", school_id)
        .eq("division_type", division_type)
        .execute()
    )
    storage_key = row.data[0].get(f"{kind}_letter_storage_key") if row.data else None
    if not storage_key:
        raise HTTPException(status_code=404, detail="לא הועלה קובץ")

    filename = row.data[0].get(f"{kind}_letter_file_name") or f"{kind}.pdf"
    content = db.storage.from_("check-files").download(storage_key)
    return Response(
        content=content,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _delete_control_letter_file(school_id: str, division_type: str, user: dict, kind: str):
    db = get_admin_client()
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")

    row = (
        db.table("control_letters")
        .update({f"{kind}_letter_storage_key": None, f"{kind}_letter_file_name": None})
        .eq("school_id", school_id)
        .eq("division_type", division_type)
        .execute()
    )
    return row.data[0] if row.data else {}


@router.post("/{school_id}/control-letters/{division_type}/original-file")
async def upload_control_letter_original_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    return _upload_control_letter_file(school_id, division_type, user, file, "original")


@router.get("/{school_id}/control-letters/{division_type}/original-file")
def download_control_letter_original_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    return _download_control_letter_file(school_id, division_type, user, "original")


@router.delete("/{school_id}/control-letters/{division_type}/original-file")
def delete_control_letter_original_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    return _delete_control_letter_file(school_id, division_type, user, "original")


@router.post("/{school_id}/control-letters/{division_type}/response-file")
async def upload_control_letter_response_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
):
    return _upload_control_letter_file(school_id, division_type, user, file, "response")


@router.get("/{school_id}/control-letters/{division_type}/response-file")
def download_control_letter_response_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    return _download_control_letter_file(school_id, division_type, user, "response")


@router.delete("/{school_id}/control-letters/{division_type}/response-file")
def delete_control_letter_response_file(
    school_id: str,
    division_type: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    return _delete_control_letter_file(school_id, division_type, user, "response")


@router.post("/collection/gefen-organized-check")
async def gefen_organized_check(
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
    academic_year: str = DEFAULT_ACADEMIC_YEAR,
):
    """Parse a "גפן-ספקים" export, sum ordered amounts per school symbol (column F),
    and compare against each relevant school's order_amount_gefen ("מחיר כולל מע"מ").
    Relevant = client_status == "active" and "private" in order_method.
    """
    import shutil
    import tempfile
    from datetime import datetime, timezone

    _require_manager(user)

    run_dir = Path(tempfile.mkdtemp(prefix="gefen_organized_"))
    try:
        dest = run_dir / (file.filename or "gefen_suppliers.xlsx")
        dest.write_bytes(await file.read())
        wb = load_workbook(dest, data_only=True, read_only=True)
        ws = wb.active

        totals: dict[str, float] = {}
        for row in ws.iter_rows(min_row=2):
            symbol_cell = row[5].value if len(row) > 5 else None  # column F
            amount_cell = row[31].value if len(row) > 31 else None  # column AF
            status_cell = row[34].value if len(row) > 34 else None  # column AI — "סטטוס הזמנה"
            if symbol_cell is None:
                continue
            if isinstance(status_cell, str) and status_cell.strip() == 'ההזמנה בוטלה ע"י המוסד':
                continue
            try:
                symbol = str(int(float(symbol_cell))).strip()
            except (TypeError, ValueError):
                continue
            try:
                amount = float(amount_cell) if amount_cell not in (None, "") else 0.0
            except (TypeError, ValueError):
                amount = 0.0
            totals[symbol] = totals.get(symbol, 0.0) + amount
        wb.close()
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

    for attempt in range(2):
        try:
            db = get_admin_client()
            schools_rows = db.table("schools").select("id, symbol").eq("status", "active").execute()
            yad_rows = (
                db.table("school_year_admin_data")
                .select("school_id, client_status, order_method, order_amount_gefen")
                .eq("academic_year", academic_year)
                .execute()
            )
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("gefen_organized_check attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("gefen_organized_check failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    yad_by_school = {r["school_id"]: r for r in (yad_rows.data or [])}
    now_iso = datetime.now(timezone.utc).isoformat()
    result_map: dict[str, dict] = {}

    for school in (schools_rows.data or []):
        yad = yad_by_school.get(school["id"], {})
        if yad.get("client_status") != "active" or "private" not in (yad.get("order_method") or []):
            continue

        symbol = str(school.get("symbol") or "").strip()
        checked_amount = totals.get(symbol, 0.0)
        order_amount = yad.get("order_amount_gefen")
        matched = bool(order_amount) and order_amount > 0 and checked_amount == order_amount

        update_data = {
            "school_id": school["id"],
            "academic_year": academic_year,
            "gefen_organized_matched": matched,
            "gefen_organized_checked_amount": checked_amount,
            "gefen_organized_checked_at": now_iso,
            "gefen_organized_uploaded_file_name": file.filename,
        }
        row = db.table("school_year_admin_data").upsert(update_data, on_conflict="school_id,academic_year").execute()
        result_map[school["id"]] = row.data[0] if row.data else update_data

    return result_map


class GefenOrganizedMismatchRow(BaseModel):
    name: str
    authority: str | None = None
    symbol: str | None = None
    order_amount: float = 0
    checked_amount: float = 0


class GefenOrganizedMismatchExportIn(BaseModel):
    rows: list[GefenOrganizedMismatchRow]
    academic_year: str = DEFAULT_ACADEMIC_YEAR


@router.post("/collection/gefen-organized-mismatches-export")
def export_gefen_organized_mismatches(
    body: GefenOrganizedMismatchExportIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)

    wb = Workbook()
    ws = wb.active
    ws.title = "גפן מסודר - אי התאמות"
    ws.sheet_view.rightToLeft = True
    ws.append(["שם מוסד", "סמל מוסד", "בעלות", 'מחיר כולל מע"מ', "גובה הזמנה בפועל", "פער"])
    for r in body.rows:
        ws.append([r.name, r.symbol, r.authority, r.order_amount, r.checked_amount, r.checked_amount - r.order_amount])
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.number_format = "#,##0"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    # HTTP headers must be ASCII — the Hebrew academic year breaks Content-Disposition
    # entirely if embedded raw. RFC 5987's filename* handles non-ASCII names correctly
    # (with an ASCII fallback for older clients).
    import urllib.parse
    filename = f"gefen-organized-mismatches-{body.academic_year}.xlsx"
    encoded_name = urllib.parse.quote(filename)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"gefen-organized-mismatches.xlsx\"; filename*=UTF-8''{encoded_name}"},
    )


# ---------------------------------------------------------------------------
# Gefen accounts (divisions)
# ---------------------------------------------------------------------------

@router.get("/{school_id}/accounts")
def list_accounts(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("gefen_accounts")
                .select("*")
                .eq("school_id", school_id)
                .order("division_type")
                .execute()
            )
            return rows.data
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_accounts attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_accounts failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/{school_id}/accounts")
def create_account(
    school_id: str,
    body: GefenAccountIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if user["role"] not in ("owner", "manager", "advisor"):
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")
    if body.division_type not in DIVISION_LABELS:
        raise HTTPException(status_code=400, detail="סוג חטיבה לא חוקי")
    db = get_admin_client()
    if user["role"] == "advisor":
        if not _check_permission(db, user, "can_add_school"):
            raise HTTPException(status_code=403, detail="אין הרשאה להוסיף חטיבות")
        assigned = db.table("advisor_schools").select("school_id").eq("advisor_id", user["id"]).eq("school_id", school_id).execute()
        if not (assigned.data):
            raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
    row = (
        db.table("gefen_accounts")
        .insert({"school_id": school_id, **body.model_dump(exclude_none=True)})
        .execute()
    )
    return row.data[0]


@router.put("/{school_id}/accounts/{account_id}")
def update_account(
    school_id: str,
    account_id: str,
    body: GefenAccountIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db = get_admin_client()
    data = {k: v for k, v in body.model_dump().items() if k != "division_type" and v is not None}
    row = (
        db.table("gefen_accounts")
        .update(data)
        .eq("id", account_id)
        .eq("school_id", school_id)
        .execute()
    )
    if not row.data:
        raise HTTPException(status_code=404, detail="החטיבה לא נמצאה")
    return row.data[0]


@router.delete("/{school_id}/accounts/{account_id}")
def delete_account(
    school_id: str,
    account_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db = get_admin_client()
    db.table("gefen_accounts").delete().eq("id", account_id).eq("school_id", school_id).execute()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Goals (יעדים)
# ---------------------------------------------------------------------------

@router.get("/goal-definitions")
def list_goal_definitions(user: Annotated[dict, Depends(get_current_user)]):
    """Ministry-wide goal list (key/kind/goal_number/label only) for the dashboard's
    advanced-filter goal dropdown — a single source of truth instead of duplicating
    GOAL_DEFINITIONS on the frontend."""
    return [
        {"key": d["key"], "kind": d["kind"], "goal_number": d["goal_number"], "label": d["label"]}
        for d in GOAL_DEFINITIONS
        if d["kind"] in ("planning", "reporting")
    ]


@router.get("/{school_id}/goals")
def list_goals(
    school_id: str,
    division_type: str,
    budget_name: str,
    academic_year: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    for attempt in range(2):
        try:
            db = get_admin_client()
            defs = sorted(
                (d for d in GOAL_DEFINITIONS if d["division"] in ("all", division_type)),
                key=lambda d: d["target_date"],
            )
            tracked_defs = [d for d in defs if d["kind"] in ("planning", "reporting")]
            date_defs = [d for d in defs if d["kind"] == "date"]

            statuses = (
                db.table("school_goals")
                .select("goal_key, met")
                .eq("school_id", school_id)
                .eq("division_type", division_type)
                .eq("budget_name", budget_name)
                .eq("academic_year", academic_year)
                .execute()
            )
            status_map = {r["goal_key"]: r["met"] for r in (statuses.data or [])}
            goals = [
                {
                    "key": d["key"],
                    "goal_type": d["kind"],
                    "goal_number": d["goal_number"],
                    "target_date": _shift_goal_date(d, academic_year),
                    "current_status": None,
                    "met": status_map.get(d["key"]),
                }
                for d in tracked_defs
            ]
            important_dates = [
                {
                    "label": d["label"].replace(DEFAULT_ACADEMIC_YEAR, academic_year),
                    "target_date": _shift_goal_date(d, academic_year),
                }
                for d in date_defs
            ]
            return {"goals": goals, "important_dates": important_dates}
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_goals attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_goals failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.patch("/{school_id}/goals")
def set_goal_status(
    school_id: str,
    body: GoalStatusIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    from datetime import datetime, timezone

    goal_def = next((d for d in GOAL_DEFINITIONS if d["key"] == body.goal_key), None)
    if not goal_def or goal_def["kind"] not in ("planning", "reporting"):
        raise HTTPException(status_code=400, detail="יעד לא נמצא")

    db = get_admin_client()
    # Was missing entirely (only Depends(get_current_user), no access check) — any authenticated
    # user, including an advisor with zero assignment to this school, could update ANY school's
    # goal. Same access rule already enforced by upsert_control_letter/get_year_admin_data.
    if user["role"] not in ("owner", "manager") and not _advisor_has_school_access(db, user, school_id):
        raise HTTPException(status_code=403, detail="אין גישה לבית ספר זה")
    row = (
        db.table("school_goals")
        .upsert(
            {
                "school_id": school_id,
                "division_type": body.division_type,
                "budget_name": body.budget_name,
                "goal_key": body.goal_key,
                "goal_type": goal_def["kind"],
                "goal_number": goal_def["goal_number"],
                "academic_year": body.academic_year,
                "met": body.met,
                "updated_at": datetime.now(timezone.utc).isoformat(),
            },
            on_conflict="school_id,division_type,budget_name,goal_key,academic_year",
        )
        .execute()
    )
    return row.data[0]


# ---------------------------------------------------------------------------
# Advisor assignments
# ---------------------------------------------------------------------------

@router.get("/{school_id}/advisors")
def list_advisors(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("advisor_schools")
                .select("advisor_id, profiles(id, email, full_name, role)")
                .eq("school_id", school_id)
                .execute()
            )
            return [r["profiles"] for r in rows.data if r.get("profiles")]
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_advisors attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_advisors failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/{school_id}/advisors")
def assign_advisor(
    school_id: str,
    body: AdvisorAssignIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db = get_admin_client()
    # temp_access_expires_at explicitly nulled — a manual/permanent assignment always wins
    # over any pending or already-active temporary grant for the same advisor+school pair
    # (see _process_temp_advisor_access below, which never touches a row with a null expiry).
    db.table("advisor_schools").upsert(
        {"advisor_id": body.advisor_id, "school_id": school_id, "temp_access_expires_at": None}
    ).execute()
    try:
        db.table("temp_advisor_access").update({"status": "superseded"}).eq(
            "advisor_id", body.advisor_id
        ).eq("school_id", school_id).in_("status", ["pending", "active"]).execute()
    except Exception as exc:
        logger.warning("marking temp_advisor_access superseded failed (non-fatal): %s", exc)
    try:
        school_row = db.table("schools").select("name").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
        _create_notifications(db, [{
            "recipient_id": body.advisor_id,
            "type": "advisor_assigned",
            "school_id": school_id,
            "data": {
                "title": f"שויכת לבית הספר {school_name}",
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "deeplink": f"/school/{school_id}",
            }
        }], pref_key="notify_advisor_assignment")
    except Exception as exc:
        logger.warning("advisor_assigned notification failed (non-fatal): %s", exc)
    return {"ok": True}


def _advisor_has_access_to_school_row(db, advisor_id: str, school: dict) -> bool:
    """Read-only mirror of the existing ad-hoc 3-way access check duplicated across this file
    (restrict_access_to null/list, or an advisor_schools row) — for a caller that already has
    the school row in hand (unlike _advisor_has_school_access above, which takes a school_id and
    fetches it itself). Named distinctly from that function (was previously an accidental
    same-name redefinition at module scope — Python silently keeps only the LAST def with a
    given name, so every one of _advisor_has_school_access's 9 call sites above was actually
    calling THIS function instead, with an incompatible (user_dict, school_id_str) argument
    order — a school_id string has no .get() method, so any non-manager advisor calling any of
    those endpoints (year-admin-data, control-letters, goals, ...) got a 500, not the intended
    403/allow. Fixed by giving this one its own name so the two stop colliding."""
    rat = school.get("restrict_access_to")
    if rat is None or advisor_id in (rat or []):
        return True
    row = (
        db.table("advisor_schools").select("advisor_id")
        .eq("advisor_id", advisor_id).eq("school_id", school["id"]).execute().data
    )
    return bool(row)


@router.get("/{school_id}/advisor-access")
def check_advisor_access(
    school_id: str,
    advisor_ids: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    """Round: advisor-access grant modal. advisor_ids is a comma-separated list; returns
    {advisor_id: bool} for whether each advisor currently has access to the school."""
    db = get_admin_client()
    school_row = db.table("schools").select("id, restrict_access_to").eq("id", school_id).execute()
    if not school_row.data:
        raise HTTPException(status_code=404, detail="בית ספר לא נמצא")
    school = school_row.data[0]
    ids = [a for a in advisor_ids.split(",") if a]
    return {aid: _advisor_has_access_to_school_row(db, aid, school) for aid in ids}


class AdvisorAccessMatrixIn(BaseModel):
    school_ids: list[str]
    advisor_ids: list[str]


@router.post("/advisor-access-matrix")
def advisor_access_matrix(
    body: AdvisorAccessMatrixIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    """Batch version of /{school_id}/advisor-access for cross-school meeting tables (ניהול-
    פגישות / אזור אישי), which render many schools' rows at once and need to know per-row
    which advisors lack access before the AdvisorCell picker can meaningfully flag them.
    Returns {school_id: {advisor_id: bool}}."""
    db = get_admin_client()
    if not body.school_ids or not body.advisor_ids:
        return {}
    schools = db.table("schools").select("id, restrict_access_to").in_("id", body.school_ids).execute().data or []
    school_ids = [s["id"] for s in schools]
    rows = (
        db.table("advisor_schools").select("advisor_id, school_id")
        .in_("school_id", school_ids).in_("advisor_id", body.advisor_ids).execute().data or []
    )
    granted: dict[str, set[str]] = {}
    for r in rows:
        granted.setdefault(r["school_id"], set()).add(r["advisor_id"])
    result = {}
    for s in schools:
        rat = s.get("restrict_access_to")
        school_granted = granted.get(s["id"], set())
        result[s["id"]] = {
            aid: (rat is None or aid in (rat or []) or aid in school_granted)
            for aid in body.advisor_ids
        }
    return result


class GrantTempAccessIn(BaseModel):
    starts_at: str
    expires_at: str
    source: str | None = None


@router.post("/{school_id}/advisors/{advisor_id}/grant-temp-access")
def grant_temp_advisor_access(
    school_id: str,
    advisor_id: str,
    body: GrantTempAccessIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db = get_admin_client()
    now = datetime.now(timezone.utc)
    starts_at = datetime.fromisoformat(body.starts_at.replace("Z", "+00:00"))
    row = {
        "school_id": school_id,
        "advisor_id": advisor_id,
        "starts_at": body.starts_at,
        "expires_at": body.expires_at,
        "status": "pending",
        "granted_by": user["id"],
        "source": body.source,
    }
    inserted = db.table("temp_advisor_access").insert(row).execute().data[0]
    if starts_at <= now:
        _activate_temp_access_row(db, inserted)
    return {"ok": True, "id": inserted["id"]}


def _format_date_dmy(iso_str: str) -> str:
    try:
        return datetime.fromisoformat(iso_str.replace("Z", "+00:00")).strftime("%d/%m/%Y")
    except Exception:
        return iso_str


def _activate_temp_access_row(db, row: dict) -> str:
    """Turns a pending temp_advisor_access row into a real advisor_schools grant. Never
    downgrades an existing permanent grant (temp_access_expires_at IS NULL) back to
    temporary — that row is left untouched and this one is marked superseded instead.
    Overlapping temp grants for the same advisor+school extend to the later expiry rather
    than being overwritten by whichever activates last. Returns the resulting status."""
    existing = (
        db.table("advisor_schools").select("temp_access_expires_at")
        .eq("advisor_id", row["advisor_id"]).eq("school_id", row["school_id"]).execute().data
    )
    if existing and existing[0].get("temp_access_expires_at") is None:
        db.table("temp_advisor_access").update({"status": "superseded"}).eq("id", row["id"]).execute()
        return "superseded"
    new_expiry = row["expires_at"]
    if existing and existing[0].get("temp_access_expires_at"):
        new_expiry = max(existing[0]["temp_access_expires_at"], row["expires_at"])
    db.table("advisor_schools").upsert(
        {"advisor_id": row["advisor_id"], "school_id": row["school_id"], "temp_access_expires_at": new_expiry}
    ).execute()
    db.table("temp_advisor_access").update({"status": "active"}).eq("id", row["id"]).execute()

    try:
        school_row = db.table("schools").select("name").eq("id", row["school_id"]).execute().data
        school_name = school_row[0]["name"] if school_row else "בית ספר"
        _create_notifications(db, [{
            "recipient_id": row["advisor_id"],
            "type": "temp_access_granted",
            "school_id": row["school_id"],
            "data": {
                "title": f"קיבלת גישה זמנית לבית הספר {school_name}, בתוקף עד {_format_date_dmy(new_expiry)}.",
                "school_name": school_name,
                "deeplink": f"/school/{row['school_id']}",
            },
        }])
    except Exception as exc:
        logger.warning("_activate_temp_access_row: notification failed (non-fatal): %s", exc)

    return "active"


def process_temp_advisor_access(db) -> dict:
    """Cron-driven (piggybacks on /tasks/process-scheduled-tasks, same 15-min tick as every
    other scheduled evaluation in this app — see task-scheduled-tasks.yml) — activates
    pending grants whose start date has arrived, and revokes advisor_schools access whose
    temp_access_expires_at has passed. Never touches a permanent (null-expiry) row."""
    now_iso = datetime.now(timezone.utc).isoformat()
    activated, expired = 0, 0

    pending = (
        db.table("temp_advisor_access").select("*")
        .eq("status", "pending").lte("starts_at", now_iso).execute().data or []
    )
    for row in pending:
        try:
            if _activate_temp_access_row(db, row) == "active":
                activated += 1
        except Exception as exc:
            logger.warning("process_temp_advisor_access: activation failed for row %s: %s", row["id"], exc)

    due = (
        db.table("advisor_schools").select("advisor_id, school_id")
        .not_.is_("temp_access_expires_at", "null").lte("temp_access_expires_at", now_iso)
        .execute().data or []
    )
    school_names = {}
    if due:
        try:
            rows = db.table("schools").select("id, name").in_("id", list({p["school_id"] for p in due})).execute().data or []
            school_names = {r["id"]: r["name"] for r in rows}
        except Exception as exc:
            logger.warning("process_temp_advisor_access: school-name lookup failed (non-fatal): %s", exc)
    for pair in due:
        try:
            db.table("advisor_schools").delete().eq("advisor_id", pair["advisor_id"]).eq("school_id", pair["school_id"]).execute()
            db.table("temp_advisor_access").update({"status": "expired"}).eq(
                "advisor_id", pair["advisor_id"]
            ).eq("school_id", pair["school_id"]).eq("status", "active").execute()
            expired += 1
            try:
                school_name = school_names.get(pair["school_id"], "בית ספר")
                _create_notifications(db, [{
                    "recipient_id": pair["advisor_id"],
                    "type": "temp_access_expired",
                    "school_id": pair["school_id"],
                    "data": {"title": f"הגישה הזמנית שלך לבית הספר {school_name} הסתיימה.", "school_name": school_name},
                }])
            except Exception as exc:
                logger.warning("process_temp_advisor_access: expiry notification failed (non-fatal): %s", exc)
        except Exception as exc:
            logger.warning("process_temp_advisor_access: expiry failed for %s: %s", pair, exc)

    return {"activated": activated, "expired": expired}


@router.delete("/{school_id}/advisors/{advisor_id}")
def unassign_advisor(
    school_id: str,
    advisor_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db = get_admin_client()
    existing = db.table("advisor_schools").select("advisor_id").eq("school_id", school_id).execute()
    current_ids = [r["advisor_id"] for r in (existing.data or [])]
    if advisor_id in current_ids and len(current_ids) <= 1:
        raise HTTPException(
            status_code=400,
            detail="לא ניתן להסיר את היועץ היחיד המשויך לבית הספר. יש לשייך יועץ אחר לפני ההסרה.",
        )
    try:
        school_row = db.table("schools").select("name").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
    except Exception:
        school_name = "בית ספר"
    db.table("advisor_schools").delete().eq("advisor_id", advisor_id).eq("school_id", school_id).execute()
    try:
        _create_notifications(db, [{
            "recipient_id": advisor_id,
            "type": "advisor_removed",
            "data": {
                "title": f"הוסרת מבית הספר {school_name}",
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
            }
        }], pref_key="notify_advisor_assignment")
    except Exception as exc:
        logger.warning("advisor_removed notification failed (non-fatal): %s", exc)
    return {"ok": True}


# ---------------------------------------------------------------------------
# Per-service-type advisor assignments ("יועץ מלווה [גפן/שוטף/מחוז]") — school_advisors_gefen/
# current/district. These are the "default advisor for this service type" lists, separate from
# advisor_schools (which stays the general access-control table). Assigning to a typed list also
# upserts into advisor_schools (so the advisor keeps/gains access); unassigning from a typed list
# only removes from advisor_schools if the advisor isn't left in any other typed list.
# ---------------------------------------------------------------------------

_TYPED_ADVISOR_TABLES = {
    "gefen": "school_advisors_gefen",
    "current": "school_advisors_current",
    "district": "school_advisors_district",
}


def _typed_advisor_table(service_type: str) -> str:
    table = _TYPED_ADVISOR_TABLES.get(service_type)
    if not table:
        raise HTTPException(status_code=400, detail="סוג שירות לא תקין")
    return table


@router.get("/{school_id}/advisors/{service_type}")
def list_typed_advisors(
    school_id: str,
    service_type: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    table = _typed_advisor_table(service_type)
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table(table)
                .select("advisor_id, profiles(id, email, full_name, role)")
                .eq("school_id", school_id)
                .execute()
            )
            return [r["profiles"] for r in rows.data if r.get("profiles")]
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_typed_advisors attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_typed_advisors failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/{school_id}/advisors/{service_type}")
def assign_typed_advisor(
    school_id: str,
    service_type: str,
    body: AdvisorAssignIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    table = _typed_advisor_table(service_type)
    db = get_admin_client()
    db.table(table).upsert({"advisor_id": body.advisor_id, "school_id": school_id}).execute()
    db.table("advisor_schools").upsert({"advisor_id": body.advisor_id, "school_id": school_id}).execute()
    try:
        school_row = db.table("schools").select("name").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
        _create_notifications(db, [{
            "recipient_id": body.advisor_id,
            "type": "advisor_assigned",
            "school_id": school_id,
            "data": {
                "title": f"שויכת לבית הספר {school_name}",
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "deeplink": f"/school/{school_id}",
            }
        }], pref_key="notify_advisor_assignment")
    except Exception as exc:
        logger.warning("advisor_assigned (typed) notification failed (non-fatal): %s", exc)
    return {"ok": True}


@router.delete("/{school_id}/advisors/{service_type}/{advisor_id}")
def unassign_typed_advisor(
    school_id: str,
    service_type: str,
    advisor_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    table = _typed_advisor_table(service_type)
    db = get_admin_client()

    db.table(table).delete().eq("advisor_id", advisor_id).eq("school_id", school_id).execute()

    # Only drop the advisor from the general advisor_schools access table if they're no longer
    # in ANY of the three typed lists for this school, and never if they're the school's last
    # advisor overall (mirrors the guard in unassign_advisor).
    still_typed = False
    for other_table in _TYPED_ADVISOR_TABLES.values():
        rows = (
            db.table(other_table)
            .select("advisor_id")
            .eq("school_id", school_id)
            .eq("advisor_id", advisor_id)
            .execute()
        )
        if rows.data:
            still_typed = True
            break

    if not still_typed:
        existing = db.table("advisor_schools").select("advisor_id").eq("school_id", school_id).execute()
        current_ids = [r["advisor_id"] for r in (existing.data or [])]
        if advisor_id in current_ids and len(current_ids) > 1:
            db.table("advisor_schools").delete().eq("advisor_id", advisor_id).eq("school_id", school_id).execute()
            try:
                school_row = db.table("schools").select("name").eq("id", school_id).execute()
                school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
                _create_notifications(db, [{
                    "recipient_id": advisor_id,
                    "type": "advisor_removed",
                    "data": {
                        "title": f"הוסרת מבית הספר {school_name}",
                        "school_name": school_name,
                        "sender_name": user.get("full_name", ""),
                    }
                }], pref_key="notify_advisor_assignment")
            except Exception as exc:
                logger.warning("advisor_removed (typed) notification failed (non-fatal): %s", exc)
        # If they're the last advisor overall, we silently leave them in advisor_schools —
        # the frontend's own "at least one advisor per active service type" validation is what
        # actually prevents unassigning someone required by the currently-selected service_type.

    return {"ok": True}


# ---------------------------------------------------------------------------
# Update requests (advisors submit → owner/manager approve)
# ---------------------------------------------------------------------------

class UpdateRequestIn(BaseModel):
    proposed_changes: dict


class ReviewRequestIn(BaseModel):
    status: str  # "approved" | "rejected"
    reviewer_note: str | None = None
    approved_fields: list[str] | None = None  # if set, only these fields are applied on approval


_approver_ids_cache: dict[str, tuple[list[str], float]] = {}
_APPROVER_TTL = 300  # 5 minutes


def invalidate_approver_ids_cache() -> None:
    global _approver_ids_cache
    _approver_ids_cache = {}


def _get_approver_ids(db, org_id: str) -> list[str]:
    """Return IDs of who should receive approval notifications within an org.
    Always includes all owners. Includes each manager only if they have
    can_approve_update_requests permission (via override or role default).
    Result cached per org for 5 minutes.
    """
    global _approver_ids_cache
    now = time.monotonic()
    cached = _approver_ids_cache.get(org_id)
    if cached and (now - cached[1]) < _APPROVER_TTL:
        return cached[0]

    all_candidates = (
        db.table("profiles")
        .select("id, role")
        .in_("role", ["owner", "manager"])
        .eq("org_id", org_id)
        .execute()
    )
    candidates = all_candidates.data or []
    owners = [r for r in candidates if r["role"] == "owner"]
    managers = [r for r in candidates if r["role"] == "manager"]
    ids = [r["id"] for r in owners]
    # Include each manager individually based on their effective permission
    for mgr in managers:
        mgr_user = {"id": mgr["id"], "role": "manager", "org_id": org_id}
        if _check_permission(db, mgr_user, "can_approve_update_requests"):
            ids.append(mgr["id"])
    _approver_ids_cache[org_id] = (ids, now)
    return ids


@router.post("/{school_id}/update-requests")
def submit_update_request(
    school_id: str,
    body: UpdateRequestIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    is_delete_action = body.proposed_changes.get("_action") == "delete_school"
    if not is_delete_action and not _check_permission(db, user, "can_edit_school_directly"):
        if not _check_permission(db, user, "can_request_school_update"):
            raise HTTPException(status_code=403, detail="אין הרשאה להגיש בקשות עריכה")
    row = db.table("school_update_requests").insert({
        "school_id": school_id,
        "requester_id": user["id"],
        "proposed_changes": body.proposed_changes,
        "status": "pending",
    }).execute()
    req_id = row.data[0]["id"]
    try:
        school_row = db.table("schools").select("*").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
        approver_ids = _get_approver_ids(db, user["org_id"])
        if is_delete_action:
            notif_title = f'{user.get("full_name", "יועץ")} הגיש בקשה למחיקת בית הספר {school_name}'
            extra_data = {"is_delete_request": True}
        else:
            notif_title = f'{user.get("full_name", "יועץ")} ביקש לערוך את פרטי {school_name}'
            # Include current field values so approvers can see what's changing from→to
            school_obj = school_row.data[0] if school_row.data else {}
            current_values = {
                k: school_obj.get(k)
                for k in body.proposed_changes
                if k != "_action"
            }
            extra_data = {"current_values": current_values}
        notif_rows = [{
            "recipient_id": aid,
            "type": "update_request_submitted",
            "ref_id": req_id,
            "school_id": school_id,
            "data": {
                "title": notif_title,
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "proposed_changes": body.proposed_changes,
                "deeplink": "/notifications",
                **extra_data,
            }
        } for aid in approver_ids if aid != user["id"]]
        _create_notifications(db, notif_rows, pref_key="notify_update_request_submitted")
    except Exception as exc:
        logger.warning("update_request_submitted notification failed (non-fatal): %s", exc)
    return row.data[0]


@router.get("/update-requests")
def list_update_requests(user: Annotated[dict, Depends(get_current_user)]):
    for attempt in range(2):
        try:
            db = get_admin_client()
            if user["role"] in ("owner", "manager"):
                school_ids_res = db.table("schools").select("id").eq("org_id", user["org_id"]).execute()
                school_ids = [r["id"] for r in (school_ids_res.data or [])]
                if not school_ids:
                    return []
                rows = (
                    db.table("school_update_requests")
                    .select("*, schools(name), requester:profiles!requester_id(full_name, email)")
                    .in_("school_id", school_ids)
                    .order("created_at", desc=True)
                    .execute()
                )
            else:
                rows = (
                    db.table("school_update_requests")
                    .select("*, schools(name), requester:profiles!requester_id(full_name, email)")
                    .eq("requester_id", user["id"])
                    .order("created_at", desc=True)
                    .execute()
                )
            return rows.data
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_update_requests attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_update_requests failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.patch("/update-requests/{req_id}")
def review_update_request(
    req_id: str,
    body: ReviewRequestIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    if body.status not in ("approved", "rejected"):
        raise HTTPException(status_code=400, detail="סטטוס לא חוקי")

    # Phase 1: read + validate only (NO status update — must happen after physical action)
    req = None
    school_name = "בית ספר"
    is_delete_req = False
    for attempt in range(2):
        try:
            db = get_admin_client()
            if not _check_permission(db, user, "can_approve_update_requests"):
                raise HTTPException(status_code=403, detail="אין הרשאה לאשר בקשות עדכון")
            req_row = db.table("school_update_requests").select("*").eq("id", req_id).execute()
            if not req_row.data:
                raise HTTPException(status_code=404, detail="הבקשה לא נמצאה")
            req = req_row.data[0]
            if req["status"] != "pending":
                raise HTTPException(status_code=400, detail="הבקשה כבר טופלה")
            if req.get("school_id"):
                school_row = db.table("schools").select("name").eq("id", req["school_id"]).execute()
                school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
            is_delete_req = req.get("proposed_changes", {}).get("_action") == "delete_school"
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("review_update_request attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("review_update_request failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    # Phase 2: apply changes (school deletion may cascade-delete many rows → may timeout)
    school_deleted = False
    if body.status == "approved" and req.get("proposed_changes"):
        changes = dict(req["proposed_changes"])
        if changes.get("_action") == "delete_school":
            for del_attempt in range(2):
                try:
                    db = get_admin_client()
                    from datetime import datetime, timezone
                    db.table("schools").update({
                        "status": "pending_deletion",
                        "deleted_at": datetime.now(timezone.utc).isoformat(),
                    }).eq("id", req["school_id"]).execute()
                    school_deleted = True
                    break
                except Exception as del_exc:
                    if del_attempt == 0:
                        logger.warning("school soft-delete attempt 1 failed: %s — resetting and retrying", del_exc)
                        reset_admin_client()
                        time.sleep(0.3)
                    else:
                        logger.error("school soft-delete failed after 2 attempts: %s", del_exc, exc_info=True)
                        raise HTTPException(status_code=500, detail="שגיאה זמנית בעדכון בית הספר — נסה שוב")
        else:
            db = get_admin_client()
            # Partial approval: if approved_fields is specified, only apply those fields
            if body.approved_fields is not None:
                changes = {k: v for k, v in changes.items() if k in body.approved_fields}
            if "add_advisor_to_school" in changes:
                advisor_id = changes.pop("add_advisor_to_school")
                db.table("advisor_schools").upsert({"advisor_id": advisor_id, "school_id": req["school_id"]}).execute()
            if changes:
                db.table("schools").update(changes).eq("id", req["school_id"]).execute()

    # Phase 3: update request status — AFTER physical action succeeds
    from datetime import datetime, timezone
    for status_attempt in range(2):
        try:
            db = get_admin_client()
            db.table("school_update_requests").update({
                "status": body.status,
                "reviewer_id": user["id"],
                "reviewer_note": body.reviewer_note,
                "resolved_at": datetime.now(timezone.utc).isoformat(),
            }).eq("id", req_id).execute()
            break
        except Exception as exc:
            if status_attempt == 0:
                logger.warning("status update attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.warning("status update failed after 2 attempts (non-fatal — action already applied): %s", exc)

    # Phase 4: notifications — each section independently non-fatal
    action_label = "מחיקת" if is_delete_req else "עריכת"
    notif_school_id = None if school_deleted else req.get("school_id")
    deeplink_advisor = "/notifications" if school_deleted else f'/school/{req.get("school_id")}'

    # Compute partial-approval metadata for advisor notification
    all_field_keys = [k for k in (req.get("proposed_changes") or {}) if k != "_action"]
    approved_fields_list = body.approved_fields  # None = full decision; list = partial
    is_partial = (
        not is_delete_req
        and approved_fields_list is not None
        and body.status == "approved"
        and len(approved_fields_list) < len(all_field_keys)
    )
    if is_partial:
        advisor_title = f'בקשתך לעריכת {school_name} אושרה באופן חלקי'
    elif body.status == "approved":
        advisor_title = f'בקשתך ל{action_label} {school_name} אושרה'
    else:
        advisor_title = f'בקשתך ל{action_label} {school_name} נדחתה'

    try:
        db = get_admin_client()
        _create_notifications(db, [{
            "recipient_id": req["requester_id"],
            "type": f"update_request_{body.status}",
            "ref_id": req_id,
            "school_id": notif_school_id,
            "data": {
                "title": advisor_title,
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "reviewer_note": body.reviewer_note,
                "deeplink": deeplink_advisor,
                "proposed_changes": req.get("proposed_changes"),
                "approved_fields": approved_fields_list,
                "is_partial": is_partial,
            }
        }], pref_key="notify_update_request_reviewed")
    except Exception as exc:
        logger.warning("advisor review notification failed (non-fatal): %s", exc)

    if user.get("role") != "owner":
        try:
            db = get_admin_client()
            req_submitter_name = ""
            try:
                sub_row = db.table("profiles").select("full_name").eq("id", req["requester_id"]).execute()
                req_submitter_name = sub_row.data[0]["full_name"] if sub_row.data else ""
            except Exception:
                pass
            owner_rows = db.table("profiles").select("id").eq("role", "owner").eq("org_id", user["org_id"]).execute()
            owner_ids = [r["id"] for r in (owner_rows.data or []) if r["id"] != user["id"]]
            if is_delete_req and body.status == "approved":
                owner_title = f'{user.get("full_name", "מנהל")} אישר את בקשת המחיקה של {school_name}'
                if req_submitter_name:
                    owner_title += f' שהוגשה על ידי {req_submitter_name}'
                owner_title += ' — הנתונים יימחקו סופית תוך 30 יום'
            elif is_delete_req:
                owner_title = f'{user.get("full_name", "מנהל")} דחה את בקשת המחיקה של {school_name}'
                if req_submitter_name:
                    owner_title += f' שהוגשה על ידי {req_submitter_name}'
            else:
                if is_partial:
                    status_verb = "אישר באופן חלקי"
                elif body.status == "approved":
                    status_verb = "אישר"
                else:
                    status_verb = "דחה"
                owner_title = f'{user.get("full_name", "מנהל")} {status_verb} את בקשת העריכה של {school_name}'
                if req_submitter_name:
                    owner_title += f' שהוגשה על ידי {req_submitter_name}'
            owner_notif_rows = [{
                "recipient_id": oid,
                "type": "update_request_result",
                "ref_id": req_id,
                "school_id": notif_school_id,
                "data": {
                    "title": owner_title,
                    "school_name": school_name,
                    "reviewer_name": user.get("full_name", ""),
                    "requester_name": req_submitter_name,
                    "status": body.status,
                    "reviewer_note": body.reviewer_note,
                    "deeplink": "/notifications",
                    "proposed_changes": req.get("proposed_changes"),
                    "approved_fields": approved_fields_list,
                    "is_partial": is_partial,
                }
            } for oid in owner_ids]
            _create_notifications(db, owner_notif_rows, pref_key="notify_update_request_result")
        except Exception as exc:
            logger.warning("owner review notification failed (non-fatal): %s", exc)

    return {"ok": True}


@router.get("/notifications")
def get_notifications(user: Annotated[dict, Depends(get_current_user)]):
    """Return all notifications for the current user from the unified notifications table."""
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("notifications")
                .select("*")
                .eq("recipient_id", user["id"])
                .order("created_at", desc=True)
                .limit(50)
                .execute()
            )
            items = rows.data or []
            # Enrich update_request_submitted notifications with the current request status
            # so the frontend can show the correct reviewed state on remount (non-fatal)
            try:
                action_ids = [n["ref_id"] for n in items
                              if n.get("type") == "update_request_submitted" and n.get("ref_id")]
                if action_ids:
                    req_rows = (db.table("school_update_requests")
                                .select("id, status")
                                .in_("id", action_ids)
                                .execute())
                    status_map = {r["id"]: r["status"] for r in (req_rows.data or [])}
                    for n in items:
                        if n.get("type") == "update_request_submitted" and n.get("ref_id"):
                            n["data"] = {**(n.get("data") or {}),
                                         "request_status": status_map.get(n["ref_id"])}
            except Exception as enrich_exc:
                logger.warning("request status enrichment failed (non-fatal): %s", enrich_exc)
            count = sum(1 for r in items if not r.get("read_at"))
            return {"count": count, "items": items}
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_notifications attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.warning("get_notifications failed after 2 attempts: %s", exc)
                return {"count": 0, "items": []}  # silent fallback — not critical


@router.patch("/notifications/read-all")
def mark_all_notifications_read(user: Annotated[dict, Depends(get_current_user)]):
    from datetime import datetime, timezone
    try:
        db = get_admin_client()
        db.table("notifications").update({"read_at": datetime.now(timezone.utc).isoformat()}).eq("recipient_id", user["id"]).is_("read_at", "null").execute()
    except Exception as exc:
        logger.warning("mark_all_notifications_read failed: %s", exc)
    return {"ok": True}


@router.patch("/notifications/{notification_id}/read")
def mark_notification_read(notification_id: str, user: Annotated[dict, Depends(get_current_user)]):
    from datetime import datetime, timezone
    try:
        db = get_admin_client()
        db.table("notifications").update({"read_at": datetime.now(timezone.utc).isoformat()}).eq("id", notification_id).eq("recipient_id", user["id"]).is_("read_at", "null").execute()
    except Exception as exc:
        logger.warning("mark_notification_read failed: %s", exc)
    return {"ok": True}


@router.get("/upcoming-meetings")
def get_upcoming_meetings(user: Annotated[dict, Depends(get_current_user)]):
    """Return today's meetings for the current user (for frontend meeting reminders)."""
    from datetime import datetime, timezone
    today = datetime.now(timezone.utc).date().isoformat()
    try:
        db = get_admin_client()
        if user["role"] in ("owner", "manager"):
            school_ids_res = db.table("schools").select("id").eq("org_id", user["org_id"]).execute()
            school_ids = [r["id"] for r in (school_ids_res.data or [])]
        else:
            assigned = db.table("advisor_schools").select("school_id").eq("advisor_id", user["id"]).execute()
            school_ids = [r["school_id"] for r in (assigned.data or [])]

        if not school_ids:
            return []

        meetings_res = (
            db.table("meetings")
            .select("id, school_id, meeting_date, start_time, end_time, status, advisor_ids, notes")
            .in_("school_id", school_ids)
            .eq("meeting_date", today)
            .execute()
        )
        meetings = meetings_res.data or []

        if user["role"] not in ("owner", "manager"):
            meetings = [m for m in meetings if user["id"] in (m.get("advisor_ids") or [])]

        if meetings:
            s_ids = list({m["school_id"] for m in meetings})
            school_names = {r["id"]: r["name"] for r in (db.table("schools").select("id, name").in_("id", s_ids).execute().data or [])}
            # Fetch school-level advisors so the frontend knows fallback recipients for status reminders
            school_advisors_rows = db.table("advisor_schools").select("school_id, advisor_id").in_("school_id", s_ids).execute().data or []
            school_advisors_map: dict = {}
            for row in school_advisors_rows:
                school_advisors_map.setdefault(row["school_id"], []).append(row["advisor_id"])
            for m in meetings:
                m["school_name"] = school_names.get(m["school_id"], "")
                m["school_advisor_ids"] = school_advisors_map.get(m["school_id"], [])
        return meetings
    except Exception as exc:
        logger.warning("get_upcoming_meetings failed (non-fatal): %s", exc)
        return []


@router.patch("/users/me/notification-preferences")
def update_notification_preferences(
    body: NotificationPreferencesIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    current = user.get("notification_preferences") or {"meeting_reminder": True, "meeting_reminder_minutes": 10}
    new_prefs = dict(current)
    if body.meeting_reminder is not None:
        new_prefs["meeting_reminder"] = body.meeting_reminder
    if body.meeting_reminder_minutes is not None:
        new_prefs["meeting_reminder_minutes"] = body.meeting_reminder_minutes
    if body.notify_call_contact_ambiguous is not None:
        new_prefs["notify_call_contact_ambiguous"] = body.notify_call_contact_ambiguous
    db.table("profiles").update({"notification_preferences": new_prefs}).eq("id", user["id"]).execute()
    invalidate_profile_cache(user["id"])
    return {"ok": True, "notification_preferences": new_prefs}


# ---------------------------------------------------------------------------
# Meetings stats (aggregate per school — used by dashboard columns)
# ---------------------------------------------------------------------------

@router.get("/meetings-stats")
def get_meetings_stats(user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()

    try:
        if user["role"] in ("owner", "manager"):
            # Owners/managers see all schools — skip the schools filter query entirely
            meetings_res = db.table("meetings").select("school_id, status, start_time, end_time").execute()
        else:
            # Advisors: fetch schools + assignments in parallel, then meetings
            with ThreadPoolExecutor(max_workers=2) as pool:
                schools_future = pool.submit(
                    lambda: db.table("schools").select("id, restrict_access_to").execute()
                )
                assigned_future = pool.submit(
                    lambda: db.table("advisor_schools").select("school_id").eq("advisor_id", user["id"]).execute()
                )
                all_schools = schools_future.result().data or []
                advisor_ids = {r["school_id"] for r in (assigned_future.result().data or [])}

            accessible = [
                s["id"] for s in all_schools
                if s.get("restrict_access_to") is None
                or user["id"] in (s.get("restrict_access_to") or [])
                or s["id"] in advisor_ids
            ]
            if not accessible:
                return {}
            meetings_res = db.table("meetings").select("school_id, status, start_time, end_time").in_("school_id", accessible).execute()
    except Exception as exc:
        logger.warning("get_meetings_stats failed: %s", exc)
        return {}
    stats: dict = {}
    for m in (meetings_res.data or []):
        sid = m["school_id"]
        if sid not in stats:
            stats[sid] = {"completed": 0, "total_minutes": 0}
        if m.get("status") == "completed":
            stats[sid]["completed"] += 1
            st, et = m.get("start_time"), m.get("end_time")
            if st and et:
                try:
                    sh, sm = map(int, st.split(":"))
                    eh, em = map(int, et.split(":"))
                    diff = (eh * 60 + em) - (sh * 60 + sm)
                    if diff > 0:
                        stats[sid]["total_minutes"] += diff
                except Exception:
                    pass
    return stats


@router.get("/meetings/all")
def list_all_meetings(
    user: Annotated[dict, Depends(get_current_user)],
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    advisor_id: str | None = None,
    school_id: str | None = None,
    search: str | None = None,
    academic_year: str | None = None,
):
    """Org-wide meetings list for the admin 'פגישות' tab. Owner/manager only."""
    _require_manager(user)
    meetings: list = []
    schools_map: dict = {}

    for attempt in range(2):
        try:
            db = get_admin_client()

            schools_q = (
                db.table("schools")
                .select("id, name, symbol, city, authority, district")
                .eq("org_id", user["org_id"])
                .eq("status", "active")
            )
            if school_id:
                schools_q = schools_q.eq("id", school_id)
            if search and search.strip():
                s = search.strip().replace(",", "")
                schools_q = schools_q.or_(f"name.ilike.%{s}%,symbol.ilike.%{s}%,city.ilike.%{s}%")
            schools_rows = schools_q.execute().data or []
            school_ids = [s["id"] for s in schools_rows]
            schools_map = {s["id"]: s for s in schools_rows}

            if not school_ids:
                meetings = []
                break

            q = db.table("meetings").select("*").in_("school_id", school_ids)
            if status:
                q = q.eq("status", status)
            if date_from:
                q = q.gte("meeting_date", date_from)
            if date_to:
                q = q.lte("meeting_date", date_to)
            if advisor_id:
                q = q.filter("advisor_ids", "cs", json.dumps([advisor_id]))
            if academic_year:
                q = q.eq("academic_year", academic_year)
            res = q.order("meeting_date", desc=True).execute()
            meetings = res.data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_all_meetings attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_all_meetings failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    for m in meetings:
        sc = schools_map.get(m.get("school_id"), {})
        m["school_name"] = sc.get("name", "")
        m["school_symbol"] = sc.get("symbol", "")
        m["school_city"] = sc.get("city", "")
        m["school_authority"] = sc.get("authority", "")
        m["school_district"] = sc.get("district", "")

    all_ids: set[str] = set()
    for m in meetings:
        for uid in (m.get("advisor_ids") or []):
            all_ids.add(uid)
        if m.get("advisor_id"):
            all_ids.add(m["advisor_id"])

    if all_ids:
        try:
            db = get_admin_client()
            profiles = db.table("profiles").select("id, full_name, email").in_("id", list(all_ids)).execute().data or []
            profiles_map = {p["id"]: p for p in profiles}
            for m in meetings:
                ids = m.get("advisor_ids") or ([m["advisor_id"]] if m.get("advisor_id") else [])
                m["advisor_profiles"] = [profiles_map[uid] for uid in ids if uid in profiles_map]
        except Exception as exc:
            logger.warning("list_all_meetings profile enrichment failed (non-fatal): %s", exc)
            for m in meetings:
                m["advisor_profiles"] = []
    else:
        for m in meetings:
            m["advisor_profiles"] = []

    return meetings


# ---------------------------------------------------------------------------
# Personal meetings (current user's own meetings — all roles)
# ---------------------------------------------------------------------------

@router.get("/meetings/my")
def list_my_meetings(
    user: Annotated[dict, Depends(get_current_user)],
    status: str | None = None,
    date_from: str | None = None,
    date_to: str | None = None,
    academic_year: str | None = None,
):
    """Return meetings where the current user is listed as an advisor. Available to all roles."""
    for attempt in range(2):
        try:
            db = get_admin_client()

            # All school IDs in the org (needed to scope the query)
            schools_res = db.table("schools").select("id, name, symbol, city, district") \
                .eq("org_id", user["org_id"]).eq("status", "active").execute()
            schools_map = {s["id"]: s for s in (schools_res.data or [])}
            school_ids = list(schools_map.keys())

            if not school_ids:
                return []

            q = db.table("meetings").select("*") \
                .in_("school_id", school_ids) \
                .filter("advisor_ids", "cs", json.dumps([user["id"]]))
            if status:
                q = q.eq("status", status)
            if date_from:
                q = q.gte("meeting_date", date_from)
            if date_to:
                q = q.lte("meeting_date", date_to)
            if academic_year:
                q = q.eq("academic_year", academic_year)
            res = q.order("meeting_date", desc=True).execute()
            meetings = res.data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_my_meetings attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_my_meetings failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    for m in meetings:
        sc = schools_map.get(m.get("school_id"), {})
        m["school_name"] = sc.get("name", "")
        m["school_symbol"] = sc.get("symbol", "")
        m["school_city"] = sc.get("city", "")
        m["school_district"] = sc.get("district", "")

    if meetings:
        try:
            db = get_admin_client()
            all_ids: set[str] = set()
            for m in meetings:
                for uid in (m.get("advisor_ids") or []):
                    all_ids.add(uid)
            if all_ids:
                profiles = db.table("profiles").select("id, full_name, email") \
                    .in_("id", list(all_ids)).execute().data or []
                profiles_map = {p["id"]: p for p in profiles}
                for m in meetings:
                    ids = m.get("advisor_ids") or []
                    m["advisor_profiles"] = [profiles_map[uid] for uid in ids if uid in profiles_map]
        except Exception as exc:
            logger.warning("list_my_meetings profile enrichment failed (non-fatal): %s", exc)
            for m in meetings:
                m.setdefault("advisor_profiles", [])
    else:
        pass

    return meetings


# ---------------------------------------------------------------------------
# Meeting status reminders (manual trigger by manager/owner)
# ---------------------------------------------------------------------------

@router.post("/meetings/{meeting_id}/send-status-reminder")
def send_status_reminder(
    meeting_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    force: bool = False,
):
    """Send a manual status-update reminder to the meeting's advisors."""
    from datetime import datetime, timezone, timedelta
    _require_manager(user)

    for attempt in range(2):
        try:
            db = get_admin_client()

            # Fetch meeting
            m_res = db.table("meetings").select("*").eq("id", meeting_id).execute()
            if not m_res.data:
                raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
            meeting = m_res.data[0]

            if meeting["status"] != "scheduled":
                raise HTTPException(status_code=400, detail="הפגישה אינה במצב 'נקבעה'")

            now = datetime.now(timezone.utc)
            cutoff = (now - timedelta(hours=72)).isoformat()

            # Check for existing recent reminder (within 72h)
            existing_res = db.table("meeting_status_reminders").select("id, sent_at, recipient_id") \
                .eq("meeting_id", meeting_id).gte("sent_at", cutoff).execute()

            if existing_res.data and not force:
                last = max(existing_res.data, key=lambda r: r["sent_at"])
                recipient_ids = list({r["recipient_id"] for r in existing_res.data})
                p_res = db.table("profiles").select("id, full_name, email") \
                    .in_("id", recipient_ids).execute()
                profiles_map = {p["id"]: p for p in (p_res.data or [])}
                return {
                    "already_sent": True,
                    "last_sent_at": last["sent_at"],
                    "recipients": [profiles_map.get(rid, {"id": rid}) for rid in recipient_ids],
                }

            # Determine recipients: meeting advisors, else school's assigned advisors
            advisor_ids = meeting.get("advisor_ids") or []
            if not advisor_ids and meeting.get("advisor_id"):
                advisor_ids = [meeting["advisor_id"]]
            if not advisor_ids:
                school_advisors = db.table("advisor_schools").select("advisor_id") \
                    .eq("school_id", meeting["school_id"]).execute()
                advisor_ids = [r["advisor_id"] for r in (school_advisors.data or [])]

            if not advisor_ids:
                raise HTTPException(status_code=400, detail="לא נמצאו יועצים מוקצים לפגישה זו")

            records = [{
                "meeting_id": meeting_id,
                "school_id": meeting["school_id"],
                "recipient_id": aid,
                "sent_by": user["id"],
            } for aid in advisor_ids]
            db.table("meeting_status_reminders").insert(records).execute()

            p_res = db.table("profiles").select("id, full_name, email") \
                .in_("id", advisor_ids).execute()
            profiles_map = {p["id"]: p for p in (p_res.data or [])}
            return {
                "ok": True,
                "already_sent": False,
                "recipients": [profiles_map.get(aid, {"id": aid}) for aid in advisor_ids],
                "sent_at": now.isoformat(),
            }
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("send_status_reminder attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("send_status_reminder failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת")


@router.get("/meetings/pending-status-reminders")
def get_pending_status_reminders(user: Annotated[dict, Depends(get_current_user)]):
    """Return pending (unshown) manual status reminders for the current user."""
    from datetime import datetime, timezone, timedelta
    for attempt in range(2):
        try:
            db = get_admin_client()
            now = datetime.now(timezone.utc)
            cutoff = (now - timedelta(hours=72)).isoformat()

            rem_res = db.table("meeting_status_reminders").select("id, meeting_id, school_id, sent_at") \
                .eq("recipient_id", user["id"]).is_("shown_at", "null").gte("sent_at", cutoff).execute()
            reminders = rem_res.data or []
            if not reminders:
                return []

            meeting_ids = list({r["meeting_id"] for r in reminders})
            m_res = db.table("meetings").select(
                "id, school_id, meeting_date, start_time, end_time, status, notes, advisor_ids"
            ).in_("id", meeting_ids).execute()
            meetings_map = {m["id"]: m for m in (m_res.data or [])}

            school_ids = list({r["school_id"] for r in reminders})
            s_res = db.table("schools").select("id, name").in_("id", school_ids).execute()
            schools_map = {s["id"]: s["name"] for s in (s_res.data or [])}

            result = []
            for r in reminders:
                m = meetings_map.get(r["meeting_id"])
                if not m:
                    continue
                if m["status"] != "scheduled":
                    try:
                        db.table("meeting_status_reminders").update({"shown_at": now.isoformat()}).eq("id", r["id"]).execute()
                    except Exception:
                        pass
                    continue
                result.append({
                    "reminder_id": r["id"],
                    "id": m["id"],
                    "school_id": m["school_id"],
                    "school_name": schools_map.get(m["school_id"], ""),
                    "meeting_date": m["meeting_date"],
                    "start_time": m["start_time"],
                    "end_time": m["end_time"],
                    "status": m["status"],
                    "notes": m.get("notes"),
                    "advisor_ids": m.get("advisor_ids") or [],
                })
            return result
        except Exception as exc:
            if attempt == 0:
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.warning("get_pending_status_reminders failed (non-fatal): %s", exc)
                return []


@router.patch("/meetings/status-reminders/{reminder_id}/mark-shown")
def mark_status_reminder_shown(
    reminder_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    """Mark a manual status reminder as shown (called after popup is dismissed)."""
    from datetime import datetime, timezone
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("meeting_status_reminders").update({"shown_at": datetime.now(timezone.utc).isoformat()}) \
                .eq("id", reminder_id).eq("recipient_id", user["id"]).execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.warning("mark_status_reminder_shown failed (non-fatal): %s", exc)
                return {"ok": False}


# ---------------------------------------------------------------------------
# Automated meeting reminder emails (daily cron-triggered, Phase 1)
# ---------------------------------------------------------------------------

CRON_SECRET = os.getenv("CRON_SECRET", "")


_HEBREW_WEEKDAY_NAMES = {
    0: "יום שני", 1: "יום שלישי", 2: "יום רביעי", 3: "יום חמישי",
    4: "יום שישי", 5: "יום שבת", 6: "יום ראשון",
}


def _due_meeting_dates(today):
    """Dates whose meetings should be reminded today. The daily trigger only
    runs Sun-Thu (Israeli business week); Thursday's run must also absorb
    Saturday and Sunday meetings since there is no Fri/Sat run to catch them."""
    from datetime import timedelta
    tomorrow = today + timedelta(days=1)
    dates = [tomorrow]
    if today.weekday() == 3:  # Thursday
        dates += [tomorrow + timedelta(days=1), tomorrow + timedelta(days=2)]  # Saturday, Sunday
    return dates


def _day_phrases(meeting_date: str, today: "date") -> tuple[str, str]:
    """Returns (ל-form, ב-form) of the meeting day, e.g. ('למחר', 'מחר') or
    ('ליום ראשון', 'ביום ראשון') — used when the reminder is sent more than
    one day ahead (Thursday's run absorbing Saturday/Sunday meetings)."""
    from datetime import date, timedelta
    d = date.fromisoformat(meeting_date)
    if d == today + timedelta(days=1):
        return "למחר", "מחר"
    day_name = _HEBREW_WEEKDAY_NAMES[d.weekday()]
    return f"ל{day_name}", f"ב{day_name}"


def _hebrew_join(names: list[str]) -> str:
    names = [n for n in names if n]
    if not names:
        return ""
    if len(names) == 1:
        return names[0]
    return ", ".join(names[:-1]) + " ו" + names[-1]


def _opt_out_footer_html(opt_out_link: str | None) -> str:
    """Shared footer line for the reminder/upload-request email builders below — mirrors
    booking_logic._opt_out_footer_html's wording/placement for the direct-coordination and
    booking-agent emails, kept as a separate copy here to avoid importing booking_logic just
    for this one string (schools_router already imports it locally where actually needed)."""
    if not opt_out_link:
        return ""
    return f'<p style="margin: 8px 0 0 0; font-size: 11px; color: #94a3b8;">' \
           f'<a href="{opt_out_link}" style="color: #94a3b8;">להסרה מרשימת התפוצה</a></p>'


def _build_reminder_email_html(recipient_name: str, when_lamed: str, when_bet: str, meeting_date: str,
                                start_time: str | None, advisor_name: str, meeting_service_type: str = "gefen",
                                opt_out_link: str | None = None) -> str:
    from datetime import date
    first_name = (recipient_name or "").strip().split(" ")[0]
    greeting = f"היי {first_name}," if first_name else "היי,"
    date_fmt = date.fromisoformat(meeting_date).strftime("%d/%m/%y")
    advisor_clause = f" עם {advisor_name}" if advisor_name else ""
    time_clause = f", בשעה {start_time}" if start_time else ""
    # Round 16 — this used to hardcode "על תקציב הגפ\"ן" regardless of the meeting's actual
    # type, so a שוטף/מחוז reminder said the wrong thing. Now reflects meeting_service_type;
    # "gefen" keeps the exact original wording (no visible change for existing recipients).
    if meeting_service_type == "current":
        body_line = f'רצינו להזכיר לך על הפגישה השוטפת על התוכנה הכספית שמתוכננת {when_lamed}, בתאריך <b>{date_fmt}</b>{time_clause}{advisor_clause}.'
    elif meeting_service_type == "district":
        body_line = f'רצינו להזכיר לך על הפגישה שמתוכננת {when_lamed}, בתאריך <b>{date_fmt}</b>{time_clause}{advisor_clause} בנושא המחוז.'
    else:
        body_line = f'רצינו להזכיר לך על הפגישה שמתוכננת {when_lamed}, בתאריך <b>{date_fmt}</b>{time_clause}{advisor_clause} על תקציב הגפ"ן.'
    opt_out_html = _opt_out_footer_html(opt_out_link)
    return f"""
<html>
<body dir="rtl" style="font-family: Arial, sans-serif; font-size: 14px; color: #1e293b;
                       background: #f8fafc; margin: 0; padding: 24px;">
  <div style="max-width: 520px; margin: 0 auto; background: white;
              border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden;">
    <div style="background: #0070F3; padding: 20px 24px;">
      <p style="margin: 0; color: white; font-size: 14px; font-weight: 700;">גפן AI</p>
      <p style="margin: 4px 0 0 0; color: rgba(255,255,255,0.8); font-size: 12px;">תזכורת פגישה</p>
    </div>
    <div style="padding: 28px 24px;">
      <p style="margin: 0 0 16px 0; font-size: 15px;">{greeting}</p>
      <p style="margin: 0 0 16px 0; color: #334155; line-height: 1.8;">
        {body_line}
      </p>
      <p style="margin: 0; color: #334155; line-height: 1.8;">
        נתראה {when_bet} :)
      </p>
    </div>
    <div style="background: #f1f5f9; padding: 12px 24px; text-align: center;">
      <p style="margin: 0; font-size: 11px; color: #94a3b8;">נשלח אוטומטית מגפן AI</p>
      {opt_out_html}
    </div>
  </div>
</body>
</html>"""


def _build_secretary_upload_email_html(recipient_name: str, when_bet: str, school_name: str,
                                        checklist_items: list[str], upload_url: str, no_baseline: bool,
                                        meeting_date: str, start_time: str | None, advisor_name: str,
                                        opt_out_link: str | None = None) -> str:
    from datetime import date
    first_name = (recipient_name or "").strip().split(" ")[0]
    greeting = f"היי {first_name}," if first_name else "היי,"
    date_fmt = date.fromisoformat(meeting_date).strftime("%d/%m/%y")
    time_clause = f", בשעה {start_time}" if start_time else ""
    advisor_clause = f" עם {advisor_name}" if advisor_name else ""
    items_html = "".join(f"<li style='margin-bottom:4px'>{item}</li>" for item in checklist_items)
    baseline_note = (
        "<p style='margin:0 0 16px 0; color:#b45309; background:#fffbeb; border:1px solid #fde68a; "
        "border-radius:8px; padding:10px 14px; font-size:13px;'>"
        "טרם בוצעה בדיקה עבור בית הספר בשנת הלימודים הנוכחית — הרשימה למטה כללית.</p>"
    ) if no_baseline else ""
    opt_out_html = _opt_out_footer_html(opt_out_link)
    return f"""
<html>
<body dir="rtl" style="font-family: Arial, sans-serif; font-size: 14px; color: #1e293b;
                       background: #f8fafc; margin: 0; padding: 24px;">
  <div style="max-width: 560px; margin: 0 auto; background: white;
              border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden;">
    <div style="background: #0070F3; padding: 20px 24px;">
      <p style="margin: 0; color: white; font-size: 14px; font-weight: 700;">גפן AI</p>
      <p style="margin: 4px 0 0 0; color: rgba(255,255,255,0.8); font-size: 12px;">בקשת קבצים לפגישה</p>
    </div>
    <div style="padding: 28px 24px;">
      <p style="margin: 0 0 16px 0; font-size: 15px;">{greeting}</p>
      <p style="margin: 0 0 16px 0; color: #334155; line-height: 1.8;">
        לקראת הפגישה שתתקיים {when_bet}, <b>{date_fmt}</b>{time_clause}{advisor_clause} על תקציב הגפ"ן,
        נשמח <b>שתעלי עוד היום</b> את הקבצים הבאים בקישור המצורף כדי שנוכל להיערך בהתאם:
      </p>
      {baseline_note}
      <ul style="margin: 0 0 20px 0; padding-inline-start: 20px; color: #334155;">{items_html}</ul>
      <div style="text-align: center; margin-bottom: 8px;">
        <a href="{upload_url}"
           style="display: inline-block; background: #0070F3; color: white;
                  font-size: 14px; font-weight: 700; padding: 12px 28px;
                  border-radius: 8px; text-decoration: none;">
          העלאת קבצים
        </a>
      </div>
    </div>
    <div style="background: #f1f5f9; padding: 12px 24px; text-align: center;">
      <p style="margin: 0; font-size: 11px; color: #94a3b8;">נשלח אוטומטית מגפן AI</p>
      {opt_out_html}
    </div>
  </div>
</body>
</html>"""


def _send_reminder_email(to_email: str, subject: str, html: str):
    gmail_user = os.getenv("GMAIL_USER", "")
    gmail_password = os.getenv("GMAIL_APP_PASSWORD", "")
    if not gmail_user or not gmail_password:
        raise RuntimeError("Gmail not configured")
    msg = MIMEMultipart()
    msg["From"] = f"גפן AI <{gmail_user}>"
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.attach(MIMEText(html, "html", "utf-8"))
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as server:
        server.ehlo()
        server.starttls()
        server.login(gmail_user, gmail_password)
        server.send_message(msg)


@router.post("/meetings/send-due-reminders")
def send_due_reminders(request: Request):
    """Cron-triggered (GitHub Actions, daily Sun-Thu). Sends a reminder email
    to each participant of meetings scheduled for the relevant 'due' date(s),
    per _due_meeting_dates(). Tracked per-recipient in meeting_reminders so a
    meeting is only ever attempted once and per-recipient failures are visible."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=401, detail="unauthorized")

    from datetime import datetime
    from zoneinfo import ZoneInfo

    today_il = datetime.now(ZoneInfo("Asia/Jerusalem")).date()
    due_dates = [d.isoformat() for d in _due_meeting_dates(today_il)]

    meetings = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            res = (
                db.table("meetings")
                .select("id, school_id, meeting_date, start_time, end_time, status, participants, advisor_ids, meeting_service_type")
                .eq("reminder_enabled", True)
                .eq("status", "scheduled")
                .in_("meeting_date", due_dates)
                .execute()
            )
            meetings = res.data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("send_due_reminders attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("send_due_reminders failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת")

    if not meetings:
        return {"ok": True, "sent": 0, "failed": 0, "skipped_meetings": 0}

    school_ids = list({m["school_id"] for m in meetings})
    schools_map = {}
    try:
        db = get_admin_client()
        s_res = db.table("schools").select("id, name, stage, finance_software, org_id").in_("id", school_ids).execute()
        schools_map = {s["id"]: s for s in (s_res.data or [])}
    except Exception as exc:
        logger.warning("send_due_reminders: school name lookup failed (non-fatal): %s", exc)

    # Round: per-org automation toggles (organizations.meeting_reminders_enabled /
    # secretary_upload_request_enabled). Missing org_id or a failed lookup fails OPEN (both
    # True) — a broken toggle lookup must never silently stop reminders that were always sent.
    org_ids = list({s.get("org_id") for s in schools_map.values() if s.get("org_id")})
    org_settings_map = {}
    if org_ids:
        try:
            db = get_admin_client()
            o_res = db.table("organizations").select(
                "id, meeting_reminders_enabled, secretary_upload_request_enabled"
            ).in_("id", org_ids).execute()
            org_settings_map = {o["id"]: o for o in (o_res.data or [])}
        except Exception as exc:
            logger.warning("send_due_reminders: org automation settings lookup failed (non-fatal, fail-open): %s", exc)

    advisor_ids = list({aid for m in meetings for aid in (m.get("advisor_ids") or [])})
    advisor_names_map = {}
    if advisor_ids:
        try:
            db = get_admin_client()
            p_res = db.table("profiles").select("id, full_name").in_("id", advisor_ids).execute()
            advisor_names_map = {p["id"]: p.get("full_name") for p in (p_res.data or [])}
        except Exception as exc:
            logger.warning("send_due_reminders: advisor name lookup failed (non-fatal): %s", exc)

    # Opt-out suppression — batched once for every participant email / school across every due
    # meeting (Architecture Invariant #7: no per-row query loop). client_status_map is reused
    # below both to decide suppression and whether to attach an opt-out link to sendable
    # reminders — fetched unconditionally since either use needs it regardless of whether
    # anyone has actually opted out yet.
    import task_logic  # local import — task_logic imports from this module at module level
    all_participant_emails = list({
        (p.get("email") or "").strip().lower()
        for m in meetings for p in (m.get("participants") or []) if (p.get("email") or "").strip()
    })
    client_status_map = {}
    opted_out_emails = set()
    try:
        db = get_admin_client()
        opted_out_emails = task_logic.fetch_opted_out_emails(db, all_participant_emails)
        year_rows = (
            db.table("school_year_admin_data").select("school_id, client_status")
            .eq("academic_year", DEFAULT_ACADEMIC_YEAR).in_("school_id", school_ids).execute().data or []
        )
        client_status_map = {r["school_id"]: r.get("client_status") for r in year_rows}
    except Exception as exc:
        logger.warning("send_due_reminders: opt-out lookup failed (non-fatal, treated as none opted out): %s", exc)

    sent, failed, skipped_meetings = 0, 0, 0
    for m in meetings:
        try:
            participants = [p for p in (m.get("participants") or []) if (p.get("email") or "").strip()]
            if not participants:
                continue

            db = get_admin_client()
            already = db.table("meeting_reminders").select("id").eq("meeting_id", m["id"]).execute()
            if already.data:
                skipped_meetings += 1
                continue

            when_lamed, when_bet = _day_phrases(m["meeting_date"], today_il)
            advisor_name = _hebrew_join([advisor_names_map.get(aid) for aid in (m.get("advisor_ids") or [])])
            subject = f"תזכורת: פגישה {when_bet}"
            school = schools_map.get(m["school_id"], {})

            # Round 16 — a שוטף meeting needs no files, so its secretary/finance reminder uses
            # the plain template (same as the principal's) instead of the file-upload-request one.
            meeting_service_type = m.get("meeting_service_type") or "gefen"
            org_settings = org_settings_map.get(school.get("org_id"), {})
            reminders_on = org_settings.get("meeting_reminders_enabled", True)
            upload_request_on = org_settings.get("secretary_upload_request_enabled", True)
            for p in participants:
                email_addr = p["email"].strip()
                is_upload_contact = p.get("key") in ("secretary", "finance") and meeting_service_type != "current"

                # Org-level automation toggle (ניהול → פגישות → אוטומציות). Skipped recipients
                # are still logged as "skipped" in meeting_reminders — same table the "already
                # attempted today" guard above reads from — so re-running the same day doesn't
                # re-evaluate them, and reporting stays consistent with actually-sent recipients.
                is_opted_out = (
                    email_addr.lower() in opted_out_emails
                    and client_status_map.get(m["school_id"]) != "active"
                )
                if (is_upload_contact and not upload_request_on) or (not is_upload_contact and not reminders_on) or is_opted_out:
                    try:
                        db.table("meeting_reminders").insert({
                            "meeting_id": m["id"], "school_id": m["school_id"],
                            "recipient_email": email_addr, "recipient_name": p.get("name"),
                            "status": "skipped",
                            "error_message": "recipient opted out (client_status not active)" if is_opted_out else None,
                        }).execute()
                    except Exception as log_exc:
                        logger.error("send_due_reminders: failed to log skipped reminder for meeting %s / %s: %s",
                                     m["id"], email_addr, log_exc)
                    continue

                opt_out_link = None
                if client_status_map.get(m["school_id"]) != "active":
                    email_lower = email_addr.lower()
                    opt_out_link = f"{os.getenv('APP_URL', '')}/tasks/opt-out?email={email_lower}&token={task_logic.make_optout_token(email_lower)}"

                if is_upload_contact:
                    try:
                        token = get_or_create_upload_token(db, m["id"], m["meeting_date"])
                        checklist = build_upload_checklist(db, school, m.get("academic_year"))
                        upload_url = f"{os.getenv('APP_URL', '')}/upload/{token}"
                        html = _build_secretary_upload_email_html(
                            recipient_name=p.get("name") or "",
                            when_bet=when_bet,
                            school_name=school.get("name", ""),
                            checklist_items=[i["label"] for i in checklist["items"]],
                            upload_url=upload_url,
                            no_baseline=checklist["no_baseline_this_year"],
                            meeting_date=m["meeting_date"],
                            start_time=m.get("start_time"),
                            advisor_name=advisor_name,
                            opt_out_link=opt_out_link,
                        )
                        status, error = "sent", None
                        try:
                            send_resend_email(email_addr, f"בקשת קבצים לפגישה {when_bet}", html)
                            sent += 1
                        except Exception as email_exc:
                            status, error = "failed", str(email_exc)
                            failed += 1
                            logger.warning("send_due_reminders: failed to send upload-request email to %s for meeting %s: %s",
                                           email_addr, m["id"], email_exc)
                    except Exception as exc:
                        status, error = "failed", str(exc)
                        failed += 1
                        logger.warning("send_due_reminders: upload-request setup failed for %s / meeting %s: %s",
                                       email_addr, m["id"], exc)
                else:
                    html = _build_reminder_email_html(
                        recipient_name=p.get("name") or "",
                        when_lamed=when_lamed,
                        when_bet=when_bet,
                        meeting_date=m["meeting_date"],
                        start_time=m.get("start_time"),
                        advisor_name=advisor_name,
                        meeting_service_type=meeting_service_type,
                        opt_out_link=opt_out_link,
                    )
                    status, error = "sent", None
                    try:
                        _send_reminder_email(email_addr, subject, html)
                        sent += 1
                    except Exception as email_exc:
                        status, error = "failed", str(email_exc)
                        failed += 1
                        logger.warning("send_due_reminders: failed to email %s for meeting %s: %s",
                                       email_addr, m["id"], email_exc)
                try:
                    db.table("meeting_reminders").insert({
                        "meeting_id": m["id"],
                        "school_id": m["school_id"],
                        "recipient_email": email_addr,
                        "recipient_name": p.get("name"),
                        "status": status,
                        "error_message": error,
                    }).execute()
                except Exception as log_exc:
                    logger.error("send_due_reminders: failed to log reminder for meeting %s / %s: %s",
                                 m["id"], email_addr, log_exc)
        except Exception as exc:
            logger.error("send_due_reminders: meeting %s failed: %s", m["id"], exc, exc_info=True)

    return {"ok": True, "sent": sent, "failed": failed, "skipped_meetings": skipped_meetings, "total_due": len(meetings)}


@router.post("/meetings/process-booking-email-queue")
def process_booking_email_queue(request: Request):
    """Cron-triggered (same GitHub Actions convention as send-due-reminders). Drains a small
    batch of pending 'סוכן ניהול' booking-request emails, rate-limiting bulk sends against
    Microsoft Graph the same way the existing reminder cron already does for its own queue."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=401, detail="unauthorized")

    import booking_logic
    import task_logic  # local import — task_logic imports from this module at module level
    from booking_draft_state import _update_draft

    db = get_admin_client()
    BATCH_SIZE = 20
    rows = (
        db.table("meeting_booking_email_queue")
        .select("*")
        .eq("status", "pending")
        .order("created_at")
        .limit(BATCH_SIZE)
        .execute()
        .data or []
    )

    sent, failed = 0, 0
    for row in rows:
        try:
            school = db.table("schools").select(
                "id, name, secretary_name, secretary_email, finance_contact_name, finance_contact_email"
            ).eq("id", row["school_id"]).execute().data
            school = school[0] if school else None
            advisor = db.table("profiles").select("id, full_name").eq("id", row["advisor_id"]).execute().data
            advisor = advisor[0] if advisor else None
            token_row = db.table("meeting_booking_tokens").select("token, months").eq("id", row["token_id"]).execute().data
            token_row = token_row[0] if token_row else None

            if not school or not advisor or not token_row:
                raise ValueError("missing school/advisor/token for queued row")

            to_email = school.get("secretary_email") or school.get("finance_contact_email")
            recipient_name = school.get("secretary_name") or school.get("finance_contact_name") or ""
            if not to_email:
                raise ValueError(f"school {row['school_id']} has no secretary/finance email")

            try:
                opted_out_map = task_logic.opted_out_recipients(
                    db, DEFAULT_ACADEMIC_YEAR, {row["school_id"]: to_email},
                )
            except Exception as exc:
                logger.warning("process_booking_email_queue: opt-out lookup failed (non-fatal): %s", exc)
                opted_out_map = {}
            if row["school_id"] in opted_out_map:
                db.table("meeting_booking_email_queue").update({
                    "status": "skipped", "attempted_at": "now()",
                    "error_message": "recipient opted out (client_status not active)",
                }).eq("id", row["id"]).execute()
                continue

            year_rows = (
                db.table("school_year_admin_data").select("client_status")
                .eq("academic_year", DEFAULT_ACADEMIC_YEAR).eq("school_id", row["school_id"]).execute().data or []
            )
            client_status = year_rows[0].get("client_status") if year_rows else None
            opt_out_link = None
            if client_status != "active":
                email_lower = to_email.strip().lower()
                opt_out_link = f"{os.getenv('APP_URL', '')}/tasks/opt-out?email={email_lower}&token={task_logic.make_optout_token(email_lower)}"

            booking_url = f"{os.getenv('APP_URL', '')}/book/{token_row['token']}"
            html = booking_logic.build_booking_request_email_html(
                recipient_name=recipient_name,
                school_name=school["name"],
                advisor_name=advisor.get("full_name") or "",
                months=token_row["months"],
                booking_url=booking_url,
                opt_out_link=opt_out_link,
            )
            subject = f"קביעת פגישה - {school['name']}"
            booking_logic.send_booking_request_email(row["org_id"], row["advisor_id"], to_email, subject, html)

            db.table("meeting_booking_email_queue").update({
                "status": "sent", "attempted_at": "now()",
            }).eq("id", row["id"]).execute()
            sent += 1
        except Exception as exc:
            logger.warning("process_booking_email_queue: failed to send row %s: %s", row["id"], exc)
            try:
                db.table("meeting_booking_email_queue").update({
                    "status": "failed", "error_message": str(exc), "attempted_at": "now()",
                }).eq("id", row["id"]).execute()
            except Exception as log_exc:
                logger.error("process_booking_email_queue: failed to mark row %s failed: %s", row["id"], log_exc)
            failed += 1

    # Once a draft has zero remaining pending rows, flip it to 'sent'.
    try:
        draft_org_map = {r["draft_id"]: r["org_id"] for r in rows}
        for draft_id, org_id in draft_org_map.items():
            remaining = (
                db.table("meeting_booking_email_queue")
                .select("id")
                .eq("draft_id", draft_id)
                .eq("status", "pending")
                .execute()
                .data or []
            )
            if not remaining:
                _update_draft(draft_id, org_id, {"status": "sent"})
    except Exception as exc:
        logger.warning("process_booking_email_queue: failed to finalize draft statuses (non-fatal): %s", exc)

    return {"ok": True, "sent": sent, "failed": failed, "batch_size": len(rows)}


def _reconcile_meeting_from_outlook(db, org_id: str, meeting: dict) -> dict:
    """Shared by the periodic poll (backup safety net) and the real-time webhook
    handler: checks every synced advisor's event for this meeting and applies any
    confirmed change (reschedule or cancellation) back onto the meeting row.

    Deliberately narrow in scope: we never pull the subject/title back from Outlook —
    our own naming logic (school/city/contact) stays authoritative for that. A meeting
    is only cancelled if its event is confirmed gone for *every* synced advisor (not
    just one, for multi-advisor meetings) — and any check that fails/errors is skipped
    rather than ever being treated as "deleted".

    Returns {"action": "updated"|"cancelled"|"skipped"|"none", "checked": N} where N
    is the number of advisor events actually queried against Graph.
    """
    sync_map = meeting.get("calendar_sync") or {}
    advisor_results = {}
    checked = 0
    for advisor_id, entry in sync_map.items():
        event_id = entry.get("external_event_id")
        if not event_id:
            continue
        checked += 1
        advisor_results[advisor_id] = graph_client.get_event(db, org_id, advisor_id, event_id)

    if not advisor_results:
        return {"action": "none", "checked": checked}
    if any(v == graph_client.EVENT_CHECK_SKIPPED for v in advisor_results.values()):
        return {"action": "skipped", "checked": checked}  # inconclusive — never guess, retry next round

    if all(v is None for v in advisor_results.values()):
        try:
            db.table("meetings").update({"status": "cancelled"}).eq("id", meeting["id"]).execute()
            graph_client.persist_calendar_sync(db, meeting["id"], {})
            return {"action": "cancelled", "checked": checked}
        except Exception as exc:
            logger.warning("_reconcile_meeting_from_outlook: failed to cancel meeting %s: %s", meeting["id"], exc)
            return {"action": "skipped", "checked": checked}

    for advisor_id, result in advisor_results.items():
        if result is None:
            continue
        new_date, new_start, new_end = result["start"][:10], result["start"][11:16], result["end"][11:16]
        if new_date != meeting.get("meeting_date") or new_start != meeting.get("start_time") or new_end != meeting.get("end_time"):
            try:
                db.table("meetings").update({
                    "meeting_date": new_date, "start_time": new_start, "end_time": new_end,
                }).eq("id", meeting["id"]).execute()
                return {"action": "updated", "checked": checked}
            except Exception as exc:
                logger.warning("_reconcile_meeting_from_outlook: failed to update meeting %s: %s", meeting["id"], exc)
                return {"action": "skipped", "checked": checked}

    return {"action": "none", "checked": checked}


@router.post("/meetings/poll-outlook-changes")
def poll_outlook_changes(request: Request):
    """Cron-triggered backup safety net (best-effort schedule, currently firing
    roughly every 1-3h in practice — see the real-time webhook handler in
    calendar_router.py for the primary, near-instant reconciliation path).
    Reconciles all synced future meetings via _reconcile_meeting_from_outlook,
    catching anything a missed/lost webhook notification would otherwise leave stale."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=401, detail="unauthorized")

    from datetime import datetime, timezone

    db = get_admin_client()
    today = datetime.now(timezone.utc).date().isoformat()
    try:
        meetings = (
            db.table("meetings")
            .select("id, school_id, meeting_date, start_time, end_time, status, calendar_sync")
            .eq("status", "scheduled")
            .gte("meeting_date", today)
            .execute()
            .data or []
        )
    except Exception as exc:
        logger.error("poll_outlook_changes failed to fetch meetings: %s", exc, exc_info=True)
        raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת")

    meetings = [m for m in meetings if m.get("calendar_sync")]
    if not meetings:
        return {"checked": 0, "updated": 0, "cancelled": 0}

    school_ids = list({m["school_id"] for m in meetings if m.get("school_id")})
    schools_map = {}
    if school_ids:
        try:
            srows = db.table("schools").select("id, org_id").in_("id", school_ids).execute().data or []
            schools_map = {s["id"]: s.get("org_id") for s in srows}
        except Exception as exc:
            logger.warning("poll_outlook_changes school lookup failed (non-fatal): %s", exc)

    checked = updated = cancelled = 0

    for m in meetings:
        org_id = schools_map.get(m.get("school_id"))
        if not org_id:
            continue
        result = _reconcile_meeting_from_outlook(db, org_id, m)
        checked += result["checked"]
        if result["action"] == "updated":
            updated += 1
        elif result["action"] == "cancelled":
            cancelled += 1

    return {"checked": checked, "updated": updated, "cancelled": cancelled}


@router.get("/meetings/{meeting_id}/reminder-status")
def get_meeting_reminder_status(
    meeting_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    """Per-recipient reminder send status for a single meeting (for the UI)."""
    for attempt in range(2):
        try:
            db = get_admin_client()
            res = (
                db.table("meeting_reminders")
                .select("recipient_email, recipient_name, status, error_message, sent_at")
                .eq("meeting_id", meeting_id)
                .order("sent_at")
                .execute()
            )
            return {"reminders": res.data or []}
        except Exception as exc:
            if attempt == 0:
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.warning("get_meeting_reminder_status failed (non-fatal): %s", exc)
                return {"reminders": []}


# ---------------------------------------------------------------------------
# Single school fetch (used when navigating via deeplink / notification)
# ---------------------------------------------------------------------------

@router.get("/{school_id}")
def get_school(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("schools")
                .select("*, gefen_accounts(*), advisor_schools(advisor_id)")
                .eq("id", school_id)
                .eq("org_id", user["org_id"])
                .execute()
            )
            if not rows.data:
                raise HTTPException(status_code=404, detail="בית ספר לא נמצא")
            school_data = rows.data[0]
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_school attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_school failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    # Access check for advisors
    is_advisor = user["role"] not in ("owner", "manager")
    if is_advisor:
        restrict = school_data.get("restrict_access_to")
        if restrict is not None and user["id"] not in (restrict or []):
            assigned = (
                db.table("advisor_schools")
                .select("school_id")
                .eq("advisor_id", user["id"])
                .eq("school_id", school_id)
                .execute()
            )
            if not assigned.data:
                raise HTTPException(status_code=403, detail="אין הרשאה לצפות בבית ספר זה")

    # Enrich advisor_schools + restrict_access_to with profile data in a single query (non-fatal)
    advisor_ids = [r["advisor_id"] for r in (school_data.get("advisor_schools") or []) if r.get("advisor_id")]
    restrict_ids = school_data.get("restrict_access_to") or []
    all_profile_ids = list(set(advisor_ids + restrict_ids))
    school_data["restrict_access_profiles"] = []
    if all_profile_ids:
        try:
            p_rows = db.table("profiles").select("id, full_name, email, role").in_("id", all_profile_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
            for row in school_data.get("advisor_schools") or []:
                row["profiles"] = profiles_map.get(row["advisor_id"])
            school_data["restrict_access_profiles"] = [profiles_map[i] for i in restrict_ids if i in profiles_map]
        except Exception as exc:
            logger.warning("get_school profiles enrichment failed (non-fatal): %s", exc)

    school_data["meeting_coordinator_contact"] = _resolve_meeting_coordinator(school_data)

    # Per-service-type advisor lists ("יועץ מלווה [גפן/שוטף/מחוז]") — needed by every role (not
    # just manager+) since meetings-tab default-advisor logic reads these regardless of who's
    # viewing the school. Non-fatal: on failure the lists are simply empty.
    for service_type, table_name in _TYPED_ADVISOR_TABLES.items():
        try:
            rows = (
                db.table(table_name)
                .select("advisor_id, profiles(id, email, full_name, role)")
                .eq("school_id", school_id)
                .execute()
            )
            school_data[f"advisors_{service_type}"] = [r["profiles"] for r in rows.data if r.get("profiles")]
        except Exception as exc:
            logger.warning("%s enrichment failed on get_school (non-fatal): %s", table_name, exc)
            school_data[f"advisors_{service_type}"] = []

    return school_data


# ---------------------------------------------------------------------------
# Users (owner/manager only)
# ---------------------------------------------------------------------------

@router.get("/users/me")
def get_me(user: Annotated[dict, Depends(get_current_user)]):
    result = dict(user)

    # org subscription info
    if user.get("org_id"):
        for attempt in range(2):
            try:
                db = get_admin_client()
                org_res = (
                    db.table("organizations")
                    .select("subscription_status, trial_started_at, trial_ends_at, name")
                    .eq("id", user["org_id"])
                    .single()
                    .execute()
                )
                result["org"] = org_res.data or {}
                break
            except Exception as exc:
                if attempt == 0:
                    logger.warning("get_me org fetch attempt 1 failed: %s — resetting", exc)
                    reset_admin_client()
                    time.sleep(0.3)
                else:
                    logger.warning("get_me org fetch failed after 2 attempts: %s", exc)
                    result["org"] = {}

    try:
        db = get_admin_client()
        result["can_delete_schools"] = _check_permission(db, user, "can_delete_schools")
        result["can_edit_school_directly"] = _check_permission(db, user, "can_edit_school_directly")
        result["can_request_school_update"] = _check_permission(db, user, "can_request_school_update")
        result["can_delete_own_meetings"] = _check_permission(db, user, "can_delete_own_meetings")
        result["can_invite_users"] = _check_permission(db, user, "can_invite_users")
        result["can_delete_users"] = _check_permission(db, user, "can_delete_users")
        result["can_manage_user_permissions"] = _check_permission(db, user, "can_manage_user_permissions")
        result["can_remove_call_from_school"] = _check_permission(db, user, "can_remove_call_from_school")
    except Exception as exc:
        logger.warning("get_me permission check failed (non-fatal): %s", exc)
        result["can_delete_schools"] = user.get("role") == "owner"
        result["can_edit_school_directly"] = user.get("role") in ("owner", "manager")
        result["can_request_school_update"] = True
        result["can_delete_own_meetings"] = user.get("role") in ("owner", "manager")
        result["can_invite_users"] = user.get("role") in ("owner", "manager")
        result["can_delete_users"] = user.get("role") == "owner"
        result["can_manage_user_permissions"] = user.get("role") == "owner"
        result["can_remove_call_from_school"] = user.get("role") in ("owner", "manager")

    return result



class OnboardingDismissIn(BaseModel):
    key: str  # "add_school" | "add_user"


@router.patch("/users/me/onboarding")
def dismiss_onboarding(
    body: OnboardingDismissIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if body.key not in {"add_school", "add_user"}:
        raise HTTPException(status_code=400, detail="מפתח לא חוקי")
    db = get_admin_client()
    merged = {**(user.get("onboarding_dismissed") or {}), body.key: True}
    db.table("profiles").update({"onboarding_dismissed": merged}).eq("id", user["id"]).execute()
    invalidate_profile_cache(user["id"])
    return {"ok": True}


class MyProfileIn(BaseModel):
    full_name: str


@router.patch("/users/me/profile")
def update_my_profile(
    body: MyProfileIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    name = body.full_name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="שם לא יכול להיות ריק")
    db = get_admin_client()
    db.table("profiles").update({"full_name": name}).eq("id", user["id"]).execute()
    invalidate_profile_cache(user["id"])
    return {"ok": True}


@router.get("/users/all")
def list_users(user: Annotated[dict, Depends(get_current_user)]):
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = db.table("profiles").select("*").eq("org_id", user["org_id"]).order("full_name").execute()
            return rows.data
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_users attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_users failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.post("/users/invite")
def invite_user(
    body: UserInviteIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    db_pre = get_admin_client()
    if not _check_permission(db_pre, user, "can_invite_users"):
        raise HTTPException(status_code=403, detail="אין הרשאה להזמין משתמשים חדשים")
    work_phone = _validate_work_phone(body.work_phone)
    app_url = os.getenv("APP_URL", "https://gefenai.co.il")
    for attempt in range(2):
        try:
            db = get_admin_client()
            result = db.auth.admin.invite_user_by_email(
                body.email,
                {
                    "data": {"full_name": body.full_name or "", "role": body.role},
                    "redirect_to": f"{app_url}/set-password",
                },
            )
            user_id = str(result.user.id)
            db.table("profiles").upsert({
                "id": user_id,
                "email": body.email,
                "full_name": body.full_name or "",
                "role": body.role,
                "org_id": user["org_id"],
                "status": "pending",
                "control_domains": body.control_domains,
                "work_phone": work_phone,
            }).execute()
            return {"ok": True, "user_id": user_id}
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("invite_user attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("invite_user failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


def _send_reinvite_email(to_email: str, full_name: str, action_link: str):
    gmail_user = os.getenv("GMAIL_USER", "")
    gmail_password = os.getenv("GMAIL_APP_PASSWORD", "")
    if not gmail_user or not gmail_password:
        logger.warning("Gmail not configured — skipping reinvite email to %s", to_email)
        return
    greeting = f"שלום {full_name}," if full_name else "שלום,"
    html = f"""
<html>
<body dir="rtl" style="font-family: Arial, sans-serif; font-size: 14px; color: #1e293b;
                       background: #f8fafc; margin: 0; padding: 24px;">
  <div style="max-width: 520px; margin: 0 auto; background: white;
              border-radius: 12px; border: 1px solid #e2e8f0; overflow: hidden;">
    <div style="background: #0070F3; padding: 20px 24px;">
      <p style="margin: 0; color: white; font-size: 14px; font-weight: 700;">גפן AI</p>
      <p style="margin: 4px 0 0 0; color: rgba(255,255,255,0.8); font-size: 12px;">הזמנה למערכת</p>
    </div>
    <div style="padding: 28px 24px;">
      <p style="margin: 0 0 16px 0; font-size: 15px;">{greeting}</p>
      <p style="margin: 0 0 24px 0; color: #334155; line-height: 1.6;">
        קיבלת הזמנה להצטרף למערכת גפן AI.<br>
        לחץ על הכפתור למטה כדי להגדיר סיסמה ולהשלים את הרישום.
      </p>
      <div style="text-align: center; margin-bottom: 24px;">
        <a href="{action_link}"
           style="display: inline-block; background: #0070F3; color: white;
                  font-size: 14px; font-weight: 700; padding: 12px 28px;
                  border-radius: 8px; text-decoration: none;">
          הגדרת סיסמה וכניסה למערכת
        </a>
      </div>
      <p style="margin: 0; font-size: 12px; color: #94a3b8; text-align: center;">
        הקישור תקף ל-24 שעות. אם לא ביקשת הזמנה זו, ניתן להתעלם ממייל זה.
      </p>
    </div>
    <div style="background: #f1f5f9; padding: 12px 24px; text-align: center;">
      <p style="margin: 0; font-size: 11px; color: #94a3b8;">נשלח מגפן AI</p>
    </div>
  </div>
</body>
</html>"""
    msg = MIMEMultipart()
    msg["From"] = f"גפן AI <{gmail_user}>"
    msg["To"] = to_email
    msg["Subject"] = "הזמנה למערכת גפן AI"
    msg.attach(MIMEText(html, "html", "utf-8"))
    with smtplib.SMTP("smtp.gmail.com", 587, timeout=15) as server:
        server.ehlo()
        server.starttls()
        server.login(gmail_user, gmail_password)
        server.send_message(msg)


@router.post("/users/{user_id}/resend-invite")
def resend_invite(
    user_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    app_url = os.getenv("APP_URL", "http://localhost:5173")
    for attempt in range(2):
        try:
            db = get_admin_client()
            target = db.table("profiles").select("email, org_id, full_name, role, status").eq("id", user_id).execute()
            if not target.data or target.data[0].get("org_id") != user["org_id"]:
                raise HTTPException(status_code=404, detail="המשתמש לא נמצא")
            profile = target.data[0]
            if profile.get("status") != "pending":
                raise HTTPException(status_code=400, detail="המשתמש כבר פעיל — לא ניתן לשלוח הזמנה מחדש")
            # Use REST API directly — works even when the user already confirmed
            # their email (invite_user_by_email fails with "User already registered")
            supabase_url = os.getenv("SUPABASE_URL", "")
            service_key = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")
            resp = httpx.post(
                f"{supabase_url}/auth/v1/admin/generate_link",
                headers={
                    "apikey": service_key,
                    "Authorization": f"Bearer {service_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "type": "recovery",
                    "email": profile["email"],
                    "redirect_to": f"{app_url}/set-password",
                },
                timeout=10,
            )
            resp.raise_for_status()
            action_link = resp.json()["action_link"]
            _send_reinvite_email(profile["email"], profile.get("full_name") or "", action_link)
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("resend_invite attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("resend_invite failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")
    return {"ok": True}


@router.post("/users/me/setup-complete")
def setup_complete(user: Annotated[dict, Depends(get_current_user)]):
    for attempt in range(2):
        try:
            db = get_admin_client()
            db.table("profiles").update({"status": "active"}).eq("id", user["id"]).execute()
            invalidate_profile_cache(user["id"])
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("setup_complete attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("setup_complete failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


@router.patch("/users/{user_id}/role")
def update_role(
    user_id: str,
    body: UserRoleIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    if body.role not in ("owner", "manager", "advisor"):
        raise HTTPException(status_code=400, detail="תפקיד לא חוקי")
    for attempt in range(2):
        try:
            db = get_admin_client()
            if user["role"] == "manager":
                if user_id == user["id"]:
                    raise HTTPException(status_code=403, detail="מנהל לא יכול לשנות את תפקיד עצמו")
                if not _check_permission(db, user, "can_change_user_role"):
                    raise HTTPException(status_code=403, detail="אין הרשאה לשנות תפקידים")
                if body.role == "owner":
                    raise HTTPException(status_code=403, detail="מנהל לא יכול להעניק תפקיד בעלים")
            target = db.table("profiles").select("id, org_id, role, full_name, email").eq("id", user_id).execute()
            if not target.data or target.data[0].get("org_id") != user["org_id"]:
                raise HTTPException(status_code=404, detail="המשתמש לא נמצא")
            current_role = target.data[0].get("role")
            if current_role == "owner" and user["role"] == "manager":
                raise HTTPException(status_code=403, detail="מנהל לא יכול לשנות תפקיד של בעלים")
            if current_role == "owner" and body.role != "owner":
                owners = db.table("profiles").select("id").eq("org_id", user["org_id"]).eq("role", "owner").execute()
                if len(owners.data or []) <= 1:
                    raise HTTPException(
                        status_code=400,
                        detail="לא ניתן לשנות את תפקיד הבעלים היחיד. יש להגדיר בעלים אחר לפני שינוי תפקיד זה."
                    )
            db.table("profiles").update({"role": body.role}).eq("id", user_id).execute()
            # Update user_metadata in Auth so the next JWT refresh reflects the new role
            try:
                db.auth.admin.update_user_by_id(user_id, {"user_metadata": {"role": body.role}})
            except Exception as meta_exc:
                logger.warning("update_role user_metadata update failed (non-fatal): %s", meta_exc)
            # Notifications (non-fatal)
            try:
                role_labels = {"owner": "בעלים", "manager": "מנהל", "advisor": "יועץ"}
                target_name = target.data[0].get("full_name") or target.data[0].get("email", "משתמש")
                changer_name = user.get("full_name") or user.get("email", "")
                old_label = role_labels.get(current_role, current_role)
                new_label = role_labels.get(body.role, body.role)
                # Notify the user whose role changed
                _create_notifications(db, [{
                    "recipient_id": user_id,
                    "type": "role_changed",
                    "data": {
                        "title": f"התפקיד שלך שונה מ{old_label} ל{new_label}",
                        "changer_name": changer_name,
                        "old_role": current_role,
                        "new_role": body.role,
                    },
                }], pref_key="notify_role_changed")
                # Notify all owners + managers (except the changer)
                mgr_rows = db.table("profiles").select("id").eq("org_id", user["org_id"]).in_("role", ["owner", "manager"]).execute()
                recipient_ids = [r["id"] for r in (mgr_rows.data or []) if r["id"] != user["id"] and r["id"] != user_id]
                if recipient_ids:
                    _create_notifications(db, [{
                        "recipient_id": rid,
                        "type": "role_changed",
                        "data": {
                            "title": f"התפקיד של {target_name} שונה מ{old_label} ל{new_label}",
                            "changer_name": changer_name,
                            "target_name": target_name,
                            "old_role": current_role,
                            "new_role": body.role,
                        },
                    } for rid in recipient_ids], pref_key="notify_role_changed")
            except Exception as notif_exc:
                logger.warning("update_role notification failed (non-fatal): %s", notif_exc)
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("update_role attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("update_role failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")
    invalidate_profile_cache(user_id)
    invalidate_approver_ids_cache()
    return {"ok": True}


class UserProfileUpdateIn(BaseModel):
    full_name: str | None = None
    control_domains: list[str] | None = None
    work_phone: str | None = None


@router.patch("/users/{user_id}")
def update_user_profile(
    user_id: str,
    body: UserProfileUpdateIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    provided = body.model_dump(exclude_unset=True)
    if "work_phone" in provided:
        provided["work_phone"] = _validate_work_phone(provided["work_phone"])
    data = {k: v for k, v in provided.items() if v is not None}
    if not data:
        return {"ok": True}
    for attempt in range(2):
        try:
            db = get_admin_client()
            target = db.table("profiles").select("org_id").eq("id", user_id).execute()
            if not target.data or target.data[0].get("org_id") != user["org_id"]:
                raise HTTPException(status_code=404, detail="המשתמש לא נמצא")
            db.table("profiles").update(data).eq("id", user_id).execute()
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("update_user_profile attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("update_user_profile failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")
    invalidate_profile_cache(user_id)
    return {"ok": True}


@router.delete("/users/{user_id}")
def delete_user(
    user_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    if user_id == user["id"]:
        raise HTTPException(status_code=400, detail="לא ניתן למחוק את המשתמש הנוכחי")
    for attempt in range(2):
        try:
            db = get_admin_client()
            if user["role"] == "manager" and not _check_permission(db, user, "can_delete_users"):
                raise HTTPException(status_code=403, detail="אין הרשאה למחוק משתמשים")
            target = db.table("profiles").select("org_id").eq("id", user_id).execute()
            if not target.data or target.data[0].get("org_id") != user["org_id"]:
                raise HTTPException(status_code=404, detail="המשתמש לא נמצא")
            db.auth.admin.delete_user(user_id)
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("delete_user attempt 1 failed: %s — resetting", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("delete_user failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")
    return {"ok": True}


@router.get("/users/{user_id}/future-meetings")
def get_user_future_meetings(user_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Future scheduled meetings for this advisor — shown before deleting them, so the
    caller can decide whether to transfer or cancel before the user itself is removed."""
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    from datetime import datetime, timezone
    db = get_admin_client()
    today = datetime.now(timezone.utc).date().isoformat()
    rows = (
        db.table("meetings")
        .select("id, meeting_date, start_time, end_time, school_id")
        .gte("meeting_date", today)
        .eq("status", "scheduled")
        .filter("advisor_ids", "cs", json.dumps([user_id]))
        .order("meeting_date")
        .execute()
    )
    meetings = rows.data or []
    school_ids = list({m["school_id"] for m in meetings if m.get("school_id")})
    schools_map = {}
    if school_ids:
        srows = db.table("schools").select("id, name, city").in_("id", school_ids).execute().data or []
        schools_map = {s["id"]: s for s in srows}
    for m in meetings:
        sch = schools_map.get(m.get("school_id"), {})
        m["school_name"] = sch.get("name")
        m["school_city"] = sch.get("city")
    return meetings


class TransferMeetingsIn(BaseModel):
    new_advisor_id: str


@router.post("/users/{user_id}/meetings/transfer")
def transfer_user_meetings(user_id: str, body: TransferMeetingsIn, user: Annotated[dict, Depends(get_current_user)]):
    """Moves this advisor's future meetings to another advisor, re-syncing Outlook
    (cancels the old advisor's event, creates one on the new advisor's calendar).
    Returns per-meeting results including whether the new advisor already had a
    conflicting commitment, so the caller can surface that in a summary."""
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    from datetime import datetime, timezone
    db = get_admin_client()
    today = datetime.now(timezone.utc).date().isoformat()
    rows = (
        db.table("meetings")
        .select("*")
        .gte("meeting_date", today)
        .eq("status", "scheduled")
        .filter("advisor_ids", "cs", json.dumps([user_id]))
        .execute()
    )
    meetings = rows.data or []
    school_ids = list({m["school_id"] for m in meetings if m.get("school_id")})
    schools_map = {}
    if school_ids:
        srows = db.table("schools").select("id, name").in_("id", school_ids).execute().data or []
        schools_map = {s["id"]: s["name"] for s in srows}

    results = []
    for m in meetings:
        new_ids = [body.new_advisor_id if aid == user_id else aid for aid in (m.get("advisor_ids") or [])]
        conflict = False
        try:
            db.table("meetings").update({"advisor_ids": new_ids}).eq("id", m["id"]).execute()
            with graph_client.calendar_sync_lock(db, m["id"]) as acquired:
                if acquired:
                    fresh = db.table("meetings").select("calendar_sync").eq("id", m["id"]).execute()
                    previous_sync = (fresh.data[0].get("calendar_sync") or {}) if fresh.data else {}
                    subject = _build_meeting_subject(db, m["school_id"], m.get("participants"), m.get("primary_contact_key"))
                    sync_map = graph_client.sync_meeting_update(db, user["org_id"], {**m, "advisor_ids": new_ids}, previous_sync, subject)
                    graph_client.persist_calendar_sync(db, m["id"], sync_map)
                    conflict = bool(sync_map.get(body.new_advisor_id, {}).get("conflict"))
        except Exception as exc:
            logger.warning("transfer meeting %s failed (non-fatal): %s", m["id"], exc)
        results.append({
            "meeting_id": m["id"],
            "meeting_date": m["meeting_date"],
            "start_time": m.get("start_time"),
            "end_time": m.get("end_time"),
            "school_name": schools_map.get(m["school_id"]),
            "conflict": conflict,
        })
    return {"transferred": len(results), "results": results}


@router.post("/users/{user_id}/meetings/cancel-future")
def cancel_user_future_meetings(user_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Deletes this advisor's future meetings entirely — from the system and from Outlook."""
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    from datetime import datetime, timezone
    db = get_admin_client()
    today = datetime.now(timezone.utc).date().isoformat()
    rows = (
        db.table("meetings")
        .select("id, calendar_sync")
        .gte("meeting_date", today)
        .eq("status", "scheduled")
        .filter("advisor_ids", "cs", json.dumps([user_id]))
        .execute()
    )
    meetings = rows.data or []
    for m in meetings:
        try:
            if m.get("calendar_sync"):
                with graph_client.calendar_sync_lock(db, m["id"]) as acquired:
                    if acquired:
                        graph_client.sync_meeting_cancel(db, user["org_id"], m["calendar_sync"])
        except Exception as exc:
            logger.warning("cancel meeting %s failed (non-fatal): %s", m["id"], exc)
        db.table("meetings").delete().eq("id", m["id"]).execute()
    return {"cancelled": len(meetings)}


def _get_user_sole_schools(db, user_id: str) -> list[dict]:
    """Schools where user_id is the only row in advisor_schools — i.e. deleting this
    user would leave the school with no advisor at all."""
    assigned = db.table("advisor_schools").select("school_id").eq("advisor_id", user_id).execute()
    school_ids = [r["school_id"] for r in (assigned.data or [])]
    if not school_ids:
        return []
    all_rows = db.table("advisor_schools").select("school_id, advisor_id").in_("school_id", school_ids).execute().data or []
    counts: dict[str, int] = {}
    for r in all_rows:
        counts[r["school_id"]] = counts.get(r["school_id"], 0) + 1
    return [sid for sid in school_ids if counts.get(sid, 0) <= 1]


@router.get("/users/{user_id}/sole-schools")
def get_user_sole_schools(user_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Schools this advisor is the sole advisor of — shown before deleting them, so the
    caller can transfer these schools to another advisor first. A school can never be
    left without at least one advisor."""
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    db = get_admin_client()
    sole_ids = _get_user_sole_schools(db, user_id)
    if not sole_ids:
        return []
    rows = db.table("schools").select("id, name, city").in_("id", sole_ids).execute().data or []
    return rows


class TransferSchoolsIn(BaseModel):
    new_advisor_id: str


@router.post("/users/{user_id}/schools/transfer")
def transfer_user_sole_schools(user_id: str, body: TransferSchoolsIn, user: Annotated[dict, Depends(get_current_user)]):
    """Reassigns this advisor's sole-advisor schools to another advisor, so the school is
    never left without at least one advisor once the user is deleted."""
    if user["role"] not in ("owner", "manager"):
        raise HTTPException(status_code=403, detail="אין הרשאה")
    db = get_admin_client()
    sole_ids = _get_user_sole_schools(db, user_id)
    if not sole_ids:
        return {"transferred": 0, "schools": []}
    srows = db.table("schools").select("id, name").in_("id", sole_ids).execute().data or []
    names_map = {s["id"]: s["name"] for s in srows}
    for sid in sole_ids:
        db.table("advisor_schools").upsert({"advisor_id": body.new_advisor_id, "school_id": sid}).execute()
    db.table("advisor_schools").delete().eq("advisor_id", user_id).in_("school_id", sole_ids).execute()
    try:
        _create_notifications(db, [{
            "recipient_id": body.new_advisor_id,
            "type": "advisor_assigned",
            "school_id": sid,
            "data": {
                "title": f"שויכת לבית הספר {names_map.get(sid, '')}",
                "school_name": names_map.get(sid, ""),
                "sender_name": user.get("full_name", ""),
                "deeplink": f"/school/{sid}",
            }
        } for sid in sole_ids], pref_key="notify_advisor_assignment")
    except Exception as exc:
        logger.warning("advisor_assigned notification failed (non-fatal): %s", exc)
    return {"transferred": len(sole_ids), "schools": [{"id": sid, "name": names_map.get(sid)} for sid in sole_ids]}


# ---------------------------------------------------------------------------
# Check logs
# ---------------------------------------------------------------------------

@router.get("/{school_id}/logs")
def list_logs(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str | None = None,
):
    logs = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            q = (
                db.table("check_logs")
                .select("*")
                .eq("school_id", school_id)
            )
            if academic_year:
                q = q.eq("academic_year", academic_year)
            rows = q.order("run_at", desc=True).execute()
            logs = rows.data or []
            break  # success
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_logs attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_logs failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    if not logs:
        return []
    # Pinned checks float to the top, most-recently-pinned first; unpinned checks
    # keep the run_at-desc order the query already returned.
    pinned = sorted((r for r in logs if r.get("pinned_at")), key=lambda r: r["pinned_at"], reverse=True)
    unpinned = [r for r in logs if not r.get("pinned_at")]
    logs = pinned + unpinned
    # Enrich with profile data (non-fatal — logs returned even if enrichment fails)
    user_ids = list({r["run_by"] for r in logs if r.get("run_by")})
    if user_ids:
        try:
            db = get_admin_client()
            p_rows = (
                db.table("profiles")
                .select("id, full_name, email")
                .in_("id", user_ids)
                .execute()
            )
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
            for row in logs:
                row["profiles"] = profiles_map.get(row.get("run_by"))
        except Exception as exc:
            logger.warning("list_logs profile enrichment failed (non-fatal): %s", exc)
    return logs


@router.get("/{school_id}/logs/{log_id}")
def get_log(
    school_id: str,
    log_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    row = db.table("check_logs").select("*").eq("id", log_id).eq("school_id", school_id).execute()
    if not row.data:
        raise HTTPException(status_code=404, detail="הבדיקה לא נמצאה")
    return row.data[0]


def _can_delete_check_log(user: dict, db) -> bool:
    return _check_permission(db, user, "can_delete_own_meetings")


@router.delete("/{school_id}/logs/{log_id}")
def delete_log(
    school_id: str,
    log_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    if not _can_delete_check_log(user, db):
        raise HTTPException(status_code=403, detail="אין הרשאה למחיקת בדיקות")
    # Delete stored files from Supabase Storage before removing the DB record
    log_row = db.table("check_logs").select("summary").eq("id", log_id).eq("school_id", school_id).execute()
    if log_row.data:
        stored_paths = (log_row.data[0].get("summary") or {}).get("stored_file_paths") or []
        if stored_paths:
            try:
                keys = [sp["path"] if isinstance(sp, dict) else sp for sp in stored_paths]
                db.storage.from_("check-files").remove(keys)
            except Exception as exc:
                logger.warning("Storage cleanup failed for log %s: %s", log_id, exc)
    db.table("check_logs").delete().eq("id", log_id).eq("school_id", school_id).execute()
    return {"ok": True}


class LogNameIn(BaseModel):
    custom_name: str


@router.patch("/{school_id}/logs/{log_id}/name")
def update_log_name(
    school_id: str,
    log_id: str,
    body: LogNameIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    if not _can_delete_check_log(user, db):
        raise HTTPException(status_code=403, detail="אין הרשאה לערוך שם בדיקה")
    name = body.custom_name.strip()
    db.table("check_logs").update({"custom_name": name or None}).eq("id", log_id).eq("school_id", school_id).execute()
    return {"ok": True}


class LogPinIn(BaseModel):
    pinned: bool


@router.patch("/{school_id}/logs/{log_id}/pin")
def update_log_pin(
    school_id: str,
    log_id: str,
    body: LogPinIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    from datetime import datetime, timezone

    db = get_admin_client()
    if not _can_delete_check_log(user, db):
        raise HTTPException(status_code=403, detail="אין הרשאה לנעוץ בדיקה")
    pinned_at = datetime.now(timezone.utc).isoformat() if body.pinned else None
    db.table("check_logs").update({"pinned_at": pinned_at}).eq("id", log_id).eq("school_id", school_id).execute()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Partial-report row updates ("דיווח חסר" — per-plan comment threads)
#
# Notes follow the plan (row_key) itself, not a specific check run, so they
# stay visible across future checks on the same division/budget/plan.
# ---------------------------------------------------------------------------

class PartialUpdatesBatchIn(BaseModel):
    division: str = "main"
    budget_name: str | None = None
    row_keys: list[str]


@router.post("/{school_id}/partial-updates/batch")
def batch_get_partial_updates(
    school_id: str,
    body: PartialUpdatesBatchIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if not body.row_keys:
        return {}
    rows = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            q = (
                db.table("partial_row_updates")
                .select("*")
                .eq("school_id", school_id)
                .eq("division", body.division)
                .in_("row_key", body.row_keys)
            )
            q = q.eq("budget_name", body.budget_name) if body.budget_name else q.is_("budget_name", "null")
            rows = q.order("created_at").execute().data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("batch_get_partial_updates attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("batch_get_partial_updates failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    # Enrich with author names + roles (non-fatal) — role is needed by the frontend
    # to decide edit/delete permissions per segment (role-hierarchy rules)
    author_ids = list({r["author_id"] for r in rows if r.get("author_id")})
    profiles_map = {}
    if author_ids:
        try:
            db = get_admin_client()
            p_rows = db.table("profiles").select("id, full_name, role").in_("id", author_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
        except Exception as exc:
            logger.warning("batch_get_partial_updates profile enrichment failed (non-fatal): %s", exc)

    # Group segments by row_key, then by group_id (visual "record")
    by_row_key: dict = {}
    for r in rows:
        author_profile = profiles_map.get(r["author_id"]) or {}
        segment = {
            "id": r["id"],
            "author_id": r["author_id"],
            "author_name": author_profile.get("full_name"),
            "author_role": author_profile.get("role"),
            "content": r["content"],
            "created_at": r["created_at"],
            "updated_at": r["updated_at"],
        }
        by_row_key.setdefault(r["row_key"], {}).setdefault(r["group_id"], []).append(segment)

    result = {}
    for row_key, groups in by_row_key.items():
        group_list = []
        for group_id, segments in groups.items():
            segments.sort(key=lambda s: s["created_at"])  # oldest segment first within a record
            group_list.append({"group_id": group_id, "segments": segments})
        group_list.sort(key=lambda g: g["segments"][0]["created_at"], reverse=True)  # newest record first
        result[row_key] = group_list
    return result


class PartialUpdateCreateIn(BaseModel):
    division: str = "main"
    budget_name: str | None = None
    row_key: str
    content: str


@router.post("/{school_id}/partial-updates")
def create_partial_update(
    school_id: str,
    body: PartialUpdateCreateIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    import uuid

    content = body.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="לא ניתן לשמור עדכון ריק")
    db = get_admin_client()
    row = db.table("partial_row_updates").insert({
        "school_id": school_id,
        "division": body.division,
        "budget_name": body.budget_name,
        "row_key": body.row_key,
        "group_id": str(uuid.uuid4()),
        "author_id": user["id"],
        "content": content,
    }).execute()
    return row.data[0]


class PartialUpdateSegmentIn(BaseModel):
    content: str


# Role hierarchy for cross-user edit/delete of another author's segment: a user
# may only act on a segment written by someone with a STRICTLY lower rank than
# their own (e.g. manager can act on advisor's segments but not on owner's).
# Acting on your own segment is always allowed regardless of rank.
_PARTIAL_UPDATE_ROLE_RANK = {"owner": 3, "manager": 2, "advisor": 1}


def _get_partial_segment_author(db, school_id: str, segment_id: str) -> tuple[str, str | None]:
    existing = db.table("partial_row_updates").select("author_id").eq("id", segment_id).eq("school_id", school_id).execute()
    if not existing.data:
        raise HTTPException(status_code=404, detail="העדכון לא נמצא")
    author_id = existing.data[0]["author_id"]
    author_role = None
    try:
        prof = db.table("profiles").select("role").eq("id", author_id).execute()
        if prof.data:
            author_role = prof.data[0].get("role")
    except Exception as exc:
        logger.warning("partial-updates author role lookup failed (fails closed on cross-user actions): %s", exc)
    return author_id, author_role


def _can_edit_partial_segment(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["id"] == author_id:
        return True
    if not author_role:
        return False  # unknown role — fail closed, only self-edit allowed
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


def _can_delete_partial_segment(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["role"] not in ("owner", "manager"):
        return False
    if user["id"] == author_id:
        return True
    if not author_role:
        return False
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


@router.patch("/{school_id}/partial-updates/segments/{segment_id}")
def edit_partial_update_segment(
    school_id: str,
    segment_id: str,
    body: PartialUpdateSegmentIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    content = body.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="לא ניתן לשמור עדכון ריק")
    db = get_admin_client()
    author_id, author_role = _get_partial_segment_author(db, school_id, segment_id)
    if not _can_edit_partial_segment(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה לערוך עדכון זה")
    from datetime import datetime, timezone

    db.table("partial_row_updates").update({
        "content": content,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", segment_id).execute()
    return {"ok": True}


@router.delete("/{school_id}/partial-updates/segments/{segment_id}")
def delete_partial_update_segment(
    school_id: str,
    segment_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    author_id, author_role = _get_partial_segment_author(db, school_id, segment_id)
    if not _can_delete_partial_segment(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק עדכון זה")
    db.table("partial_row_updates").delete().eq("id", segment_id).eq("school_id", school_id).execute()
    return {"ok": True}


# ---------------------------------------------------------------------------
# School notes ("הערות" + "הערות רבעוניות" — accessed from the school card's
# "הערות" button, and quarterly notes additionally surfaced as columns in
# "ניהול בתי ספר"). Modeled directly on the partial_row_updates pattern above,
# with note_type/quarter as the discriminator instead of division/row_key.
# Quarterly notes are hidden entirely from advisors (server-side, not just UI).
# ---------------------------------------------------------------------------

def _group_school_note_rows(rows: list[dict], profiles_map: dict) -> list[dict]:
    by_group: dict = {}
    for r in rows:
        author_profile = profiles_map.get(r["author_id"]) or {}
        segment = {
            "id": r["id"],
            "author_id": r["author_id"],
            "author_name": author_profile.get("full_name"),
            "author_role": author_profile.get("role"),
            "content": r["content"],
            "created_at": r["created_at"],
            "updated_at": r["updated_at"],
        }
        by_group.setdefault(r["group_id"], []).append(segment)
    group_list = []
    for group_id, segments in by_group.items():
        segments.sort(key=lambda s: s["created_at"])  # oldest segment first within a record
        group_list.append({"group_id": group_id, "segments": segments})
    group_list.sort(key=lambda g: g["segments"][0]["created_at"], reverse=True)  # newest record first
    return group_list


@router.get("/{school_id}/notes")
def get_school_notes(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    rows = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = db.table("school_notes").select("*").eq("school_id", school_id).order("created_at").execute().data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_school_notes attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_school_notes failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    author_ids = list({r["author_id"] for r in rows if r.get("author_id")})
    profiles_map = {}
    if author_ids:
        try:
            db = get_admin_client()
            p_rows = db.table("profiles").select("id, full_name, role").in_("id", author_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
        except Exception as exc:
            logger.warning("get_school_notes profile enrichment failed (non-fatal): %s", exc)

    general_rows = [r for r in rows if r["note_type"] == "general"]
    result = {"general": _group_school_note_rows(general_rows, profiles_map), "quarterly": {"1": [], "2": [], "3": [], "4": []}}

    # Quarterly notes are entirely inaccessible to advisors — enforced server-side, not just UI
    if user["role"] != "advisor":
        for q in (1, 2, 3, 4):
            q_rows = [r for r in rows if r["note_type"] == "quarterly" and r["quarter"] == q]
            result["quarterly"][str(q)] = _group_school_note_rows(q_rows, profiles_map)

    return result


class SchoolNoteCreateIn(BaseModel):
    note_type: str
    quarter: int | None = None
    content: str


@router.post("/{school_id}/notes")
def create_school_note(
    school_id: str,
    body: SchoolNoteCreateIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    import uuid

    if body.note_type not in ("general", "quarterly"):
        raise HTTPException(status_code=400, detail="סוג הערה לא תקין")
    if body.note_type == "quarterly":
        if body.quarter not in (1, 2, 3, 4):
            raise HTTPException(status_code=400, detail="יש לבחור רבעון בין 1 ל-4")
        if user["role"] == "advisor":
            raise HTTPException(status_code=403, detail="אין הרשאה להוסיף הערה רבעונית")
    elif body.quarter is not None:
        raise HTTPException(status_code=400, detail="הערה כללית לא יכולה להיות משויכת לרבעון")

    content = body.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="לא ניתן לשמור הערה ריקה")
    db = get_admin_client()
    row = db.table("school_notes").insert({
        "school_id": school_id,
        "note_type": body.note_type,
        "quarter": body.quarter,
        "group_id": str(uuid.uuid4()),
        "author_id": user["id"],
        "content": content,
    }).execute()
    return row.data[0]


class SchoolNoteSegmentIn(BaseModel):
    content: str


def _get_school_note_author(db, school_id: str, segment_id: str) -> tuple[str, str | None, str]:
    existing = db.table("school_notes").select("author_id, note_type").eq("id", segment_id).eq("school_id", school_id).execute()
    if not existing.data:
        raise HTTPException(status_code=404, detail="ההערה לא נמצאה")
    author_id = existing.data[0]["author_id"]
    note_type = existing.data[0]["note_type"]
    author_role = None
    try:
        prof = db.table("profiles").select("role").eq("id", author_id).execute()
        if prof.data:
            author_role = prof.data[0].get("role")
    except Exception as exc:
        logger.warning("school-notes author role lookup failed (fails closed on cross-user actions): %s", exc)
    return author_id, author_role, note_type


def _can_edit_school_note(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["id"] == author_id:
        return True
    if not author_role:
        return False
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


def _can_delete_school_note(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["role"] not in ("owner", "manager"):
        return False
    if user["id"] == author_id:
        return True
    if not author_role:
        return False
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


@router.patch("/{school_id}/notes/segments/{segment_id}")
def edit_school_note_segment(
    school_id: str,
    segment_id: str,
    body: SchoolNoteSegmentIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    content = body.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="לא ניתן לשמור הערה ריקה")
    db = get_admin_client()
    author_id, author_role, note_type = _get_school_note_author(db, school_id, segment_id)
    if note_type == "quarterly" and user["role"] == "advisor":
        raise HTTPException(status_code=403, detail="אין הרשאה לערוך הערה רבעונית")
    if not _can_edit_school_note(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה לערוך הערה זו")
    from datetime import datetime, timezone

    db.table("school_notes").update({
        "content": content,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", segment_id).execute()
    return {"ok": True}


@router.delete("/{school_id}/notes/segments/{segment_id}")
def delete_school_note_segment(
    school_id: str,
    segment_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    author_id, author_role, note_type = _get_school_note_author(db, school_id, segment_id)
    if note_type == "quarterly" and user["role"] == "advisor":
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק הערה רבעונית")
    if not _can_delete_school_note(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק הערה זו")
    db.table("school_notes").delete().eq("id", segment_id).eq("school_id", school_id).execute()
    return {"ok": True}


class QuarterlyNotesSummaryBatchIn(BaseModel):
    school_ids: list[str]


@router.post("/notes/quarterly-summary/batch")
def batch_get_quarterly_notes_summary(
    body: QuarterlyNotesSummaryBatchIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    if not body.school_ids or user["role"] == "advisor":
        return {}
    rows = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("school_notes")
                .select("*")
                .eq("note_type", "quarterly")
                .in_("school_id", body.school_ids)
                .order("created_at")
                .execute()
                .data or []
            )
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("batch_get_quarterly_notes_summary attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("batch_get_quarterly_notes_summary failed after 2 attempts: %s", exc, exc_info=True)
                return {}

    author_ids = list({r["author_id"] for r in rows if r.get("author_id")})
    profiles_map = {}
    if author_ids:
        try:
            db = get_admin_client()
            p_rows = db.table("profiles").select("id, full_name, role").in_("id", author_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
        except Exception as exc:
            logger.warning("batch_get_quarterly_notes_summary profile enrichment failed (non-fatal): %s", exc)

    # Group by school_id -> quarter -> group_id (a "record"), so the count reflects
    # distinct notes rather than raw segments.
    by_school: dict = {}
    for r in rows:
        by_school.setdefault(r["school_id"], {}).setdefault(r["quarter"], {}).setdefault(r["group_id"], []).append(r)

    result = {}
    for school_id, quarters in by_school.items():
        result[school_id] = {}
        for q, groups in quarters.items():
            group_ids = list(groups.keys())
            count = len(group_ids)
            latest_segment = None
            if count == 1:
                only_group_rows = groups[group_ids[0]]
                latest_row = max(only_group_rows, key=lambda r: r["created_at"])
                author_profile = profiles_map.get(latest_row["author_id"]) or {}
                latest_segment = {
                    "id": latest_row["id"],
                    "author_id": latest_row["author_id"],
                    "author_name": author_profile.get("full_name"),
                    "author_role": author_profile.get("role"),
                    "content": latest_row["content"],
                    "created_at": latest_row["created_at"],
                    "updated_at": latest_row["updated_at"],
                }
            result[school_id][str(q)] = {"count": count, "latest_segment": latest_segment}
    return result


# ---------------------------------------------------------------------------
# School files ("קבצים" — file attachments with a description, shown right
# below "הערות" on the school card). One row = one file; unlike school_notes
# there's no group/segment threading since a file is replaced only via
# delete+re-upload, not edited in place. Visible to everyone (no advisor
# restriction), matching "הערות".
# ---------------------------------------------------------------------------

def _enrich_school_files_authors(db, rows: list[dict]) -> list[dict]:
    author_ids = list({r["author_id"] for r in rows if r.get("author_id")})
    profiles_map = {}
    if author_ids:
        try:
            p_rows = db.table("profiles").select("id, full_name, role").in_("id", author_ids).execute()
            profiles_map = {p["id"]: p for p in (p_rows.data or [])}
        except Exception as exc:
            logger.warning("school-files profile enrichment failed (non-fatal): %s", exc)
    for r in rows:
        author_profile = profiles_map.get(r["author_id"]) or {}
        r["author_name"] = author_profile.get("full_name")
        r["author_role"] = author_profile.get("role")
    return rows


@router.get("/{school_id}/files")
def get_school_files(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    rows = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = (
                db.table("school_files")
                .select("*")
                .eq("school_id", school_id)
                .order("created_at", desc=True)
                .execute()
                .data or []
            )
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_school_files attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_school_files failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    db = get_admin_client()
    return _enrich_school_files_authors(db, rows)


@router.post("/{school_id}/files")
async def upload_school_file(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    file: UploadFile = File(...),
    description: str = Form(""),
):
    import secrets
    import shutil
    import tempfile

    db = get_admin_client()
    run_dir = Path(tempfile.mkdtemp(prefix=f"school_file_{school_id}_"))
    try:
        suffix = Path(file.filename or "").suffix
        dest = run_dir / f"upload{suffix}"
        dest.write_bytes(await file.read())
        storage_key = f"school-files/{school_id}/{secrets.token_hex(8)}{suffix}"
        db.storage.from_("check-files").upload(storage_key, dest.read_bytes())
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)

    row = db.table("school_files").insert({
        "school_id": school_id,
        "author_id": user["id"],
        "description": description.strip() or None,
        "storage_key": storage_key,
        "file_name": file.filename or "file",
    }).execute()
    created = row.data[0]
    created["author_name"] = user.get("full_name")
    created["author_role"] = user.get("role")
    return created


class SchoolFileUpdateIn(BaseModel):
    description: str


def _get_school_file_author(db, school_id: str, file_id: str) -> tuple[str, str | None, str, str]:
    existing = db.table("school_files").select("author_id, storage_key, file_name").eq("id", file_id).eq("school_id", school_id).execute()
    if not existing.data:
        raise HTTPException(status_code=404, detail="הקובץ לא נמצא")
    author_id = existing.data[0]["author_id"]
    storage_key = existing.data[0]["storage_key"]
    file_name = existing.data[0]["file_name"]
    author_role = None
    try:
        prof = db.table("profiles").select("role").eq("id", author_id).execute()
        if prof.data:
            author_role = prof.data[0].get("role")
    except Exception as exc:
        logger.warning("school-files author role lookup failed (fails closed on cross-user actions): %s", exc)
    return author_id, author_role, storage_key, file_name


def _can_edit_school_file(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["id"] == author_id:
        return True
    if not author_role:
        return False
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


def _can_delete_school_file(user: dict, author_id: str, author_role: str | None) -> bool:
    if user["role"] not in ("owner", "manager"):
        return False
    if user["id"] == author_id:
        return True
    if not author_role:
        return False
    return _PARTIAL_UPDATE_ROLE_RANK.get(user["role"], 0) > _PARTIAL_UPDATE_ROLE_RANK.get(author_role, 0)


@router.patch("/{school_id}/files/{file_id}")
def edit_school_file(
    school_id: str,
    file_id: str,
    body: SchoolFileUpdateIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    author_id, author_role, _storage_key, _file_name = _get_school_file_author(db, school_id, file_id)
    if not _can_edit_school_file(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה לערוך קובץ זה")
    from datetime import datetime, timezone

    db.table("school_files").update({
        "description": body.description.strip() or None,
        "updated_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", file_id).execute()
    return {"ok": True}


@router.delete("/{school_id}/files/{file_id}")
def delete_school_file(
    school_id: str,
    file_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    author_id, author_role, storage_key, _file_name = _get_school_file_author(db, school_id, file_id)
    if not _can_delete_school_file(user, author_id, author_role):
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק קובץ זה")
    db.table("school_files").delete().eq("id", file_id).eq("school_id", school_id).execute()
    try:
        db.storage.from_("check-files").remove([storage_key])
    except Exception as exc:
        logger.warning("school-files storage cleanup failed (non-fatal, DB row already removed): %s", exc)
    return {"ok": True}


@router.get("/{school_id}/files/{file_id}/download")
def download_school_file(
    school_id: str,
    file_id: str,
    user: Annotated[dict, Depends(get_current_user)],
):
    db = get_admin_client()
    row = db.table("school_files").select("storage_key, file_name").eq("id", file_id).eq("school_id", school_id).execute()
    if not row.data:
        raise HTTPException(status_code=404, detail="הקובץ לא נמצא")
    storage_key = row.data[0]["storage_key"]
    file_name = row.data[0]["file_name"] or "file"
    content = db.storage.from_("check-files").download(storage_key)

    # RFC 5987 filename* — Hebrew filenames render as gibberish or get dropped
    # entirely if embedded raw in a plain filename= parameter.
    import urllib.parse

    ext = Path(file_name).suffix or ""
    ascii_fallback = f"file{ext}"
    encoded_name = urllib.parse.quote(file_name)
    return Response(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded_name}"},
    )


# ---------------------------------------------------------------------------
# Bulk import from Excel
# ---------------------------------------------------------------------------

DIVISION_HEB_MAP = {
    "חטיבה עליונה": "tikkon", "תיכון": "tikkon", "tikkon": "tikkon",
    "חטיבת ביניים": "beinayim", "ביניים": "beinayim", "beinayim": "beinayim",
    "יסודי": "yesodi", "yesodi": "yesodi",
    "אחר": "other", "other": "other",
}


@router.post("/import")
async def import_schools(
    file: UploadFile,
    user: Annotated[dict, Depends(get_current_user)],
):
    _require_manager(user)
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "יש להעלות קובץ Excel בלבד (.xlsx)")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"imported": 0, "errors": ["הקובץ ריק"]}

    # Skip header row if first cell looks like a column title
    first_cell = str(rows[0][0] or "").strip()
    start = 1 if first_cell in ("שם בית ספר", "שם", "name", "Name") else 0

    db = get_admin_client()
    imported = 0
    errors = []

    for i, row in enumerate(rows[start:], start=start + 2):
        name = str(row[0] or "").strip() if len(row) > 0 else ""
        symbol = str(row[1] or "").strip().split(".")[0] if len(row) > 1 else ""
        city = str(row[2] or "").strip() if len(row) > 2 else ""
        notes = str(row[3] or "").strip() if len(row) > 3 else ""
        divisions_raw = str(row[4] or "").strip() if len(row) > 4 else ""

        if not name:
            continue

        if not symbol or not symbol.isdigit() or len(symbol) not in (5, 6):
            errors.append(f"שורה {i}: סמל מוסד לא תקין — '{symbol}'")
            continue

        try:
            school_data: dict = {"name": name, "symbol": symbol, "org_id": user["org_id"]}
            if city:
                school_data["city"] = city
            if notes:
                school_data["notes"] = notes

            res = db.table("schools").insert(school_data).execute()
            school_id = res.data[0]["id"]

            if divisions_raw:
                for div_str in divisions_raw.split(","):
                    div_type = DIVISION_HEB_MAP.get(div_str.strip())
                    if div_type:
                        try:
                            db.table("gefen_accounts").insert(
                                {"school_id": school_id, "division_type": div_type}
                            ).execute()
                        except Exception:
                            pass

            imported += 1
        except Exception as exc:
            errors.append(f"שורה {i}: {str(exc)[:80]}")

    return {"imported": imported, "errors": errors}


# ---------------------------------------------------------------------------
# Meetings
# ---------------------------------------------------------------------------

def _build_meeting_subject(db, school_id: str, participants: list[dict] | None, primary_contact_key: str | None) -> str:
    """Outlook event subject: "<school>, <city> - <contact name> - <contact phone>".

    The contact is resolved from `participants` (selected in the meeting row): if exactly
    one principal-like contact (principal / principal_chativa) is among them, it wins;
    otherwise `primary_contact_key` (explicitly chosen by the user via a disambiguation
    dialog on the frontend — e.g. multiple participants with no single principal, or both
    principals selected together for a six-year school) picks which one. Falls back to
    just "<school>, <city>" if there's no participant to resolve.
    """
    try:
        row = (
            db.table("schools")
            .select("name, city, principal_name, principal_phone, principal_chativa_name, "
                     "principal_chativa_phone, secretary_name, secretary_phone, "
                     "finance_contact_name, finance_contact_phone, extra_contacts")
            .eq("id", school_id)
            .execute()
        )
        school = row.data[0] if row.data else {}
    except Exception:
        school = {}

    school_name = school.get("name") or ""
    city = school.get("city") or ""
    base = f"{school_name}, {city}" if city else school_name

    contact_map = {
        "principal": (school.get("principal_name"), school.get("principal_phone")),
        "principal_chativa": (school.get("principal_chativa_name"), school.get("principal_chativa_phone")),
        "secretary": (school.get("secretary_name"), school.get("secretary_phone")),
        "finance": (school.get("finance_contact_name"), school.get("finance_contact_phone")),
    }
    for i, ec in enumerate(school.get("extra_contacts") or []):
        contact_map[f"extra_{i}"] = (ec.get("name"), ec.get("phone"))

    participants = participants or []
    key = primary_contact_key
    if not key:
        principal_matches = [p.get("key") for p in participants if p.get("key") in ("principal", "principal_chativa")]
        if len(principal_matches) == 1:
            key = principal_matches[0]
        elif len(participants) == 1:
            key = participants[0].get("key")

    name, phone = contact_map.get(key, (None, None)) if key else (None, None)
    if name:
        return f"{base} - {name} - {phone or ''}".rstrip(" -")
    return base


class MeetingSubjectPreviewIn(BaseModel):
    participants: list[dict] | None = None
    primary_contact_key: str | None = None


@router.post("/{school_id}/meeting-subject-preview")
def preview_meeting_subject(school_id: str, body: MeetingSubjectPreviewIn, user: Annotated[dict, Depends(get_current_user)]):
    """What the Outlook event subject will look like for this school + participant
    selection — used by the frontend's conflict-warning dialog so it shows the same
    text that will actually be sent to Outlook, not a stale hardcoded guess."""
    db = get_admin_client()
    subject = _build_meeting_subject(db, school_id, body.participants, body.primary_contact_key)
    return {"subject": subject}


@router.get("/{school_id}/calls")
def list_school_calls(
    school_id: str,
    user: Annotated[dict, Depends(get_current_user)],
    academic_year: str | None = None,
):
    """Calls (from Voicenter) matched to this school's contacts — same visibility as the
    rest of the school card (advisor assigned / restrict_access_to / manager+)."""
    from academic_years import get_academic_year_date_range
    from routers.voicenter_router import _pull_org_calls

    for attempt in range(2):
        try:
            db = get_admin_client()
            school_row = db.table("schools").select("id, restrict_access_to").eq("id", school_id).eq("org_id", user["org_id"]).execute()
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_school_calls attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_school_calls failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    if not school_row.data:
        raise HTTPException(status_code=404, detail="בית ספר לא נמצא")
    if user["role"] not in ("owner", "manager") and not _advisor_has_access_to_school_row(db, user["id"], school_row.data[0]):
        raise HTTPException(status_code=403, detail="אין הרשאה לצפות בבית ספר זה")

    start_date, end_date = get_academic_year_date_range(academic_year or DEFAULT_ACADEMIC_YEAR)
    date_from = f"{start_date.isoformat()}T00:00:00"
    date_to = f"{end_date.isoformat()}T23:59:59"

    try:
        result = _pull_org_calls(user["org_id"], date_from, date_to)
    except HTTPException as exc:
        if exc.status_code == 400:
            # Voicenter not configured/enabled for this org — friendly empty state, not an error
            return {"calls": [], "voicenter_enabled": False}
        raise

    calls = [
        c for c in result["calls"]
        if school_id not in (c.get("excluded_school_ids") or [])
        and (c.get("school_id") == school_id or school_id in (c.get("linked_school_ids") or []))
    ]
    return {"calls": calls, "voicenter_enabled": True}


# ---------------------------------------------------------------------------
# Meeting "בפועל" activity — calls attributed to a specific meeting + manually
# logged offline work, feeding both display and (opt-in) auto-completion.
# ---------------------------------------------------------------------------

def _call_time_to_israel_hm(call_time_iso: str | None) -> str | None:
    """Voicenter's `date`/call_time field is a UTC ISO timestamp (e.g. "...T07:12:43Z").
    Naively slicing characters [11:16] grabs the raw UTC wall-clock time, not Israel local
    time — off by 2-3 hours depending on DST. Every other place that shows this same data
    (שיחות tab, ביצועים tab) converts via a JS `Date` object, which localizes correctly to
    the browser's Israel timezone; this does the equivalent conversion on the backend for
    values that get stored/compared as plain "HH:MM" (meeting.calls_start_time, and the
    time-bucket attribution fallback below)."""
    if not call_time_iso:
        return None
    from zoneinfo import ZoneInfo
    try:
        dt = datetime.fromisoformat(call_time_iso.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(ZoneInfo("Asia/Jerusalem")).strftime("%H:%M")
    except Exception:
        return call_time_iso[11:16] if len(call_time_iso) >= 16 else None


def _attribute_calls_to_meetings(calls: list[dict], meetings_for_day: list[dict]) -> dict:
    """Two-tier attribution for a school with possibly more than one meeting the same day:
    1) role/stage match (מנהל/ת -&gt; tichon, מנהל/ת חט"ב -&gt; chativa) IF it uniquely matches
       exactly one of that day's meetings' stage_scope;
    2) otherwise, time-bucket fallback — earliest meeting whose end_time &gt;= the call's time,
       or the day's last meeting as a catch-all. Returns {call_id: meeting_id}."""
    if not meetings_for_day:
        return {}
    meetings_sorted = sorted(meetings_for_day, key=lambda m: m.get("end_time") or "23:59")
    assignments: dict = {}
    for call in calls:
        role = call.get("contact_role")
        stage_signal = "tichon" if role == "מנהל/ת" else ("chativa" if role == "מנהל/ת חט\"ב" else None)
        target = None
        if stage_signal:
            matches = [m for m in meetings_sorted if m.get("stage_scope") in (stage_signal, "both")]
            if len(matches) == 1:
                target = matches[0]
        if target is None:
            call_time_hm = _call_time_to_israel_hm(call.get("start_time")) or ""
            candidates = [m for m in meetings_sorted if m.get("end_time") and m["end_time"] >= call_time_hm]
            target = candidates[0] if candidates else meetings_sorted[-1]
        assignments[call["call_id"]] = target["id"]
    return assignments


def _recompute_meeting_aggregates(db, meeting_id: str) -> None:
    """Recomputes calls_start_time/calls_duration_seconds from whatever meeting_call_links
    rows currently point at this meeting (both auto and manual) — no Voicenter call, cheap."""
    rows = db.table("meeting_call_links").select("call_time, duration_seconds").eq("meeting_id", meeting_id).execute().data or []
    if rows:
        earliest = min(rows, key=lambda r: r.get("call_time") or "")
        calls_start_time = _call_time_to_israel_hm(earliest.get("call_time"))
        calls_duration_seconds = sum(r.get("duration_seconds") or 0 for r in rows)
    else:
        calls_start_time = None
        calls_duration_seconds = 0
    db.table("meetings").update({
        "calls_start_time": calls_start_time,
        "calls_duration_seconds": calls_duration_seconds,
        "calls_synced_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", meeting_id).execute()


def _recompute_meeting_call_activity(db, org_id: str, school_id: str, meeting_date: str, calls_for_school: list[dict]) -> list[str]:
    """Attributes calls_for_school (already filtered to this school for this date, from
    _pull_org_calls) to the day's meetings for this school, upserts meeting_call_links
    (never overwriting source='manual' rows), and refreshes each meeting's cached aggregate
    columns. Returns the affected meeting ids (all of that day's meetings for this school,
    even ones with zero calls — marks them as synced)."""
    meetings_for_day = (
        db.table("meetings").select("id, stage_scope, end_time, status")
        .eq("school_id", school_id).eq("meeting_date", meeting_date).execute()
    ).data or []
    if not meetings_for_day:
        return []

    call_ids = [c["call_id"] for c in calls_for_school]
    existing = []
    if call_ids:
        existing = (
            db.table("meeting_call_links").select("call_id, source")
            .eq("org_id", org_id).in_("call_id", call_ids).execute()
        ).data or []
    manual_call_ids = {r["call_id"] for r in existing if r.get("source") == "manual"}

    attributable = [
        c for c in calls_for_school
        if c["call_id"] not in manual_call_ids and not c.get("pending_school_resolution")
    ]
    assignments = _attribute_calls_to_meetings(attributable, meetings_for_day)

    rows = []
    for c in attributable:
        meeting_id = assignments.get(c["call_id"])
        if not meeting_id:
            continue
        rows.append({
            "org_id": org_id,
            "meeting_id": meeting_id,
            "school_id": school_id,
            "call_id": c["call_id"],
            "call_time": c.get("start_time"),
            "duration_seconds": c.get("duration_seconds") or 0,
            "contact_name": c.get("contact_name"),
            "contact_role": c.get("contact_role"),
            "counterpart_phone": c.get("counterpart_phone"),
            "advisor_id": c.get("advisor_id"),
            "source": "auto",
        })
    if rows:
        db.table("meeting_call_links").upsert(rows, on_conflict="org_id,call_id").execute()

    meeting_ids = [m["id"] for m in meetings_for_day]
    for mid in meeting_ids:
        _recompute_meeting_aggregates(db, mid)
    return meeting_ids


def _sum_offline_seconds(entries: list[dict] | None) -> int:
    total = 0
    for e in (entries or []):
        st, et = e.get("start_time"), e.get("end_time")
        if not st or not et:
            continue
        try:
            sh, sm = (int(x) for x in st.split(":")[:2])
            eh, em = (int(x) for x in et.split(":")[:2])
            total += max(0, (eh * 60 + em) - (sh * 60 + sm)) * 60
        except Exception:
            continue
    return total


def _maybe_auto_complete_meeting(db, org_id: str, meeting_id: str) -> None:
    """Auto-flips a meeting to 'completed' when: the org's automation toggle is on, the
    meeting is still 'scheduled', its scheduled end_time has already passed, and its total
    invested time (calls + offline work) reaches 5 minutes. Never touches cancelled/postponed
    meetings, and never reverses a completion that already happened."""
    try:
        org_res = db.table("organizations").select("auto_complete_meetings_from_activity_enabled").eq("id", org_id).execute()
        org = org_res.data[0] if org_res.data else None
        if not org or not org.get("auto_complete_meetings_from_activity_enabled"):
            return
        m_res = db.table("meetings").select(
            "id, school_id, status, meeting_date, end_time, calls_duration_seconds, offline_work_entries"
        ).eq("id", meeting_id).execute()
        m = m_res.data[0] if m_res.data else None
        if not m or m.get("status") != "scheduled" or not m.get("meeting_date") or not m.get("end_time"):
            return
        from zoneinfo import ZoneInfo
        now_il = datetime.now(ZoneInfo("Asia/Jerusalem"))
        try:
            meeting_end = datetime.fromisoformat(f"{m['meeting_date']}T{m['end_time']}:00").replace(tzinfo=ZoneInfo("Asia/Jerusalem"))
        except ValueError:
            return
        if now_il < meeting_end:
            return
        total = (m.get("calls_duration_seconds") or 0) + _sum_offline_seconds(m.get("offline_work_entries"))
        if total < 300:
            return
        _apply_meeting_patch(db, org_id, m["school_id"], meeting_id, {"status": "completed"})
    except Exception as exc:
        logger.warning("_maybe_auto_complete_meeting failed (non-fatal) for meeting %s: %s", meeting_id, exc)


def _require_school_access_for_meeting_activity(db, user: dict, school_id: str) -> dict:
    school_row = db.table("schools").select("id, restrict_access_to").eq("id", school_id).eq("org_id", user["org_id"]).execute()
    if not school_row.data:
        raise HTTPException(status_code=404, detail="בית ספר לא נמצא")
    if user["role"] not in ("owner", "manager") and not _advisor_has_access_to_school_row(db, user["id"], school_row.data[0]):
        raise HTTPException(status_code=403, detail="אין הרשאה לפעולה זו")
    return school_row.data[0]


@router.get("/{school_id}/users-with-access")
def list_users_with_access(school_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """All users (any role) who currently have access to this school — owners/managers
    (implicit org-wide access) plus advisors covered by restrict_access_to / advisor_schools.
    Used to populate the offline-work user multi-select in MeetingActualDetail."""
    for attempt in range(2):
        try:
            db = get_admin_client()
            school = _require_school_access_for_meeting_activity(db, user, school_id)

            profiles = db.table("profiles").select("id, full_name, role").eq("org_id", user["org_id"]).execute().data or []
            adv_rows = db.table("advisor_schools").select("advisor_id").eq("school_id", school_id).execute().data or []
            explicit_ids = {r["advisor_id"] for r in adv_rows}
            rat = school.get("restrict_access_to")

            result = []
            for p in profiles:
                if p["role"] in ("owner", "manager") or rat is None or p["id"] in (rat or []) or p["id"] in explicit_ids:
                    result.append({"id": p["id"], "full_name": p["full_name"], "role": p["role"]})
            return result
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_users_with_access attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.1)
            else:
                logger.error("list_users_with_access failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")


@router.get("/{school_id}/meetings/{meeting_id}/actual-detail")
def get_meeting_actual_detail(school_id: str, meeting_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Powers the meeting row's expanded 'בפועל' panel: calls attributed to this meeting +
    manually logged offline-work entries.

    For today's/yesterday's meetings, ALWAYS re-syncs from Voicenter before returning — same
    "pull fresh on open" behavior as the שיחות and ביצועים tabs, so a call placed minutes ago
    shows up immediately instead of waiting for the 10-minute cron (recompute-call-activity)
    to catch up. Older meetings only bootstrap once (first-ever open) since their calls are
    long settled and re-pulling them on every view would be pure waste."""
    from zoneinfo import ZoneInfo

    db = get_admin_client()
    _require_school_access_for_meeting_activity(db, user, school_id)

    m_res = db.table("meetings").select("*").eq("id", meeting_id).eq("school_id", school_id).execute()
    if not m_res.data:
        raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
    meeting = m_res.data[0]

    today_il_date = datetime.now(ZoneInfo("Asia/Jerusalem")).date()
    today_il = today_il_date.isoformat()
    yesterday_il = (today_il_date - timedelta(days=1)).isoformat()
    meeting_date = meeting.get("meeting_date")
    is_recent = meeting_date in (today_il, yesterday_il)
    should_sync = meeting_date and meeting_date <= today_il and (is_recent or not meeting.get("calls_synced_at"))
    if should_sync:
        try:
            from routers.voicenter_router import _pull_org_calls
            d = meeting["meeting_date"]
            result = _pull_org_calls(user["org_id"], f"{d}T00:00:00", f"{d}T23:59:59")
            calls_for_school = [
                c for c in result["calls"]
                if school_id not in (c.get("excluded_school_ids") or [])
                and (c.get("school_id") == school_id or school_id in (c.get("linked_school_ids") or []))
            ]
            _recompute_meeting_call_activity(db, user["org_id"], school_id, d, calls_for_school)
            refreshed = db.table("meetings").select("*").eq("id", meeting_id).execute()
            if refreshed.data:
                meeting = refreshed.data[0]
        except Exception as exc:
            logger.warning("actual-detail on-demand recompute failed (non-fatal) for meeting %s: %s", meeting_id, exc)

    link_rows = db.table("meeting_call_links").select("*").eq("meeting_id", meeting_id).order("call_time").execute().data or []

    # AI summary/transcript-availability are always looked up fresh here (never cached on
    # meeting_call_links) so late-arriving transcripts show up the moment a user opens the row.
    ai_by_call_id: dict = {}
    call_ids = [r["call_id"] for r in link_rows]
    if call_ids:
        try:
            ai_rows = (
                db.table("voicenter_call_ai").select("call_id, summary, transcript_path")
                .eq("org_id", user["org_id"]).in_("call_id", call_ids).execute()
            ).data or []
            ai_by_call_id = {r["call_id"]: r for r in ai_rows}
        except Exception as exc:
            logger.warning("actual-detail AI enrichment failed (non-fatal): %s", exc)

    offline_entries = meeting.get("offline_work_entries") or []

    # Non-fatal enrichment: resolve advisor_id (calls) and user_ids/created_by (offline entries)
    # to display names in one batched profiles query.
    names_by_id: dict = {}
    try:
        needed_ids = {r["advisor_id"] for r in link_rows if r.get("advisor_id")}
        for e in offline_entries:
            for uid in (e.get("user_ids") or ([e["created_by"]] if e.get("created_by") else [])):
                needed_ids.add(uid)
        if needed_ids:
            p_rows = db.table("profiles").select("id, full_name").in_("id", list(needed_ids)).execute().data or []
            names_by_id = {p["id"]: p["full_name"] for p in p_rows}
    except Exception as exc:
        logger.warning("actual-detail user-name enrichment failed (non-fatal): %s", exc)

    calls_out = []
    for r in link_rows:
        ai = ai_by_call_id.get(r["call_id"])
        calls_out.append({
            "call_id": r["call_id"],
            "call_time": r["call_time"],
            "duration_seconds": r["duration_seconds"],
            "contact_name": r["contact_name"],
            "contact_role": r["contact_role"],
            "counterpart_phone": r.get("counterpart_phone"),
            "notes": r.get("notes"),
            "source": r["source"],
            "advisor_id": r.get("advisor_id"),
            "advisor_name": names_by_id.get(r.get("advisor_id")),
            "ai_summary": ai["summary"] if ai else None,
            "ai_transcript_available": bool(ai and ai.get("transcript_path")),
        })

    for e in offline_entries:
        user_ids = e.get("user_ids") or ([e["created_by"]] if e.get("created_by") else [])
        e["users"] = [{"id": uid, "full_name": names_by_id.get(uid)} for uid in user_ids]

    other_meetings_same_day = []
    if meeting.get("meeting_date"):
        other_meetings_same_day = (
            db.table("meetings").select("id, start_time, end_time, stage_scope")
            .eq("school_id", school_id).eq("meeting_date", meeting["meeting_date"])
            .neq("id", meeting_id).execute()
        ).data or []

    offline_seconds = _sum_offline_seconds(offline_entries)
    calls_seconds = meeting.get("calls_duration_seconds") or 0

    return {
        "calls_start_time": meeting.get("calls_start_time"),
        "calls_duration_seconds": calls_seconds,
        "offline_duration_seconds": offline_seconds,
        "total_seconds": calls_seconds + offline_seconds,
        "calls": calls_out,
        "offline_entries": offline_entries,
        "other_meetings_same_day": other_meetings_same_day,
    }


class CallLinkNotesIn(BaseModel):
    notes: str | None = None


@router.patch("/meetings/{meeting_id}/call-links/{call_id}/notes")
def update_call_link_notes(meeting_id: str, call_id: str, body: CallLinkNotesIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    link_res = db.table("meeting_call_links").select("id, school_id").eq("meeting_id", meeting_id).eq("call_id", call_id).eq("org_id", user["org_id"]).execute()
    if not link_res.data:
        raise HTTPException(status_code=404, detail="שיחה לא נמצאה עבור פגישה זו")
    link = link_res.data[0]
    _require_school_access_for_meeting_activity(db, user, link["school_id"])
    db.table("meeting_call_links").update({"notes": body.notes}).eq("id", link["id"]).execute()
    return {"ok": True}


class ReassignCallIn(BaseModel):
    target_meeting_id: str


@router.patch("/meetings/{meeting_id}/call-links/{call_id}/reassign")
def reassign_call_link(meeting_id: str, call_id: str, body: ReassignCallIn, user: Annotated[dict, Depends(get_current_user)]):
    """Moves a call from one meeting to another — must be the same school, same day. Marks
    the link source='manual' so future recomputes never move it back automatically. Updates
    both meetings' cached aggregates immediately and re-checks auto-completion for the
    DESTINATION only — never un-completes the source meeting."""
    db = get_admin_client()
    link_res = db.table("meeting_call_links").select("*").eq("meeting_id", meeting_id).eq("call_id", call_id).eq("org_id", user["org_id"]).execute()
    if not link_res.data:
        raise HTTPException(status_code=404, detail="שיחה לא נמצאה עבור פגישה זו")
    link = link_res.data[0]
    _require_school_access_for_meeting_activity(db, user, link["school_id"])

    source_res = db.table("meetings").select("id, school_id, meeting_date").eq("id", meeting_id).execute()
    target_res = db.table("meetings").select("id, school_id, meeting_date").eq("id", body.target_meeting_id).execute()
    if not source_res.data or not target_res.data:
        raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
    sm, tm = source_res.data[0], target_res.data[0]
    if sm["school_id"] != tm["school_id"] or sm["meeting_date"] != tm["meeting_date"]:
        raise HTTPException(status_code=400, detail="ניתן להעביר שיחה רק לפגישה אחרת של אותו בית ספר באותו יום")

    db.table("meeting_call_links").update({
        "meeting_id": body.target_meeting_id,
        "source": "manual",
        "linked_by": user["id"],
        "linked_at": datetime.now(timezone.utc).isoformat(),
    }).eq("id", link["id"]).execute()

    _recompute_meeting_aggregates(db, meeting_id)
    _recompute_meeting_aggregates(db, body.target_meeting_id)
    _maybe_auto_complete_meeting(db, user["org_id"], body.target_meeting_id)
    return {"ok": True}


class OfflineWorkIn(BaseModel):
    start_time: str
    end_time: str
    notes: str


class OfflineWorkPatchIn(BaseModel):
    start_time: str | None = None
    end_time: str | None = None
    notes: str | None = None
    user_ids: list[str] | None = None


@router.post("/{school_id}/meetings/{meeting_id}/offline-work")
def add_offline_work(school_id: str, meeting_id: str, body: OfflineWorkIn, user: Annotated[dict, Depends(get_current_user)]):
    import uuid
    if len(body.notes.strip()) < 10:
        raise HTTPException(status_code=400, detail="יש להוסיף הערה של לפחות 10 תווים")
    db = get_admin_client()
    _require_school_access_for_meeting_activity(db, user, school_id)
    m_res = db.table("meetings").select("id, offline_work_entries").eq("id", meeting_id).eq("school_id", school_id).execute()
    if not m_res.data:
        raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
    entries = m_res.data[0].get("offline_work_entries") or []
    entry = {
        "id": str(uuid.uuid4()),
        "start_time": body.start_time,
        "end_time": body.end_time,
        "notes": body.notes,
        "created_by": user["id"],
        "user_ids": [user["id"]],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    entries.append(entry)
    db.table("meetings").update({"offline_work_entries": entries}).eq("id", meeting_id).execute()
    _maybe_auto_complete_meeting(db, user["org_id"], meeting_id)
    return {"ok": True, "entry": entry}


@router.patch("/{school_id}/meetings/{meeting_id}/offline-work/{entry_id}")
def update_offline_work(school_id: str, meeting_id: str, entry_id: str, body: OfflineWorkPatchIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    _require_school_access_for_meeting_activity(db, user, school_id)
    m_res = db.table("meetings").select("id, offline_work_entries").eq("id", meeting_id).eq("school_id", school_id).execute()
    if not m_res.data:
        raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
    if body.user_ids is not None and len(body.user_ids) == 0:
        raise HTTPException(status_code=400, detail="חייב להישאר לפחות משתמש אחד ברשומת עבודה עצמאית")
    entries = m_res.data[0].get("offline_work_entries") or []
    found = False
    for e in entries:
        if e.get("id") == entry_id:
            if body.start_time is not None: e["start_time"] = body.start_time
            if body.end_time is not None: e["end_time"] = body.end_time
            if body.notes is not None: e["notes"] = body.notes
            if body.user_ids is not None: e["user_ids"] = body.user_ids
            found = True
            break
    if not found:
        raise HTTPException(status_code=404, detail="רשומת עבודה עצמאית לא נמצאה")
    db.table("meetings").update({"offline_work_entries": entries}).eq("id", meeting_id).execute()
    _maybe_auto_complete_meeting(db, user["org_id"], meeting_id)
    return {"ok": True}


@router.delete("/{school_id}/meetings/{meeting_id}/offline-work/{entry_id}")
def delete_offline_work(school_id: str, meeting_id: str, entry_id: str, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    _require_school_access_for_meeting_activity(db, user, school_id)
    m_res = db.table("meetings").select("id, offline_work_entries").eq("id", meeting_id).eq("school_id", school_id).execute()
    if not m_res.data:
        raise HTTPException(status_code=404, detail="פגישה לא נמצאה")
    entries = [e for e in (m_res.data[0].get("offline_work_entries") or []) if e.get("id") != entry_id]
    db.table("meetings").update({"offline_work_entries": entries}).eq("id", meeting_id).execute()
    _maybe_auto_complete_meeting(db, user["org_id"], meeting_id)
    return {"ok": True}


@router.post("/meetings/recompute-call-activity")
def recompute_meeting_call_activity(request: Request):
    """Scheduled job (every 10 min, see .github/workflows/meeting-call-activity-recompute.yml):
    for every org, refreshes today's + yesterday's meetings' 'בפועל' call data (one Voicenter
    pull per org per day — cheap regardless of how many schools/meetings that covers) and
    runs the opt-in auto-completion check. Sequential — no ThreadPoolExecutor on the shared
    Supabase client."""
    if not CRON_SECRET or request.headers.get("X-Cron-Secret") != CRON_SECRET:
        raise HTTPException(status_code=403, detail="אין הרשאה")

    from zoneinfo import ZoneInfo
    from routers.voicenter_router import _pull_org_calls

    today_il = datetime.now(ZoneInfo("Asia/Jerusalem")).date()
    dates = [today_il.isoformat(), (today_il - timedelta(days=1)).isoformat()]

    for attempt in range(2):
        try:
            db = get_admin_client()
            orgs = db.table("organizations").select("id").execute().data or []
            break
        except Exception as exc:
            if attempt == 0:
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("recompute_meeting_call_activity: failed to load orgs: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת")

    processed = 0
    for org in orgs:
        org_id = org["id"]
        try:
            db = get_admin_client()
            school_ids = [s["id"] for s in (db.table("schools").select("id").eq("org_id", org_id).execute().data or [])]
        except Exception as exc:
            logger.warning("recompute_meeting_call_activity: failed to load schools for org %s: %s", org_id, exc)
            continue
        if not school_ids:
            continue

        for date_str in dates:
            try:
                db = get_admin_client()
                meeting_rows = (
                    db.table("meetings").select("school_id")
                    .eq("meeting_date", date_str).in_("school_id", school_ids).execute()
                ).data or []
            except Exception as exc:
                logger.warning("recompute_meeting_call_activity: failed to load meetings org=%s date=%s: %s", org_id, date_str, exc)
                continue
            schools_with_meetings = list({m["school_id"] for m in meeting_rows})
            if not schools_with_meetings:
                continue

            try:
                result = _pull_org_calls(org_id, f"{date_str}T00:00:00", f"{date_str}T23:59:59")
            except HTTPException as exc:
                if exc.status_code != 400:
                    logger.warning("recompute_meeting_call_activity: pull failed org=%s date=%s: %s", org_id, date_str, exc.detail)
                continue
            except Exception as exc:
                logger.warning("recompute_meeting_call_activity: pull failed org=%s date=%s: %s", org_id, date_str, exc)
                continue

            calls = result["calls"]
            for school_id in schools_with_meetings:
                calls_for_school = [
                    c for c in calls
                    if school_id not in (c.get("excluded_school_ids") or [])
                    and (c.get("school_id") == school_id or school_id in (c.get("linked_school_ids") or []))
                ]
                try:
                    db = get_admin_client()
                    affected = _recompute_meeting_call_activity(db, org_id, school_id, date_str, calls_for_school)
                except Exception as exc:
                    logger.warning("recompute_meeting_call_activity: recompute failed org=%s school=%s date=%s: %s", org_id, school_id, date_str, exc)
                    continue
                for mid in affected:
                    _maybe_auto_complete_meeting(db, org_id, mid)
                processed += 1

    return {"ok": True, "processed_school_days": processed}


@router.get("/{school_id}/meetings")
def list_meetings(school_id: str, user: Annotated[dict, Depends(get_current_user)], academic_year: str | None = None):
    meetings = []
    for attempt in range(2):
        try:
            db = get_admin_client()
            q = db.table("meetings").select("*").eq("school_id", school_id)
            if academic_year:
                q = q.eq("academic_year", academic_year)
            res = q.order("created_at", desc=True).execute()
            meetings = res.data or []
            break
        except Exception as exc:
            if attempt == 0:
                logger.warning("list_meetings attempt 1 failed: %s — resetting client and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("list_meetings failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    # Collect all referenced advisor IDs (new array field + legacy single field)
    all_ids: set[str] = set()
    for m in meetings:
        for uid in (m.get("advisor_ids") or []):
            all_ids.add(uid)
        if m.get("advisor_id"):
            all_ids.add(m["advisor_id"])

    if all_ids:
        try:
            db = get_admin_client()
            profiles = db.table("profiles").select("id, full_name, email").in_("id", list(all_ids)).execute().data or []
            profiles_map = {p["id"]: p for p in profiles}
            for m in meetings:
                ids = m.get("advisor_ids") or []
                if not ids and m.get("advisor_id"):  # backward compat: single advisor
                    ids = [m["advisor_id"]]
                m["advisor_profiles"] = [profiles_map[uid] for uid in ids if uid in profiles_map]
        except Exception as exc:
            logger.warning("list_meetings profile enrichment failed (non-fatal): %s", exc)
            for m in meetings:
                m["advisor_profiles"] = []
    else:
        for m in meetings:
            m["advisor_profiles"] = []

    return meetings


@router.post("/{school_id}/meetings")
def create_meeting(school_id: str, body: MeetingIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    data = {
        "school_id": school_id,
        "created_by": user["id"],
        "status": body.status or "scheduled",
        "reminder_enabled": body.reminder_enabled if body.reminder_enabled is not None else False,
        "participants": body.participants if body.participants is not None else [],
        "academic_year": body.academic_year or DEFAULT_ACADEMIC_YEAR,
    }
    if body.meeting_date: data["meeting_date"] = body.meeting_date
    if body.start_time: data["start_time"] = body.start_time
    if body.end_time: data["end_time"] = body.end_time
    # Defaults to "remote" ("מרחוק") when not explicitly chosen — server-side safety net so
    # every meeting-creation path lands on the same default, not just whichever ones the
    # frontend happens to hardcode it for.
    data["meeting_type"] = body.meeting_type or "remote"
    if body.meeting_service_type is not None: data["meeting_service_type"] = body.meeting_service_type
    if body.actual_duration: data["actual_duration"] = body.actual_duration
    if body.notes: data["notes"] = body.notes
    if body.primary_contact_key is not None: data["primary_contact_key"] = body.primary_contact_key
    if body.stage_scope is not None: data["stage_scope"] = body.stage_scope
    # advisor_ids takes precedence; fall back to legacy advisor_id
    if body.advisor_ids is not None:
        data["advisor_ids"] = body.advisor_ids
    elif body.advisor_id:
        data["advisor_ids"] = [body.advisor_id]
    else:
        data["advisor_ids"] = []
    try:
        res = db.table("meetings").insert(data).execute()
        meeting = res.data[0]
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאת DB: {str(e)}")

    try:
        with graph_client.calendar_sync_lock(db, meeting["id"]) as acquired:
            if acquired:
                subject = _build_meeting_subject(db, school_id, meeting.get("participants"), meeting.get("primary_contact_key"))
                sync_map = graph_client.sync_meeting_create(db, user["org_id"], meeting, subject)
                if sync_map:
                    graph_client.persist_calendar_sync(db, meeting["id"], sync_map)
                    meeting["calendar_sync"] = sync_map
    except Exception as exc:
        logger.warning("calendar sync failed for new meeting %s (non-fatal): %s", meeting.get("id"), exc)

    return meeting


_DIRECT_COORDINATION_SERVICE_TYPES = {"gefen": "גפן", "current": "שוטף", "district": "מחוז"}
_DIRECT_COORDINATION_DURATIONS = set(range(30, 181, 15))


@router.post("/{school_id}/meetings/direct-coordination")
def send_direct_coordination_request(
    school_id: str, body: DirectCoordinationIn, user: Annotated[dict, Depends(get_current_user)],
):
    """Sends a 'תיאום ישיר' email to the school's meeting_coordinator with a token-gated
    booking link covering one or more requested date ranges (one meeting request each),
    each with its own service type / advisors' availability / participants chosen up front
    by the manager. Mirrors the AI-agent bulk booking-link flow's building blocks
    (booking_token_logic.py / booking_logic.py / meeting_booking_router.py) but always mints
    a fresh token (no reuse) since this is a one-off targeted action, not a recurring batch."""
    _require_manager(user)
    if not body.advisor_ids:
        raise HTTPException(status_code=400, detail="יש לבחור לפחות יועץ אחד")
    if not body.ranges:
        raise HTTPException(status_code=400, detail="יש להוסיף לפחות טווח תאריכים אחד")

    db = get_admin_client()

    school_res = db.table("schools").select("*").eq("id", school_id).eq("org_id", user["org_id"]).execute()
    if not school_res.data:
        raise HTTPException(status_code=404, detail="בית הספר לא נמצא")
    school = school_res.data[0]

    coordinator = _resolve_meeting_coordinator(school)
    if not coordinator or not coordinator.get("email"):
        raise HTTPException(status_code=400, detail="יש להגדיר אחראי/ת לתיאום פגישות עם כתובת מייל בפרטי בית הספר לפני שליחה")

    import task_logic  # local import — task_logic imports from this module at module level
    try:
        opted_out_map = task_logic.opted_out_recipients(
            db, DEFAULT_ACADEMIC_YEAR, {school_id: coordinator.get("email")},
        )
    except Exception as exc:
        logger.warning("send_direct_coordination_request: opt-out lookup failed (non-fatal): %s", exc)
        opted_out_map = {}
    if school_id in opted_out_map:
        raise HTTPException(
            status_code=400,
            detail="לא ניתן לשלוח — בית הספר ביקש הסרה מרשימת התפוצה, עד שסטטוס הלקוח שלו יהפוך ל'פעיל'",
        )

    advisor_rows = (
        db.table("profiles").select("id, full_name, email")
        .eq("org_id", user["org_id"]).in_("id", body.advisor_ids).execute().data or []
    )
    if len(advisor_rows) != len(set(body.advisor_ids)):
        raise HTTPException(status_code=400, detail="אחד או יותר מהיועצים שנבחרו אינם תקינים")
    advisor_names_map = {a["id"]: (a.get("full_name") or a.get("email") or "") for a in advisor_rows}
    advisor_names = [advisor_names_map[aid] for aid in body.advisor_ids]

    type_counts: dict[str, int] = {}
    for r in body.ranges:
        if r.start_date > r.end_date:
            raise HTTPException(status_code=400, detail="טווח תאריכים לא תקין: תאריך ההתחלה מאוחר מתאריך הסיום")
        if r.meeting_service_type not in _DIRECT_COORDINATION_SERVICE_TYPES:
            raise HTTPException(status_code=400, detail="יש לבחור סוג פגישה (גפן/שוטף/מחוז) לכל טווח")
        if r.duration_minutes not in _DIRECT_COORDINATION_DURATIONS:
            raise HTTPException(status_code=400, detail="משך פגישה לא תקין")
        if not r.participants:
            raise HTTPException(status_code=400, detail="יש לבחור לפחות משתתף אחד לכל טווח/פגישה")
        type_counts[r.meeting_service_type] = type_counts.get(r.meeting_service_type, 0) + 1

    type_seen: dict[str, int] = {}
    ranges_data = []
    for i, r in enumerate(body.ranges):
        type_seen[r.meeting_service_type] = type_seen.get(r.meeting_service_type, 0) + 1
        base_label = f"פגישת {_DIRECT_COORDINATION_SERVICE_TYPES[r.meeting_service_type]}"
        label = base_label if type_counts[r.meeting_service_type] == 1 else f"{base_label} ({type_seen[r.meeting_service_type]})"
        ranges_data.append({
            "key": f"r{i}-{secrets.token_urlsafe(6)}",
            "start_date": r.start_date,
            "end_date": r.end_date,
            "service_type": r.meeting_service_type,
            "duration_minutes": r.duration_minutes,
            "label": label,
            "participants": [p.model_dump() for p in r.participants],
        })

    import booking_logic
    import booking_token_logic

    token_row = booking_token_logic.create_direct_booking_token(db, user["org_id"], school_id, body.advisor_ids, ranges_data)
    booking_url = f"{os.getenv('APP_URL', '')}/book/{token_row['token']}"
    opt_out_link = None
    if coordinator.get("email"):
        year_rows = (
            db.table("school_year_admin_data").select("client_status")
            .eq("academic_year", DEFAULT_ACADEMIC_YEAR).eq("school_id", school_id).execute().data or []
        )
        client_status = year_rows[0].get("client_status") if year_rows else None
        if client_status != "active":
            email_lower = coordinator["email"].strip().lower()
            opt_out_link = f"{os.getenv('APP_URL', '')}/tasks/opt-out?email={email_lower}&token={task_logic.make_optout_token(email_lower)}"
    html = booking_logic.build_direct_coordination_email_html(
        coordinator["name"], school["name"], advisor_names, ranges_data, booking_url, opt_out_link,
    )
    subject = f"בקשה לתיאום פגישה - {school['name']}"
    try:
        booking_logic.send_booking_request_email(user["org_id"], body.advisor_ids[0], coordinator["email"], subject, html)
    except Exception as exc:
        logger.error("send_direct_coordination_request: email send failed for school %s: %s", school_id, exc, exc_info=True)
        raise HTTPException(status_code=502, detail="שליחת המייל נכשלה, נסה שוב")

    return {"ok": True, "booking_url": booking_url}


@router.put("/{school_id}/meetings/{meeting_id}")
def update_meeting(school_id: str, meeting_id: str, body: MeetingIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    data = {
        "status": body.status or "scheduled",
        "reminder_enabled": body.reminder_enabled if body.reminder_enabled is not None else False,
        "participants": body.participants if body.participants is not None else [],
    }
    if body.meeting_date: data["meeting_date"] = body.meeting_date
    if body.start_time: data["start_time"] = body.start_time
    if body.end_time: data["end_time"] = body.end_time
    if body.meeting_type: data["meeting_type"] = body.meeting_type
    if body.meeting_service_type is not None: data["meeting_service_type"] = body.meeting_service_type
    if body.actual_duration: data["actual_duration"] = body.actual_duration
    if body.notes: data["notes"] = body.notes
    if body.academic_year: data["academic_year"] = body.academic_year
    if body.primary_contact_key is not None: data["primary_contact_key"] = body.primary_contact_key
    if body.stage_scope is not None: data["stage_scope"] = body.stage_scope
    # advisor_ids takes precedence; fall back to legacy advisor_id
    if body.advisor_ids is not None:
        data["advisor_ids"] = body.advisor_ids
    elif body.advisor_id:
        data["advisor_ids"] = [body.advisor_id]
    else:
        data["advisor_ids"] = []

    try:
        res = db.table("meetings").update(data).eq("id", meeting_id).eq("school_id", school_id).execute()
        meeting = res.data[0] if res.data else {}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"שגיאת DB: {str(e)}")

    try:
        with graph_client.calendar_sync_lock(db, meeting_id) as acquired:
            if acquired:
                existing = db.table("meetings").select("calendar_sync").eq("id", meeting_id).execute()
                previous_sync = (existing.data[0].get("calendar_sync") or {}) if existing.data else {}
                if meeting.get("status") in ("cancelled", "postponed"):
                    graph_client.sync_meeting_cancel(db, user["org_id"], previous_sync)
                    sync_map = {}
                else:
                    subject = _build_meeting_subject(db, school_id, meeting.get("participants"), meeting.get("primary_contact_key"))
                    sync_map = graph_client.sync_meeting_update(db, user["org_id"], {**meeting, "id": meeting_id}, previous_sync, subject)
                graph_client.persist_calendar_sync(db, meeting_id, sync_map)
                meeting["calendar_sync"] = sync_map
    except Exception as exc:
        logger.warning("calendar sync failed for updated meeting %s (non-fatal): %s", meeting_id, exc)

    return meeting


def _apply_meeting_patch(db, org_id: str, school_id: str, meeting_id: str, patch_dict: dict) -> dict:
    """Shared body of the lightweight meeting PATCH — used by the `patch_meeting` route AND
    internal callers (auto-complete-from-activity job/checks) that have no Request/user to go
    through HTTP with. Includes the same Outlook calendar-sync side effect as the route always
    has had — do not strip that out when calling this from a new internal context."""
    if not patch_dict:
        return {"ok": True}
    try:
        db.table("meetings").update(patch_dict).eq("id", meeting_id).eq("school_id", school_id).execute()
    except Exception:
        raise HTTPException(status_code=500, detail="שגיאה בעדכון פגישה")

    try:
        with graph_client.calendar_sync_lock(db, meeting_id) as acquired:
            if acquired:
                row = db.table("meetings").select("*").eq("id", meeting_id).execute()
                meeting = row.data[0] if row.data else None
                if meeting:
                    previous_sync = meeting.get("calendar_sync") or {}
                    if meeting.get("status") in ("cancelled", "postponed"):
                        graph_client.sync_meeting_cancel(db, org_id, previous_sync)
                        sync_map = {}
                    else:
                        subject = _build_meeting_subject(db, school_id, meeting.get("participants"), meeting.get("primary_contact_key"))
                        sync_map = graph_client.sync_meeting_update(db, org_id, meeting, previous_sync, subject)
                    graph_client.persist_calendar_sync(db, meeting_id, sync_map)
    except Exception as exc:
        logger.warning("calendar sync failed for patched meeting %s (non-fatal): %s", meeting_id, exc)

    return {"ok": True}


@router.patch("/{school_id}/meetings/{meeting_id}")
def patch_meeting(school_id: str, meeting_id: str, body: MeetingStatusPatchIn, user: Annotated[dict, Depends(get_current_user)]):
    """Partial update — only updates the fields provided. Does not touch advisor_ids, participants, etc."""
    db = get_admin_client()
    data = {}
    if body.status is not None: data["status"] = body.status
    if body.notes is not None: data["notes"] = body.notes
    if body.start_time is not None: data["start_time"] = body.start_time
    if body.end_time is not None: data["end_time"] = body.end_time
    return _apply_meeting_patch(db, user["org_id"], school_id, meeting_id, data)


@router.delete("/{school_id}/meetings/{meeting_id}")
def delete_meeting(school_id: str, meeting_id: str, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    if not _check_permission(db, user, "can_delete_own_meetings"):
        raise HTTPException(status_code=403, detail="אין הרשאה למחוק פגישות")
    try:
        with graph_client.calendar_sync_lock(db, meeting_id) as acquired:
            if acquired:
                row = db.table("meetings").select("calendar_sync").eq("id", meeting_id).execute()
                previous_sync = (row.data[0].get("calendar_sync") or {}) if row.data else {}
                graph_client.sync_meeting_cancel(db, user["org_id"], previous_sync)
    except Exception as exc:
        logger.warning("calendar cancel-sync failed before deleting meeting %s (non-fatal): %s", meeting_id, exc)
    db.table("meetings").delete().eq("id", meeting_id).eq("school_id", school_id).execute()
    return {"ok": True}


class MeetingReassignSchoolIn(BaseModel):
    new_school_id: str


@router.patch("/meetings/{meeting_id}/reassign-school")
def reassign_meeting_school(meeting_id: str, body: MeetingReassignSchoolIn, user: Annotated[dict, Depends(get_current_user)]):
    """Move an existing meeting to a different school. Used by the admin 'פגישות' tab school-picker cell.

    Clears `participants`/`primary_contact_key` on reassignment — they refer to the
    *old* school's staff, and re-syncing without clearing them would build the Outlook
    subject from a mismatched name (old school's contact) + phone (looked up fresh from
    the new school), which is worse than just showing the school name until participants
    are re-selected for the new school.
    """
    _require_manager(user)
    db = None
    meeting = None
    for attempt in range(2):
        try:
            db = get_admin_client()
            sch = db.table("schools").select("id").eq("id", body.new_school_id).eq("org_id", user["org_id"]).execute()
            if not sch.data:
                raise HTTPException(status_code=404, detail="בית ספר לא נמצא")
            res = db.table("meetings").update({
                "school_id": body.new_school_id,
                "participants": [],
                "primary_contact_key": None,
            }).eq("id", meeting_id).execute()
            meeting = res.data[0] if res.data else {}
            break
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("reassign_meeting_school attempt 1 failed: %s — resetting and retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("reassign_meeting_school failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב בעוד מספר שניות")

    try:
        with graph_client.calendar_sync_lock(db, meeting_id) as acquired:
            if acquired:
                existing = db.table("meetings").select("calendar_sync").eq("id", meeting_id).execute()
                previous_sync = (existing.data[0].get("calendar_sync") or {}) if existing.data else {}
                subject = _build_meeting_subject(db, body.new_school_id, [], None)
                sync_map = graph_client.sync_meeting_update(db, user["org_id"], {**meeting, "id": meeting_id}, previous_sync, subject)
                graph_client.persist_calendar_sync(db, meeting_id, sync_map)
    except Exception as exc:
        logger.warning("calendar sync failed for reassigned meeting %s (non-fatal): %s", meeting_id, exc)

    return {"ok": True}


class MentionIn(BaseModel):
    mentioned_user_ids: list[str]
    note_preview: str | None = None


@router.post("/{school_id}/meetings/{meeting_id}/mentions")
def create_mentions(school_id: str, meeting_id: str, body: MentionIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    recipients = [uid for uid in body.mentioned_user_ids if uid != user["id"]]
    if not recipients:
        return {"ok": True, "sent": 0}
    try:
        school_row = db.table("schools").select("name").eq("id", school_id).execute()
        school_name = school_row.data[0]["name"] if school_row.data else "בית ספר"
        rows = [{
            "recipient_id": uid,
            "type": "mention",
            "ref_id": meeting_id,
            "school_id": school_id,
            "data": {
                "title": f'{user.get("full_name", "משתמש")} תייג אותך בהערה ב-{school_name}',
                "school_name": school_name,
                "sender_name": user.get("full_name", ""),
                "note_preview": body.note_preview,
                "deeplink": f"/school/{school_id}?tab=meetings&meeting={meeting_id}",
            }
        } for uid in recipients]
        _create_notifications(db, rows, pref_key="notify_mention")
    except Exception as exc:
        logger.warning("create_mentions failed (non-fatal): %s", exc)
    return {"ok": True, "sent": len(recipients)}


# ---------------------------------------------------------------------------
# Permissions — role defaults + per-user overrides
# ---------------------------------------------------------------------------

# All supported permission keys with their role defaults
PERMISSION_DEFAULTS: dict[str, dict[str, bool]] = {
    "can_approve_update_requests":  {"manager": False, "advisor": False},
    "can_invite_users":             {"manager": True,  "advisor": False},
    "can_delete_users":             {"manager": False, "advisor": False},
    "can_change_user_role":         {"manager": False, "advisor": False},
    "can_delete_schools":           {"manager": False, "advisor": False},
    "can_add_school":               {"manager": True,  "advisor": False},
    "can_edit_school_directly":     {"manager": True,  "advisor": False},
    "can_request_school_update":    {"manager": True,  "advisor": True},
    "can_delete_own_meetings":      {"manager": True,  "advisor": True},
    "can_manage_user_permissions":  {"manager": False, "advisor": False},
    "can_view_billing":             {"manager": False, "advisor": False},
    "can_manage_billing":           {"manager": False, "advisor": False},
    "can_edit_meeting_automations": {"manager": True,  "advisor": False},
    "can_remove_call_from_school":  {"manager": True,  "advisor": False},
}

PERMISSION_LABELS: dict[str, str] = {
    "can_approve_update_requests":  "לאשר בקשות עריכת פרטים",
    "can_invite_users":             "להוסיף משתמש חדש",
    "can_delete_users":             "למחוק משתמש קיים",
    "can_change_user_role":         "לשנות תפקיד של משתמש",
    "can_delete_schools":           "למחוק בית ספר קיים",
    "can_add_school":               "להוסיף בית ספר חדש",
    "can_edit_school_directly":     "לערוך בית ספר קיים ישירות (מבלי להגיש בקשה לעריכת פרטים)",
    "can_request_school_update":    "להגיש בקשה לעריכת פרטי בית ספר",
    "can_delete_own_meetings":      "למחוק נתוני פגישה של בית ספר",
    "can_manage_user_permissions":  "לערוך הרשאות של יועצים",
    "can_view_billing":             "לצפות באזור 'חיובים' של הארגון",
    "can_manage_billing":           "לנהל את אזור 'חיובים' (לרבות אמצעי תשלום)",
    "can_edit_meeting_automations": "לערוך אוטומציות של פגישות",
    "can_remove_call_from_school":  "להסיר שיחה מטאב 'שיחות' בכרטיס בית ספר",
}


class PermissionSettingIn(BaseModel):
    role: str
    permission: str
    allowed: bool


class UserPermissionOverrideIn(BaseModel):
    permission: str
    allowed: bool | None  # None = remove override (revert to role default)


def _get_owner_id(db, user: dict) -> str:
    """Return the owner's profile ID for the current tenant.
    org_id on profiles is the organizations table ID, NOT the owner's profile ID.
    To find the owner: query profiles WHERE role='owner' AND org_id=user['org_id'].
    """
    if user["role"] == "owner":
        return user["id"]
    if user.get("org_id"):
        rows = db.table("profiles").select("id").eq("role", "owner").eq("org_id", user["org_id"]).execute().data or []
        if rows:
            return rows[0]["id"]
    # final fallback (single-tenant dev only)
    rows = db.table("profiles").select("id").eq("role", "owner").execute().data or []
    if not rows:
        raise HTTPException(status_code=404, detail="לא נמצא בעלים בארגון")
    return rows[0]["id"]


def _check_permission(db, user: dict, permission: str) -> bool:
    """Resolve effective permission for a user: override → role default → system default.
    Owner always returns True. Errors fall through to system default (fail-open for non-destructive,
    but callers decide whether to raise 403 or silently skip).
    """
    if user["role"] == "owner":
        return True

    # Layer 1: per-user override
    try:
        ov = db.table("user_permission_overrides").select("allowed") \
            .eq("user_id", user["id"]).eq("permission", permission).execute()
        if ov.data:
            return ov.data[0]["allowed"]
    except Exception as exc:
        logger.warning("_check_permission override lookup failed (non-fatal): %s", exc)

    # Layer 2: org role default
    try:
        owner_id = _get_owner_id(db, user)
        rs = db.table("permission_settings").select("allowed") \
            .eq("owner_id", owner_id).eq("role", user["role"]).eq("permission", permission).execute()
        if rs.data:
            return rs.data[0]["allowed"]
    except Exception as exc:
        logger.warning("_check_permission role default lookup failed (non-fatal): %s", exc)

    # Layer 3: system default
    return PERMISSION_DEFAULTS.get(permission, {}).get(user["role"], False)


@router.get("/permissions/defaults")
def get_permission_defaults(user: Annotated[dict, Depends(get_current_user)]):
    """Return role defaults for this org merged with system defaults."""
    for attempt in range(2):
        try:
            db = get_admin_client()
            owner_id = _get_owner_id(db, user)
            saved = db.table("permission_settings").select("role,permission,allowed").eq("owner_id", owner_id).execute().data or []
            saved_map = {(r["role"], r["permission"]): r["allowed"] for r in saved}

            result = {}
            for perm, role_defaults in PERMISSION_DEFAULTS.items():
                result[perm] = {
                    "label": PERMISSION_LABELS[perm],
                    "manager": saved_map.get(("manager", perm), role_defaults["manager"]),
                    "advisor": saved_map.get(("advisor", perm), role_defaults["advisor"]),
                }
            return result
        except HTTPException:
            raise
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_permission_defaults attempt 1 failed: %s — retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_permission_defaults failed after 2 attempts: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


@router.put("/permissions/defaults")
def set_permission_default(body: PermissionSettingIn, user: Annotated[dict, Depends(get_current_user)]):
    """Owner saves a role-level default. Manager with can_manage_user_permissions can edit advisor defaults only."""
    if user["role"] == "manager":
        db_pre = get_admin_client()
        if not _check_permission(db_pre, user, "can_manage_user_permissions"):
            raise HTTPException(status_code=403, detail="אין הרשאה לעריכת הרשאות")
        if body.role != "advisor":
            raise HTTPException(status_code=403, detail="מנהל יכול לערוך הרשאות של יועצים בלבד")
    elif user["role"] != "owner":
        raise HTTPException(status_code=403, detail="פעולה זו מיועדת לבעלים בלבד")

    if body.permission not in PERMISSION_DEFAULTS:
        raise HTTPException(status_code=400, detail="הרשאה לא מוכרת")
    if body.role not in ("manager", "advisor"):
        raise HTTPException(status_code=400, detail="תפקיד לא תקין")

    for attempt in range(2):
        try:
            db = get_admin_client()
            owner_id = _get_owner_id(db, user)
            db.table("permission_settings").upsert(
                {"owner_id": owner_id, "role": body.role, "permission": body.permission, "allowed": body.allowed},
                on_conflict="owner_id,role,permission",
            ).execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("set_permission_default attempt 1 failed: %s — retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("set_permission_default failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


def _require_can_edit_automations(db, user: dict) -> None:
    if user["role"] == "owner":
        return
    if user["role"] == "manager" and _check_permission(db, user, "can_edit_meeting_automations"):
        return
    raise HTTPException(status_code=403, detail="אין הרשאה לעריכת אוטומציות")


@router.get("/meetings/automations")
def get_meeting_automations(user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    _require_can_edit_automations(db, user)
    org = db.table("organizations").select(
        "meeting_reminders_enabled, secretary_upload_request_enabled, auto_complete_meetings_from_activity_enabled"
    ).eq("id", user["org_id"]).single().execute().data or {}
    return {
        "meeting_reminders_enabled": org.get("meeting_reminders_enabled", True),
        "secretary_upload_request_enabled": org.get("secretary_upload_request_enabled", True),
        # Defaults to False, unlike the other automations above — higher-impact (auto-changes
        # meeting status), must stay opt-in until an owner/manager explicitly turns it on.
        "auto_complete_meetings_from_activity_enabled": org.get("auto_complete_meetings_from_activity_enabled", False),
    }


class MeetingAutomationsIn(BaseModel):
    meeting_reminders_enabled: bool | None = None
    secretary_upload_request_enabled: bool | None = None
    auto_complete_meetings_from_activity_enabled: bool | None = None


@router.put("/meetings/automations")
def set_meeting_automations(body: MeetingAutomationsIn, user: Annotated[dict, Depends(get_current_user)]):
    db = get_admin_client()
    _require_can_edit_automations(db, user)
    patch = {k: v for k, v in body.model_dump().items() if v is not None}
    if not patch:
        return {"ok": True}
    db.table("organizations").update(patch).eq("id", user["org_id"]).execute()
    return {"ok": True}


@router.get("/permissions/overrides/counts")
def get_override_counts(user: Annotated[dict, Depends(get_current_user)]):
    """Return override count per user_id for all users in the org (manager+ only)."""
    _require_manager(user)
    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = db.table("user_permission_overrides").select("user_id").execute().data or []
            counts: dict[str, int] = {}
            for r in rows:
                uid = r["user_id"]
                counts[uid] = counts.get(uid, 0) + 1
            return counts
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_override_counts attempt 1 failed: %s — retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_override_counts failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


@router.get("/permissions/overrides/{user_id}")
def get_user_overrides(user_id: str, user: Annotated[dict, Depends(get_current_user)]):
    """Return per-user overrides for a specific user (manager+ or self)."""
    if user["role"] not in ("owner", "manager") and user["id"] != user_id:
        raise HTTPException(status_code=403, detail="אין הרשאה")

    for attempt in range(2):
        try:
            db = get_admin_client()
            rows = db.table("user_permission_overrides").select("permission,allowed").eq("user_id", user_id).execute().data or []
            return {r["permission"]: r["allowed"] for r in rows}
        except Exception as exc:
            if attempt == 0:
                logger.warning("get_user_overrides attempt 1 failed: %s — retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("get_user_overrides failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


@router.put("/permissions/overrides/{user_id}")
def set_user_override(user_id: str, body: UserPermissionOverrideIn, user: Annotated[dict, Depends(get_current_user)]):
    """Set or remove a per-user permission override (owner, or manager with can_manage_user_permissions)."""
    _require_manager(user)
    if user["role"] != "owner":
        db_pre = get_admin_client()
        if not _check_permission(db_pre, user, "can_manage_user_permissions"):
            raise HTTPException(status_code=403, detail="אין הרשאה לעריכת הרשאות משתמשים")
        # managers can only edit advisor overrides, not other managers
        target_rows = db_pre.table("profiles").select("role").eq("id", user_id).execute().data or []
        if not target_rows or target_rows[0]["role"] != "advisor":
            raise HTTPException(status_code=403, detail="מנהל יכול לערוך הרשאות של יועצים בלבד")
    if body.permission not in PERMISSION_DEFAULTS:
        raise HTTPException(status_code=400, detail="הרשאה לא מוכרת")

    for attempt in range(2):
        try:
            db = get_admin_client()
            owner_id = _get_owner_id(db, user)
            if body.allowed is None:
                db.table("user_permission_overrides").delete().eq("user_id", user_id).eq("permission", body.permission).execute()
            else:
                db.table("user_permission_overrides").upsert(
                    {"user_id": user_id, "owner_id": owner_id, "permission": body.permission, "allowed": body.allowed},
                    on_conflict="user_id,permission",
                ).execute()
            return {"ok": True}
        except Exception as exc:
            if attempt == 0:
                logger.warning("set_user_override attempt 1 failed: %s — retrying", exc)
                reset_admin_client()
                time.sleep(0.3)
            else:
                logger.error("set_user_override failed: %s", exc, exc_info=True)
                raise HTTPException(status_code=503, detail="שגיאה זמנית בשרת — נסה שוב")


# ---------------------------------------------------------------------------
# Schools list export — PDF
# ---------------------------------------------------------------------------

class SchoolsPdfExportIn(BaseModel):
    title: str = "רשימת בתי ספר"
    headers: list[str]
    rows: list[list[str]]


@router.post("/export-pdf")
def export_schools_pdf(
    body: SchoolsPdfExportIn,
    user: Annotated[dict, Depends(get_current_user)],
):
    pdf_bytes = _build_schools_pdf(body.title, body.headers, body.rows)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="schools.pdf"'},
    )
