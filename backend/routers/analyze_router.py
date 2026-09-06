import logging
import re
import shutil
import tempfile
import traceback
import uuid
from pathlib import Path
from typing import Annotated

from pydantic import BaseModel

logger = logging.getLogger(__name__)

import pandas as pd
from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import Response

from academic_years import DEFAULT_ACADEMIC_YEAR
from auth import get_current_user
from supabase_client import get_admin_client
from logic.excel_exporter import export
from logic.pdf_exporter import export_pdf
from logic.file_identifier import identify_file
from plan_roster import extract_plan_roster
from logic.gefen_processor import load_gefen, normalize_amount
from logic.kesafim_processor import load_kesafim
from logic.payscool_processor import canonical_payscool_invoice, load_payscool
from logic.schoolcash_processor import load_schoolcash
from logic.reconciler import BEINAYIM_ONLY, TIKKON_ONLY, reconcile
from logic.tikhnun_processor import load_tikhnun, cross_reference_doch, build_tikhnun_result
from logic.tikhnun_exporter import export_tikhnun_excel, export_tikhnun_pdf

router = APIRouter()

_INTERNAL_KEYS = {"_school_ctx"}


def _update_run(run_id: str, run_data: dict) -> None:
    """Upsert run state to Supabase run_states. Non-fatal on failure."""
    try:
        db = get_admin_client()
        db.table("run_states").upsert(
            {"id": run_id, "status": run_data.get("status", "processing"), "result": run_data},
            on_conflict="id",
        ).execute()
    except Exception as exc:
        logger.warning("run_states upsert failed for %s: %s", run_id, exc)


def _get_run(run_id: str) -> dict | None:
    """Fetch fresh run state from Supabase run_states. Returns None if not found."""
    try:
        db = get_admin_client()
        row = db.table("run_states").select("result, status").eq("id", run_id).single().execute()
        if row.data:
            result = row.data.get("result") or {}
            result["status"] = row.data["status"]
            return result
    except Exception:
        pass
    return None


def _strip_for_response(run: dict) -> dict:
    """Remove internal keys before sending to frontend."""
    result = {k: v for k, v in run.items() if k not in _INTERNAL_KEYS}
    for key in ("tikhnun", "tikhnun_tikkon", "tikhnun_beinayim"):
        if result.get(key) and isinstance(result[key], dict) and "_zihuy_ctx" in result[key]:
            result[key] = {k: v for k, v in result[key].items() if k != "_zihuy_ctx"}
    return result

# Unified JSON column names — same for both finance types and gefen side
_DISPLAY_COLS = ["קוד דיווח", "שם ספק", "מספר אסמכתה", "תאריך", "סכום", "תיאור"]

# Column maps: list of (source_col, display_col, transform_fn | None)
_PAYSCOOL_COL_MAP = [
    ("קוד דיווח",      "קוד דיווח",   None),
    ("שם ספק",         "שם ספק",       None),
    ("מספר חשבונית",   "מספר אסמכתה", normalize_amount),
    ("תאריך חשבונית",  "תאריך",        _norm_date := None),  # assigned below
    ('סה"כ לסעיף',     "סכום",         None),
    ("תיאור",           "תיאור",        None),
]

_SCHOOLCASH_COL_MAP = [
    ("קוד דיווח",              "קוד דיווח",   None),
    ("שם ספק",                 "שם ספק",       None),
    ("מספר חשבונית",           "מספר אסמכתה", normalize_amount),
    ("תאריך חשבונית",          "תאריך",        None),  # patched below
    ("סכום",                   "סכום",         None),  # patched below
    ("תאור שורה בחשבונית",    "תיאור",        None),
]

_KESAFIM_COL_MAP = [
    ("קוד דיווח",      "קוד דיווח",   None),
    ("שם ספק",         "שם ספק",       None),
    ("מספר חשבונית",   "מספר אסמכתה", None),
    ("תאריך חשבונית",  "תאריך",        None),  # patched below
    ("סכום",            "סכום",         None),
    ("תיאור",           "תיאור",        None),
]

_GEFEN_COL_MAP = [
    ("report_code",    "קוד דיווח",   None),
    ("קוד ושם ספק",    "שם ספק",       None),
    ("מספר חשבונית",   "מספר אסמכתה", normalize_amount),
    ("תאריך חשבונית",  "תאריך",        None),  # patched below
    ("סכום פריט",      "סכום",         normalize_amount),
    ("מהות ההוצאה",    "תיאור",        None),
]

# Same as _GEFEN_COL_MAP but last column is "סיבת הדחייה" (extracted from col M)
_GEFEN_REJECTED_COL_MAP = [
    ("report_code",    "קוד דיווח",     None),
    ("קוד ושם ספק",    "שם ספק",         None),
    ("מספר חשבונית",   "מספר אסמכתה",  normalize_amount),
    ("תאריך חשבונית",  "תאריך",          None),  # patched below
    ("סכום פריט",      "סכום",           None),   # patched below
    ("סיבת הדחייה",   "סיבת הדחייה",   None),
]

# Columns to strip before writing Excel (internal/computed)
_STRIP_COLS = {"ichud", "supplier_number", "amount", "report_code"}

# Known data columns in a "דיווח ביצוע" sheet (used for no-PDF completeness check)
_GEFEN_DATA_COLS = [
    "מספר חשבונית", "תאריך חשבונית", "קוד ושם ספק",
    "מהות ההוצאה", "מספר פריט בחשבונית", "כמות",
    "תיאור פריט", "סכום פריט",
]
# Status/PDF column names as they appear in the gefen sheet
_GEFEN_STATUS_COL = "סטטוס חשבונית"
_GEFEN_PDF_COL    = "האם קיים קובץ"

# Hebrew display names for kesafim2000 English column names
_KESAFIM_RENAME = {
    "report_code":    "קוד דיווח",
    "supplier":       "ספק",
    "supplier_name":  "שם ספק",
    "invoice_date":   "תאריך חשבונית",
    "invoice_number": "מספר חשבונית",
    "voucher":        "שובר",
    "item_number":    "מספר פריט",
    "item_name":      "שם פריט",
    "description":    "תיאור",
    "amount_raw":     "סכום",
    "total":          'סה"כ',
    "status":         "סטטוס",
}


# ---------------------------------------------------------------------------
# Value normalizers
# ---------------------------------------------------------------------------

def _normalize_date(val: str) -> str:
    """Normalize any date format to DD/MM/YYYY."""
    s = str(val).strip()
    if not s or s == "nan":
        return ""
    # DD/MM/YYYY — already correct
    if re.match(r"^\d{1,2}/\d{1,2}/\d{4}$", s):
        d, m, y = s.split("/")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    # DD-MM-YYYY
    if re.match(r"^\d{1,2}-\d{1,2}-\d{4}$", s):
        d, m, y = s.split("-")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    # DD.MM.YY or DD.MM.YYYY
    if re.match(r"^\d{1,2}\.\d{1,2}\.\d{2,4}$", s):
        parts = s.split(".")
        d, m, y = parts[0], parts[1], parts[2]
        if len(y) == 2:
            y = "20" + y
        return f"{int(d):02d}/{int(m):02d}/{y}"
    # YYYY-MM-DD ... (ISO or pandas Timestamp with time component)
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        y, m, d = s[:10].split("-")
        return f"{int(d):02d}/{int(m):02d}/{y}"
    return s


def _format_display_amount(val: str) -> str:
    """Format a numeric string with thousands comma separator for display (e.g. 2500 → 2,500)."""
    s = str(val).strip().replace(",", "")
    if not s or s == "nan":
        return ""
    try:
        f = float(s)
        if f == int(f):
            return f"{int(f):,}"
        return f"{f:,.2f}".rstrip("0").rstrip(".")
    except ValueError:
        return val


# Patch date normalizer and amount formatter into all maps
_PAYSCOOL_COL_MAP[3]    = ("תאריך חשבונית", "תאריך", _normalize_date)
_PAYSCOOL_COL_MAP[4]    = ('סה"כ לסעיף',    "סכום",   _format_display_amount)
_SCHOOLCASH_COL_MAP[3]  = ("תאריך חשבונית", "תאריך", _normalize_date)
_SCHOOLCASH_COL_MAP[4]  = ("סכום",          "סכום",   _format_display_amount)
_KESAFIM_COL_MAP[3]     = ("תאריך חשבונית", "תאריך", _normalize_date)
_GEFEN_COL_MAP[3]       = ("תאריך חשבונית", "תאריך", _normalize_date)
_GEFEN_COL_MAP[4]       = ("סכום פריט",     "סכום",   _format_display_amount)
_GEFEN_REJECTED_COL_MAP[3] = ("תאריך חשבונית", "תאריך", _normalize_date)
_GEFEN_REJECTED_COL_MAP[4] = ("סכום פריט",     "סכום",   _format_display_amount)


def _sum_display_amounts(rows: list) -> float:
    """Sum the 'סכום' field from a list of display-record dicts."""
    total = 0.0
    for row in (rows or []):
        try:
            total += float(str(row.get("סכום", "0")).replace(",", "").strip() or 0)
        except (ValueError, TypeError):
            pass
    return total


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/upload")
async def upload(
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    school_id: str | None = Form(None),
    gefen_account_id: str | None = Form(None),
    update_log_id: str | None = Form(None),
    academic_year: str | None = Form(None),
    user: dict = Depends(get_current_user),
):
    run_id = str(uuid.uuid4())
    run_dir = Path(tempfile.mkdtemp(prefix=f"gefen_{run_id}_"))

    saved: list[Path] = []
    for uf in files:
        dest = run_dir / uf.filename
        dest.write_bytes(await uf.read())
        saved.append(dest)

    _update_run(run_id, {"status": "processing"})
    background_tasks.add_task(_process, run_id, saved, run_dir, user["id"], school_id, gefen_account_id, update_log_id, academic_year)
    return {"run_id": run_id}


@router.post("/save-for-account")
async def save_for_account(
    run_id: str = Form(...),
    school_id: str = Form(...),
    gefen_account_id: str | None = Form(None),
    academic_year: str | None = Form(None),
    user: dict = Depends(get_current_user),
):
    """Save a completed run under a different gefen account (used when division mismatch is detected)."""
    run = _get_run(run_id)
    if not run or run.get("status") != "done":
        raise HTTPException(status_code=404, detail="הריצה לא נמצאה או טרם הושלמה")
    _save_check_log(run_id, user["id"], school_id, gefen_account_id, run_data=run, academic_year=academic_year)
    return {"ok": True}


@router.post("/add-file/{log_id}")
async def add_file_to_check(
    log_id: str,
    background_tasks: BackgroundTasks,
    files: list[UploadFile] = File(...),
    user: dict = Depends(get_current_user),
):
    """Re-run a check with original stored files + new uploaded file(s)."""
    db = get_admin_client()
    log_result = db.table("check_logs").select("*").eq("id", log_id).single().execute()
    log = log_result.data
    if not log:
        raise HTTPException(status_code=404, detail="הבדיקה לא נמצאה")

    stored_paths = (log.get("summary") or {}).get("stored_file_paths") or []
    if not stored_paths:
        raise HTTPException(status_code=400, detail="אין קבצים מקוריים שמורים לבדיקה זו")

    run_id = str(uuid.uuid4())
    run_dir = Path(tempfile.mkdtemp(prefix=f"gefen_{run_id}_"))
    all_paths: list[Path] = []

    for sp in stored_paths:
        # New format: {"path": "run_id/file_00.xlsx", "name": "original_name.xlsx"}
        # Legacy format: plain string path (ASCII filenames only from before the dict format)
        if isinstance(sp, dict):
            storage_key = sp["path"]
            fname = sp["name"]
        else:
            storage_key = sp
            fname = Path(sp).name
        dest = run_dir / fname
        try:
            dest.write_bytes(db.storage.from_("check-files").download(storage_key))
            all_paths.append(dest)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"שגיאה בהורדת הקובץ המקורי: {fname}")

    for uf in files:
        dest = run_dir / uf.filename
        dest.write_bytes(await uf.read())
        all_paths.append(dest)

    _update_run(run_id, {"status": "processing"})
    background_tasks.add_task(
        _process, run_id, all_paths, run_dir,
        user["id"], log["school_id"], log.get("gefen_account_id"), log_id, log.get("academic_year"),
    )
    return {"run_id": run_id}


@router.post("/meetings/{meeting_id}/run-check-from-uploads")
async def run_check_from_uploads(
    meeting_id: str,
    background_tasks: BackgroundTasks,
    user: dict = Depends(get_current_user),
):
    """Run the real reconciliation using files already uploaded via the
    meeting-upload portal (Phase 2) — no re-upload needed. Mirrors
    add_file_to_check's storage-download pattern exactly."""
    db = get_admin_client()
    meeting_res = db.table("meetings").select("id, school_id, academic_year").eq("id", meeting_id).execute()
    if not meeting_res.data:
        raise HTTPException(status_code=404, detail="הפגישה לא נמצאה")
    meeting = meeting_res.data[0]

    files_res = db.table("meeting_upload_files").select("storage_key, original_filename").eq("meeting_id", meeting_id).execute()
    stored_files = files_res.data or []
    if not stored_files:
        raise HTTPException(status_code=400, detail="לא נמצאו קבצים שהועלו עבור פגישה זו")

    run_id = str(uuid.uuid4())
    run_dir = Path(tempfile.mkdtemp(prefix=f"gefen_{run_id}_"))
    all_paths: list[Path] = []
    for f in stored_files:
        fname = f.get("original_filename") or Path(f["storage_key"]).name
        dest = run_dir / fname
        try:
            dest.write_bytes(db.storage.from_("check-files").download(f["storage_key"]))
            all_paths.append(dest)
        except Exception:
            raise HTTPException(status_code=500, detail=f"שגיאה בהורדת הקובץ: {fname}")

    _update_run(run_id, {"status": "processing"})
    background_tasks.add_task(
        _process, run_id, all_paths, run_dir,
        user["id"], meeting["school_id"], None, None, meeting.get("academic_year"),
    )
    return {"run_id": run_id}


class ComparePlansRequest(BaseModel):
    newer_log_id: str
    older_log_id: str


def _download_plan_rosters(db, log: dict) -> dict | None:
    """Download this check's stored files, identify + parse any planning
    file(s) among them, and return a merged {budget_norm: [plans]} roster.
    Returns None if no stored files / no tikhnun file could be found.
    Isolated from the /add-file + _process pipeline — read-only, no reconciliation.
    """
    stored_paths = (log.get("summary") or {}).get("stored_file_paths") or []
    if not stored_paths:
        return None

    run_dir = Path(tempfile.mkdtemp(prefix=f"compareplans_{log['id']}_"))
    try:
        roster: dict = {}
        found_tikhnun = False
        for sp in stored_paths:
            if isinstance(sp, dict):
                storage_key = sp["path"]
                fname = sp["name"]
            else:
                storage_key = sp
                fname = Path(sp).name
            dest = run_dir / fname
            try:
                dest.write_bytes(db.storage.from_("check-files").download(storage_key))
            except Exception as exc:
                logger.warning("compare-plans: failed to download %s: %s", fname, exc)
                continue
            if identify_file(str(dest)) != "tikhnun":
                continue
            found_tikhnun = True
            try:
                file_roster = extract_plan_roster(str(dest))
            except Exception as exc:
                logger.warning("compare-plans: failed to parse %s: %s", fname, exc)
                continue
            for budget_norm, plans in file_roster.items():
                roster.setdefault(budget_norm, []).extend(plans)

        return roster if found_tikhnun else None
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)


@router.post("/compare-plans")
def compare_plans(body: ComparePlansRequest, _user: dict = Depends(get_current_user)):
    """Compare the full plan roster (מענים) of two checks' planning files.
    Independent from "דיווח חסר" — includes fully-reported plans too."""
    db = get_admin_client()

    newer_log = db.table("check_logs").select("*").eq("id", body.newer_log_id).single().execute().data
    older_log = db.table("check_logs").select("*").eq("id", body.older_log_id).single().execute().data
    if not newer_log or not older_log:
        raise HTTPException(status_code=404, detail="אחת הבדיקות לא נמצאה")

    newer_roster = _download_plan_rosters(db, newer_log)
    older_roster = _download_plan_rosters(db, older_log)

    missing = {"newer": newer_roster is None, "older": older_roster is None}
    newer_roster = newer_roster or {}
    older_roster = older_roster or {}

    budget_names = sorted(set(newer_roster) | set(older_roster))
    budgets = []
    for name in budget_names:
        newer_plans = {p["key"]: p for p in newer_roster.get(name, [])}
        older_plans = {p["key"]: p for p in older_roster.get(name, [])}

        added = [p for k, p in newer_plans.items() if k not in older_plans]
        removed = [p for k, p in older_plans.items() if k not in newer_plans]
        updated = []
        for k, new_p in newer_plans.items():
            old_p = older_plans.get(k)
            if old_p is None:
                continue
            if round(new_p["tikhnun"], 2) != round(old_p["tikhnun"], 2):
                updated.append({
                    "key": k, "mispnum": new_p["mispnum"], "rcode": new_p["rcode"], "name": new_p["name"],
                    "oldAmount": old_p["tikhnun"], "newAmount": new_p["tikhnun"],
                    "diff": new_p["tikhnun"] - old_p["tikhnun"],
                })

        budgets.append({"name": name, "added": added, "removed": removed, "updated": updated})

    return {"missing": missing, "budgets": budgets}


@router.get("/result/{run_id}")
def get_result(run_id: str, _user: dict = Depends(get_current_user)):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    return _strip_for_response(run)


@router.get("/logs/{log_id}/source-file")
def download_log_source_file(log_id: str, name: str, _user: dict = Depends(get_current_user)):
    """Download one of the original files uploaded for a saved check — used by
    the filename links in the "קבצים שזוהו" tooltip in SchoolPage.jsx."""
    db = get_admin_client()
    log_result = db.table("check_logs").select("summary").eq("id", log_id).single().execute()
    log = log_result.data
    if not log:
        raise HTTPException(status_code=404, detail="הבדיקה לא נמצאה")

    stored_paths = (log.get("summary") or {}).get("stored_file_paths") or []
    match = None
    for sp in stored_paths:
        sp_name = sp.get("name") if isinstance(sp, dict) else Path(sp).name
        if sp_name == name:
            match = sp.get("path") if isinstance(sp, dict) else sp
            break
    if not match:
        raise HTTPException(status_code=404, detail="הקובץ לא נמצא")

    try:
        content = db.storage.from_("check-files").download(match)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"שגיאה בהורדת הקובץ: {exc}")

    # HTTP headers must be ASCII — a raw Hebrew filename in Content-Disposition
    # breaks the response entirely. RFC 5987's filename* handles non-ASCII names
    # correctly (with an ASCII fallback for older clients).
    import urllib.parse
    ext = Path(name).suffix or ".xlsx"
    ascii_fallback = f"file{ext}"
    encoded_name = urllib.parse.quote(name)
    return Response(
        content=content,
        media_type="application/octet-stream",
        headers={"Content-Disposition": f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{encoded_name}"},
    )


@router.get("/download/{run_id}")
def download(run_id: str, _user: dict = Depends(get_current_user)):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="File not ready")
    excel_key = run.get("excel_storage_key")
    if not excel_key:
        raise HTTPException(status_code=503, detail="הקובץ אינו זמין")
    try:
        excel_bytes = get_admin_client().storage.from_("check-files").download(excel_key)
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"שגיאה בהורדת הקובץ: {exc}")
    return Response(
        content=excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="hashvaa-gefen-ksafim.xlsx"'},
    )


@router.get("/pdf/{run_id}")
def download_pdf(run_id: str, section: str = "hashva", _user: dict = Depends(get_current_user)):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="Run not complete")
    pdf_bytes = export_pdf(run, section=section)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="hashvaa-gefen-kesafim.pdf"'},
    )


@router.get("/download-tikhnun/{run_id}")
def download_tikhnun_excel(
    run_id: str,
    section: str = "sikar",
    multiplier: str = "03",
    _user: dict = Depends(get_current_user),
):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="Run not complete")
    tikhnun = run.get("tikhnun")
    if not tikhnun:
        raise HTTPException(status_code=400, detail="No tikhnun data for this run")
    xlsx_bytes = export_tikhnun_excel(tikhnun, section, multiplier)
    filename = f"tikhnun-{section}.xlsx"
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf-tikhnun/{run_id}")
def download_tikhnun_pdf(
    run_id: str,
    section: str = "sikar",
    multiplier: str = "03",
    _user: dict = Depends(get_current_user),
):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="Run not complete")
    tikhnun = run.get("tikhnun")
    if not tikhnun:
        raise HTTPException(status_code=400, detail="No tikhnun data for this run")
    pdf_bytes = export_tikhnun_pdf(tikhnun, section, multiplier)
    filename = f"tikhnun-{section}.pdf"
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/pdf-combined/{run_id}")
def download_combined_pdf(
    run_id: str,
    sections: str = "hashva",
    multiplier: str = "03",
    _user: dict = Depends(get_current_user),
):
    from logic.combined_exporter import export_combined_pdf
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="Run not complete")
    section_list = [s.strip() for s in sections.split(",") if s.strip()]
    pdf_bytes = export_combined_pdf(run, section_list, multiplier)
    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="gefen-combined.pdf"'},
    )


@router.get("/excel-combined/{run_id}")
def download_combined_excel(
    run_id: str,
    sections: str = "hashva",
    multiplier: str = "03",
    _user: dict = Depends(get_current_user),
):
    from logic.combined_exporter import export_combined_excel
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="Run not found")
    if run.get("status") != "done":
        raise HTTPException(status_code=400, detail="Run not complete")
    section_list = [s.strip() for s in sections.split(",") if s.strip()]
    xlsx_bytes = export_combined_excel(run, section_list, multiplier)
    return Response(
        content=xlsx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="gefen-combined.xlsx"'},
    )


class ClassifyRequest(BaseModel):
    classifications: list[dict]  # [{"union_key": str, "budget": str, "stage": str}]
    skipped: list[str]           # union_keys to skip
    division: str = "main"       # "main" | "tikkon" | "beinayim"


@router.post("/classify/{run_id}")
async def classify_rows(
    run_id: str,
    req: ClassifyRequest,
    user: dict = Depends(get_current_user),
):
    """Apply manual budget/stage classifications to pending rows, then finalize tikhnun metrics."""
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="הריצה לא נמצאה")

    if req.division == "tikkon":
        tikhnun_key = "tikhnun_tikkon"
    elif req.division == "beinayim":
        tikhnun_key = "tikhnun_beinayim"
    else:
        tikhnun_key = "tikhnun"

    tikhnun = run.get(tikhnun_key)
    if not tikhnun or not tikhnun.get("pending_identification"):
        raise HTTPException(status_code=400, detail="אין שורות הממתינות לסיווג")

    ctx = tikhnun.get("_zihuy_ctx")
    if not ctx:
        raise HTTPException(status_code=400, detail="חסר הקשר זיהוי")

    results_clean = ctx["results_clean"]
    budgets_raw   = ctx["budgets_raw"]
    perut_rows    = ctx["perut_rows"]

    classify_map = {c["union_key"]: c for c in req.classifications}
    skip_set     = set(req.skipped)

    for r in results_clean:
        key = r.get("union_key", "")
        if key in skip_set:
            r["_skip"] = True
        elif key in classify_map:
            r["budget"] = classify_map[key].get("budget", "")
            r["stage"]  = classify_map[key].get("stage", "")

    results_clean = [r for r in results_clean if not r.get("_skip")]

    _finalize_tikhnun_metrics(tikhnun, results_clean, budgets_raw, perut_rows)
    if tikhnun.get("budgets"):
        try:
            _compute_per_budget_yozma(tikhnun["budgets"], perut_rows, results_clean)
        except Exception as _ye:
            logger.warning("per-budget yozma (classify) failed: %s", _ye)
        try:
            _propagate_meshuyakh_to_root(tikhnun)
        except Exception as _pe:
            logger.warning("meshuyakh propagation (classify) failed: %s", _pe)
        try:
            _build_and_attach_yozma_breakdown(tikhnun["budgets"], perut_rows, results_clean)
        except Exception as _ybe:
            logger.warning("per-budget yozma breakdown (classify) failed: %s", _ybe)
        try:
            _build_and_attach_nihul_breakdown(tikhnun["budgets"], results_clean)
        except Exception as _nhe:
            logger.warning("per-budget nihul breakdown (classify) failed: %s", _nhe)
    try:
        # Always attach, even when empty ({} means "checked, zero found" — a
        # falsy-dict skip here would make that indistinguishable from "never
        # computed" on the frontend, which needs to tell the two apart).
        tikhnun["per_budget_rejected"] = _build_rejected_from_results_clean(results_clean)
    except Exception as rej_exc:
        logger.warning("per_budget_rejected build failed (classify): %s", rej_exc)
    try:
        tikhnun["per_budget_no_pdf"] = _build_no_pdf_from_results_clean(results_clean)
    except Exception as nopdf_exc:
        logger.warning("per_budget_no_pdf build failed (classify): %s", nopdf_exc)

    gefen_paths_cl   = [Path(p) for p in ctx.get("gefen_paths", [])]
    finance_paths_cl = [Path(p) for p in ctx.get("finance_paths", [])]
    finance_type_cl  = ctx.get("finance_type") or run.get("finance_type")
    if gefen_paths_cl and finance_paths_cl and finance_type_cl:
        try:
            _cl_stage = _get_school_stage(run.get("_school_ctx", {}).get("gefen_account_id"))
            df_gefen_cl, _, _ = _load_gefen_files(gefen_paths_cl)
            per_combo = _run_per_combo_reconciliation(
                df_gefen_cl, finance_paths_cl, finance_type_cl,
                results_clean, _get_finance_col_map(finance_type_cl),
                school_stage=_cl_stage,
            )
            if per_combo:
                run["per_combo_results"] = per_combo
        except Exception as exc:
            logger.warning("per_combo_results in classify_rows failed: %s", exc)

    for k in ("pending_identification", "unidentified_rows", "available_budgets", "_zihuy_ctx",
              "needs_finance_upload", "missing_budgets", "covered_budgets", "finance_invalid_reason"):
        tikhnun.pop(k, None)

    run[tikhnun_key] = tikhnun
    if tikhnun_key != "tikhnun":
        run["tikhnun"] = run.get("tikhnun_tikkon") or run.get("tikhnun_beinayim")

    school_ctx = run.get("_school_ctx", {})
    if school_ctx.get("school_id"):
        saved_log_id = run.get("saved_log_id")
        _save_check_log(
            run_id,
            school_ctx.get("user_id", ""),
            school_ctx["school_id"],
            school_ctx.get("gefen_account_id"),
            saved_log_id or school_ctx.get("update_log_id"),
            run_data=run,
            academic_year=school_ctx.get("academic_year"),
        )

    _update_run(run_id, run)
    return {"ok": True, "tikhnun": tikhnun, "per_combo_results": run.get("per_combo_results")}


@router.post("/retry-finance/{run_id}")
async def retry_finance(
    run_id: str,
    division: str = Form("main"),
    files: list[UploadFile] = File(...),
    user: dict = Depends(get_current_user),
):
    run = _get_run(run_id)
    if not run:
        raise HTTPException(status_code=404, detail="הריצה לא נמצאה")

    tikhnun_key = {"main": "tikhnun", "tikkon": "tikhnun_tikkon", "beinayim": "tikhnun_beinayim"}.get(division, "tikhnun")
    tikhnun = run.get(tikhnun_key)
    if not tikhnun or not tikhnun.get("pending_identification"):
        raise HTTPException(status_code=400, detail="אין שורות הממתינות לסיווג עבור חלוקה זו")

    ctx = tikhnun.get("_zihuy_ctx")
    if not ctx:
        raise HTTPException(status_code=400, detail="חסר הקשר זיהוי")

    import tempfile, shutil as _shutil
    tmp_paths: list[Path] = []
    finance_type: str | None = None
    for upload in files:
        suffix = Path(upload.filename).suffix if upload.filename else ".xlsx"
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp_path = Path(tmp.name)
            _shutil.copyfileobj(upload.file, tmp)
        tmp_paths.append(tmp_path)
        if finance_type is None:
            finance_type = identify_file(str(tmp_path))

    try:
        if finance_type not in ("kesafim2000", "payscool", "schoolcash"):
            raise HTTPException(status_code=400, detail="הקובץ שהועלה אינו מזוהה כקובץ תוכנת כספים. ודא שהקובץ תקין.")

        budgets_raw: list = ctx.get("budgets_raw", [])
        plan_budget_norms = [b["norm_name"] for b in budgets_raw]

        ichud_map, _fin_warnings, new_covered, invalid_reason = _build_finance_ichud_budget_map(
            tmp_paths, finance_type, plan_budget_norms
        )

        results_clean: list = ctx.get("results_clean", [])
        if ichud_map:
            for r in results_clean:
                if not r.get("budget"):
                    mapped = ichud_map.get(r.get("union_key", ""))
                    if mapped:
                        r["budget"] = mapped
                        try:
                            rc = int(str(r["orig"][1]).strip())
                        except Exception:
                            rc = 0
                        if rc in TIKKON_ONLY:
                            r["stage"] = "תיכון"
                        elif rc in BEINAYIM_ONLY:
                            r["stage"] = "חטיבת ביניים"
                        else:
                            r["stage"] = ""

        existing_covered: list = tikhnun.get("covered_budgets", [])
        merged_covered = list(set(existing_covered) | set(new_covered))
        tikhnun["covered_budgets"] = merged_covered

        unidentified = [r for r in results_clean if not r.get("budget")]
        if not unidentified:
            perut_rows = ctx.get("perut_rows", [])
            _finalize_tikhnun_metrics(tikhnun, results_clean, budgets_raw, perut_rows)
            if tikhnun.get("budgets"):
                try:
                    _compute_per_budget_yozma(tikhnun["budgets"], perut_rows, results_clean)
                except Exception as _ye:
                    logger.warning("per-budget yozma (classify2) failed: %s", _ye)
                try:
                    _propagate_meshuyakh_to_root(tikhnun)
                except Exception as _pe:
                    logger.warning("meshuyakh propagation (classify2) failed: %s", _pe)
                try:
                    _build_and_attach_yozma_breakdown(tikhnun["budgets"], perut_rows, results_clean)
                except Exception as _ybe:
                    logger.warning("per-budget yozma breakdown (classify2) failed: %s", _ybe)
                try:
                    _build_and_attach_nihul_breakdown(tikhnun["budgets"], results_clean)
                except Exception as _nhe:
                    logger.warning("per-budget nihul breakdown (classify2) failed: %s", _nhe)
            try:
                # Always attach, even when empty — see comment in the "classify" branch above.
                tikhnun["per_budget_rejected"] = _build_rejected_from_results_clean(results_clean)
            except Exception as rej_exc:
                logger.warning("per_budget_rejected build failed (classify2): %s", rej_exc)
            try:
                tikhnun["per_budget_no_pdf"] = _build_no_pdf_from_results_clean(results_clean)
            except Exception as nopdf_exc:
                logger.warning("per_budget_no_pdf build failed (classify2): %s", nopdf_exc)
            gefen_paths_rc   = [Path(p) for p in ctx.get("gefen_paths", [])]
            finance_paths_rc = [Path(p) for p in ctx.get("finance_paths", [])]
            finance_type_rc  = ctx.get("finance_type") or run.get("finance_type")
            if gefen_paths_rc and finance_paths_rc and finance_type_rc:
                try:
                    _rc_stage = _get_school_stage(run.get("_school_ctx", {}).get("gefen_account_id"))
                    df_gefen_rc, _, _ = _load_gefen_files(gefen_paths_rc)
                    per_combo = _run_per_combo_reconciliation(
                        df_gefen_rc, finance_paths_rc, finance_type_rc,
                        results_clean, _get_finance_col_map(finance_type_rc),
                        school_stage=_rc_stage,
                    )
                    if per_combo:
                        run["per_combo_results"] = per_combo
                except Exception as exc:
                    logger.warning("per_combo_results in retry_finance failed: %s", exc)

            for k in ("pending_identification", "unidentified_rows", "available_budgets", "_zihuy_ctx",
                      "needs_finance_upload", "missing_budgets", "covered_budgets", "finance_invalid_reason"):
                tikhnun.pop(k, None)
            run[tikhnun_key] = tikhnun
            if tikhnun_key != "tikhnun":
                run["tikhnun"] = run.get("tikhnun_tikkon") or run.get("tikhnun_beinayim")
            school_ctx = run.get("_school_ctx", {})
            if school_ctx.get("school_id"):
                _save_check_log(
                    run_id,
                    school_ctx.get("user_id", ""),
                    school_ctx["school_id"],
                    school_ctx.get("gefen_account_id"),
                    run.get("saved_log_id") or school_ctx.get("update_log_id"),
                    run_data=run,
                    academic_year=school_ctx.get("academic_year"),
                )
            _update_run(run_id, run)
            return {"pending": False, "tikhnun": tikhnun, "per_combo_results": run.get("per_combo_results")}

        covered_set = set(merged_covered)
        missing_budgets = [b for b in plan_budget_norms if b not in covered_set]
        needs_upload = bool(missing_budgets) or bool(invalid_reason)
        tikhnun["needs_finance_upload"] = needs_upload
        tikhnun["missing_budgets"] = missing_budgets
        tikhnun["finance_invalid_reason"] = invalid_reason
        tikhnun["unidentified_rows"] = [
            {
                "union_key":   r.get("union_key", ""),
                "invoice":     str(r["orig"][4]).strip()  if r.get("orig") and r["orig"][4]  is not None else "",
                "date":        str(r["orig"][5]).strip()  if r.get("orig") and r["orig"][5]  is not None else "",
                "supplier":    str(r["orig"][6]).strip()  if r.get("orig") and r["orig"][6]  is not None else "",
                "amount":      str(r["orig"][11]).strip() if r.get("orig") and r["orig"][11] is not None else "",
                "report_code": str(r["orig"][1]).strip()  if r.get("orig") and r["orig"][1]  is not None else "",
            }
            for r in unidentified
        ]
        ctx["results_clean"] = results_clean
        tikhnun["_zihuy_ctx"] = ctx
        run[tikhnun_key] = tikhnun
        _update_run(run_id, run)
        return {"pending": True, "tikhnun": tikhnun}

    finally:
        for p in tmp_paths:
            try:
                p.unlink()
            except Exception:
                pass


# ---------------------------------------------------------------------------
# Background processing pipeline
# ---------------------------------------------------------------------------

def _save_check_log(run_id: str, user_id: str, school_id: str, gefen_account_id: str | None, update_log_id: str | None = None, run_data: dict | None = None, academic_year: str | None = None) -> None:
    run = run_data if run_data is not None else (_get_run(run_id) or {})
    if run.get("status") not in ("done", "saving") or run.get("tikhnun_only"):
        return
    summary = run.get("summary", {})
    try:
        gefen_file_names = [f.get("filename") for f in summary.get("gefen_files", []) if f.get("filename")]
        finance_file_name = summary.get("finance_file", {}).get("filename") if not run.get("gefen_only") else None

        summary_to_save = {**summary}
        summary_to_save["gefen_only"] = run.get("gefen_only", False)
        summary_to_save["finance_type"] = run.get("finance_type")
        summary_to_save["run_id"] = run_id
        summary_to_save["stored_file_paths"] = run.get("stored_file_paths")
        summary_to_save["per_combo_results"] = run.get("per_combo_results")

        tikhnun = run.get("tikhnun")
        if tikhnun and isinstance(tikhnun, dict):
            tikhnun_clean = {k: v for k, v in tikhnun.items() if k != "_zihuy_ctx"}
            if "overview" in tikhnun_clean:
                summary_to_save["tikhnun_overview"] = tikhnun_clean["overview"]
            summary_to_save["tikhnun_result"] = tikhnun_clean
        summary_to_save["has_tikhnun"] = bool(tikhnun and isinstance(tikhnun, dict) and not tikhnun.get("error"))

        tikhnun_tikkon = run.get("tikhnun_tikkon")
        if tikhnun_tikkon:
            summary_to_save["tikhnun_tikkon_result"] = {k: v for k, v in tikhnun_tikkon.items() if k != "_zihuy_ctx"}

        tikhnun_beinayim = run.get("tikhnun_beinayim")
        if tikhnun_beinayim:
            summary_to_save["tikhnun_beinayim_result"] = {k: v for k, v in tikhnun_beinayim.items() if k != "_zihuy_ctx"}

        tikhnun_filenames = run.get("tikhnun_filenames", [])
        if tikhnun_filenames:
            summary_to_save["tikhnun_filenames"] = tikhnun_filenames

        rows_rejected = run.get("rows_gefen_rejected")
        if rows_rejected is not None:
            summary_to_save["rows_gefen_rejected"] = rows_rejected

        rows_no_pdf = run.get("rows_gefen_no_pdf")
        if rows_no_pdf is not None:
            summary_to_save["rows_gefen_no_pdf"] = rows_no_pdf

        db = get_admin_client()
        rows_fn = run.get("rows_finance_not_gefen")
        rows_gn = run.get("rows_gefen_not_finance")
        log_fields = {
            "finance_file_name": finance_file_name,
            "gefen_file_names": gefen_file_names,
            "in_finance_not_gefen_count": summary.get("in_finance_not_gefen", 0),
            "in_gefen_not_finance_count": summary.get("in_gefen_not_finance", 0),
            "in_finance_not_gefen_sum": _sum_display_amounts(rows_fn),
            "in_gefen_not_finance_sum": _sum_display_amounts(rows_gn),
            "summary": summary_to_save,
            "rows_finance_not_gefen": rows_fn,
            "rows_gefen_not_finance": rows_gn,
        }

        if update_log_id:
            # Safety: if this update is gefen-only (no finance file), preserve any existing
            # finance data from the original log so a partial re-upload can't wipe kasafim data.
            if run.get("gefen_only"):
                try:
                    orig = db.table("check_logs").select(
                        "finance_file_name, in_finance_not_gefen_count, in_finance_not_gefen_sum, rows_finance_not_gefen"
                    ).eq("id", update_log_id).single().execute()
                    if orig.data and orig.data.get("finance_file_name"):
                        log_fields["finance_file_name"] = orig.data["finance_file_name"]
                        log_fields["in_finance_not_gefen_count"] = orig.data.get("in_finance_not_gefen_count") or 0
                        log_fields["in_finance_not_gefen_sum"] = orig.data.get("in_finance_not_gefen_sum") or 0.0
                        log_fields["rows_finance_not_gefen"] = orig.data.get("rows_finance_not_gefen")
                except Exception as merge_exc:
                    logger.warning("Could not merge original finance data for %s: %s", update_log_id, merge_exc)
            db.table("check_logs").update(log_fields).eq("id", update_log_id).execute()
        else:
            result = db.table("check_logs").insert({
                "school_id": school_id,
                "gefen_account_id": gefen_account_id,
                "run_by": user_id,
                "academic_year": academic_year or DEFAULT_ACADEMIC_YEAR,
                **log_fields,
            }).execute()
            if result.data:
                run["saved_log_id"] = result.data[0]["id"]
                _update_run(run_id, run)

        _save_check_metrics(db, school_id, gefen_account_id, academic_year or DEFAULT_ACADEMIC_YEAR, run)
    except Exception as exc:
        logger.error("Failed to save check_log for run %s: %s", run_id, exc)


_STAGE_TO_DIVISION_TYPE = {"תיכון": "tikkon", "חטיבת ביניים": "beinayim", "יסודי": "yesodi"}


def _parse_amount_str(s) -> float:
    """Mirrors the frontend's sumRowsAmount() (ResultsView.jsx) — parses a Hebrew-formatted
    'סכום' display string (e.g. '1,234.56') into a float, defaulting to 0 on any failure."""
    try:
        return float(str(s or "0").replace(",", "")) or 0.0
    except (TypeError, ValueError):
        return 0.0


def _find_combo_for_budget(per_combo_results: dict, budget_name: str, division_type: str) -> dict | None:
    """Finds the reconciliation combo matching (budget_name, division_type). Combos are keyed
    by normalized budget + Hebrew stage label, not by division_type directly, so we map stage
    to division_type first; if that's ambiguous, fall back to the sole budget-name match."""
    if not per_combo_results:
        return None
    candidates = [c for c in per_combo_results.values() if c.get("budget") == budget_name]
    if not candidates:
        return None
    for c in candidates:
        if _STAGE_TO_DIVISION_TYPE.get(c.get("stage")) == division_type:
            return c
    return candidates[0] if len(candidates) == 1 else None


def _compute_check_metrics_rows(school_id: str, gefen_account_id: str | None, academic_year: str, run: dict, db=None) -> list[dict]:
    """Pure computation of check_metrics rows (one per division+budget) from a run/summary-like
    dict — shared by the live save path (_save_check_metrics) and the one-time backfill script,
    so both stay perfectly consistent. `db` is only needed to resolve a single-division school's
    division_type from gefen_accounts (or, failing that, run["summary"]["division"] / school.stage)
    when not already known via tikhnun_tikkon/tikhnun_beinayim."""
    division_results = []
    tikkon = run.get("tikhnun_tikkon")
    beinayim = run.get("tikhnun_beinayim")
    if tikkon or beinayim:
        if tikkon:
            division_results.append(("tikkon", tikkon))
        if beinayim:
            division_results.append(("beinayim", beinayim))
    else:
        tikhnun = run.get("tikhnun")
        if tikhnun and isinstance(tikhnun, dict) and not tikhnun.get("error"):
            division_type = run.get("division_type")
            if not division_type and db is not None and gefen_account_id:
                acc = db.table("gefen_accounts").select("division_type").eq("id", gefen_account_id).single().execute()
                division_type = (acc.data or {}).get("division_type")

            school_stage = None
            if not division_type and db is not None and school_id:
                sch = db.table("schools").select("stage").eq("id", school_id).single().execute()
                school_stage = (sch.data or {}).get("stage")

            # Some checks (saved before a school had a gefen_accounts row, or run through the
            # older single-division-at-a-time flow) carry no gefen_account_id at all — but the
            # gefen file's own division was still detected and stored in summary.division
            # ("tikkon"/"beinayim"). That detector has no concept of "יסודי" at all (יסודי
            # schools reuse ביניים's report codes and get bucketed there by convention), so for
            # a school whose stage is genuinely yesodi this signal would be actively wrong —
            # only trust it when the school isn't declared single-division-יסודי.
            if not division_type and school_stage != "yesodi":
                run_division = (run.get("summary") or {}).get("division")
                if run_division in ("tikkon", "beinayim"):
                    division_type = run_division

            # Last resort: only safe when the school has at most one real gefen_account — with
            # 2+ accounts we can't tell which division an orphaned check belongs to, so we skip
            # it rather than misattribute it to an arbitrary one (school.stage isn't even a valid
            # single division for "sheshshnati" schools).
            if not division_type and db is not None and school_id:
                accs = db.table("gefen_accounts").select("division_type").eq("school_id", school_id).execute()
                acc_rows = accs.data or []
                if len(acc_rows) == 1:
                    division_type = acc_rows[0].get("division_type")
                elif len(acc_rows) == 0 and school_stage in ("tikkon", "beinayim", "yesodi", "other"):
                    division_type = school_stage

            if division_type:
                division_results.append((division_type, tikhnun))

    per_combo_results = run.get("per_combo_results") or {}

    rows_to_save = []
    for division_type, tikhnun_result in division_results:
        if not isinstance(tikhnun_result, dict) or tikhnun_result.get("error"):
            continue
        budgets = tikhnun_result.get("budgets")
        overview = tikhnun_result.get("overview") or {}
        per_budget_rejected = tikhnun_result.get("per_budget_rejected") or {}
        per_budget_no_pdf = tikhnun_result.get("per_budget_no_pdf") or {}
        partial_rows = tikhnun_result.get("partial_rows") or []

        def _extra_metrics(budget_name: str) -> dict:
            rejected = per_budget_rejected.get(budget_name, [])
            no_pdf = per_budget_no_pdf.get(budget_name, [])
            partial = [r for r in partial_rows if r.get("budget") == budget_name]
            combo = _find_combo_for_budget(per_combo_results, budget_name, division_type)
            fin_not_gefen = combo.get("in_finance_not_gefen", []) if combo else []
            gefen_not_fin = combo.get("in_gefen_not_finance", []) if combo else []
            return {
                "rejected_count": len(rejected),
                "rejected_sum": sum(_parse_amount_str(r.get("סכום")) for r in rejected),
                "no_pdf_count": len(no_pdf),
                "no_pdf_sum": sum(_parse_amount_str(r.get("סכום")) for r in no_pdf),
                "partial_count": len(partial),
                "partial_sum": sum(r.get("hefresh") or 0 for r in partial),
                "finance_not_gefen_count": len(fin_not_gefen),
                "finance_not_gefen_sum": sum(_parse_amount_str(r.get("סכום")) for r in fin_not_gefen),
                "gefen_not_finance_count": len(gefen_not_fin),
                "gefen_not_finance_sum": sum(_parse_amount_str(r.get("סכום")) for r in gefen_not_fin),
            }

        if budgets:
            for bud in budgets:
                bud_overview = bud.get("overview") or {}
                budget_name = bud.get("name") or "כללי"
                row = {
                    "school_id": school_id,
                    "gefen_account_id": gefen_account_id,
                    "division_type": division_type,
                    "budget_name": budget_name,
                    "academic_year": academic_year,
                    # Plan-file-derived — safe to always update, regardless of whether this
                    # run's doch/finance files happened to cover this specific budget.
                    "pct_plan": bud_overview.get("pct_plan"),
                    "pct_divuach": bud_overview.get("pct_divuach"),
                    "budget_amount": bud_overview.get("budget"),
                    "planned_amount": bud_overview.get("planned"),
                    "fixed_gap_abs": bud_overview.get("fixed_gap_abs"),
                    "flexible_remaining": bud_overview.get("flexible_remaining"),
                    "sum_chayav": bud_overview.get("sum_chayav"),
                    "sum_divuach": bud_overview.get("sum_divuach"),
                }
                # Doch/finance-run-scoped fields — only include (and thus only overwrite in
                # the upsert) when THIS run actually analyzed this budget; otherwise omit them
                # entirely so the previous good values already in check_metrics are preserved
                # instead of being silently zeroed out.
                if bud_overview.get("doch_analyzed", True):
                    row["pct_tanuz"] = bud_overview.get("pct_tanuz")
                    row.update(_extra_metrics(budget_name))
                rows_to_save.append(row)
        elif overview:
            budget_amt = overview.get("budget") or 0
            planned_amt = overview.get("planned") or 0
            budget_name = "כללי"
            row = {
                "school_id": school_id,
                "gefen_account_id": gefen_account_id,
                "division_type": division_type,
                "budget_name": budget_name,
                "academic_year": academic_year,
                "pct_plan": (planned_amt / budget_amt) if budget_amt else None,
                "pct_divuach": overview.get("pct_divuach"),
                "budget_amount": overview.get("budget"),
                "planned_amount": overview.get("planned"),
                "fixed_gap_abs": overview.get("fixed_gap_abs"),
                "flexible_remaining": overview.get("flexible_remaining"),
                "sum_chayav": overview.get("sum_chayav"),
                "sum_divuach": overview.get("sum_divuach"),
            }
            if overview.get("doch_analyzed", True):
                row["pct_tanuz"] = overview.get("pct_tanuz")
                row.update(_extra_metrics(budget_name))
            rows_to_save.append(row)

    return rows_to_save


def _save_check_metrics(db, school_id: str, gefen_account_id: str | None, academic_year: str, run: dict) -> None:
    """Upsert flattened planning/reporting/reconciliation metrics into check_metrics, one row
    per division+budget, so the dashboard can query them directly instead of parsing
    check_logs.summary JSON. Non-fatal: must never break check_log saving."""
    try:
        rows_to_save = _compute_check_metrics_rows(school_id, gefen_account_id, academic_year, run, db=db)
        for row in rows_to_save:
            db.table("check_metrics").upsert(
                row, on_conflict="school_id,division_type,budget_name,academic_year"
            ).execute()
    except Exception as exc:
        logger.warning("check_metrics enrichment failed (non-fatal) for school %s: %s", school_id, exc)


def _any_tikhnun_pending(run_dict: dict) -> bool:
    return any(
        (run_dict.get(k) or {}).get("pending_identification")
        for k in ("tikhnun", "tikhnun_tikkon", "tikhnun_beinayim")
    )


def _build_finance_ichud_budget_map(
    finance_paths: list[Path], finance_type: str | None, plan_budgets: list[str]
) -> tuple[dict[str, str], list[str], list[str], str | None]:
    """Build a mapping from ichud key → normalized budget name using finance file(s).
    Returns (ichud_to_budget, warnings, covered_budgets, invalid_reason).
    covered_budgets: budget names (normalized) that the uploaded file(s) cover.
    invalid_reason: human-readable reason why the file cannot be used for cross-ref (else None).
    """
    from zihuy_core import normalize_budget_name as _norm_bname

    ichud_to_budget: dict[str, str] = {}
    warnings: list[str] = []
    covered_set: set[str] = set()
    invalid_reason: str | None = None

    if not finance_paths or not finance_type:
        return ichud_to_budget, warnings, [], None

    def _norm_key_amount(val) -> str:
        if not val:
            return ""
        s = str(val).replace(",", "").strip()
        try:
            f = float(s)
            return str(int(f)) if f == int(f) else f"{f:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return s

    if finance_type == "kesafim2000":
        for fpath in finance_paths:
            try:
                with open(str(fpath), "rb") as fh:
                    raw = fh.read().decode("iso-8859-8")
                lines = raw.splitlines()
                if not lines:
                    continue
                # Row 0 = header: E1 at index 4
                header = lines[0].split("\t")
                e1_val = header[4].strip() if len(header) > 4 else ""
                budget_norm = _norm_bname(e1_val) if e1_val else None

                if not budget_norm:
                    if len(plan_budgets) == 1:
                        budget_norm = plan_budgets[0]
                    else:
                        missing = ", ".join(plan_budgets)
                        warnings.append(
                            f"קובץ כספים2000 שהועלה אינו מסונן לפי תקציב. "
                            f"לזיהוי ודאי נדרשים קבצים נפרדים עבור: {missing}"
                        )
                        if not invalid_reason:
                            invalid_reason = "קובץ כספים2000 אינו מסונן לפי תקציב — יש להעלות קובץ נפרד לכל תקציב."
                        continue

                covered_set.add(budget_norm)

                for line in lines[1:]:
                    row = line.split("\t")
                    if len(row) < 11:
                        continue
                    supplier    = str(row[0]).strip()
                    report_code = str(row[1]).strip()
                    invoice     = str(row[3]).strip()
                    amount      = _norm_key_amount(row[10])
                    if supplier and report_code and invoice and amount:
                        key = f"{supplier}-{invoice}-{report_code}-{amount}"
                        ichud_to_budget[key] = budget_norm
            except Exception as exc:
                logger.error("Failed to build ichud map from kesafim2000 %s: %s", fpath, exc)

    elif finance_type == "payscool":
        from openpyxl import load_workbook as _lw

        for fpath in finance_paths:
            try:
                wb = _lw(str(fpath), read_only=True)
                for sheet_name in wb.sheetnames:
                    budget_norm = _norm_bname(sheet_name)
                    if not budget_norm or budget_norm == sheet_name:
                        continue

                    covered_set.add(budget_norm)
                    ws = wb[sheet_name]
                    rows = list(ws.iter_rows(values_only=True))
                    if not rows:
                        continue

                    header_idx = None
                    for i, row in enumerate(rows[:6]):
                        if row and any(str(c).strip() == "סעיף" for c in row if c is not None):
                            header_idx = i
                            break
                    if header_idx is None:
                        continue

                    hdr = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
                    try:
                        sif_col = hdr.index("סעיף")
                        hp_col  = hdr.index("ח.פ")
                        inv_col = hdr.index("מספר חשבונית")
                        amt_col = next(
                            i for i, h in enumerate(hdr) if 'סה"כ' in h and 'לסעיף' in h
                        )
                    except (ValueError, StopIteration):
                        continue

                    stat_col = hdr.index("סטטוס חשבונית") if "סטטוס חשבונית" in hdr else None

                    for row in rows[header_idx + 1:]:
                        if not row or all(c is None for c in row):
                            continue
                        if stat_col is not None and row[stat_col] == "מבוטלת":
                            continue
                        sif = str(row[sif_col]).strip() if row[sif_col] is not None else ""
                        m   = re.search(r'\((\d+)\)', sif)
                        if not m:
                            continue
                        report_code = m.group(1)
                        supplier    = str(row[hp_col]).strip()  if row[hp_col]  is not None else ""
                        invoice     = canonical_payscool_invoice(row[inv_col])
                        amount      = _norm_key_amount(row[amt_col])
                        if supplier and invoice and report_code and amount:
                            key = f"{supplier}-{invoice}-{report_code}-{amount}"
                            ichud_to_budget[key] = budget_norm
                wb.close()
            except Exception as exc:
                logger.error("Failed to build ichud map from payscool %s: %s", fpath, exc)

    elif finance_type == "schoolcash":
        from openpyxl import load_workbook as _lw

        for fpath in finance_paths:
            try:
                wb = _lw(str(fpath), read_only=True)
                ws = wb.active
                rows_list = list(ws.iter_rows(values_only=True))
                if not rows_list:
                    wb.close()
                    continue

                hdr = [str(c).strip() if c is not None else "" for c in rows_list[0]]
                supplier_col = inv_col = sif_col = amt_col = None
                budget_col   = 8  # column I (0-indexed)

                for i, h in enumerate(hdr):
                    if h == "עוסק מורשה":    supplier_col = i
                    elif h == "מספר חשבונית": inv_col = i
                    elif "סעיף" in h and "גפ" in h and "תיאור" not in h: sif_col = i
                    elif h == "סכום":        amt_col = i

                if any(c is None for c in [supplier_col, inv_col, sif_col, amt_col]):
                    wb.close()
                    continue

                budgets_seen: set[str] = set()
                for row in rows_list[1:]:
                    if not row or all(c is None for c in row):
                        continue
                    budget_raw  = str(row[budget_col]).strip() if len(row) > budget_col and row[budget_col] is not None else ""
                    budget_norm = _norm_bname(budget_raw) if budget_raw else None
                    if budget_norm:
                        budgets_seen.add(budget_norm)
                    supplier    = str(row[supplier_col]).strip() if row[supplier_col] is not None else ""
                    invoice     = str(row[inv_col]).strip()      if row[inv_col]       is not None else ""
                    report_code = str(row[sif_col]).strip()      if row[sif_col]       is not None else ""
                    amount      = _norm_key_amount(row[amt_col])
                    if supplier and invoice and report_code and amount and budget_norm:
                        key = f"{supplier}-{invoice}-{report_code}-{amount}"
                        ichud_to_budget[key] = budget_norm

                covered_set.update(budgets_seen)
                missing = [b for b in plan_budgets if b not in budgets_seen]
                if missing and budgets_seen:
                    warnings.append(
                        f"קובץ סקולקאש מכיל תקציב/ים: {', '.join(budgets_seen)}. "
                        f"לזיהוי ודאי ייתכן שנדרש קובץ גם עבור: {', '.join(missing)}"
                    )
                wb.close()
            except Exception as exc:
                logger.error("Failed to build ichud map from schoolcash %s: %s", fpath, exc)

    return ichud_to_budget, warnings, list(covered_set), invalid_reason


def _finalize_tikhnun_metrics(
    tikhnun_result: dict,
    results_clean: list,
    budgets_raw: list,
    perut_rows: list,
) -> None:
    """Compute per-budget nikuy, pct_tanuz, partial_rows, and budgets list.
    Modifies tikhnun_result in-place.
    """
    from zihuy_core import normalize_budget_name as _norm_bname

    def _to_f(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    doch_budget_norms = {
        _norm_bname(r["budget"]) for r in results_clean if r.get("budget")
    }

    # Per-budget nikuy (only when doch results are available)
    if results_clean:
        for bud in budgets_raw:
            norm = bud["norm_name"]
            nadche = lelo_k = overlap = 0.0
            for r in results_clean:
                rb = _norm_bname(r["budget"]) if r.get("budget") else None
                if rb != norm:
                    continue
                orig = r["orig"]
                try:
                    l_val = float(str(orig[11]).replace(",", "").strip()) if orig[11] else 0.0
                except Exception:
                    l_val = 0.0
                if not orig[11] or l_val == 0:
                    continue
                stat   = str(orig[12]).strip() if orig[12] else ""
                fexist = str(orig[13]).strip() if orig[13] else ""
                is_n = stat.startswith("נדחה:")
                is_l = fexist == "לא"
                if is_n: nadche  += l_val
                if is_l: lelo_k  += l_val
                if is_n and is_l: overlap += l_val
            nikuy = nadche + lelo_k - overlap
            sc = bud["sum_chayav"]
            bud["nikuy"]     = nikuy
            bud["pct_tanuz"] = (bud["S"] - nikuy) / sc if sc > 0 else None
            # Explicit marker for callers (e.g. check_metrics) that need to know whether
            # THIS run actually analyzed this budget's doch/finance files — pct_tanuz alone
            # isn't a reliable signal, since it computes a real (non-null) ratio here even
            # for budgets with zero matching rows this run.
            bud["doch_analyzed"] = norm in doch_budget_norms

    # Update overview for doch-relevant budgets
    if doch_budget_norms and "overview" in tikhnun_result:
        relevant = [b for b in budgets_raw if b["norm_name"] in doch_budget_norms]
        if relevant:
            total_sc    = sum(b["sum_chayav"] for b in relevant)
            total_nikuy = sum(b.get("nikuy") or 0.0 for b in relevant)
            total_s     = sum(b["S"] for b in relevant)
            tikhnun_result["overview"]["sum_chayav"] = total_sc
            tikhnun_result["overview"]["pct_tanuz"]  = (
                (total_s - total_nikuy) / total_sc if total_sc > 0 else None
            )
    if "overview" in tikhnun_result:
        tikhnun_result["overview"]["doch_analyzed"] = bool(doch_budget_norms)

    # Rebuild partial_rows
    if doch_budget_norms and results_clean and "partial_rows" in tikhnun_result:
        from collections import defaultdict as _defdict

        def _pk_perut(r):
            j = str(r[9]).strip() if len(r) > 9 and r[9] else ""
            if j and j != " ":
                return j
            rcode = str(r[17]).strip() if len(r) > 17 and r[17] else ""
            iname = str(r[8]).strip()  if len(r) > 8  and r[8]  else ""
            return f"{rcode}-{iname}"

        # Per-budget exec sums: {budget_norm: {plan_key: amount}}
        # Must be per-budget to avoid cross-budget contamination when two budgets
        # share the same plan key (same report code + no plan number).
        exec_sums: dict[str, dict[str, float]] = {}
        for _zr in results_clean:
            _orig     = _zr["orig"]
            _a        = str(_orig[0]).strip() if _orig[0] else ""
            if "מענה משרדי" in _a:
                continue
            _bud_norm = _norm_bname(_zr.get("budget") or "") if _zr.get("budget") else ""
            if not _bud_norm:
                continue
            _d    = str(_orig[3]).strip() if _orig[3] else ""
            _b    = str(_orig[1]).strip() if _orig[1] else ""
            _left = _d.split("-")[0].strip()
            _pk   = _left if _left.isdigit() else f"{_b}-{_d}"
            bud_sums = exec_sums.setdefault(_bud_norm, {})
            bud_sums[_pk] = bud_sums.get(_pk, 0.0) + (_to_f(_orig[11]) if _orig[11] is not None else 0.0)

        _fseen: set = set()
        _filtered_plans = []
        for row in perut_rows[1:]:
            raw_bname = str(row[0]).strip() if row[0] else ""
            if _norm_bname(raw_bname) not in doch_budget_norms:
                continue
            col_r = row[17] if len(row) > 17 else None
            if not col_r or not str(col_r).strip():
                continue
            _col13 = row[13] if len(row) > 13 else None
            if _col13 is not None:
                key10 = (
                    str(row[9] or "").strip()  if len(row) > 9  else "",
                    str(row[10] or "").strip() if len(row) > 10 else "",
                    str(row[11] or "").strip() if len(row) > 11 else "",
                    str(row[12] or "").strip() if len(row) > 12 else "",
                    _col13,
                    str(row[17] or "").strip(),
                )
            else:
                key10 = (tuple(str(x).strip() if x else "" for x in row[:10]),
                         str(row[17] or "").strip())
            if key10 in _fseen:
                continue
            _fseen.add(key10)
            _filtered_plans.append(row)

        # Group rows by (A,B,C,D,E,G,H,I,J,O,R) — skipping only col F (funding basket)
        # and col P (partial amount). Same plan split across two baskets will have
        # identical O (total cost) but different F and P values.
        _plan_groups: dict = {}
        for r in _filtered_plans:
            gk = tuple(
                str(r[i]).strip() if (len(r) > i and r[i] is not None) else ""
                for i in [0, 1, 2, 3, 4, 6, 7, 8, 9, 14, 17]
            )
            if gk not in _plan_groups:
                budget_norm = _norm_bname(str(r[0]).strip() if r[0] else "")
                pkey        = _pk_perut(r)
                rcode       = str(r[17]).strip() if r[17] else ""
                name        = str(r[8]).strip()  if r[8]  else ""
                j_val       = str(r[9]).strip()  if len(r) > 9 and r[9] else ""
                mispnum     = j_val if (j_val and j_val != " ") else "אין"
                _plan_groups[gk] = {
                    "tikhnun": 0.0, "rcode": rcode, "name": name,
                    "mispnum": mispnum, "budget_norm": budget_norm, "pkey": pkey,
                }
            _plan_groups[gk]["tikhnun"] += (_to_f(r[13]) if (len(r) > 13 and r[13] is not None) else _to_f(r[15])) if len(r) > 13 else (_to_f(r[15]) if len(r) > 15 else 0.0)

        _new_partial = []
        for g in _plan_groups.values():
            tikhnun_v   = g["tikhnun"]
            budget_norm = g["budget_norm"]
            pkey        = g["pkey"]
            divuach     = exec_sums.get(budget_norm, {}).get(pkey, 0.0)
            hefresh     = tikhnun_v - divuach
            if hefresh < 1:
                continue
            pct = divuach / tikhnun_v if tikhnun_v > 0 else 0.0
            _new_partial.append({
                "key": pkey, "rcode": g["rcode"], "name": g["name"], "mispnum": g["mispnum"],
                "tikhnun": tikhnun_v, "divuach": divuach, "hefresh": hefresh,
                "pct": pct, "budget": budget_norm,
            })

        _new_partial.sort(key=lambda x: (round(x["pct"], 4), -x["tikhnun"]))
        tikhnun_result["partial_rows"]        = _new_partial
        tikhnun_result["sum_hefresh_partial"] = sum(pr["hefresh"] for pr in _new_partial)
        tikhnun_result["partial_has_issues"]  = bool(_new_partial)

    # Build budgets list
    budgets_list = []
    for bud in budgets_raw:
        pct_plan = bud["L"] / bud["H"] if bud["H"] > 0 else 0.0
        budgets_list.append({
            "name":     bud["norm_name"],
            "raw_name": bud["raw_name"],
            "overview": {
                "budget":             bud["H"],
                "planned":            bud["L"],
                "pct_plan":           pct_plan,
                "sum_divuach":        bud["S"],
                "pct_divuach":        bud["T"],
                "sum_chayav":         bud["sum_chayav"],
                "pct_tanuz":          bud.get("pct_tanuz"),
                # Approximate: H - L (kvua correction per-budget is computed separately when available)
                "flexible_remaining": max(bud["H"] - bud["L"], 0.0),
                "doch_analyzed":      bud.get("doch_analyzed", False),
            },
        })

    tikhnun_result["budgets"] = budgets_list

    # Compute per-budget fixed_gap_abs and correct flexible_remaining using kvua_rows
    kvua_rows = tikhnun_result.get("kvua_rows") or []
    if kvua_rows and budgets_list:
        from zihuy_core import normalize_budget_name as _nb_kvua
        kvua_by_budget: dict = {}
        for kr in kvua_rows:
            bt = str(kr.get("budget_type") or "").strip()
            nb = _nb_kvua(bt)
            kr["budget_norm"] = nb
            kvua_by_budget[nb] = kvua_by_budget.get(nb, 0.0) + float(kr.get("hefresh") or 0)
        for bdict in budgets_list:
            nb = bdict["name"]
            fixed_gap = abs(kvua_by_budget.get(nb, 0.0))
            ov = bdict.get("overview", {})
            H_i = float(ov.get("budget") or 0)
            L_i = float(ov.get("planned") or 0)
            ov["fixed_gap_abs"]      = fixed_gap
            ov["flexible_remaining"] = H_i - L_i - fixed_gap


# ---------------------------------------------------------------------------
# Per-combo (budget × stage) reconciliation helpers
# ---------------------------------------------------------------------------

def _build_gefen_combo_map(results_clean: list) -> dict:
    """Returns {ichud_key: (budget_norm, stage)} for every identified doch row.
    Rows with empty stage (SHARED report codes 157/159/163) fall back to the
    budget's dominant stage so they land in a concrete combo bucket.
    """
    from zihuy_core import normalize_budget_name as _nb
    from logic.gefen_processor import normalize_amount as _na

    # First pass: collect definite stage per budget_norm
    budget_stage_fallback: dict = {}
    for r in results_clean:
        if r.get("budget") and r.get("stage"):
            budget_stage_fallback.setdefault(_nb(r["budget"]), r["stage"])

    combo_map: dict = {}
    for r in results_clean:
        budget  = r.get("budget")
        orig    = r.get("orig")
        sup_num = r.get("sup_num")
        if not (budget and orig and sup_num):
            continue
        budget_n  = _nb(budget)
        eff_stage = r.get("stage") or budget_stage_fallback.get(budget_n, "")
        if not eff_stage:
            continue
        key = f"{sup_num}-{_na(orig[4])}-{orig[1]}-{_na(orig[11])}"
        combo_map[key] = (budget_n, eff_stage)
    return combo_map


def _split_gefen_by_combo(df_gefen, combo_map: dict) -> dict:
    """Returns {(budget_norm, stage): sub_df} by looking up each row's ichud in combo_map."""
    groups: dict = {}
    for idx, row in df_gefen.iterrows():
        combo = combo_map.get(row["ichud"])
        if combo is None:
            continue
        groups.setdefault(combo, []).append(idx)
    return {
        combo: df_gefen.loc[indices].reset_index(drop=True)
        for combo, indices in groups.items()
    }


def _stage_from_report_code(report_code, fallback_stage: str | None) -> str | None:
    """TIKKON_ONLY codes → 'תיכון'; BEINAYIM_ONLY → fallback_stage (resolved from doch)."""
    from logic.reconciler import TIKKON_ONLY, BEINAYIM_ONLY
    try:
        c = int(str(report_code).strip())
    except (ValueError, TypeError):
        return fallback_stage
    if c in set(TIKKON_ONLY):
        return "תיכון"
    if c in set(BEINAYIM_ONLY):
        return fallback_stage
    return fallback_stage  # SHARED or unknown → defer to doch identification


def _get_fallback_stage(results_clean: list, budget_norm: str) -> str | None:
    """Returns the stage that results_clean identified for this budget (first match)."""
    from zihuy_core import normalize_budget_name as _nb
    for r in results_clean:
        if r.get("budget") and _nb(r["budget"]) == budget_norm and r.get("stage"):
            return r["stage"]
    return None


def _split_finance_kesafim(finance_paths: list, results_clean: list) -> dict:
    """Returns {(budget_norm, stage): df} for Kesafim2000 files."""
    import pandas as pd
    from zihuy_core import normalize_budget_name as _nb
    from logic.kesafim_processor import load_kesafim
    all_groups: dict = {}
    for fpath in finance_paths:
        try:
            with open(str(fpath), "rb") as fh:
                raw = fh.read().decode("iso-8859-8")
            lines = raw.splitlines()
            if not lines:
                continue
            header = lines[0].split("\t")
            e1_val = header[4].strip() if len(header) > 4 else ""
            budget_norm = _nb(e1_val) if e1_val else None
            if not budget_norm:
                continue
            fallback = _get_fallback_stage(results_clean, budget_norm)
            df = load_kesafim(str(fpath))
            df["_stage"] = df["report_code"].apply(
                lambda rc: _stage_from_report_code(rc, fallback)
            )
            for stage_val, sub_df in df.groupby("_stage"):
                if stage_val:
                    all_groups.setdefault((budget_norm, stage_val), []).append(
                        sub_df.drop(columns=["_stage"])
                    )
        except Exception as exc:
            logger.warning("_split_finance_kesafim error for %s: %s", fpath, exc)
    return {
        combo: pd.concat(dfs, ignore_index=True).drop_duplicates()
        for combo, dfs in all_groups.items()
    }


def _load_payscool_sheet_df(fpath, sheet_name: str):
    """Load one PaySchool sheet as a reconcilable DataFrame (report_code + ichud).
    Replicates load_payscool logic for a named sheet. Returns None if unusable.
    """
    import re, pandas as pd
    from logic.gefen_processor import normalize_amount as _na

    def _rc(v):
        m = re.search(r'\((\d+)\)', str(v)) if pd.notna(v) else None
        return m.group(1) if m else None

    try:
        df = pd.read_excel(str(fpath), sheet_name=sheet_name, header=None)
        if df.empty:
            return None
        header_row = None
        for i in range(min(6, len(df))):
            if "סעיף" in {str(v).strip() for v in df.iloc[i] if pd.notna(v)}:
                header_row = i
                break
        if header_row is None:
            return None
        df.columns = df.iloc[header_row]
        df = df.iloc[header_row + 1:].reset_index(drop=True)
        if "סטטוס חשבונית" in df.columns:
            df = df[df["סטטוס חשבונית"] != "מבוטלת"].copy()
        if "סעיף" not in df.columns:
            return None
        df["report_code"] = df["סעיף"].apply(_rc)
        df = df[df["report_code"].notna()].copy()
        if df.empty:
            return None
        amt_col = next((c for c in df.columns if 'סה"כ' in str(c) and 'לסעיף' in str(c)), None)
        if amt_col is None or "ח.פ" not in df.columns or "מספר חשבונית" not in df.columns:
            return None
        df["ichud"] = (
            df["ח.פ"].apply(_na) + "-"
            + df["מספר חשבונית"].apply(canonical_payscool_invoice).apply(_na) + "-"
            + df["report_code"].astype(str) + "-"
            + df[amt_col].apply(_na)
        )
        return df
    except Exception as exc:
        logger.warning("_load_payscool_sheet_df %s/%s: %s", fpath, sheet_name, exc)
        return None


def _split_finance_payscool(finance_paths: list, results_clean: list) -> dict:
    """Returns {(budget_norm, stage): df} for PaySchool files (one sheet per budget)."""
    import pandas as pd
    from zihuy_core import normalize_budget_name as _nb
    import openpyxl
    all_groups: dict = {}
    for fpath in finance_paths:
        try:
            wb = openpyxl.load_workbook(str(fpath), read_only=True)
            sheet_names = wb.sheetnames
            wb.close()
            for sname in sheet_names:
                budget_norm = _nb(sname)
                if not budget_norm:
                    continue
                df = _load_payscool_sheet_df(fpath, sname)
                if df is None or df.empty:
                    continue
                fallback = _get_fallback_stage(results_clean, budget_norm)
                df["_stage"] = df["report_code"].apply(
                    lambda rc: _stage_from_report_code(rc, fallback)
                )
                for stage_val, sub_df in df.groupby("_stage"):
                    if stage_val:
                        all_groups.setdefault((budget_norm, stage_val), []).append(
                            sub_df.drop(columns=["_stage"])
                        )
        except Exception as exc:
            logger.warning("_split_finance_payscool error for %s: %s", fpath, exc)
    return {
        combo: pd.concat(dfs, ignore_index=True).drop_duplicates()
        for combo, dfs in all_groups.items()
    }


def _split_finance_schoolcash(finance_paths: list, results_clean: list) -> dict:
    """Returns {(budget_norm, stage): df} for SchoolCash files (budget in column index 8)."""
    import pandas as pd
    from zihuy_core import normalize_budget_name as _nb
    from logic.schoolcash_processor import load_schoolcash
    all_groups: dict = {}
    for fpath in finance_paths:
        try:
            df = load_schoolcash(str(fpath))
            df_raw = pd.read_excel(str(fpath), header=0)
            if len(df_raw.columns) <= 8:
                continue
            budget_col_name = df_raw.columns[8]
            report_col = 'סעיף תקציב גפ"ן'
            df_raw_filt = df_raw.dropna(subset=["עוסק מורשה", "מספר חשבונית", "סכום"]).copy()
            df_raw_filt["_rc_tmp"] = df_raw_filt[report_col].apply(
                lambda x: int(float(x)) if pd.notna(x) and str(x).strip() not in ("", "nan") else None
            )
            df_raw_filt = df_raw_filt[df_raw_filt["_rc_tmp"].notna()].reset_index(drop=True)
            df = df.reset_index(drop=True)
            if len(df) != len(df_raw_filt):
                logger.warning("_split_finance_schoolcash: row count mismatch for %s", fpath)
                continue
            df["_budget_norm"] = df_raw_filt[budget_col_name].apply(
                lambda x: _nb(str(x).strip()) if pd.notna(x) and str(x).strip() else None
            ).values
            df["_stage"] = df.apply(
                lambda row: _stage_from_report_code(
                    row["report_code"],
                    _get_fallback_stage(results_clean, row["_budget_norm"]) if row["_budget_norm"] else None,
                ),
                axis=1,
            )
            for (budget_norm, stage_val), sub_df in df.groupby(["_budget_norm", "_stage"]):
                if budget_norm and stage_val:
                    all_groups.setdefault((budget_norm, stage_val), []).append(
                        sub_df.drop(columns=["_budget_norm", "_stage"])
                    )
        except Exception as exc:
            logger.warning("_split_finance_schoolcash error for %s: %s", fpath, exc)
    return {
        combo: pd.concat(dfs, ignore_index=True).drop_duplicates()
        for combo, dfs in all_groups.items()
    }


def _get_finance_col_map(finance_type: str) -> list:
    if finance_type == "kesafim2000":
        return _KESAFIM_COL_MAP
    if finance_type == "schoolcash":
        return _SCHOOLCASH_COL_MAP
    return _PAYSCOOL_COL_MAP


def _get_school_stage(gefen_account_id: str | None) -> str | None:
    """Fetch division_type from gefen_accounts for authoritative stage resolution."""
    if not gefen_account_id:
        return None
    try:
        from supabase_client import get_admin_client as _gac
        acc = _gac().table("gefen_accounts").select("division_type").eq("id", gefen_account_id).single().execute()
        return (acc.data or {}).get("division_type")
    except Exception:
        return None


def _run_per_combo_reconciliation(
    df_gefen,
    finance_paths: list,
    finance_type: str,
    results_clean: list,
    finance_col_map: list,
    school_stage: str | None = None,
) -> dict | None:
    """Splits doch + finance by (budget, stage) and reconciles each combo separately.
    Returns None when fewer than 2 combos are found or identification is incomplete.
    """
    from zihuy_core import normalize_budget_name as _nb

    if not results_clean:
        return None
    if not all(r.get("budget") for r in results_clean):
        return None

    # For יסודי accounts: BEINAYIM_ONLY codes default to "חטיבת ביניים" when no plan
    # stage marker is available, but every such code must be "יסודי" in a yesodi school.
    # Override here so that gefen_splits, finance_splits, and combos all use one stage.
    if school_stage == "yesodi":
        results_clean = [
            {**r, "stage": "יסודי"} if r.get("stage") == "חטיבת ביניים" else r
            for r in results_clean
        ]

    # Build per-budget fallback stage for SHARED code rows (stage="")
    budget_stage_fallback: dict = {}
    for r in results_clean:
        if r.get("budget") and r.get("stage"):
            budget_stage_fallback.setdefault(_nb(r["budget"]), r["stage"])

    def _eff_stage(r):
        s = r.get("stage", "")
        return s or budget_stage_fallback.get(_nb(r["budget"]), "") if r.get("budget") else ""

    combos = set((_nb(r["budget"]), _eff_stage(r)) for r in results_clean if r.get("budget") and _eff_stage(r))
    if len(combos) == 0:
        return None

    combo_map     = _build_gefen_combo_map(results_clean)
    gefen_splits  = _split_gefen_by_combo(df_gefen, combo_map)

    if finance_type == "kesafim2000":
        finance_splits = _split_finance_kesafim(finance_paths, results_clean)
    elif finance_type == "payscool":
        finance_splits = _split_finance_payscool(finance_paths, results_clean)
    else:
        finance_splits = _split_finance_schoolcash(finance_paths, results_clean)

    # Build global ichud sets once for cross-budget mismatch filtering.
    # A finance/gefen row that is globally matched but under a different budget
    # should not be reported as a per-combo discrepancy.
    global_gefen_ichud: set = set(df_gefen["ichud"].dropna()) if "ichud" in df_gefen.columns else set()
    global_fin_ichud: set = set()
    for _sub_df in finance_splits.values():
        if "ichud" in _sub_df.columns:
            global_fin_ichud.update(_sub_df["ichud"].dropna())

    per_combo: dict = {}
    for combo in sorted(combos):
        budget_norm, stage = combo
        key      = f"{budget_norm}-{stage}"
        doch_sub = gefen_splits.get(combo)
        fin_sub  = finance_splits.get(combo)

        if doch_sub is None or doch_sub.empty:
            per_combo[key] = {
                "budget": budget_norm, "stage": stage,
                "not_checked": False,
                "in_finance_not_gefen": [], "in_gefen_not_finance": [],
                "division": "both", "finance_rows_checked": 0,
            }
            continue

        if fin_sub is None or fin_sub.empty:
            # Finance file doesn't cover this budget — mark as not checked (not a false "0 discrepancies")
            per_combo[key] = {
                "budget": budget_norm, "stage": stage,
                "not_checked": True,
                "not_checked_reason": "missing_finance",
                "not_checked_text": f"לא נמצאו נתוני כספים עבור תקציב {budget_norm}",
                "in_finance_not_gefen": [],
                "in_gefen_not_finance": [],
                "division": "both",
                "finance_rows_checked": 0,
            }
            continue

        try:
            in_fng, in_gnf, division, fin_checked = reconcile(doch_sub, fin_sub)

            # Remove cross-budget attribution rows: ichud present globally but assigned to a
            # different budget/stage in the other system. Not real discrepancies.
            if global_gefen_ichud and not in_fng.empty and "ichud" in in_fng.columns:
                in_fng = in_fng[~in_fng["ichud"].isin(global_gefen_ichud)].copy()
            if global_fin_ichud and not in_gnf.empty and "ichud" in in_gnf.columns:
                in_gnf = in_gnf[~in_gnf["ichud"].isin(global_fin_ichud)].copy()

            if finance_type == "kesafim2000":
                in_fng = in_fng.rename(columns=_KESAFIM_RENAME)
            else:
                in_fng = in_fng.rename(columns={"report_code": "קוד דיווח"})
            per_combo[key] = {
                "budget": budget_norm, "stage": stage,
                "not_checked": False,
                "in_finance_not_gefen": _build_display_records(in_fng, finance_col_map),
                "in_gefen_not_finance": _build_display_records(in_gnf, _GEFEN_COL_MAP),
                "division": division,
                "finance_rows_checked": fin_checked,
            }
        except Exception as exc:
            logger.warning("per-combo reconciliation failed for %s: %s", key, exc)
            per_combo[key] = {
                "budget": budget_norm, "stage": stage,
                "not_checked": False,
                "in_finance_not_gefen": [], "in_gefen_not_finance": [],
                "division": "both", "finance_rows_checked": 0,
            }

    return per_combo if len(per_combo) >= 1 else None


# ---------------------------------------------------------------------------
# Per-budget yozma (initiatives/needs) helpers
# ---------------------------------------------------------------------------

_YOZMA_CODES_TIKKON = {
    "106": ("נלוות",         1.0),
    "107": ("רכוש קבוע",    0.5),
    "105": ("כיבוד",         0.15),
    "108": ("תיקונים קלים", 0.10),
}
_YOZMA_CODES_BEINAYIM = {
    "69": ("נלוות",         1.0),
    "70": ("רכוש קבוע",    0.5),
    "68": ("כיבוד",         0.15),
    "71": ("תיקונים קלים", 0.10),
}


def _compute_per_budget_yozma(
    budgets_list: list,
    perut_rows: list,
    results_clean: list | None,
) -> None:
    """Mutates each budget dict in budgets_list to add yozma_03 and yozma_04."""
    from zihuy_core import normalize_budget_name as _nb_yoz

    def _to_f(v):
        if v is None:
            return 0.0
        try:
            return float(str(v).replace(",", "").strip())
        except Exception:
            return 0.0

    for bdict in budgets_list:
        budget_norm = bdict.get("name", "")
        H = float((bdict.get("overview") or {}).get("budget") or 0)
        if not H:
            continue

        # Determine stage from results_clean (first matching row)
        stage = None
        if results_clean:
            for r in results_clean:
                if _nb_yoz(r.get("budget", "")) == budget_norm and r.get("stage"):
                    stage = r["stage"]
                    break
        yozma_map = _YOZMA_CODES_TIKKON if stage == "תיכון" else _YOZMA_CODES_BEINAYIM

        # Dedup by component identity (plan+supplier+type+name+code) so each
        # physical component is counted once even when it appears in multiple baskets.
        component_seen: set = set()
        betikhnun_total = 0.0
        detail_betikhnun: dict = {c: 0.0 for c in yozma_map}
        for row in perut_rows[1:]:
            if not row[0] or _nb_yoz(str(row[0])) != budget_norm:
                continue
            code = str(row[17] or "").strip()
            if code not in yozma_map:
                continue
            _c13 = row[13] if len(row) > 13 else None
            if _c13 is not None:
                comp_key = (
                    str(row[9] or "").strip()  if len(row) > 9  else "",
                    str(row[10] or "").strip() if len(row) > 10 else "",
                    str(row[11] or "").strip() if len(row) > 11 else "",
                    str(row[12] or "").strip() if len(row) > 12 else "",
                    _c13,
                    code,
                )
            else:
                comp_key = (tuple(str(x).strip() if x else "" for x in row[:10]), code)
            if comp_key in component_seen:
                continue
            component_seen.add(comp_key)
            amt = _to_f(_c13) if _c13 is not None else _to_f(row[15])
            betikhnun_total += amt
            detail_betikhnun[code] = detail_betikhnun.get(code, 0.0) + amt

        # Sum doch execution amounts per yozma code (for "משויך" column)
        code_exec: dict = {c: 0.0 for c in yozma_map}
        if results_clean:
            for r in results_clean:
                if _nb_yoz(r.get("budget", "")) != budget_norm:
                    continue
                orig = r.get("orig", ())
                if len(orig) < 12:
                    continue
                cd = str(orig[1] or "").strip()
                if cd not in yozma_map:
                    continue
                code_exec[cd] = code_exec.get(cd, 0.0) + _to_f(orig[11])

        flexible = float((bdict.get("overview") or {}).get("flexible_remaining") or 0)

        for mult, pct in (("03", 0.3), ("04", 0.4)):
            mx = H * pct
            detail = []
            for code, (label, cap_pct) in yozma_map.items():
                cap = mx * cap_pct
                bt = detail_betikhnun.get(code, 0.0)
                diff = cap - bt
                hefresh_shanim = min(diff, flexible) if diff >= 0 else diff
                detail.append({
                    "label":      label,
                    "code":       code,
                    "cap":        int(cap),
                    "betikhnun":  int(bt),
                    "meshuyakh":  int(code_exec.get(code, 0.0)),
                    "hefresh":    int(hefresh_shanim),
                })
            bdict[f"yozma_{mult}"] = {
                "max":         int(mx),
                "betikhnun":   int(betikhnun_total),
                "hefresh":     int(mx - betikhnun_total),
                "is_negative": (mx - betikhnun_total) < 0,
                "detail":      detail,
            }


def _propagate_meshuyakh_to_root(tikhnun_result: dict) -> None:
    """Copy meshuyakh from per-budget yozma detail back to root yozma_03/yozma_04 detail.

    _compute_per_budget_yozma writes meshuyakh onto each tikhnun_result["budgets"][i]["yozma_03/04"].
    The frontend reads the ROOT tikhnun_result["yozma_03/04"] for single-budget schools (no pills).
    This function merges the meshuyakh values into the root detail by matching report codes.
    """
    budgets = tikhnun_result.get("budgets") or []
    if not budgets:
        return
    for mult in ("03", "04"):
        root_detail = (tikhnun_result.get(f"yozma_{mult}") or {}).get("detail") or []
        if not root_detail:
            continue
        code_to_idx = {item["code"]: i for i, item in enumerate(root_detail)}
        for item in root_detail:
            item["meshuyakh"] = 0
        for bdict in budgets:
            for bitem in (bdict.get(f"yozma_{mult}") or {}).get("detail") or []:
                cd = bitem.get("code", "")
                if cd in code_to_idx:
                    root_detail[code_to_idx[cd]]["meshuyakh"] += bitem.get("meshuyakh") or 0


def _build_and_attach_yozma_breakdown(budgets_list: list, perut_rows: list, results_clean: list) -> None:
    """Add yozma_breakdown to each budget dict — supplier totals and transactions per (plan, code)."""
    import re as _re
    from collections import defaultdict as _dd
    from zihuy_core import normalize_budget_name as _nb_yoz

    if not results_clean or not budgets_list:
        return

    def _to_f(v):
        if v is None: return 0.0
        try: return float(str(v).replace(",", "").strip())
        except Exception: return 0.0

    # (norm_budget, plan_num, code) → initiative name from פירוט המענים
    plan_code_to_name: dict = {}
    for row in perut_rows[1:]:
        if not row[0]:
            continue
        nb = _nb_yoz(str(row[0]))
        pn = str(row[9] or "").strip() if len(row) > 9 else ""
        cd = str(row[17] or "").strip() if len(row) > 17 else ""
        nm = str(row[8] or "").strip() if len(row) > 8 else ""
        if pn and cd and (nb, pn, cd) not in plan_code_to_name:
            plan_code_to_name[(nb, pn, cd)] = nm

    # Group doch rows: (norm_budget, plan_num, code) → sup_num → transaction list
    combo_sups: dict = _dd(lambda: _dd(list))
    for r in results_clean:
        if not r.get("budget"):
            continue
        nb = _nb_yoz(str(r["budget"]))
        pn = str(r.get("plan_num") or "")
        if not pn:
            continue
        orig = r.get("orig", ())
        if len(orig) < 12:
            continue
        cd = str(orig[1] or "").strip()
        sup_num = str(r.get("sup_num") or "").strip()
        sup_raw = str(orig[6] or "")
        sup_name = _re.sub(r'^\s*\d+\s*-\s*', '', sup_raw).strip()
        combo_sups[(nb, pn, cd)][sup_num].append({
            "date":          _normalize_date(str(orig[5])) if orig[5] is not None else "",
            "invoice":       str(orig[4] or "").strip(),
            "description":   str(orig[10] or "").strip(),
            "amount":        int(round(_to_f(orig[11]))),
            "supplier_name": sup_name,
        })

    for bdict in budgets_list:
        budget_norm = bdict.get("name", "")
        stage = None
        for r in results_clean:
            if _nb_yoz(r.get("budget", "")) == budget_norm and r.get("stage"):
                stage = r["stage"]
                break
        yozma_map = _YOZMA_CODES_TIKKON if stage == "תיכון" else _YOZMA_CODES_BEINAYIM

        items: list = []
        seen: set = set()
        for (nb, pn, cd), sup_dict in combo_sups.items():
            if nb != budget_norm or cd not in yozma_map:
                continue
            if (pn, cd) in seen:
                continue
            seen.add((pn, cd))
            init_name = plan_code_to_name.get((budget_norm, pn, cd), "")
            suppliers: list = []
            combo_total = 0.0
            for sup_num, txns in sup_dict.items():
                sup_total = sum(_to_f(t["amount"]) for t in txns)
                combo_total += sup_total
                suppliers.append({
                    "supplier_number": sup_num,
                    "supplier_name":   txns[0]["supplier_name"],
                    "total_amount":    int(round(sup_total)),
                    "transactions": [
                        {"date": t["date"], "invoice": t["invoice"],
                         "description": t["description"], "amount": t["amount"]}
                        for t in txns
                    ],
                })
            suppliers.sort(key=lambda x: -x["total_amount"])
            items.append({
                "plan_number":     pn,
                "code":            cd,
                "initiative_name": init_name,
                "total_amount":    int(round(combo_total)),
                "suppliers":       suppliers,
            })
        items.sort(key=lambda x: (int(x["code"]) if x["code"].isdigit() else 0, x["plan_number"]))
        if items:
            bdict["yozma_breakdown"] = items


_NIHUL_CODE_TIKKON   = "104"
_NIHUL_CODE_BEINAYIM = "67"


def _build_and_attach_nihul_breakdown(budgets_list: list, results_clean: list) -> None:
    """Add nihul_breakdown to each budget dict — supplier totals per (code 104 or 67)."""
    import re as _re
    from collections import defaultdict as _dd
    from zihuy_core import normalize_budget_name as _nb_nihul

    if not results_clean or not budgets_list:
        return

    def _to_f(v):
        if v is None: return 0.0
        try: return float(str(v).replace(",", "").strip())
        except Exception: return 0.0

    # Group all doch rows by (norm_budget, code) → sup_num → transaction list
    code_sups: dict = _dd(lambda: _dd(list))
    for r in results_clean:
        if not r.get("budget"):
            continue
        orig = r.get("orig", ())
        if len(orig) < 12:
            continue
        nb  = _nb_nihul(str(r["budget"]))
        cd  = str(orig[1] or "").strip()
        sup_num  = str(r.get("sup_num") or "").strip()
        sup_raw  = str(orig[6] or "")
        sup_name = _re.sub(r'^\s*\d+\s*-\s*', '', sup_raw).strip()
        code_sups[(nb, cd)][sup_num].append({
            "date":          _normalize_date(str(orig[5])) if orig[5] is not None else "",
            "invoice":       str(orig[4] or "").strip(),
            "description":   str(orig[10] or "").strip(),
            "amount":        int(round(_to_f(orig[11]))),
            "supplier_name": sup_name,
        })

    for bdict in budgets_list:
        budget_norm = bdict.get("name", "")
        # Determine stage to pick the correct nihul code
        stage = None
        for r in results_clean:
            if _nb_nihul(r.get("budget", "")) == budget_norm and r.get("stage"):
                stage = r["stage"]
                break
        nihul_code = _NIHUL_CODE_TIKKON if stage == "תיכון" else _NIHUL_CODE_BEINAYIM

        sup_dict = code_sups.get((budget_norm, nihul_code), {})
        if not sup_dict:
            continue

        suppliers: list = []
        total = 0.0
        for sup_num, txns in sup_dict.items():
            sup_total = sum(_to_f(t["amount"]) for t in txns)
            total += sup_total
            suppliers.append({
                "supplier_number": sup_num,
                "supplier_name":   txns[0]["supplier_name"],
                "total_amount":    int(round(sup_total)),
                "transactions": [
                    {"date": t["date"], "invoice": t["invoice"],
                     "description": t["description"], "amount": t["amount"]}
                    for t in txns
                ],
            })
        suppliers.sort(key=lambda x: -x["total_amount"])
        bdict["nihul_breakdown"] = [{
            "code":            nihul_code,
            "initiative_name": "ניהול ותפעול",
            "plan_number":     "",
            "total_amount":    int(round(total)),
            "suppliers":       suppliers,
        }]


def _compute_multi_budget_tikhnun(
    tikhnun_result: dict,
    plan_fpath: str,
    doch_paths: list,
    finance_paths: list | None = None,
    finance_type: str | None = None,
) -> tuple[dict, list | None]:
    """
    Extract per-budget metrics from a planning file.
    When rows remain unidentified after zihuy + finance cross-reference,
    sets pending_identification=True and stores _zihuy_ctx for the classify endpoint.
    Returns (tikhnun_result, results_clean) — results_clean is non-None only when
    all rows are identified with certainty.
    """
    try:
        from openpyxl import load_workbook as _lw
        from zihuy_core import identify as _zihuy_identify, normalize_budget_name as _norm_bname

        wb = _lw(plan_fpath, read_only=True)
        hakol_ws_name = perut_ws_name = None
        for sh in wb.sheetnames:
            sh_s = sh.strip()
            if sh_s == "הכל":
                hakol_ws_name = sh
            elif sh_s == "פירוט המענים":
                perut_ws_name = sh
        if not hakol_ws_name or not perut_ws_name:
            for sh in wb.sheetnames:
                if sh == hakol_ws_name or sh == perut_ws_name:
                    continue
                peek = list(wb[sh].iter_rows(min_row=1, max_row=1, values_only=True))
                if not peek:
                    continue
                hdr = peek[0]
                if not hakol_ws_name and len(hdr) > 14 and hdr[10] == "השתתפות רשות/ בעלות":
                    hakol_ws_name = sh
                elif not perut_ws_name and len(hdr) > 15 and hdr[14] == "עלות מענה כוללת":
                    perut_ws_name = sh

        if not hakol_ws_name or not perut_ws_name:
            wb.close()
            return tikhnun_result, None

        hakol_rows = [list(r) for r in wb[hakol_ws_name].iter_rows(values_only=True)]
        perut_rows = [list(r) for r in wb[perut_ws_name].iter_rows(values_only=True)]
        wb.close()

        def _to_f(v):
            if v is None:
                return 0.0
            try:
                return float(str(v).replace(",", "").strip())
            except Exception:
                return 0.0

        def _fmt_pct(v):
            try:
                s = str(v).replace("%", "").strip()
                f = float(s)
                return f / 100 if f > 1 else f
            except Exception:
                return 0.0

        # Extract budgets from הכל
        budgets_raw = []
        seen_names: set = set()
        for row in hakol_rows[1:]:
            name = row[0] if row else None
            if not name:
                continue
            name_s = str(name).strip()
            if not name_s or "סה''כ" in name_s:
                continue
            if name_s in seen_names:
                continue
            seen_names.add(name_s)
            h_num = _to_f(row[7]) if len(row) > 7 else 0.0
            if h_num <= 0:
                continue
            budgets_raw.append({
                "raw_name":   name_s,
                "norm_name":  _norm_bname(name_s),
                "H":          h_num,
                "L":          _to_f(row[11]) if len(row) > 11 else 0.0,
                "S":          _to_f(row[18]) if len(row) > 18 else 0.0,
                "T":          _fmt_pct(row[19]) if len(row) > 19 else 0.0,
                "sum_chayav": 0.0,
                "nikuy":      None,
                "pct_tanuz":  None,
            })

        if len(budgets_raw) == 0:
            return tikhnun_result, None

        # sum_chayav per budget from פירוט המענים
        name_to_idx    = {b["raw_name"]: i for i, b in enumerate(budgets_raw)}
        norm_to_idx    = {b["norm_name"]: i for i, b in enumerate(budgets_raw)}
        seen_plan_keys: set = set()
        for row in perut_rows[1:]:
            raw_bname = str(row[0]).strip() if row[0] else ""
            col_r     = row[17] if len(row) > 17 else None
            if not col_r:
                continue
            idx = name_to_idx.get(raw_bname)
            if idx is None:
                idx = norm_to_idx.get(_norm_bname(raw_bname))
            if idx is None:
                continue
            _sk13 = row[13] if len(row) > 13 else None
            if _sk13 is not None:
                key = (
                    str(row[9] or "").strip()  if len(row) > 9  else "",
                    str(row[10] or "").strip() if len(row) > 10 else "",
                    str(row[11] or "").strip() if len(row) > 11 else "",
                    str(row[12] or "").strip() if len(row) > 12 else "",
                    _sk13,
                    str(row[17] or "").strip(),
                )
            else:
                key = (tuple(str(x).strip() if x else "" for x in row[:10]),
                       str(row[17] or "").strip())
            if key in seen_plan_keys:
                continue
            seen_plan_keys.add(key)
            budgets_raw[idx]["sum_chayav"] += _to_f(_sk13) if _sk13 is not None else (_to_f(row[15]) if len(row) > 15 else 0.0)

        # No doch files — build budgets list only (no nikuy/pct_tanuz)
        if not doch_paths:
            _finalize_tikhnun_metrics(tikhnun_result, [], budgets_raw, perut_rows)
            if tikhnun_result.get("budgets"):
                try:
                    _compute_per_budget_yozma(tikhnun_result["budgets"], perut_rows, None)
                except Exception as _ye:
                    logger.warning("per-budget yozma (no-doch) failed: %s", _ye)
                try:
                    _propagate_meshuyakh_to_root(tikhnun_result)
                except Exception as _pe:
                    logger.warning("meshuyakh propagation (no-doch) failed: %s", _pe)
            return tikhnun_result, None

        # Run zihuy identification
        plan_budget_norms = [b["norm_name"] for b in budgets_raw]
        results_clean: list = []
        try:
            doch_str = [str(p) for p in doch_paths]
            results_clean, _warn, _miss, _all = _zihuy_identify(doch_str, [plan_fpath])
        except Exception as exc:
            logger.error("Zihuy identification failed: %s", exc)
            _finalize_tikhnun_metrics(tikhnun_result, [], budgets_raw, perut_rows)
            return tikhnun_result, None

        # Finance cross-reference for unidentified rows
        covered_budgets: list[str] = []
        invalid_reason: str | None = None
        unidentified = [r for r in results_clean if not r.get("budget")]
        ichud_map, _fin_warnings, covered_budgets, invalid_reason = _build_finance_ichud_budget_map(
            list(finance_paths), finance_type, plan_budget_norms
        )
        if unidentified and ichud_map:
            for r in results_clean:
                if not r.get("budget"):
                    mapped = ichud_map.get(r.get("union_key", ""))
                    if mapped:
                        r["budget"] = mapped
                        try:
                            rc = int(str(r["orig"][1]).strip())
                        except Exception:
                            rc = 0
                        if rc in TIKKON_ONLY:
                            r["stage"] = "תיכון"
                        elif rc in BEINAYIM_ONLY:
                            r["stage"] = "חטיבת ביניים"
                        else:
                            r["stage"] = ""
            unidentified = [r for r in results_clean if not r.get("budget")]

        # Still unidentified — block tikhnun and store context for classify endpoint
        if unidentified:
            covered_set = set(covered_budgets)
            missing_budgets = [b for b in plan_budget_norms if b not in covered_set]
            needs_upload = bool(missing_budgets) or bool(invalid_reason)
            tikhnun_result["pending_identification"] = True
            tikhnun_result["needs_finance_upload"] = needs_upload
            tikhnun_result["missing_budgets"] = missing_budgets
            tikhnun_result["covered_budgets"] = covered_budgets
            tikhnun_result["finance_invalid_reason"] = invalid_reason
            tikhnun_result["unidentified_rows"] = [
                {
                    "union_key":   r.get("union_key", ""),
                    "invoice":     str(r["orig"][4]).strip()  if r.get("orig") and r["orig"][4]  is not None else "",
                    "date":        str(r["orig"][5]).strip()  if r.get("orig") and r["orig"][5]  is not None else "",
                    "supplier":    str(r["orig"][6]).strip()  if r.get("orig") and r["orig"][6]  is not None else "",
                    "amount":      str(r["orig"][11]).strip() if r.get("orig") and r["orig"][11] is not None else "",
                    "report_code": str(r["orig"][1]).strip()  if r.get("orig") and r["orig"][1]  is not None else "",
                }
                for r in unidentified
            ]
            tikhnun_result["available_budgets"] = plan_budget_norms
            tikhnun_result["_zihuy_ctx"] = {
                "results_clean": results_clean,
                "budgets_raw":   budgets_raw,
                "perut_rows":    perut_rows,
                "gefen_paths":   [str(p) for p in (doch_paths or [])],
                "finance_paths": [str(p) for p in (finance_paths or [])],
                "finance_type":  finance_type,
            }
            return tikhnun_result, None

        # All identified — compute full metrics
        _finalize_tikhnun_metrics(tikhnun_result, results_clean, budgets_raw, perut_rows)
        if tikhnun_result.get("budgets"):
            try:
                _compute_per_budget_yozma(tikhnun_result["budgets"], perut_rows, results_clean)
            except Exception as _ye:
                logger.warning("per-budget yozma failed: %s", _ye)
            try:
                _propagate_meshuyakh_to_root(tikhnun_result)
            except Exception as _pe:
                logger.warning("meshuyakh propagation failed: %s", _pe)
            try:
                _build_and_attach_yozma_breakdown(tikhnun_result["budgets"], perut_rows, results_clean)
            except Exception as _ybe:
                logger.warning("per-budget yozma breakdown failed: %s", _ybe)
            try:
                _build_and_attach_nihul_breakdown(tikhnun_result["budgets"], results_clean)
            except Exception as _nhe:
                logger.warning("per-budget nihul breakdown failed: %s", _nhe)
        try:
            # Always attach, even when empty — see comment in the "classify" branch above.
            tikhnun_result["per_budget_rejected"] = _build_rejected_from_results_clean(results_clean)
        except Exception as rej_exc:
            logger.warning("per_budget_rejected build failed: %s", rej_exc)
        try:
            tikhnun_result["per_budget_no_pdf"] = _build_no_pdf_from_results_clean(results_clean)
        except Exception as nopdf_exc:
            logger.warning("per_budget_no_pdf build failed: %s", nopdf_exc)
        return tikhnun_result, results_clean

    except Exception as exc:
        logger.error("Multi-budget tikhnun computation failed: %s", exc)
        return tikhnun_result, None


def _upload_excel_to_storage(run_id: str, excel_path: str) -> str | None:
    """Upload generated Excel file to Supabase Storage. Returns storage key or None on failure."""
    key = f"excel/{run_id}/hashvaa-gefen-ksafim.xlsx"
    try:
        db = get_admin_client()
        db.storage.from_("check-files").upload(key, Path(excel_path).read_bytes())
        logger.info("Excel uploaded to Storage: %s", key)
        return key
    except Exception as exc:
        logger.warning("Excel Storage upload failed for %s: %s", run_id, exc)
        return None


def _upload_files_to_storage(paths: list[Path], run_id: str) -> list[dict]:
    """Upload check files to Supabase Storage.
    Returns list of {"path": storage_path, "name": original_filename} dicts.
    Uses index-based ASCII storage keys to avoid Hebrew filename issues.
    Non-fatal on failure.
    """
    stored: list[dict] = []
    logger.info("Storage upload: starting for run %s, %d file(s): %s", run_id, len(paths), [p.name for p in paths])
    try:
        db = get_admin_client()
        for i, p in enumerate(paths):
            safe_key = f"file_{i:02d}{p.suffix.lower()}"
            storage_path = f"{run_id}/{safe_key}"
            try:
                db.storage.from_("check-files").upload(storage_path, p.read_bytes())
                stored.append({"path": storage_path, "name": p.name})
                logger.info("Storage upload: OK → %s (%s)", storage_path, p.name)
            except Exception as exc:
                logger.warning("Storage upload FAILED for %s (%s): %s", storage_path, p.name, exc)
    except Exception as exc:
        logger.warning("Storage upload skipped (client error): %s", exc)
    logger.info("Storage upload: done, %d/%d succeeded", len(stored), len(paths))
    return stored


def _process(run_id: str, paths: list[Path], run_dir: Path, user_id: str = "", school_id: str | None = None, gefen_account_id: str | None = None, update_log_id: str | None = None, academic_year: str | None = None) -> None:
    run_data: dict = {"status": "processing"}
    try:
        # Upload source files to Supabase Storage for future "add file" retrieval
        stored_file_paths = _upload_files_to_storage(paths, run_id)
        if stored_file_paths:
            run_data["stored_file_paths"] = stored_file_paths

        # Fetch division_type for authoritative stage resolution in per-combo reconciliation
        school_stage: str | None = _get_school_stage(gefen_account_id)

        gefen_paths, finance_paths, finance_type, tikhnun_paths = _classify_files(paths)
        finance_path = finance_paths[0] if finance_paths else None

        # ── Tikhnun-only run (no gefen doch) ─────────────────────────────────
        if not gefen_paths and tikhnun_paths:
            if len(tikhnun_paths) == 2:
                td0 = load_tikhnun(str(tikhnun_paths[0]))
                td1 = load_tikhnun(str(tikhnun_paths[1]))
                td0["filename"] = tikhnun_paths[0].name
                td1["filename"] = tikhnun_paths[1].name
                tikkon_data, beinayim_data = _assign_tikhnun_pair(td0, td1)
                tikkon_fpath   = str(tikhnun_paths[0]) if td0 is tikkon_data   else str(tikhnun_paths[1])
                beinayim_fpath = str(tikhnun_paths[0]) if td0 is beinayim_data else str(tikhnun_paths[1])
                tikkon_result   = build_tikhnun_result(tikkon_data)   if tikkon_data   else None
                beinayim_result = build_tikhnun_result(beinayim_data) if beinayim_data else None
                if tikkon_result:
                    tikkon_result, _   = _compute_multi_budget_tikhnun(tikkon_result,   tikkon_fpath,   [])
                if beinayim_result:
                    beinayim_result, _ = _compute_multi_budget_tikhnun(beinayim_result, beinayim_fpath, [])
                run_data = {
                    "status": "done",
                    "tikhnun_only": True,
                    "tikhnun": tikkon_result or beinayim_result,
                    "tikhnun_tikkon": tikkon_result,
                    "tikhnun_beinayim": beinayim_result,
                    "stored_file_paths": stored_file_paths or None,
                    "_school_ctx": {"user_id": user_id, "school_id": school_id, "gefen_account_id": gefen_account_id, "update_log_id": update_log_id, "academic_year": academic_year},
                }
            else:
                tikhnun_data = load_tikhnun(str(tikhnun_paths[0]))
                tikhnun_data["filename"] = tikhnun_paths[0].name
                tikhnun_result_only = build_tikhnun_result(tikhnun_data)
                tikhnun_result_only, _ = _compute_multi_budget_tikhnun(tikhnun_result_only, str(tikhnun_paths[0]), [])
                run_data = {
                    "status": "done",
                    "tikhnun_only": True,
                    "tikhnun": tikhnun_result_only,
                    "stored_file_paths": stored_file_paths or None,
                    "_school_ctx": {"user_id": user_id, "school_id": school_id, "gefen_account_id": gefen_account_id, "update_log_id": update_log_id, "academic_year": academic_year},
                }
            _update_run(run_id, run_data)
            return

        # ── Normal run: gefen files present ──────────────────────────────────
        df_gefen, gefen_file_stats, gefen_merge_note = _load_gefen_files(gefen_paths)
        in_gefen_rejected, in_gefen_no_pdf = _extract_gefen_only_results(df_gefen)
        excel_path = str(run_dir / "hashvaa-gefen-ksafim.xlsx")

        # Process tikhnun if present (cross-reference with matching-division gefen doch)
        tikhnun_result        = None
        tikhnun_tikkon_result  = None
        tikhnun_beinayim_result = None
        results_clean_for_recon: list | None = None

        if len(tikhnun_paths) == 1:
            try:
                tikhnun_data = load_tikhnun(str(tikhnun_paths[0]))
                tikhnun_data["filename"] = tikhnun_paths[0].name
                tikhnun_data = cross_reference_doch(tikhnun_data, str(gefen_paths[0]))
                tikhnun_result = build_tikhnun_result(tikhnun_data)
                tikhnun_result, results_clean_for_recon = _compute_multi_budget_tikhnun(
                    tikhnun_result, str(tikhnun_paths[0]), gefen_paths,
                    finance_paths=finance_paths, finance_type=finance_type,
                )
            except Exception as exc:
                logger.error("Tikhnun processing error for run %s: %s", run_id, exc)
                tikhnun_result = {"error": str(exc)}
        elif len(tikhnun_paths) == 2:
            try:
                td0 = load_tikhnun(str(tikhnun_paths[0]))
                td1 = load_tikhnun(str(tikhnun_paths[1]))
                td0["filename"] = tikhnun_paths[0].name
                td1["filename"] = tikhnun_paths[1].name
                tikkon_data, beinayim_data = _assign_tikhnun_pair(td0, td1)
                tikkon_fpath   = str(tikhnun_paths[0]) if td0 is tikkon_data   else str(tikhnun_paths[1])
                beinayim_fpath = str(tikhnun_paths[0]) if td0 is beinayim_data else str(tikhnun_paths[1])
                rc_t: list | None = None
                rc_b: list | None = None
                if tikkon_data:
                    tikkon_gpath = _find_gefen_path_for_division(gefen_paths, gefen_file_stats, "tikkon")
                    if tikkon_gpath:
                        tikkon_data = cross_reference_doch(tikkon_data, str(tikkon_gpath))
                    tikhnun_tikkon_result = build_tikhnun_result(tikkon_data)
                    tikhnun_tikkon_result, rc_t = _compute_multi_budget_tikhnun(
                        tikhnun_tikkon_result, tikkon_fpath,
                        [tikkon_gpath] if tikkon_gpath else gefen_paths,
                        finance_paths=finance_paths, finance_type=finance_type,
                    )
                if beinayim_data:
                    beinayim_gpath = _find_gefen_path_for_division(gefen_paths, gefen_file_stats, "beinayim")
                    if beinayim_gpath:
                        beinayim_data = cross_reference_doch(beinayim_data, str(beinayim_gpath))
                    tikhnun_beinayim_result = build_tikhnun_result(beinayim_data)
                    tikhnun_beinayim_result, rc_b = _compute_multi_budget_tikhnun(
                        tikhnun_beinayim_result, beinayim_fpath,
                        [beinayim_gpath] if beinayim_gpath else gefen_paths,
                        finance_paths=finance_paths, finance_type=finance_type,
                    )
                results_clean_for_recon = (rc_t or []) + (rc_b or []) or None
                tikhnun_result = tikhnun_tikkon_result or tikhnun_beinayim_result
            except Exception as exc:
                logger.error("Tikhnun processing error for run %s: %s", run_id, exc)
                tikhnun_result = {"error": str(exc)}

        # Gefen-only run (no finance) — skip reconciliation
        if finance_path is None:
            export(
                _for_excel(df_gefen),
                None,
                None,
                None,
                excel_path,
                finance_label=None,
                in_gefen_rejected=_for_excel(in_gefen_rejected),
                in_gefen_no_pdf=_for_excel(in_gefen_no_pdf),
                gefen_only=True,
            )
            excel_storage_key = _upload_excel_to_storage(run_id, excel_path)
            run_data = {
                "status": "saving",
                "gefen_only": True,
                "finance_type": None,
                "tikhnun": tikhnun_result,
                "tikhnun_tikkon": tikhnun_tikkon_result,
                "tikhnun_beinayim": tikhnun_beinayim_result,
                "tikhnun_filenames": [p.name for p in tikhnun_paths],
                "stored_file_paths": stored_file_paths or None,
                "excel_storage_key": excel_storage_key,
                "summary": {
                    "gefen_rows": len(df_gefen),
                    "in_gefen_rejected": len(in_gefen_rejected),
                    "in_gefen_no_pdf": len(in_gefen_no_pdf),
                    "division": _detect_gefen_division(df_gefen),
                    "gefen_files": gefen_file_stats,
                    "gefen_merge_note": gefen_merge_note,
                },
                "rows_gefen_rejected": _build_display_records(in_gefen_rejected, _GEFEN_REJECTED_COL_MAP),
                "rows_gefen_no_pdf": _build_display_records(in_gefen_no_pdf, _GEFEN_COL_MAP),
                "_school_ctx": {"user_id": user_id, "school_id": school_id, "gefen_account_id": gefen_account_id, "update_log_id": update_log_id, "academic_year": academic_year},
            }
            if school_id and not _any_tikhnun_pending(run_data):
                _save_check_log(run_id, user_id, school_id, gefen_account_id, update_log_id, run_data=run_data, academic_year=academic_year)
            run_data["status"] = "done"
            _update_run(run_id, run_data)
            return

        # Load raw finance df — kesafim2000 still has English column names here
        # so that reconciler._filter_by_division can access "report_code"
        df_finance_raw, finance_label, finance_file_stats = _load_finance_raw(finance_paths, finance_type)
        in_finance_not_gefen, in_gefen_not_finance, division, finance_rows_checked = reconcile(df_gefen, df_finance_raw)

        # Rename report_code → קוד דיווח for all finance types after reconciliation.
        # Kesafim also renames its other English columns to Hebrew display names.
        if finance_type == "kesafim2000":
            df_finance = df_finance_raw.rename(columns=_KESAFIM_RENAME)
            in_finance_not_gefen = in_finance_not_gefen.rename(columns=_KESAFIM_RENAME)
        else:
            _payscool_rename = {"report_code": "קוד דיווח"}
            df_finance = df_finance_raw.rename(columns=_payscool_rename)
            in_finance_not_gefen = in_finance_not_gefen.rename(columns=_payscool_rename)

        export(
            _for_excel(df_gefen),
            _for_excel(df_finance),
            _for_excel(in_finance_not_gefen),
            _for_excel(in_gefen_not_finance),
            excel_path,
            finance_label=finance_label,
            in_gefen_rejected=_for_excel(in_gefen_rejected),
            in_gefen_no_pdf=_for_excel(in_gefen_no_pdf),
        )

        if finance_type == "kesafim2000":
            finance_col_map = _KESAFIM_COL_MAP
        elif finance_type == "schoolcash":
            finance_col_map = _SCHOOLCASH_COL_MAP
        else:
            finance_col_map = _PAYSCOOL_COL_MAP

        per_combo_results: dict | None = None
        if results_clean_for_recon:
            try:
                per_combo_results = _run_per_combo_reconciliation(
                    df_gefen, finance_paths, finance_type,
                    results_clean_for_recon, finance_col_map,
                    school_stage=school_stage,
                )
            except Exception as exc:
                logger.warning("per_combo_results failed in _process: %s", exc)

        excel_storage_key = _upload_excel_to_storage(run_id, excel_path)
        run_data = {
            "status": "saving",
            "gefen_only": False,
            "finance_type": finance_type,
            "tikhnun": tikhnun_result,
            "tikhnun_tikkon": tikhnun_tikkon_result,
            "tikhnun_beinayim": tikhnun_beinayim_result,
            "tikhnun_filenames": [p.name for p in tikhnun_paths],
            "stored_file_paths": stored_file_paths or None,
            "excel_storage_key": excel_storage_key,
            "summary": {
                "gefen_rows": len(df_gefen),
                "finance_rows_total": len(df_finance_raw),
                "finance_rows_checked": finance_rows_checked,
                "in_finance_not_gefen": len(in_finance_not_gefen),
                "in_gefen_not_finance": len(in_gefen_not_finance),
                "in_gefen_rejected": len(in_gefen_rejected),
                "in_gefen_no_pdf": len(in_gefen_no_pdf),
                "division": division,
                "gefen_files": gefen_file_stats,
                "gefen_merge_note": gefen_merge_note,
                "finance_file": {
                    **finance_file_stats,
                    "rows_total": len(df_finance_raw),
                    "rows_checked": finance_rows_checked,
                },
            },
            "rows_finance_not_gefen": _build_display_records(in_finance_not_gefen, finance_col_map),
            "rows_gefen_not_finance": _build_display_records(in_gefen_not_finance, _GEFEN_COL_MAP),
            "rows_gefen_rejected": _build_display_records(in_gefen_rejected, _GEFEN_REJECTED_COL_MAP),
            "rows_gefen_no_pdf": _build_display_records(in_gefen_no_pdf, _GEFEN_COL_MAP),
            "per_combo_results": per_combo_results,
            "_school_ctx": {"user_id": user_id, "school_id": school_id, "gefen_account_id": gefen_account_id, "update_log_id": update_log_id, "academic_year": academic_year},
        }
        if school_id and not _any_tikhnun_pending(run_data):
            _save_check_log(run_id, user_id, school_id, gefen_account_id, update_log_id, run_data=run_data, academic_year=academic_year)
        run_data["status"] = "done"
        _update_run(run_id, run_data)

    except UnicodeDecodeError as exc:
        tb = traceback.format_exc()
        logger.error("Run %s encoding error:\n%s", run_id, tb)
        _update_run(run_id, {
            "status": "error",
            "user_message": (
                "המערכת לא הצליחה לעבד את קובץ כספים2000. "
                "במידה והקובץ אינו הקובץ הגולמי כפי שהורד מהמערכת, יש לנסות מחדש עם הקובץ הגולמי."
            ),
            "error": str(exc),
        })
    except ValueError as exc:
        logger.error("Run %s validation error: %s", run_id, exc)
        _update_run(run_id, {"status": "error", "error": str(exc)})
    except Exception as exc:
        tb = traceback.format_exc()
        logger.error("Run %s unexpected error:\n%s", run_id, tb)
        _update_run(run_id, {"status": "error", "error": f"שגיאה פנימית: {exc}", "traceback": tb})
    finally:
        shutil.rmtree(run_dir, ignore_errors=True)


def _extract_gefen_only_results(df_gefen: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Extract rejected and no-PDF rows from a Gefen dataframe.

    Looks up columns by name (not by positional index) so the function is
    robust against files that have extra/blank columns or a different column
    count than the standard export.
    """
    # ── Rejected rows ────────────────────────────────────────────────────────
    if _GEFEN_STATUS_COL in df_gefen.columns:
        in_gefen_rejected = df_gefen[
            df_gefen[_GEFEN_STATUS_COL].astype(str).str.startswith("נדחה:")
        ].copy()
        in_gefen_rejected["סיבת הדחייה"] = (
            in_gefen_rejected[_GEFEN_STATUS_COL]
            .astype(str)
            .str.replace(r"^נדחה:\s*", "", regex=True)
        )
    else:
        in_gefen_rejected = df_gefen.iloc[0:0].copy()
        in_gefen_rejected["סיבת הדחייה"] = pd.Series([], dtype=str)

    # ── No-PDF rows ──────────────────────────────────────────────────────────
    if _GEFEN_PDF_COL in df_gefen.columns:
        data_cols = [c for c in _GEFEN_DATA_COLS if c in df_gefen.columns]
        if data_cols:
            has_data = (
                df_gefen[data_cols].notna().all(axis=1)
                & df_gefen[data_cols]
                .apply(lambda col: col.astype(str).str.strip() != "")
                .all(axis=1)
            )
        else:
            has_data = pd.Series(True, index=df_gefen.index)
        no_pdf = df_gefen[_GEFEN_PDF_COL].astype(str).str.strip() == "לא"
        in_gefen_no_pdf = df_gefen[has_data & no_pdf]
    else:
        in_gefen_no_pdf = df_gefen.iloc[0:0]

    return in_gefen_rejected, in_gefen_no_pdf


def _build_rejected_from_results_clean(results_clean: list) -> dict[str, list[dict]]:
    """Build per-budget rejected display records from identified zihuy results.

    Returns {normalized_budget_name: [display_record, ...]} where each record
    matches the 6-field structure of _GEFEN_REJECTED_COL_MAP.
    Only includes rows where orig[12] (status) starts with 'נדחה:'.
    """
    from zihuy_core import normalize_budget_name as _norm_bname
    per_budget: dict[str, list[dict]] = {}
    for r in results_clean:
        orig = r.get("orig")
        if not orig or len(orig) < 13:
            continue
        col_stat = str(orig[12]) if orig[12] is not None else ""
        if not col_stat.startswith("נדחה:"):
            continue
        budget_raw = r.get("budget") or ""
        budget_norm = _norm_bname(budget_raw) if budget_raw else budget_raw
        record = {
            "קוד דיווח":   str(orig[1]) if orig[1] is not None else "",
            "שם ספק":      str(orig[6]) if orig[6] is not None else "",
            "מספר אסמכתה": normalize_amount(orig[4]),
            "תאריך":       _normalize_date(str(orig[5])) if orig[5] is not None else "",
            "סכום":        _format_display_amount(str(orig[11])) if orig[11] is not None else "",
            "סיבת הדחייה": col_stat[len("נדחה:"):].strip(),
        }
        per_budget.setdefault(budget_norm, []).append(record)
    return per_budget


def _build_no_pdf_from_results_clean(results_clean: list) -> dict[str, list[dict]]:
    """Build per-budget no-PDF display records from identified zihuy results.

    Returns {normalized_budget_name: [display_record, ...]} where each record
    matches the 6-field structure of UNIFIED_COLS (frontend).
    Only includes rows where orig[13] (file_exists / 'האם קיים קובץ') == 'לא'.
    """
    from zihuy_core import normalize_budget_name as _norm_bname
    per_budget: dict[str, list[dict]] = {}
    for r in results_clean:
        orig = r.get("orig")
        if not orig or len(orig) < 14:
            continue
        col_file = str(orig[13]).strip() if orig[13] is not None else ""
        if col_file != "לא":
            continue
        if not normalize_amount(orig[11]):
            continue
        budget_raw = r.get("budget") or ""
        budget_norm = _norm_bname(budget_raw) if budget_raw else budget_raw
        record = {
            "קוד דיווח":   str(orig[1]) if orig[1] is not None else "",
            "שם ספק":      str(orig[6]) if orig[6] is not None else "",
            "מספר אסמכתה": normalize_amount(orig[4]),
            "תאריך":       _normalize_date(str(orig[5])) if orig[5] is not None else "",
            "סכום":        _format_display_amount(str(orig[11])) if orig[11] is not None else "",
            "תיאור":       str(orig[7]) if orig[7] is not None else "",
        }
        per_budget.setdefault(budget_norm, []).append(record)
    return per_budget


def _classify_files(paths: list[Path]) -> tuple[list[Path], list[Path], str | None, list[Path]]:
    gefen: list[Path] = []
    finance_list: list[Path] = []
    finance_type: str | None = None
    tikhnun_paths: list[Path] = []

    for p in paths:
        ftype = identify_file(str(p))
        if ftype == "gefen":
            gefen.append(p)
        elif ftype in ("kesafim2000", "payscool", "schoolcash"):
            if finance_list:
                # Multiple kesafim2000 and schoolcash files are allowed and will be merged; payscool must be single
                if ftype != finance_type or ftype not in ("kesafim2000", "schoolcash"):
                    raise ValueError("התקבלו קבצי כספים מסוגים שונים. אנא העלה קבצי כספים מאותו סוג בלבד.")
            finance_list.append(p)
            finance_type = ftype
        elif ftype == "tikhnun":
            tikhnun_paths.append(p)
            if len(tikhnun_paths) > 2:
                raise ValueError("התקבלו יותר משני קבצי תכנון. אנא העלה עד שני קבצי תכנון.")
        else:
            raise ValueError(
                f"הקובץ '{p.name}' אינו בצורתו הגולמית כפי שהורד מהמערכת. "
                "אנא העלה את הקבצים בצורתם הגולמית כפי שהורדו מהמערכות השונות, ללא שינויים."
            )

    # tikhnun only (with or without finance) — treat as tikhnun-only
    if tikhnun_paths and not gefen:
        return [], [], None, tikhnun_paths

    if not gefen and finance_list:
        raise ValueError("לא ניתן לבצע את הבדיקה עם קובץ מתוכנת הכספים בלבד.")
    if not gefen:
        raise ValueError("לא קיבלתי קבצים מזוהים.")
    if len(gefen) > 5:
        raise ValueError("התקבלו יותר מחמישה קבצי גפן. אנא העלה עד חמישה קבצי גפן.")

    return gefen, finance_list, finance_type, tikhnun_paths


def _assign_tikhnun_pair(td0: dict, td1: dict) -> tuple[dict | None, dict | None]:
    """Assign two loaded tikhnun dicts to (tikkon, beinayim) based on school_stage."""
    tikkon, beinayim = None, None
    for td in (td0, td1):
        if td.get("school_stage") == "תיכון":
            tikkon = td
        else:
            beinayim = td
    return tikkon, beinayim


def _find_gefen_path_for_division(
    gefen_paths: list[Path], stats: list[dict], division: str
) -> Path | None:
    """Return the gefen path whose division matches the requested division, or None."""
    for i, stat in enumerate(stats):
        if stat["division"] == division or stat["division"] == "both":
            return gefen_paths[i]
    return None


def _detect_gefen_division(df: pd.DataFrame) -> str:
    codes = set(df["report_code"].dropna().astype(int).tolist())
    has_tikkon   = bool(codes & set(TIKKON_ONLY))
    has_beinayim = bool(codes & set(BEINAYIM_ONLY))
    if has_tikkon and not has_beinayim:
        return "tikkon"
    if has_beinayim and not has_tikkon:
        return "beinayim"
    return "both"


def _load_gefen_files(paths: list[Path]) -> tuple[pd.DataFrame, list[dict], dict | None]:
    loaded    = [load_gefen(str(p)) for p in paths]
    dfs       = [df for df, _ in loaded]
    dedup_flags = [was_dedup for _, was_dedup in loaded]

    per_file_stats = [
        {
            "filename": p.name,
            "division": _detect_gefen_division(df),
            "rows": len(df),
            "was_deduplicated": was_dedup,
        }
        for p, df, was_dedup in zip(paths, dfs, dedup_flags)
    ]

    if len(dfs) == 1:
        return dfs[0], per_file_stats, None

    if len(dfs) == 2:
        # Two gefen files — compute overlap, merge with dedup
        set0    = set(dfs[0]["ichud"])
        set1    = set(dfs[1]["ichud"])
        overlap = len(set0 & set1)

        if set0 >= set1:
            merged = dfs[0]
        elif set1 >= set0:
            merged = dfs[1]
        else:
            merged = (
                pd.concat([dfs[0], dfs[1]], ignore_index=True)
                .drop_duplicates(subset=["ichud"])
                .reset_index(drop=True)
            )

        merge_note = {
            "overlap": overlap,
            "unique": len(set0 | set1),
            "file0_rows": len(dfs[0]),
            "file1_rows": len(dfs[1]),
        }
        return merged, per_file_stats, merge_note

    # Three or more gefen files — merge all with dedup
    merged = (
        pd.concat(dfs, ignore_index=True)
        .drop_duplicates(subset=["ichud"])
        .reset_index(drop=True)
    )
    merge_note = {"unique": len(merged)}
    return merged, per_file_stats, merge_note


def _load_finance_raw(paths: list[Path], ftype: str) -> tuple[pd.DataFrame, str, dict]:
    """Load finance file(s) without renaming columns — reconciler needs 'report_code' intact.
    Multiple paths are supported only for kesafim2000; the DataFrames are concatenated and deduplicated.
    """
    if ftype == "kesafim2000":
        dfs = [load_kesafim(str(p)) for p in paths]
        df = pd.concat(dfs, ignore_index=True).drop_duplicates()
        filename = ", ".join(p.name for p in paths)
        stats = {"filename": filename, "software": "כספים2000", "cancelled_rows": None}
        return df, "כספים", stats
    if ftype == "schoolcash":
        dfs = [load_schoolcash(str(p)) for p in paths]
        df = pd.concat(dfs, ignore_index=True).drop_duplicates()
        filename = ", ".join(p.name for p in paths)
        stats = {"filename": filename, "software": "סקולקאש", "cancelled_rows": None}
        return df, "סקולקאש", stats
    df, cancelled = load_payscool(str(paths[0]))
    paren_normalized = 0
    if "מספר חשבונית" in df.columns:
        paren_normalized = int(df["מספר חשבונית"].apply(
            lambda v: canonical_payscool_invoice(v) != ("" if v is None else str(v).strip())
        ).sum())
    stats = {
        "filename": paths[0].name,
        "software": "פייסקול",
        "cancelled_rows": cancelled,
        "paren_invoice_normalized": paren_normalized,
    }
    return df, "פייסקול", stats


def _for_excel(df: pd.DataFrame) -> pd.DataFrame:
    """Return df with internal/computed columns removed.

    Only strip the raw internal names (report_code, ichud, etc.) — NOT their
    Hebrew renamed equivalents like קוד דיווח, which are display columns that
    belong in the Excel output.
    """
    keep = [c for c in df.columns if c not in _STRIP_COLS]
    return df[keep]


def _build_display_records(
    df: pd.DataFrame,
    col_map: list[tuple],
) -> list[dict]:
    """Build JSON records with unified display column names and value transforms."""
    result: dict[str, list] = {}
    for src_col, display_col, transform in col_map:
        if src_col in df.columns:
            series = df[src_col].fillna("").astype(str).replace("nan", "")
            if transform:
                series = series.apply(lambda v: transform(v) if v else "")
        else:
            series = pd.Series([""] * len(df))
        result[display_col] = series.tolist()

    # Transpose to list of dicts
    keys = [display_col for _, display_col, _ in col_map]
    return [
        {k: result[k][i] for k in keys}
        for i in range(len(df))
    ]
